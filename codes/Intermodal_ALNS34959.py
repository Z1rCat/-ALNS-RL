import pandas as pd
import numpy as np
from collections import Counter
import copy
import re
import random
import matplotlib.pyplot as plt
import timeit
import time, datetime, functools
import math
from pandas.util import hash_pandas_object
from itertools import groupby
from itertools import compress
from pathlib import Path
from time import process_time
import networkx as nx
import shutil
from sympy.solvers import solve
from sympy import Symbol, exp
import sys
from collections import defaultdict
# This import registers the 3D projection, but is otherwise unused.
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401 unused import
import os
import wrapt
from line_profiler import LineProfiler
from numba import jit
import os.path

# kernprof -l Intermodal_ALNS_new_operators_20201005.py
# python -m line_profiler Intermodal_ALNS_new_operators_20201005.py.lprof
# import cProfile
# cProfile.run('foo()')
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor
import pickle
import hashlib
import orjson
import json
import skfuzzy as fuzz
from skfuzzy import control as ctrl
#may cause bug list:
#1. request_flow_t is not updated when finding best solution by hash table
import fuzzy_HP
import psutil
import openpyxl
from shutil import copyfile
from openpyxl import load_workbook
from emission_models import Trip
import warnings
import dynamic_RL34959
import Dynamic_ALNS_RL34959
# warnings.filterwarnings("error") #need to un-comment this if want to capture warning
def print_k_that_serve_routes():
    for k in routes.keys():
        if len(routes[k][0]) > 2:
            print(k, routes[k])


def check_T_k_record_and_R():
    if len(T_k_record) != len(R):
        print('error find')

def check_routes_and_CP_R_pool_repeat():
    #need modifaction based on exps, etc
    CP_R_pool = pd.read_excel(CP_R_pool_path, 'CP_R_pool', index_col=0)
    for carrier in range(1, 4):
        print('carrier', carrier)
        path = "/data/yimeng/Figures/experiment12739/percentage0parallel_number" + str(
            carrier) + "dynamic0/best_routespercentage0parallel_number" + str(carrier) + "dynamic0_12739.xlsx"
        routes_local = pd.read_excel(path, None, index_col=0)
        R_pool = copy.deepcopy(np.array(CP_R_pool))
        served_R_number = 0
        all_served_r = []
        for k in routes_local.keys():
            if len(routes_local[k].columns) > 2:
                for m in routes_local[k].columns:
                    if new_getLetters(m) == 'pickup':
                        served_R_number = served_R_number + 1
                        # if get_all == 1:
                        request_number = get_numbers(m)
                        if request_number in all_served_r:
                            print('caught it')
                        else:
                            all_served_r.append(request_number)
        for i in all_served_r:
            if i in R_pool[:, 7]:
                print(i, 'caught')


def get_time_of_wenjing():
    for request_number in [5, 10, 20, 30, 50, 100, 200, 400, 700, 1300, 1600]:
        routes_file = pd.ExcelFile('/data/yimeng/Case study/vs. Wenjing/Wenjing matching/routes/' + str(request_number) + 'r_result_correct_right.xlsx')
        routes = pd.read_excel(routes_file)
        overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost, overall_transshipment_cost, overall_un_load_cost, overall_emission_cost, overall_storage_cost, overall_delay_penalty, overall_number_transshipment, overall_average_speed, overall_average_time_ratio, overall_emission_transshipment = overall_obj(routes)
        obj_record = pd.DataFrame(index=[0],
                                  columns=['overall_distance', 'overall_cost', 'overall_time', 'overall_profit',
                                           'overall_emission', 'served_requests', 'overall_request_cost',
                                           'overall_vehicle_cost', 'overall_wait_cost', 'overall_transshipment_cost',
                                           'overall_un_load_cost', 'overall_emission_cost', 'overall_storage_cost',
                                           'overall_delay_penalty'])
        obj_record = obj_record.append(pd.series([overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost, overall_transshipment_cost, overall_un_load_cost, overall_emission_cost, overall_storage_cost, overall_delay_penalty],index = obj_record.columns), ignore_index=True)
        with pd.ExcelWriter('/data/yimeng/Case study/vs. Wenjing/Wenjing matching/routes/' + str(request_number) + 'objs.xlsx') as writer:  # doctest: +SKIP
            obj_record.to_excel(writer, sheet_name='obj_record')

def get_all_digits_of_a_number(n,all_digits):

    if n < 10:
        all_digits.append(n)
    else:
        get_all_digits_of_a_number(n // 10, all_digits)
        all_digits.append(n % 10)
    return all_digits

def has_handle(fpath):
    for proc in psutil.process_iter():
        try:
            for item in proc.open_files():
                if fpath == item.path:
                    return True
        except Exception:
            pass

    return False

#change the values' type in a list to int
def foo(l, dtype=int):
    return list(map(dtype, l))

def find_unchecked_r_preference(r_list):
    # if request_flow_t[index_r][5] - request_flow_t[index_r][1] > 10:
    #     print(2)
    for r in r_list:
        used_k = find_used_k(r)
        cost = 0
        emissions = 0
        for k in used_k:
            if k != -1:
                cost = cost + objective_value_i(r, k, routes[k])[0]
                emissions = emissions + objective_value_i(r, k, routes[k])[2]
        

def check_relevant_try_not_in_routes():
    global relevant_try
    for k in relevant_try.keys():
        if isinstance(relevant_try[k][0] == routes[k], bool):
            print(8)


def check_repeat_r_in_R_pool():
    global R_pool
    # served=check_served_R()
    served, all_served_r=check_served_R(0,-1,1)
    for i in all_served_r:
        if i in R_pool[:,7]:
            print('caught')
    if len(R_pool)+served>len(R):
        print('sf')
    if len(R_pool)+served<len(R):
        print('R pool lost r')
    inr=[]
    for x in range(len(R_pool[:, 7])):
        r=R_pool[x, 7]
        if r in inr:
            print(R_pool[x])
        else:
            inr.append(r)


def check_capacity(routes):
    for k in routes:
        if isinstance(capacity_constraints(has_end_depot, K, R, k, routes[k]),bool):
            print('fwfe')

def check_satisfy_constraints():
    for v in range(len(K)):

        route = routes[v]
        if len(route[0]) <= 2:
            continue
        request_number = get_numbers(route[4,1])
        if isinstance(
            satisfy_constraints(routes, has_end_depot, R, v, route, fixed_vehicles_percentage, K,
                                no_route_barge, no_route_truck, request_number, 1), bool):
            satisfied = 0
            break

def get_numbers(col):
    return int(''.join(filter(str.isdigit, col)))


def get_mode_share(index_in_all_routes):
    number_used_vehicles, barge_seved_r_number, train_seved_r_number, truck_seved_r_number = 0, 0, 0, 0
    for key, value in all_routes[index_in_all_routes].items():
        seved_r_number = len(value[4]) / 2 - 1
        if K[key, 5] == 1:
            barge_seved_r_number = barge_seved_r_number + seved_r_number
        if K[key, 5] == 2:
            train_seved_r_number = train_seved_r_number + seved_r_number
        if K[key, 5] == 3:
            truck_seved_r_number = truck_seved_r_number + seved_r_number
        if len(value[4]) > 2:
            number_used_vehicles = number_used_vehicles + 1
    all_number = truck_seved_r_number + train_seved_r_number + barge_seved_r_number
    if all_number == 0:
        barge_seved_r_portion, train_seved_r_portion, truck_seved_r_portion = 0,0,0
    else:
        barge_seved_r_portion = barge_seved_r_number / all_number * 100
        train_seved_r_portion = train_seved_r_number / all_number * 100
        truck_seved_r_portion = truck_seved_r_number / all_number * 100
    return number_used_vehicles, all_number, barge_seved_r_portion, train_seved_r_portion, truck_seved_r_portion


def find_r_served_by_k(new_try):
    if has_end_depot == 1:
        try:
            end = len(new_try[0])-1
        except:
            sys.exit(-2)
    else:
        end = len(new_try[0])
    r_served_by_k = []
    for m in range(1,end):
        col = new_try[4,m]
        request_number = get_numbers(col)
        r_served_by_k.append(request_number)
    return set(r_served_by_k)

def get_all_served_r():
    global routes
    served_r = []
    for k in routes.keys():
        served_r.extend(list(find_r_served_by_k(routes[k])))
    return served_r

def find_used_k(r, current_k = -1):
    global routes
    used_k = []
    operations = []
    if current_k != -1:
        used_k.append(current_k)
    for k in routes.keys():
       if len(routes[k][0])>2:
           for i in routes[k][4][1:-1]:
                if r == int(''.join(filter(str.isdigit, i))):
                    used_k.append(k)
                    operations.append(new_getLetters(i))
                    break

    if len(used_k) == 0:
        return [-1,-1,-1]
    #remove duplicates in list and keep order
    # b = []
    # for i in used_k:
    #     # Add to the new list
    #     # only if not present
    #     if i not in b:
    #         b.append(i)
    # used_k = b
    z = []
    for x in range(len(operations)):
        if operations[x] == 'pickup':
            z.append(used_k[x])
            break
    for x in range(len(operations)):
        if operations[x] == 'Tp':
            z.append(used_k[x])
            break
    for x in range(len(operations)):
        if operations[x] == 'secondTp':
            z.append(used_k[x])
            break
    used_k = z
    if len(used_k) == 1:
        used_k = [used_k[0],-1,-1]
    else:
        if len(used_k) == 2:
            used_k = [used_k[0], used_k[1], -1]
    return used_k


def find_used_k_2(r, current_k=-1):
    global routes
    used_k = []
    operations = []
    if current_k != -1:
        used_k.append(current_k)
    for k in routes.keys():
        if len(routes[k][0]) > 2:
            for i in routes[k][4][1:-1]:
                if r == int(''.join(filter(str.isdigit, i))):
                    used_k.append(k)
                    operations.append(new_getLetters(i))

    if len(used_k) == 0:
        return [-1, -1, -1]
    # remove duplicates in list and keep order
    # b = []
    # for i in used_k:
    #     # Add to the new list
    #     # only if not present
    #     if i not in b:
    #         b.append(i)
    # used_k = b
    z = []
    for x in range(len(operations)):
        if operations[x] == 'pickup':
            z.append(used_k[x])

    for x in range(len(operations)):
        if operations[x] == 'Tp':
            z.append(used_k[x])

    for x in range(len(operations)):
        if operations[x] == 'secondTp':
            z.append(used_k[x])

    used_k = z
    if len(used_k) == 1:
        used_k = [used_k[0], -1, -1]
    else:
        if len(used_k) == 2:
            used_k = [used_k[0], used_k[1], -1]
    return used_k


# import cPickle
def save_obj(objs, names, what='pickle'):
    time_s = timeit.default_timer()
    for o in range(len(objs)):
        if names[o] == 'request_number_in_R':
            with open('obj/' + str(exp_number - 1) + names[o] + '.pkl', 'wb') as f:
                if what == 'pickle':
                    pickle.dump(objs[o], f, pickle.HIGHEST_PROTOCOL)
                else:
                    ujson.dump(objs[o], f)
        else:
            with open('obj/' + str(exp_number - 1) + names[o] + str(parallel_number) + '.pkl', 'wb') as f:
                if what == 'pickle':
                    pickle.dump(objs[o], f, pickle.HIGHEST_PROTOCOL)
                else:
                    ujson.dump(objs[o], f)
    print('save', timeit.default_timer() - time_s)


def my_deepcopy(routes_new):
    return pickle.loads(pickle.dumps(routes_new))

def parallel_load(name):
    with open('obj/' + str(exp_number - 1) + name + str(parallel_number) + '.pkl', 'rb') as f:
        return pickle.load(f)


def load_obj(names):
    time_s = timeit.default_timer()
    a = 1
    while a == 1:
        # try:
        if len(names) == 1:
            if names == ['request_number_in_R']:
                with open('obj/' + str(exp_number - 1) + names[0] + '.pkl', 'rb') as f:
                    return [pickle.load(f)]
        objs = []
        with ThreadPoolExecutor() as e:
            results = e.map(parallel_load, names)
        for result in results:
            objs.append(result)
        print('load_success', timeit.default_timer() - time_s)
        return objs
        # except:
        #     print('load_fail', timeit.default_timer() - time_s)
        #     continue


if 'builtins' not in dir() or not hasattr(builtins, 'profile'):
    import builtins

def profile(func):
    def inner(*args, **kwargs):
        return func(*args, **kwargs)

    return inner


builtins.__dict__['profile'] = profile


def profile():

    lp = LineProfiler()

    @wrapt.decorator
    def wrapper(func, instance, args, kwargs):
        # global lp
        lp_wrapper = lp(func)
        res = lp_wrapper(*args, **kwargs)
        lp.print_stats()
        # lp.dump_stats(path + current_save + '/better_obj_record' + current_save + '.txt')
        return res

    return wrapper


# @profile()
def time_me(info="used"):
    global functions_time
    num = 0
    overall_function_time = 0

    def _time_me(func):
        global functions_time
        nonlocal overall_function_time

        @functools.wraps(func)
        def _wrapper(*args, **kwargs):
            global functions_time
            nonlocal num, overall_function_time
            num += 1

            if sys.version[0] == "3":
                start = time.perf_counter()
            else:
                start = time.clock()
            if sys.version[0] == "3":
                end = time.perf_counter()
            else:
                end = time.clock()

            str(datetime.timedelta(seconds=end - start))

            if func.__name__ not in functions_time:
                functions_time[func.__name__] = 0
            functions_time[func.__name__][0] = num
            print("%s This function is called %s times, %s" % (func.__name__, num, info))

            start = timeit.default_timer()
            func_output = func(*args, **kwargs)
            end = timeit.default_timer()
            overall_function_time = overall_function_time + end - start
            if func.__name__ not in functions_time:
                functions_time[func.__name__] = 0
            functions_time[func.__name__][1] = float(overall_function_time)
            print("time and overall time are %s %s" % (end - start, overall_function_time))
            return func_output

        # logging.info("%s %s %s\n"%(func.__name__, info, str(datetime.timedelta(seconds = end - start))))
        print("%s This function is called %s times, %s %s" % (func.__name__, num, info, str(datetime.timedelta(seconds=end - start))))
        return _wrapper

    return _time_me


# @profile()
def set_fun(func):
    num = 0

    def call_fun(*args, **kwargs):
        nonlocal num
        start = timeit.default_timer()
        num += 1
        func(*args, **kwargs)
        print(func.__name__)
        end = timeit.default_timer()
        longtime = end - start
        print("This function is called %s times，this call spend：%s" % (num, longtime))

    return call_fun


# delete duplicate elements while reserve order

# @time_me()
# @profile()
def unique(sequence):
    seen = set()
    return [x for x in sequence if not (x in seen or seen.add(x))]


# @profile()
# @time_me()
def hasNumbers(inputString):
    return inputString[0].isdigit()


# @profile()
# @time_me()
def getLetters(string):
    ## initializing a new string to apppend only alphabets
    only_alpha = ""

    ## looping through the string to find out alphabets
    for char in string:

        ## ord(chr) returns the ascii value
        ## CHECKING FOR UPPER CASE
        if ord(char) >= 65 and ord(char) <= 90:
            only_alpha += char
        ## checking for lower case
        elif ord(char) >= 97 and ord(char) <= 122:
            only_alpha += char
    return only_alpha

def new_getLetters(s):
    return ''.join([i for i in s if not i.isdigit()])


# @profile()
# @time_me()
def value_in_df_output_index(value, df):
    index_list = []
    for column in df.columns:
        if True in (df[column] == value):
            index_list.extend(df.index[df[column] == value].tolist())
    index_list = list(dict.fromkeys(index_list))
    return index_list


# @profile()
# @time_me()
def assign_time(k, route, inserted_r,insert_position1):
    global relevant_request_position_number, check_start_position
    relevant_request_position_number = {}
    check_start_position = insert_position1

    bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k, route, inserted_r)
    return bool_or_route


# @profile()
# @time_me()
def get_routes_tuple(routes):
    routes_list = []
    for k in routes.keys():
        if len(routes[k][4]) > 2:
            # must has ,k, otherwise top_hash will cause wrong matches
            routes_list.extend([df_tuple(routes[k], k), k])
    return tuple(routes_list)


# @profile()
# @time_me()
def load_emission_cost(k1, d, i):
    index_r = list(R[:,7]).index(i)
    if K[k1, 5] == 1 or K[k1, 5] == 2:
        load_unload_cost = R[index_r, 6] * 18 * 2
        if K[k1, 5] == 1:
            emission_cost = d * K[k1, 4] * R[index_r, 6] / 1000 * 8
        else:
            emission_cost = d * K[k1, 4] * R[index_r, 6] / 1000 * 8
    else:
        load_unload_cost = R[index_r, 6] * 3 * 2
        emission_cost = d * K[k1, 4] * R[index_r, 6] / 1000 * 8
    return load_unload_cost, emission_cost

def get_r_basic_cost_unit(i,k,n1,n2):
    index_r = list(R[:, 7]).index(i)
    d = D[k][n2,n1]
    request_cost = (K[k, 3] * d + K[k, 2] * d / K[k, 1]) * R[index_r, 6]
    load_unload_cost, emission_cost = load_emission_cost(k, d, i)
    return w1 * (request_cost + load_unload_cost) + w3 * (emission_cost)
# @profile()
# @time_me()
def get_r_basic_cost(p, d, i, k1, k2=-1, T=-1, k3 = -1, T2 = -1):

    if k1 == -1:
        return 9999999999999999999999
    if T != -1:
        if T2 == -1:
            r_basic_cost = get_r_basic_cost_unit(i,k1,p,T) + get_r_basic_cost_unit(i,k2,T,d)
        else:
            #3k
            r_basic_cost = get_r_basic_cost_unit(i,k1,p,T) + get_r_basic_cost_unit(i,k2,T,T2) + get_r_basic_cost_unit(i,k3,T2,d)
    else:
        r_basic_cost = get_r_basic_cost_unit(i, k1, p, d)
    return r_basic_cost


# @profile()
# @time_me()
def check_served_R(final = 0, routes_input = -1, get_all = 0):
    global routes
    if final == 0:
        routes_local = my_deepcopy(routes)
    else:
        routes_local = my_deepcopy(routes_input)
    served_R_number = 0
    all_served_r = []
    for k in routes_local.keys():
        if len(routes_local[k][4]) > 2:
            for m in routes_local[k][4]:
                if new_getLetters(m) == 'pickup':
                    served_R_number = served_R_number + 1
                    if get_all == 1:
                        request_number = get_numbers(m)
                        if request_number in all_served_r:
                            print(request_number, 'caught it')
                        else:
                            all_served_r.append(request_number)
    # if served_R_number < len(R):
    #     print('wrong')
    # if served_R_number > len(R):
    #     print('wrong_exceed')
    # print(served_R_number)
    if get_all == 1:
        return served_R_number, all_served_r
    else:
        return served_R_number


def lost_r():
    served_R_number = check_served_R()
    if served_R_number + len(R_pool) != len(R):
        return 'lost'


def create_routes():
    routes = {}
    for k in range(len(K)):
        routes[k] = np.array(np.empty(shape=(5, 0)), dtype='object')
        routes[k] = np.insert(routes[k], 0, [o[k,0], o[k,0], o[k,0], o[k,0], 'begin_depot'], axis=1)
        # no end depot
        if has_end_depot == 1:
            routes[k] = np.insert(routes[k], 1, [o[k,1], o[k,1], o[k,1], o[k,1], 'end_depot'], axis=1)
    return routes

def revert_names(type='str'):
    # return {'Delta': 1, 'Euromax': 2, 'HOME': 3, 'Moerdijk': 4, 'Venlo': 5, 'Duisburg': 6,
    #                          'Willebroek': 7, 'Neuss': 8, 'Dortmund': 9, 'Nuremberg': 10}
    if different_companies == 1:
        if parallel_number == 1:
            if type == 'str':
                return {'Delta': 0, 'Euromax': 1, 'HOME': 2, 'Moerdijk': 3, 'Venlo': 4, 'Duisburg': 5,
                        'Willebroek': 6, 'Neuss': 7, 'Dortmund': 8, 'Nuremberg': 9}
            else:
                return {1: 0, 2: 1, 3: 2, 4: 3, 5: 4, 6: 5, 7: 6, 8: 7, 9: 8, 10: 9}
        else:
            # []
            return {'Delta': 0, 'Euromax': 1, 'HOME': 2, 'Moerdijk': 3, 'Venlo': 4, 'Duisburg': 5,
                    'Willebroek': 6, 'Neuss': 7, 'Dortmund': 8, 'Nuremberg': 9, 'Antwerp': 10, 'Emmelsum': 11, 'Bonn': 12, 'Koblenz': 13, 'Gustavsburg': 14, 'Frankfurt West': 15, 'Frankfurt Ost': 16,
             'Worms': 17, 'Mannheim': 18, 'Ludwigshafen': 19, 'Worth': 20, 'Karlsruhe': 21, 'Strasbourg': 22, 'Ottmarsheim': 23, 'Weil am Rhein': 24,
             'Basel': 25, 'Lauterbourg': 26, 'Kehl': 27, 'Andernach': 28, 'APM': 29, 'RWG': 30,	'Emmerich': 31,	'Bruay-sur-lEscaut': 32}

    else:
        if Demir == 1:
            # return {'Budapest Port': 1, 'Budapest BILK': 2, 'Wien Freudenau': 3, 'Wien NWB': 4, 'Linz': 5,
            #         'Regensburg': 6, 'Munich': 7, 'Wels': 8, 'Praha Zizkov': 9, 'Salzburg': 10, 'Villach': 11,
            #         'Trieste':12, 'Koper':13, 'Nurnberg':14, 'Duisburg': 15, 'Dunajska Streda': 16,
            #         'Ceska Trebova': 17, 'Ostrava': 18, 'Zlin': 19, 'Plzen': 20,
            #         'Vienna Port': 21, 'Vienna Rail': 22, 'Prague': 23}
            return {'Budapest Port': 0, 'Vienna Port': 1, 'Linz': 2, 'Budapest BILK': 3, 'Vienna Rail': 4,
                    'Prague': 5, 'Munich': 6, 'Regensburg': 7, 'Wels': 8, 'Salzburg': 9}
        else:
            if type == 'str':
                return {'Delta': 0, 'Euromax': 1, 'HOME': 2, 'Moerdijk': 3, 'Venlo': 4, 'Duisburg': 5,
                        'Willebroek': 6, 'Neuss': 7, 'Dortmund': 8, 'Nuremberg': 9}
            else:
                return {1: 0, 2: 1, 3: 2, 4: 3, 5: 4, 6: 5, 7: 6, 8: 7, 9: 8, 10: 9}

# @profile()
# @time_me()
def read_data():
    global routes,not_initial_in_CP
    #    K = pd.read_excel(Data, 'K')
    #    o = pd.read_excel(Data, 'o')
    #    K = K.set_index('K')
    #    o = o.set_index('K')
    #    R = pd.read_excel(Data, 'R')

    # N_origin = pd.read_excel(Data, 'N')
    #    N = pd.read_excel(Data, 'N')
    #    T = pd.read_excel(Data, 'T')
    #    T_all = pd.read_excel(Data, 'T_all')
    D, D_origin_All = read_D('D_All', K)
    no_route_barge, no_route_truck = read_no_route()
    #        D[k][o[k,0]][o[k,1]] = 0

    # S = {}

    # R_pool = R.copy()

    R_pool_2v = {}
    R_pool_3v = {}

    for r_index in range(len(R)):
        R_i = tuple(zip(R[r_index], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
        if R_i in R_pool_2v.keys():
            pass
        else:
            R_pool_2v[R_i] = {}
            # R_change = R.copy()
            for T_change in T:
                request = R[r_index]
                if T_change == request[0] or T_change == request[1]:
                    continue
                first_segment_r, second_segment_r = segment_request(request, T_change)

                R_pool_2v[R_i][T_change] = pd.concat([first_segment_r, second_segment_r], axis=1).T
                # R_pool_2v[R_i][T_change].columns=['p','d','ap','bp','ad','bd','qr','r']
                # R_pool_2v[R_i][T_change].index = [0,1]
                R_pool_2v[R_i][T_change] = R_pool_2v[R_i][T_change].values
    if len(T) >= 2:
        for r_index in range(len(R)) :
            # danger this break should be removed if 2T is considered
            if two_T == 0:
                break
            R_i = tuple(zip(R[r_index], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
            if R_i in R_pool_3v.keys():
                pass
            else:
                R_pool_3v[R_i] = {}
                for T_change in T:
                    first_segment_r, original_second_segment_r = R_pool_2v[R_i][T_change][0], \
                                                                 R_pool_2v[R_i][T_change][1]
                    T_2 = T.copy()
                    for T_change2 in T_2:
                        if T_change2 == T_change:
                            continue
                        second_segment_r, third_segment_r = segment_request(original_second_segment_r,
                                                                            T_change2)
                        R_pool_3v[R_i][(T_change, T_change2)] = pd.DataFrame(
                            [first_segment_r, second_segment_r, third_segment_r]).values
    if not (dynamic == 1 and dynamic_t > 0):
        if not_initial_in_CP == 0:
            routes = create_routes()
    elif VCP_coordination == 1:
        routes = -1
    return D, routes, R_pool_2v, R_pool_3v, no_route_barge, no_route_truck, D_origin_All


# @profile()
# @time_me()
def bundle():
    bundle_R = {}
    for index_r in range(len(R)) :
        key = tuple([R[index_r, 0], R[index_r, 1]])
        # R_r = np.append(R[index_r],r)
        if key not in bundle_R.keys():
            bundle_R[key] = R[index_r]
        else:
            bundle_R[key] = np.vstack([bundle_R[key], R[index_r]])
    return bundle_R


# insert_terminals includes 2 terminals for no T, 3 for 1T, 4 for 2T; positions conclude all positions for different k;
# @profile()
# @time_me()
def insert_bundle(i, key, number_T, used_k, insert_terminals):
    # bundle_this_insert = pd.DataFrame(columns=['p','d','ap','bp','ad','bd','qr','r'])
    # bundle_this_insert = np.array(np.empty(shape=(0, 7)), dtype='object')
    if len(np.shape(bundle_R[key])) > 1:
        bundle_this_insert = bundle_R[key][~(bundle_R[key][:, 7] == i)]
    else:
        insert_r = bundle_R[key][7]
        if insert_r == i:
            return
        else:
            if insert_r in R_pool[:, 7]:
                insert_a_r(0, insert_r, used_k, 0, 0, 'mark', 0, insert_terminals, 0, 1)
                return
    #find_unchecked_r_preference([6,45])
    not_in = []
    for r in bundle_this_insert[:, 7]:
        if r not in R_pool[:, 7]:
            not_in.append(list(bundle_this_insert[:, 7]).index(r))
    bundle_this_insert = np.delete(bundle_this_insert, not_in, axis=0)

    if bundle_this_insert.size != 0:
        #find_unchecked_r_preference([6,45])
        # sort r depending on load
        bundle_this_insert = bundle_this_insert[bundle_this_insert[:, 6].argsort()[::-1]]
        random_bundle = 0
        if random_bundle == 1:
            # distribution
            p = []
            p_i = 1
            for p_i_i in range(0, len(bundle_this_insert)):
                #$\varsigma$ = 1.1
                p_i = p_i / 1.1
                p.append(p_i)
            insert_number = random.choices(range(len(bundle_this_insert), 0, -1), p)[0]

            for r in range(insert_number):
                insert_r_index = random.choices(range(len(bundle_this_insert)), p)[0]
                insert_r = bundle_this_insert[insert_r_index, 7]
                bundle_this_insert = np.delete(bundle_this_insert, insert_r_index, axis=0)
                del p[insert_r_index]
                capacity_full = insert_a_r(0, insert_r, used_k, 0, 0, 'mark', 0, insert_terminals, 0, 1)[2]
                if capacity_full == 1:
                    break
            #find_unchecked_r_preference([6,45])
        else:
            #find_unchecked_r_preference([6,45])
            for insert_r_index in range(len(bundle_this_insert)):

                insert_r = bundle_this_insert[insert_r_index, 7]
                # bundle_this_insert = np.delete(bundle_this_insert, insert_r_index, axis=0)

                capacity_full = insert_a_r(0, insert_r, used_k, 0, 0, 'mark', 0, insert_terminals, 0, 1)[2]
                if capacity_full == 1:
                    break

# @profile()
# @time_me()
def insert_bundle_pre(i, key, number_T, best_T, top_key, k):
    index_r = list(R[:, 7]).index(i)
    if isinstance(best_T, (int, np.integer)):
        best_T = [best_T]
    #find_unchecked_r_preference([6,45])
    insert_bundle_or_not = 1


    # used_k = pd.DataFrame(index=[i], columns=['k1', 'k2', 'k3'])
    # used_k columns=['k1', 'k2', 'k3',i]
    used_k = np.array(np.empty(shape=(1, 4)), dtype='object')
    used_k[:] = -1
    used_k[0] = -1,-1,-1,i
    
    index = 0
    if number_T != -1:
        if number_T > 0:

            if best_T[0] != -1:
                if number_T == 1:

                    insert_terminals = [key[0], best_T[0], key[1]]

                else:

                    insert_terminals = [key[0], best_T[0], best_T[1], key[1]]

                if top_key in hash_top.keys():
                    print('top_bundle',len(R_pool))

                    used_k[index,0] = hash_top[top_key]['k'][0]
                    used_k[index,1] = hash_top[top_key]['k'][1]
                    if len(hash_top[top_key]['k']) == 3:
                        used_k[index,2] = hash_top[top_key]['k'][2]
                else:
                    insert_bundle_or_not = 0
            else:
                insert_bundle_or_not = 0

        else:
            if k != -1:
                used_k[index,0] = k
                insert_terminals = [R[index_r, 0], R[index_r, 1]]
            else:
                insert_bundle_or_not = 0
        if insert_bundle_or_not == 1:
            r_cost = get_r_cost_in_all_routes(i)[0]

            if used_k[0,1]==-1:
                r_basic_cost = get_r_basic_cost(R[index_r, 0], R[index_r, 1], i, used_k[0,0])
            else:
                #danger 2T is not considered
                if used_k[0,2]==-1:
                    r_basic_cost = get_r_basic_cost(R[index_r, 0], R[index_r, 1], i, used_k[0,0], used_k[0,1], best_T[0])
            try:
                if r_cost < r_basic_cost + R[index_r, 6] * 2 * c_storage - 0.1:
                    #find_unchecked_r_preference([6,45])
                    insert_bundle(i, key, number_T, used_k, insert_terminals)
                    #find_unchecked_r_preference([6,45])
            except:
                print('ew')

# @profile()
# @time_me()
def segment_request(request, T_change):
    request_change = pd.DataFrame(request.copy()).transpose()

    request_change.at[0, 1] = T_change
    first_segment_r = request_change.loc[0]
    request_change = pd.DataFrame(request.copy()).transpose()
    request_change.at[0, 0] = T_change
    second_segment_r = request_change.loc[0]

    return first_segment_r, second_segment_r


# @profile()
# @time_me()
def get_fix_k_0_ap(k, fixed_vehicles_percentage, Fixed):
    if k in fixed_vehicles_percentage:
        fix_k_ap = str(Fixed[k][:,1])
        fix_k_bp = str(Fixed[k][:,2])
    else:
        fix_k_ap = -1
        fix_k_bp = -1

    return fix_k_ap, fix_k_bp


# @profile()
# @time_me()
def get_key_1k(R_i_1, original_route_no_columns_1, k_1, fixed_vehicles_percentage, Fixed, K):
    fix_k_ap, fix_k_bp = get_fix_k_0_ap(k_1, fixed_vehicles_percentage, Fixed)
    key = (R_i_1, original_route_no_columns_1, K[k_1, 0], K[k_1, 1], fix_k_ap, fix_k_bp)
    return key


# @profile()
# @time_me()
def remove_a_request(request_number, routes_local, R_pool_local):
    global check_start_position, R_change_dynamic
    #check_satisfy_constraints()
    #find_unchecked_r_preference([6,45])

    v_has_r = [-1, -1, -1]
    used_T = [-1, -1]
    # if dynamic == 1 and dynamic_t > 0:
    #     #if not changed r, the other r's routes can't be changed, but schedules can be changed by inserting changed r
    #     if request_number not in R_change_dynamic[:,7]:
    #         return routes_local, R_pool_local, v_has_r, used_T
    routes_save = my_deepcopy(routes_local)
    request_number = int(request_number)

    # remove the request from all vehicles
    satisfied = 1
    #check whether I can delete request_number
    for v in range(len(K)):
        if len(routes[v][4]) <= 2:
            continue
        # both routes and routes_local will change, but if constraints are not be satisfied, the routes_save will be returned
        new_try = copy.copy(routes_local[v])
        new_try_copy = my_deepcopy(new_try)
        droped = 0
        check_number = 0
        for col in new_try[4]:

            request_number_col = ''.join(filter(str.isdigit, col))
            if str(request_number) == request_number_col:
                if check_number == 0:
                    check_start_position = list(new_try[4]).index(col)
                    check_number = 1
                droped = 1
                new_try = np.delete(new_try, list(new_try[4]).index(col), 1)
        #if after iteration in CP only consider eco-labels, just delete it and not check constraints, because the I use average load factor during iteration, it will cause many infeasible solutions when consider real load factor after iteration, so just delete one by one
        # danger, here CP can only use fixed vehicles, because I didn't consider the time constraints after removing under flexible vehicles
        if not (during_iteration == 0 and percentage == 0 and only_eco_label == 1 and heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0 and emission_preference_constraints_after_iteration == 1):
            if droped == 1:

                if isinstance(satisfy_constraints(routes, has_end_depot, R, v, new_try, fixed_vehicles_percentage, K,
                                                  no_route_barge, no_route_truck,request_number,1)[0], bool):
                    satisfied = 0
                    break
                if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
                    #I should check all r except request_number (which is removed) in the k and also relevant k
                    preference_final_ok_or1 = preference_relevant(v, new_try, request_number, 1, new_try_copy)
                    if preference_final_ok_or1 == 0:
                        satisfied = 0
                        break
    if satisfied == 0 and (dynamic_t == 0 or stochastic == 0):
        #when there is no stochastic parameter or initial solution, not satisfy constraints, do nothing
        routes_local = my_deepcopy(routes_save)
    else:
        #delete request_number
        for v in range(len(K)):
            new_try_copy = my_deepcopy(routes_local[v])
            droped = 0
            for col in routes_local[v][4]:
                request_number_col = ''.join(filter(str.isdigit, col))
                if str(request_number) == request_number_col:

                    string = new_getLetters(col)
                    if string == 'pickup':
                        v_has_r[0] = v
                    else:
                        if  string == 'Tp':
                            v_has_r[1] = v
                            used_T[0] = routes_local[v][0][list(routes_local[v][4]).index(col)]
                        else:
                            if string == 'secondTp':
                                v_has_r[2] = v
                                used_T[1] = routes_local[v][0][list(routes_local[v][4]).index(col)]
                    droped = 1
                    routes_local[v] = np.delete(routes_local[v], list(routes_local[v][4]).index(col), 1)
            if not (
                    during_iteration == 0 and percentage == 0 and only_eco_label == 1 and heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0 and emission_preference_constraints_after_iteration == 1):
                if droped == 1:

                    #only check the vehicle which serve the request_number
                    if isinstance(
                            satisfy_constraints(routes, has_end_depot, R, v, routes_local[v], fixed_vehicles_percentage, K,
                                                no_route_barge, no_route_truck, request_number, 1)[0], bool):
                        satisfied = 0
                        break
                    if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
                        # I should check all r except request_number (which is removed) in the k and also relevant k
                        preference_final_ok_or1 = preference_relevant(v, new_try_copy, request_number)
                        if preference_final_ok_or1 == 0:
                            satisfied = 0
                            break

        if satisfied == 0 and (dynamic_t == 0 or stochastic == 0):
            routes_local = my_deepcopy(routes_save)
            v_has_r = [-1, -1, -1]
            used_T = [-1, -1]
        else:
            R_pool_local = np.vstack([R_pool_local, R[(R[:, 7] == request_number)]])

    #find_unchecked_r_preference([6,45])
    return routes_local, R_pool_local, v_has_r, used_T


# @profile()
# @time_me()
def length_a_route(k, res):
    if len(res) >= 2:
        res2 = list(zip(res, res[1:] + res[:1]))
        del res2[-1]
        length = 0
        for pair in res2:
            length = length + D[k][int(pair[1]),int(pair[0])]
    else:
        length = 0
    return length


# @profile()
# @time_me()
##@jit
def route_no_columns(route):
    # for df which contains letters, the df.values.tobytes() results in different processors are different, try simple df with only numbers, results are same
    # for df which only contains numbers, the hash(df.values.tobytes()) results in different processors are different
    # str(route.tolist()) will lost requests (found in 5r instance), no matter multiple processors or not, don't know why
    ########version1#######
    # for k, v in hash_table_route_no_columns.items():
    #     if v.equals(route):
    #         return k
    # aa = list(hash_pandas_object(route))
    # aa.append(tuple(route[4]))
    # route_hash = tuple(aa)
    # hash_table_route_no_columns[route_hash] = route
    ########version1#######
    # if parallel == 1:
        ########version2#######
        # 20201111 if no hash_df_table then tuple(hash_pandas_object(route)) tahn value.equals(route)
        # for key, value in hash_df_table.items():
        #     if type(value) is not list:
        #         if value.equals(route):
        #             return key
        # route_hash = tuple(hash_pandas_object(route))
        # hash_df_table[route_hash] = route
        # return route_hash
        ########version2#######

        ########version4#######
        # 20201111 if no hash_df_table then tuple(hash_pandas_object(route)) tahn value.equals(route)
        # for key, value in hash_df_table.items():
        #     if type(value) is not list:
        #         if value.equals(route):
        #             return key
        # os.environ['PYTHONHASHSEED'] = '0'
        # return hash(str(route.tolist()))
        # return hashlib.md5(route.tobytes()).hexdigest()
        # for key, value in hash_df_table.items():
        #     if type(value) is not list:
        #         if value.equals(route):
        #             route_hash = hashlib.sha256(str(route.tolist()).encode('utf-8')).hexdigest()
        #             if route_hash != key:
        #                 print('findwrong')
        # print(route)
        # route_hash = hashlib.sha256(str(route.tolist()).encode('utf-8')).hexdigest()
        # print(route_hash)
        # hash_df_table[route_hash] = route
    return hashlib.sha256(str(route.tolist()).encode('utf-8')).hexdigest()
        # return str(route.tolist())
        # hash_df_table[route_hash] = route
        # return hash(tuple(tuple(i) for i in route))
        # return route.tobytes()
        #######version4#######


    # else:
    #     ########version3#######
    #     #this version will generate different value when using deepcopy
    #     return hash(route.tobytes())
        ########version3#######


# hash_table_top didn't consider insert the columns after find the results in hashtable, so it must totally same, so add columns in this function
# but I don't know why befter optimize codes, in route_no_columns, there is also original_route_no_columns1.append(tuple(route[4]))

# @profile()
# @time_me()
def df_tuple(route, k):
    ###############
    # for key, value in hash_df_table.items():
    #     if type(value) is list:
    #         if value[0].equals(route) and value[1] == k:
    #             return key
    # aa = list(hash_pandas_object(route))
    # aa.append(tuple(route[4]))
    # aa.append(k)
    # route_hash = tuple(aa)
    # hash_df_table[route_hash] = [route, k]
    ############
    route_hash_no_columns = route_no_columns(route)
    if isinstance(route, np.ndarray):
        return route_hash_no_columns
    else:
        route_hash = tuple([route_hash_no_columns, tuple(route[4])])
    return route_hash


# @profile()
# @time_me()
def update_r_best_obj_record(i, cost_inserted_request,v_has_r,used_T):
    global r_best_obj_record

    index_r = list(R[:, 7]).index(i)

    if pd.isnull(r_best_obj_record[index_r,0]):
        r_best_obj_record[index_r,0:3] = [cost_inserted_request,v_has_r,used_T]

        return 1
    else:
        if cost_inserted_request <= r_best_obj_record[index_r,0]:
            if cost_inserted_request < r_best_obj_record[index_r,0]:
                r_best_obj_record[index_r,0:3] = [cost_inserted_request,v_has_r,used_T]
            return 1
        else:
            return 0


# @profile()
# @time_me()
def update_r_best_obj_in_insertion(i, len1, old_overall_cost,v_has_r,used_T):
    global routes, R_pool
    len_final = len(R_pool[:, 7])
    if len_final < len1:
        overall_cost = overall_obj(routes)[1]
        cost_inserted_request = overall_cost - old_overall_cost
        update_r_best_obj_record(i, cost_inserted_request,v_has_r,used_T)


# @profile()
# @time_me()
def get_r_cost_in_all_routes(request_number, history_removal_mark=0, r_hasbeen_caculated=[]):
    global routes, R_pool
    routes_local = my_deepcopy(routes)
    # R_pool_local = copy.copy(R_pool)
    routes_after_removed, R_pool_after_removed, v_has_r, used_T = remove_a_request(request_number,
                                                                           routes_local,
                                                                           R_pool)
    old_cost = 0
    new_cost = 0
    if v_has_r[0] != -1:
        for j in v_has_r:
            if j != -1:
                old_cost = old_cost + objective_value_k(j, routes[j])[0]
                new_cost = new_cost + objective_value_k(j, routes_after_removed[j])[0]
        r_cost_in_all_routes = old_cost - new_cost
        if history_removal_mark == 0:
            update_r_best_obj_record(request_number, r_cost_in_all_routes, v_has_r, used_T)
        r_hasbeen_caculated.append(request_number)
    else:
        # if removing this r from rotues makes the routes infeasible, then mark it
        r_cost_in_all_routes = 1000000000
        r_hasbeen_caculated.append(request_number)
        # sys.exit('Error! I dont know why I check v_has_r empty before but I set a remind if it happens20201005')
    return r_cost_in_all_routes, r_hasbeen_caculated, routes_after_removed, R_pool_after_removed,v_has_r, used_T


# @profile()
# @time_me()
def get_remove_number(delete_node_or_not):
    K_serve_r = []
    for k in routes.keys():
        if len(routes[k][4]) > 2:
            if delete_node_or_not == 1 and percentage != 0:
                if k not in fixed_vehicles_percentage:
                    K_serve_r.append(k)
            else:
                K_serve_r.append(k)
    p = []
    p_i = 1
    for p_i_i in range(1, len(K_serve_r) + 1):
        p_i = p_i / 1.3
        p.append(p_i)
    probability_choose_k = []
    if len(K_serve_r) != 0:
        remove_number = random.choices(range(1, len(K_serve_r) + 1), p)[0]
        left_capacity = []
        for k in K_serve_r:
            if K[k, 5] == 3:
                left_capacity.append(0)
            else:
                if K[k, 5] == 1:
                    #increase the probability of remove barge
                    left_capacity.append(capacity_constraints(has_end_depot, K, R, k, routes[k], 0, 1)[1] + 50)
                else:
                    left_capacity.append(capacity_constraints(has_end_depot, K, R, k, routes[k], 0, 1)[1])
        sum_left_capacity = sum(left_capacity)
        if sum_left_capacity == 0:
            length = len(left_capacity)
            for pro_index in range(length):
                probability_choose_k.append(1 / length)
        else:
            for k in range(len(K_serve_r)):
                probability_choose_k.append(left_capacity[k] / sum_left_capacity)
    else:
        remove_number = 0

    return remove_number, K_serve_r, probability_choose_k

def adjust_probability(probability_choose_k):
    sum_pro = sum(probability_choose_k)
    if sum_pro == 0:
        length = len(probability_choose_k)
        for pro_index in range(len(probability_choose_k)):
            probability_choose_k[pro_index] = 1/length
    else:
        for pro_index in range(len(probability_choose_k)):
            probability_choose_k[pro_index] = probability_choose_k[pro_index]/sum_pro
    return probability_choose_k

# @profile()
# @time_me()
def delete_node():
    global routes, R_pool
    remove_number, K_serve_r, probability_choose_k = get_remove_number(1)
    for n in range(remove_number):
        # k = random.choice(K_serve_r)

        k = int(np.random.choice(K_serve_r, size=(1,), p=probability_choose_k))

        probability_choose_k.pop(K_serve_r.index(k))
        probability_choose_k = adjust_probability(probability_choose_k)
        K_serve_r.remove(k)
        transposed_route = routes[k][0].T
        res = [x[0] for x in groupby(transposed_route.tolist())]
        original_length = length_a_route(k, res)
        res_copy2 = copy.copy(res)
        del res_copy2[0]

        new_try = routes[k]
        for item in res:
            node = random.choice(res_copy2)
            res_copy2.remove(node)
            res_copy = copy.copy(res)
            res_copy.remove(node)
            new_length = length_a_route(k, res_copy)
            if new_length < original_length:
                for col in new_try[4]:
                    if hasNumbers(col):
                        if new_try[0, list(new_try[4]).index(col)] == node:
                            request_number = get_numbers(col)
                            # routes_local = my_deepcopy(routes)
                            # R_pool_local = copy.copy(R_pool)
                            routes, R_pool = remove_a_request(request_number, routes, R_pool)[0:2]
                            #lost_r()
                break
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    return routes, R_pool


# @profile()
# @time_me()
def random_removal():
    global routes, R_pool
    deleted_r = []
    K_serve_r = []
    for k in range(len(K)):
        if len(routes[k][4]) > 2:
            K_serve_r.append(k)
    if not K_serve_r:
        return routes, R_pool
    remove_k_number = max(int(0.4 * len(K_serve_r)), 1)
    for number in range(0, remove_k_number):
        k = random.choice(K_serve_r)
        K_serve_r.remove(k)
        # maybe there is a k only serve half r
        if len(routes[k][4]) <= 2:
            continue
        random_int = random.randrange(1, len(routes[k][4]) - 1, 1)
        request_string = routes[k][4, random_int]
        request_number = int(''.join(filter(str.isdigit, request_string)))
        if request_number not in deleted_r:
            # routes_local = my_deepcopy(routes)
            # R_pool_local = copy.copy(R_pool)
            routes, R_pool = remove_a_request(request_number, routes, R_pool)[0:2]
            #lost_r()
            deleted_r.append(request_number)
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    return routes, R_pool


# @profile()
# @time_me()
def clear_a_route():
    global routes, R_pool, o, bundle_R
    #    routes = {}
    #    R_pool = R.copy()
    remove_number, K_serve_r, probability_choose_k = get_remove_number(0)
    for n in range(remove_number):
        # k = random.choice(K_serve_r)
        k = int(np.random.choice(K_serve_r, size=(1,), p=probability_choose_k))

        probability_choose_k.pop(K_serve_r.index(k))
        probability_choose_k = adjust_probability(probability_choose_k)
        K_serve_r.remove(k)
        deleted_r = []
        for col in routes[k][4]:
            if hasNumbers(col):
                request_number = get_numbers(col)
                if request_number not in deleted_r:
                    # routes_local = my_deepcopy(routes)
                    # R_pool_local = copy.copy(R_pool)
                    routes, R_pool = remove_a_request(request_number, routes, R_pool)[0:2]
                    #lost_r()
                    #find_unchecked_r_preference([6,45])
                    deleted_r.append(request_number)
        #in the meantime remove some r has the same OD with k
        key=tuple([o[k,0], o[k, 1]])
        if key in bundle_R.keys():
            try_remove_r = bundle_R[key]
            delete_index=[]
            if try_remove_r.size > 17:
                for r_list_index in range(len(try_remove_r)):
                    r_list = try_remove_r[r_list_index]
                    if r_list[7] in R_pool[:,7]:
                        delete_index.append(r_list_index)
                try_remove_r = np.delete(try_remove_r, delete_index, axis=0)
            else:
                r_list = try_remove_r
                r_list_index = 0
                try:
                    if r_list[7] in R_pool[:, 7]:
                        delete_index.append(r_list_index)
                        try_remove_r = np.array(np.empty(shape=(8,0)))
                except:
                    print('ss')
            if try_remove_r.size > 0:
                for remove_r_number in range(max(1,int(0.1 * len(try_remove_r)))):
                    remove_r_list_index = random.choice(range(len(try_remove_r)))
                    if try_remove_r.size > 17:
                        routes, R_pool = remove_a_request(try_remove_r[remove_r_list_index][7], routes, R_pool)[0:2]
                        # find_unchecked_r_preference([6, 45])
                    else:
                        try:
                            routes, R_pool = remove_a_request(try_remove_r[7], routes, R_pool)[0:2]
                        except:
                            routes, R_pool = remove_a_request(try_remove_r[0,7], routes, R_pool)[0:2]
                    # find_unchecked_r_preference([6, 45])
                    #lost_r()
                    try_remove_r = np.delete(try_remove_r, remove_r_list_index, axis=0)
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    return routes, R_pool


# @profile()
# @time_me()
def remove_all():
    # routes = {}
    R_pool = R.copy()
    routes = create_routes()
    return routes, R_pool

# @profile()
# @time_me()
def worst_removal():
    global routes, R_pool
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    # print(len(R_pool))
    # cost_of_r = pd.DataFrame(index=R[:,7] , columns=['cost of r'])
    cost_of_r = np.array(np.empty(shape=(len(R),2)))
    cost_of_r[:] = np.NaN
    cost_of_r[:,1]=R[:,7]
    r_hasbeen_caculated = []
    for k in range(len(K)):
        if has_end_depot == 1:
            length = len(routes[k][4])
        else:
            length = len(routes[k][4]) + 1
        if length > 2:
            if has_end_depot == 0:
                length = length - 1
            for h in range(0, length - 1):
                
                hasNumbers(routes[k][4, h])
                
                if hasNumbers(routes[k][4, h]):
                    request_string = routes[k][4, h]
                    request_number = int(''.join(filter(str.isdigit, request_string)))
                    index_r = list(R[:, 7]).index(request_number)
                    if request_number not in r_hasbeen_caculated:
                        r_cost_in_all_routes, r_hasbeen_caculated = get_r_cost_in_all_routes(request_number, 0,
                                                                                             r_hasbeen_caculated)[0:2]
                        if r_cost_in_all_routes != 1000000000:
                            cost_of_r[index_r,0] = r_cost_in_all_routes / (
                                    R[index_r, 6] * D_origin_All[R[index_r, 0]][
                                R[index_r, 1]])
                        else:
                            cost_of_r[index_r, 0] = -100
    # cost_of_r.dropna(inplace=True)
    cost_of_r = cost_of_r[~(cost_of_r[:,0]==-100)]
    cost_of_r = cost_of_r[~np.isnan(cost_of_r[:, 0])]
    removal_number = min(int(0.3 * len(cost_of_r)), 50)
    # If I have a bad luck, all requests are removed, and no request was inserted in (may because the random_insert), then cost_of_r is empty (nan), just return the input
    if len(cost_of_r[:,0]) > 1:
        # cost_of_r['cost of r'] = pd.to_numeric(cost_of_r['cost of r'])
        for x in range(0, removal_number):
            # worst_r = cost_of_r['cost of r'].idxmax(skipna=True)
            worst_r_index = np.argmax(cost_of_r[:,0])
            worst_r = cost_of_r[worst_r_index,1]
            # routes_local2 = my_deepcopy(routes)
            # R_pool_local = copy.copy(R_pool)
            routes, R_pool = remove_a_request(worst_r, routes, R_pool)[0:2]
            #lost_r()
            cost_of_r = np.delete(cost_of_r,worst_r_index,axis=0)
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    return routes, R_pool


# @profile()
# @time_me()
def related_removal():
    global routes, R_pool
    r_in_routes = []
    for r in R[:,7] :
        if r not in R_pool[:, 7]:
            r_in_routes.extend([r])
    if not r_in_routes:
        return routes, R_pool
    random_r = random.choice(r_in_routes)
    index_random_r = list(R[:,7]).index(random_r)
    # routes_local = my_deepcopy(routes)
    # R_pool_local = copy.copy(R_pool)
    routes, R_pool = remove_a_request(random_r, routes, R_pool)[0:2]
    # find_unchecked_r_preference([6, 45])
    #lost_r()
    remove_number = int(max(0.05 * len(R_pool), 1))
    # calculate relateness
    other_r = copy.copy(r_in_routes)
    other_r.remove(random_r)
    # relateness_value = pd.DataFrame(columns=['relateness_value'], index=other_r)
    relateness_value = np.array(np.empty(shape=(len(other_r),2)))
    relateness_value[:,1] = other_r
    theta_distance, theta_time, theta_load, theta_vehicle = 0.25, 0.25, 0.25, 0.25
    index_random = list(ok_K_canpickr[len(K)]).index(random_r)
    for r in other_r:
        index = list(ok_K_canpickr[len(K)]).index(r)
        index_r = list(R[:,7]).index(r)
        K_canpick_both = [x for x in ok_K_canpickr[~np.isnan(ok_K_canpickr[:,index])][:-1,index] if x in ok_K_canpickr[~np.isnan(ok_K_canpickr[:,index_random])][:-1,index_random]]
        relateness_value[list(relateness_value[:,1]).index(r),0] = theta_distance * (
                D_origin_All[R[index_r, 0]][R[index_random_r, 0]] + D_origin_All[R[index_r, 1]][R[index_random_r, 1]]) / \
                                                  D_origin_All[R[index_random_r, 0]][R[index_random_r, 1]] + \
                                                  theta_time * (abs(
            request_flow_t[index_r,0] - request_flow_t[index_random_r,0]) + abs(
            request_flow_t[index_r,5] - request_flow_t[index_random_r,5])) / (
                                                          request_flow_t[index_random_r,5] -
                                                          request_flow_t[index_random_r,0]) + \
                                                  theta_load * abs(R[index_r, 6] - R[index_random_r, 6]) / R[index_random_r, 6] + \
                                                  theta_vehicle * (
                                                          len(K_canpick_both) / min(len(ok_K_canpickr[~np.isnan(ok_K_canpickr[:,index])][:-1,index]),
                                                                                    len(ok_K_canpickr[~np.isnan(ok_K_canpickr[:,index_random])][:-1,index_random])))
    # not remove r which is total same in routes
    for r in relateness_value[:,1]:
        if relateness_value[list(relateness_value[:,1]).index(r),0] == 0:
            relateness_value = np.delete(relateness_value,list(relateness_value[:,1]).index(r), axis=0)
    if len(relateness_value) == 0:
        return routes, R_pool
    relateness_value =  relateness_value[np.argsort(relateness_value[:, 0])]
    remove_number = int(max(0.2 * len(relateness_value), 1))
    for r in relateness_value[:,1][0:remove_number]:
        # routes_local = my_deepcopy(routes)
        # R_pool_local = copy.copy(R_pool)

        routes, R_pool = remove_a_request(r, routes, R_pool)[0:2]
        # find_unchecked_r_preference([6, 45])
        #lost_r()
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    #find_unchecked_r_preference([6,45])
    return routes, R_pool


# @profile()
# @time_me()
def history_removal(swap=0):
    global routes, R_pool, R
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    r_in_routes = list(R[:,7])
    for r in R_pool[:, 7]:
        try:
            r_in_routes.remove(r)
        except:
            sys.exit(-9)


    # r_cost_gap = pd.DataFrame(columns=['cost_gap'], index=r_in_routes)
    r_cost_gap = np.array(np.empty(shape=(len(r_in_routes), 2)))
    r_cost_gap[:] = np.NaN
    r_cost_gap[:, 1] = r_in_routes
    for r in r_in_routes:
        index_r = list(R[:, 7]).index(r)
        result = get_r_cost_in_all_routes(r, 1)
        current_cost,v_has_r, used_T = result[0],result[4],result[5]
        if current_cost != 1000000000:
            if not np.isnan(r_best_obj_record[index_r,0]):
                r_cost_gap[list(r_cost_gap[:,1]).index(r),0] = current_cost - r_best_obj_record[index_r,0]
            update_r_best_obj_record(r, current_cost,v_has_r, used_T)
    for r in r_cost_gap[:,1]:
        # if the current insert is the best or it hasn't been record before, do nothing
        index = list(r_cost_gap[:,1]).index(r)
        if r_cost_gap[index,0] <= 0.01 or pd.isnull(r_cost_gap[index,0]):
            r_cost_gap = np.delete(r_cost_gap, index, axis=0)
    if len(r_cost_gap) == 0:
        #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
        if swap == 1:
            return r_cost_gap
        else:
            return routes, R_pool
    # r_cost_gap = r_cost_gap.sort_values(by=['cost_gap'], ascending=False)
    r_cost_gap = r_cost_gap[np.argsort(-r_cost_gap[:, 0])]
    if swap == 1:
        #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
        return r_cost_gap
    remove_number = int(max(len(r_cost_gap), 1))
    for r in r_cost_gap[:,1][0:remove_number]:
        # if r in [15, 31, 43, 46, 50, 65]:
        #     print('wfw')
        # routes_local = my_deepcopy(routes)
        # R_pool_local = copy.copy(R_pool)
        routes, R_pool = remove_a_request(r, routes, R_pool)[0:2]
        #lost_r()
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    return routes, R_pool


# @profile()
# @time_me()
def ok_distance(m, n, k_change, T_change):
    m,n = int(m),int(n)
    original_distance = D_origin_All[m][n]
    new_distance = D[k_change][T_change,m] + D[k_change][n,T_change]
    if new_distance > 1.3 * original_distance:
        return 0
    else:
        return 1


# @profile()
# @time_me()
def func_ok_K_canpickr():
    # ok_K_canpickr = pd.DataFrame(columns=R[:,7], index=range(len(range(len(K)))))
    ok_K_canpickr = np.array(np.empty(shape=(len(K)+1,len(R))))
    ok_K_canpickr[:] = np.NaN
    ok_K_canpickr[len(K)] = R[:,7]
    for r in R[:,7]:
        n = 0
        index = list(R[:,7]).index(r)
        index_k = list(ok_K_canpickr[len(K)]).index(r)
        for k in range(len(K)):
            # capacity < load
            
            if K[k, 0] >= R[index, 6]:
                arrive_time = D[k][R[index, 0],o[k,0]] / K[k, 1]

                if k in fixed_vehicles_percentage:
                    # for fixed k, it can't wait out of time window of terminal
                    if K[k, 5] == 3:
                        if Fixed[k][0, 2] < R[index, 2]:
                            continue
                    else:

                        if Fixed[k][0, 2] <= R[index, 2]:
                            continue

                    arrive_time = arrive_time + Fixed[k][0, 1]
                else:
                    # for free k, it's begin terminal can't far away from pickup terminal, but I need to consider that a k serve multiple r and one r is in the middle of routes,
                    # so the far value is very large
                    if D[k][R[index, 0],o[k,0]] > 400:
                        continue

                # this is used to remove k which can't departure earlier than final pickup time, although there is storage, but containers can only be stored in the pickup time window
                if Demir == 1:
                    #in Demir, the time is based on number of containers, so I just don't add it
                    departure_time = arrive_time
                else:
                    if K[k, 5] == 1 or K[k, 5] == 2:
                        departure_time = arrive_time + service_time
                    else:
                        departure_time = arrive_time
                if departure_time > R[index, 3] and Demir != 1:
                    pass
                else:
                    if k in fixed_vehicles_percentage:
                        # this considers the fixed k which has more than two fixed terminals, then the pickup terminal may not the begin depot
                        if departure_time <= Fixed[k][0,2]:
                            ok_K_canpickr[n,index_k] = k
                            n = n + 1
                    else:
                        ok_K_canpickr[n,index_k] = k
                        n = n + 1
    return ok_K_canpickr

def get_K_R_unit(r,k,l,k_is_first,l_is_last):
    index_r = list(R[:,7]).index(r)
    # forbid transshipment between trucks
    if forbid_T_trucks == 1 and percentage != 0:
        if K[k, 5] == 3 and K[l, 5] == 3:
            return 0
    if k in fixed_vehicles_percentage and len(Fixed[k]) == 2 and l in fixed_vehicles_percentage and len(
            Fixed[l]) == 2:
        if not (Fixed[k][1, 0] == Fixed[l][0, 0]):
            return 0
    if k_is_first == 1:
        if k in fixed_vehicles_percentage:
            if len(Fixed[k]) == 2:
                if not R[index_r, 0] == Fixed[k][0, 0]:
                    return 0
            else:
                if not (R[index_r, 0] in list(Fixed[k][:, 0]) and R[index_r, 0] != list(Fixed[k][:, 0])[-1]):
                    return 0
    if l_is_last == 1:
        if l in fixed_vehicles_percentage:
            if len(Fixed[l]) == 2:
                if not R[index_r, 1] == Fixed[l][1, 0]:
                    return 0
            else:
                if not (R[index_r, 1] in list(Fixed[l][:, 0]) and R[index_r, 1] != list(Fixed[l][:, 0])[0]):
                    return 0
        else:
            # danger it may lose some potential solution
            if D[l][o[l, 1], R[index_r, 1]] > 300:
                return 0
    return 1

def get_K_R_unit2(r):
    global K_R, ok_K_canpickr
    index_r = list(R[:, 7]).index(r)
    K_R['1k'][r] = []
    K_R['2k'][r] = []
    K_R['3k'][r] = []
    index = list(ok_K_canpickr[len(K)]).index(r)
    for k in ok_K_canpickr[~np.isnan(ok_K_canpickr[:, index])][:-1, index]:

        k = int(k)

        if D[k][R[index_r, 1], R[index_r, 0]] > 10000:
            continue
        route = np.array(np.empty(shape=(5, 0)), dtype='object')
        route = np.insert(route, 0, [R[index_r, 0], R[index_r, 0], R[index_r, 0], R[index_r, 0], 'begin_depot'], axis=1)
        route = np.insert(route, 1, [R[index_r, 1], R[index_r, 1], R[index_r, 1], R[index_r, 1], 'end_depot'], axis=1)
        if k not in fixed_vehicles_percentage:
            # danger it may lose some potential solution
            if D[k][o[k, 1], R[index_r, 1]] > 300:
                continue
            if forbid_much_delay == 1:
                # can't delay more than 2 h
                delay_time = R[index_r, 2] + D[k][R[index_r, 1], R[index_r, 0]] / K[k, 1] - R[index_r, 5]
                if delay_time > 1:
                    continue
        else:
            if forbid_much_delay == 1:
                # can't delay more than 2 h
                delay_time = Fixed[k][1, 1] - R[index_r, 5]
                if delay_time > 1:
                    continue

        if Fixed_route(k, route) == False:
            continue
        else:
            K_R['1k'][r].append(k)
    index = list(ok_K_canpickr[len(K)]).index(r)
    # only one k is enough because another k may not able to pickup it but can transfer it
    if len(ok_K_canpickr[~np.isnan(ok_K_canpickr[:, index])][:-1, index]) >= 1:
        for k in ok_K_canpickr[~np.isnan(ok_K_canpickr[:, index])][:-1, index]:
            k = int(k)
            l_list = list(range(len(K)))
            l_list.remove(k)
            for l in l_list:
                if get_K_R_unit(r, k, l, 1, 1) == 1:
                    K_R['2k'][r].append([k, l])
    # only two k is enough because another k may not able to pickup it but can transfer it
    if two_T == 1 and len(ok_K_canpickr[~np.isnan(ok_K_canpickr[:, index])][:-1, index]) >= 2:
        for k in ok_K_canpickr[~np.isnan(ok_K_canpickr[:, index])][:-1, index]:
            k = int(k)
            l_list = list(range(len(K)))
            l_list.remove(k)
            for l in l_list:
                v_list = list(range(len(K)))
                v_list.remove(k)
                v_list.remove(l)
                for v in v_list:
                    if get_K_R_unit(r, k, l, 1, 0) == 1 and get_K_R_unit(r, l, v, 0, 1) == 1:
                        K_R['3k'][r].append([k, l, v])
# this function used to delete k which is unsuitable for r
# @profile()
# @time_me()
def get_K_R():
    global K_R
    K_R = {}
    K_R['1k'] = {}
    K_R['2k'] = {}
    K_R['3k'] = {}
    for r in R[:,7]:
        get_K_R_unit2(r)
    return K_R

def ok_TK_unit(i,T_change,k_change,l_change,k_first=1,l_last=1):
    index_r = list(R[:, 7]).index(i)
    if o[k_change, 0] == T_change:
        return 0

    # if fixed k's fixed terminals not in the terminals of request and T, then not be considered. But this is only for fixed k with two terminals
    if k_change in fixed_vehicles_percentage:
        if len(Fixed[k_change]) == 2:
            if not (o[k_change, 1] == T_change):
                return 0
        else:
            if T_change not in list(Fixed[k_change][:, 0]):
                return 0
    # if T is far away from k's end depot, but danger, it will lose some pential solution, but it's ok if I set the far value very large
    else:
        if D[k_change][T_change, o[k_change, 1]] > 300:
            return 0

    # else:
    #     if D[k_change][o[k_change,0]][R[index_r,0]] / K[k_change,1] > R[index_r,3]:
    #         return 0

    if o[l_change, 1] == T_change:
        return 0
    if (k_first == 1 and D[k_change][T_change, R[index_r, 0]] > 10000) or (l_last == 1 and D[l_change][R[index_r, 1], T_change] > 10000):
        return 0
    # if fixed k's fixed terminals not in the terminals of request and T, then not be considered. But this is only for fixed k with two terminals
    if l_change in fixed_vehicles_percentage:
        if len(Fixed[l_change]) == 2:
            if not (o[l_change, 0] == T_change):
                return 0
        else:
            if T_change not in list(Fixed[l_change][:, 0]):
                return 0

    # if T is far away from l's begin depot, but danger, it will lose some pential solution, but it's ok if I set the far value very large
    else:
        if D[l_change][T_change, o[l_change, 0]] > 300:
            return 0

    # if both k in fixed k, if k and l are not truck, then l's latest departure time can't be earlier than k's earliest arrival time at T
    if k_change in fixed_vehicles_percentage and l_change in fixed_vehicles_percentage:
        if Fixed[l_change][0, 2] < Fixed[k_change][1, 1] and K[l_change, 5] != 3 and K[k_change, 5] != 3:
            return 0

    # else:
    #     if D[k_change][o[k_change,0]][R[index_r, 0]] / K[k_change,1] + service_time * 4 + (D[l_change][o[l_change,0]][T_change] + D[l_change][T_change][R[index_r, 1]])/ K[l_change,1] > R[index_r,5] + 10:
    #         return 0
    #        route_terminals = routes[k][0].tolist()

    # if begin node of k_change very close to the pickup node and end node of l_change very clcose to delivery node, the vehicle may designed for this request, no matter how far it to the T
    # if end node of k_change begin node of l_change very close to the T terminal, it also considered
    if (k_first == 1 and l_last == 1 and D[k_change][R[index_r, 0], o[k_change, 0]] <= 100 and D[k_change][
        R[index_r, 1], o[l_change, 1]] <= 100) or (
            D[k_change][T_change, o[k_change, 1]] <= 100 and D[l_change][T_change, o[l_change, 0]] <= 100):
        return 1
    # can't delay more than 2 h
    if l_last == 1:
        if l_change in fixed_vehicles_percentage and K[l_change, 5] != 3:
            if forbid_much_delay == 1:
                delay_time = Fixed[l_change][1, 2] - R[index_r, 5]
                if delay_time > 1:
                    return 0
        else:
            if k_first == 1:
                if k_change in fixed_vehicles_percentage and K[k_change, 5] != 3:
                    # danger fixed k only two terminals
                    arrive_T_time = Fixed[k_change][1, 2]
                else:
                    arrive_T_time = R[index_r, 2] + D[k_change][T_change, R[index_r, 0]] / K[k_change, 1] - R[index_r, 5]
                if Demir == 0:
                    if K[k_change, 5] != 3:
                        arrive_T_time = arrive_T_time + 1
                    if K[l_change, 5] != 3:
                        arrive_T_time = arrive_T_time + 1
                if forbid_much_delay == 1:
                    delay_time = arrive_T_time + D[l_change][R[index_r, 1], T_change] / K[l_change, 1] - R[index_r, 5]
                    if delay_time > 1:
                        return 0
    return 1
# @profile()
# @time_me()
##@jit
def ok_TK(i):
    global K_R
    index_r = list(R[:, 7]).index(i)
    all_ok_TK_i = {}
    for T_change in T:

        all_ok_TK_i_list = []
        if R[index_r, 0] == T_change or R[index_r, 1] == T_change:
            continue

        if D_origin_All[R[index_r, 0]][T_change] + D_origin_All[T_change][R[index_r, 1]] > 1.3 * D_origin_All[R[index_r, 0]][
            R[index_r, 1]]:
            continue
        #        all_ok_K = ok_K(T_change, all_ok_K)
        for index in range(0, len(K_R['2k'][i])):
            k_change = K_R['2k'][i][index][0]
            l_change = K_R['2k'][i][index][1]
            if ok_TK_unit(i,T_change, k_change, l_change) == 1:
                all_ok_TK_i_list.append([k_change, l_change])
        if all_ok_TK_i_list:
            # all_ok_TK_i[T_change] = pd.DataFrame(all_ok_TK_i_list, columns=['k', 'l'])
            all_ok_TK_i[T_change] = np.array(all_ok_TK_i_list)
    if two_T == 1 and len(T) > 1:
        for T_change1 in T:
            for T_change2 in T:
                if T_change1 != T_change2:

                    all_ok_TK_i_list = []
                    if R[index_r, 0] == T_change1 or R[index_r, 1] == T_change2:
                        continue

                    if D_origin_All[R[index_r, 0]][T_change1] + D_origin_All[T_change1][T_change2] + D_origin_All[T_change2][R[index_r, 1]] > 1.5 * D_origin_All[R[index_r, 0]][
                        R[index_r, 1]]:
                        continue
                    #        all_ok_K = ok_K(T_change, all_ok_K)
                    for index in range(0, len(K_R['3k'][i])):
                        k_change,l_change,v_change = K_R['3k'][i][index]

                        if ok_TK_unit(i,T_change1, k_change, l_change,1,0) == 1 and ok_TK_unit(i,T_change2, l_change, v_change,0,1) == 1:
                            all_ok_TK_i_list.append([k_change, l_change, v_change])
                    if all_ok_TK_i_list:
                        # all_ok_TK_i[T_change] = pd.DataFrame(all_ok_TK_i_list, columns=['k', 'l'])
                        all_ok_TK_i[tuple([T_change1,T_change2])] = np.array(all_ok_TK_i_list)
    return all_ok_TK_i


# @profile()
# @time_me()
def ok_k(k_change, T_change):
    global routes
    if parallel == 1:
        routes = load_obj(['routes'])[0]
    transposed_route = routes[k_change][0].T
    res = [x[0] for x in groupby(transposed_route.tolist())]

    if T_change in res:
        return 1
    else:
        res2 = list(zip(res, res[1:] + res[:1]))
        del res2[-1]
        for pair in res2:
            #                    distance_all = D[k_change][pair[0]][T_change] + D[k_change][T_change][pair[1]]
            ok_or = ok_distance(pair[0], pair[1], k_change, T_change)
            if ok_or == 1:
                return 1
    return 0


def delete_k_local(args):
    global D, D_origin_All, K, exp_number, parallel_number, parallel
    k_change, l_change, T_change, request_number_in_R, exp_number, parallel_number, parallel = args
    delete_K_pair_T_change_local = []
    K = read_R_K(request_number_in_R, 'K')
    D, D_origin_All = read_D('D_All', K)
    if ok_k(k_change, T_change) == 1 and ok_k(l_change, T_change) == 1:
        delete_K_pair_T_change_local.append([k_change, l_change])
    return delete_K_pair_T_change_local


# #@time_me
# @profile()
# @time_me()
def delete_k(i):
    delete_K_pair = {}
    if parallel == 1:
        save_obj([routes], ['routes'])
    for T_change in all_ok_TK[i]:
        #        all_ok_K = ok_K(T_change, all_ok_K)
        if isinstance(T_change,int):
            if parallel == 1:
                iterate_what = []
                for index in range(0, len(all_ok_TK[i][T_change])):
                    k_change, l_change = all_ok_TK[i][T_change][index]
                    iterate_what.append(
                        [k_change, l_change, T_change, request_number_in_R, exp_number, parallel_number, parallel])
                # save_obj([all_ok_k_pair, T_change],['all_ok_k_pair', 'T_change'])
                chunksize = int(max(os.cpu_count() * 2, len(all_ok_TK[i][T_change]) / os.cpu_count()))
                with ProcessPoolExecutor(initializer=init_routes) as e:
                    results = e.map(delete_k_local, iterate_what, chunksize=chunksize)
                for result in results:
                    delete_K_pair[T_change] = result

            else:
                delete_K_pair[T_change] = []
                for index in range(0, len(all_ok_TK[i][T_change])):
                    k_change, l_change = all_ok_TK[i][T_change][index]
                    #        route_terminals = routes[k][0].tolist()
                    if ok_k(k_change, T_change) == 1 and ok_k(l_change, T_change) == 1:
                        delete_K_pair[T_change].append([k_change, l_change])
        else:
            delete_K_pair[T_change] = []
            for index in range(0, len(all_ok_TK[i][T_change])):
                k_change, l_change, v_change = all_ok_TK[i][T_change][index]
                #        route_terminals = routes[k][0].tolist()
                T_change1, T_change2 = T_change
                if ok_k(k_change, T_change1) == 1 and ok_k(l_change, T_change1) == 1 and ok_k(l_change, T_change2) == 1 and ok_k(v_change, T_change2) == 1:
                    delete_K_pair[T_change].append([k_change, l_change, v_change])
    return delete_K_pair


# @profile()
# @time_me()

def ok_position(seq, length):
    all_no_position = []
    all_position = list(range(1, length))

    tally = defaultdict(list)
    for i, item in enumerate(seq):
        tally[item].append(i)
    for dup in sorted((key, locs) for key, locs in tally.items() if len(locs) > 1):
        del dup[1][0]
        all_no_position.extend(dup[1])
    all_ok_position = [x for x in all_position if x not in all_no_position]
    return all_ok_position


def get_best_position(obj_list):
    # obj_df_one_column = pd.DataFrame(obj_list, columns=['one_column'])
    obj_df = pd.DataFrame(obj_list,
                          columns=['k', 'original_route', 'original_route_no_columns', 'cost_inserted_request',
                                   'dict_a_request_best_position'])
    obj_best = obj_df.loc[obj_df['cost_inserted_request'] == obj_df['cost_inserted_request'].min()]

    dict_a_request_best_position = obj_best.iloc[0]['dict_a_request_best_position']
    return dict_a_request_best_position


# @profile()
# @time_me()
##@jit
def one_position_insert_delivery(routes2, R, has_end_depot, i, k, j, h, R_i, original_route_no_columns,
                                 all_obj_positions_list, Trans, Trans_Td,
                                 Trans_secondTp, Trans_secondTd, new_try, fixed_vehicles_percentage, Fixed, K,
                                 hash_table_1v_all_fail, hash_table_1v_all, no_route_barge, no_route_truck, segment_in_dynamic, check_uncertainty_in_insertion_by_RL_or_not=0, new_row=-1, finish_or_begin=-1, uncertainty_index=-1, store_routes_for_another_k=-1, store_R_pool_for_another_k=-1,duration=-1, congestion_link=-1,
                                 congestion_node=-1, index = -1):
    global check_start_position, routes
    routes = routes2
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    position = [j, h]
    position = tuple(position)

    key = get_key_1k(R_i, original_route_no_columns, k, fixed_vehicles_percentage, Fixed, K)

    if key in hash_table_1v_all_fail.keys():
        if position in hash_table_1v_all_fail[key].keys():
            return all_obj_positions_list, new_try
    if key in hash_table_1v_all.keys():
        if position in hash_table_1v_all[key].keys():
            all_obj_positions_list.append([position, hash_table_1v_all[key][position]['cost_inserted_request']])
    else:

        if Trans == 1 and Trans_Td == 1:
            operation = 'Td'
        elif Trans == 1 and Trans_secondTd == 1:
            operation = 'secondTd'
        else:
            operation = 'delivery'

        new_try = np.insert(new_try, h, [R_i[1][0], R_i[1][0], R_i[1][0], R_i[1][0], str(i) + operation], axis=1)

        check_start_position = j
        result, wrong_time = satisfy_constraints(routes, has_end_depot, R, k, new_try, fixed_vehicles_percentage, K, no_route_barge,
                                     no_route_truck,i)
        if stochastic == 1 and dynamic_t > 0 and ALNS_greedy_under_unknown_duration_assume_duration != 0 and K[k, 5] == influenced_mode_by_current_event:
            for index2, element in enumerate(new_try[0]):
                if index2 != 0 and index2 != len(new_try[0]) - 1:
                    if element == congestion_node and i == get_numbers(new_try[4, index2]):
                        col_congestion = index2
                        vehicle_stop_time = get_vehicle_stop_time(R_change_dynamic_travel_time, index, new_try,
                                                                  col_congestion, duration)
                        add_duration(R_change_dynamic_travel_time, index, k, congestion_link, congestion_node, uncertainty_index, new_try, col_congestion,
                                     duration,
                                     vehicle_stop_time,i)
                        result, wrong_time = satisfy_constraints(routes, has_end_depot, R, k, new_try,
                                                                 fixed_vehicles_percentage, K, no_route_barge,
                                                                 no_route_truck, i,0,1)
                        break
        # 20211121 here I should not check preferences, because this function also used by insert2vehicles, and when only insert the first segment, the preferences may not be satisfied, but when consider the whole journey, the preferences can be satisfied
            # and in preference_constraints function, the distance is the overall distance of request r, not request segment, so the calculation is wrong
        # if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
        #     update_request_flow_t(new_try)
        #     satisfy_preference = preference_constraints(i,k,-1,-1,new_try,-1,-1)
        #     if satisfy_preference == 1:
        #         if preference_relevant(k,new_try,i) != 1:
        #             satisfy_preference = 0
        feasible = 0
        if isinstance(result, np.ndarray):
            feasible = 1
            # # 20211121 commented, see above
            # if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
            #     if satisfy_preference != 1:
            #         feasible = 0

        if feasible == 1:

            if stochastic == 1 and add_RL == 0 and dynamic_t > 0 and K[k, 5] == influenced_mode_by_current_event:
                routes_copy = my_deepcopy(routes)
                routes_copy[k] = result
                # store the routes, action for evaluate the performance of ALNS during implementation
                if (uncertainty_index, i) not in ALNS_insertion_implementation_store.keys():
                    ALNS_insertion_implementation_store[
                        (uncertainty_index, i)] = {}
                ALNS_insertion_implementation_store[(uncertainty_index, i)][
                    k] = store_all(0, -1, 0, routes_copy)
                # routes = my_deepcopy(store_unchanged_routes_for_RL_insertion)
            if check_uncertainty_in_insertion_by_RL_or_not == 1 and K[k, 5] == influenced_mode_by_current_event:
                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                store_unchanged_routes_for_RL_insertion = my_deepcopy(routes)
                routes[k] = result
                feasibility_or_route = RL_insertion_constraints(index, result, i, new_row, finish_or_begin, uncertainty_index, k, store_routes_for_another_k, store_R_pool_for_another_k,duration, congestion_link,
                                 congestion_node)
                routes = my_deepcopy(store_unchanged_routes_for_RL_insertion)
                # for key_test in routes.keys():
                #     if np.array_equal(routes[key_test], result, equal_nan=False):
                #         print('caught it!')

                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                if not isinstance(feasibility_or_route, np.ndarray):
                    return all_obj_positions_list, new_try
            new_try = result
            cost_inserted_request, cost_all_requests = objective_value_i(i, k, new_try)[0:2]
            dict_a_request_a_position = {'route': copy.copy(new_try),
                                         'cost_inserted_request': cost_inserted_request,
                                         'cost_all_requests': cost_all_requests}
            if key not in hash_table_1v_all.keys():
                hash_table_1v_all[key] = {}
            hash_table_1v_all[key][position] = dict_a_request_a_position
            all_obj_positions_list.append([position, hash_table_1v_all[key][position]['cost_inserted_request']])
        else:

            if stochastic == 1 and add_RL == 0 and dynamic_t > 0 and K[k, 5] == influenced_mode_by_current_event:
                if wrong_time == 'wrong_time':
                    #only store the k which not satisfies time_constraints
                    routes_copy = my_deepcopy(routes)
                    routes_copy[k] = new_try
                    # store the routes, action for evaluate the performance of ALNS during implementation
                    if (uncertainty_index, i) not in ALNS_insertion_implementation_store.keys():
                        ALNS_insertion_implementation_store[
                            (uncertainty_index, i)] = {}
                    ALNS_insertion_implementation_store[(uncertainty_index, i)][
                        k] = store_all(0, -1, 1, routes_copy)
            if key not in hash_table_1v_all_fail.keys():
                hash_table_1v_all_fail[key] = {}
            hash_table_1v_all_fail[key][position] = {}

        if Trans == 1 and Trans_Td == 1:
            new_try = np.delete(new_try, list(new_try[4]).index(str(i) + operation), 1)
        elif Trans == 1 and Trans_secondTd == 1:
            new_try = np.delete(new_try, list(new_try[4]).index(str(i) + operation), 1)
        else:
            new_try = np.delete(new_try, list(new_try[4]).index(str(i) + operation), 1)
    return all_obj_positions_list, new_try


# @profile()
# @time_me()
##@jit
# #@set_fun
def one_position_insert(R, has_end_depot, routes2, j, random_position, i, k, Trans, Trans_Tp, Trans_Td, Trans_secondTp,
                        Trans_secondTd, R_i,
                        original_route_no_columns, all_obj_positions_list, all_ok_position, fixed_vehicles_percentage,
                        Fixed, K, hash_table_1v_all_fail, hash_table_1v_all, no_route_barge, no_route_truck, segment_in_dynamic = 0, check_uncertainty_in_insertion_by_RL_or_not=0, new_row=-1, finish_or_begin=-1, uncertainty_index=-1, store_routes_for_another_k=-1, store_R_pool_for_another_k=-1,duration=-1, congestion_link=-1,
                                 congestion_node=-1, index = -1):
    global routes
    routes = routes2
    new_try = copy.copy(routes[k])
    #    new_try.iloc[1] = 'string'
    #    new_try.iloc[2] = 'string'
    #    new_try.iloc[3] = 'string'
    # kth routes, jth position, column name is ith request, insert request's pick up node

    if Trans == 1 and Trans_Tp == 1:
        operation = 'Tp'
    elif Trans == 1 and Trans_secondTp == 1:
        operation = 'secondTp'
    else:
        operation = 'pickup'
    new_try = np.insert(new_try, j, [R_i[0][0], R_i[0][0], R_i[0][0], R_i[0][0], str(i) + operation], axis=1)
    if new_subtour_constraints(new_try[0]) == False or capacity_constraints(has_end_depot, K, R, k, new_try) == False:
        return all_obj_positions_list

    #    if has_end_depot==1:
    #        length = len(new_try[4])
    #    else:
    #        length = len(new_try[4])+1
    all_ok_position2 = []
    if j == max(all_ok_position):
        all_ok_position2.append(j + 1)
    else:
        for x in all_ok_position:
            if x >= j:
                y = x + 1
                all_ok_position2.append(y)
    # after the pick up node is added, it should be j+1
    if random_position == 0:
        for h in all_ok_position2:
            if h > j:
                # if check_uncertainty_in_insertion_by_RL_or_not == 0:
                #     all_obj_positions_list, new_try = one_position_insert_delivery(routes, R, has_end_depot, i, k, j, h,
                #                                                                    R_i,
                #                                                                    original_route_no_columns,
                #                                                                    all_obj_positions_list, Trans, Trans_Td,
                #                                                                    Trans_secondTp, Trans_secondTd, new_try,
                #                                                                    fixed_vehicles_percentage, Fixed, K,
                #                                                                    hash_table_1v_all_fail,
                #                                                                    hash_table_1v_all, no_route_barge,
                #                                                                    no_route_truck,segment_in_dynamic)
                # else:
                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                all_obj_positions_list, new_try = one_position_insert_delivery(routes, R, has_end_depot, i, k, j, h,
                                                                               R_i,
                                                                               original_route_no_columns,
                                                                               all_obj_positions_list, Trans,
                                                                               Trans_Td,
                                                                               Trans_secondTp, Trans_secondTd,
                                                                               new_try,
                                                                               fixed_vehicles_percentage, Fixed, K,
                                                                               hash_table_1v_all_fail,
                                                                               hash_table_1v_all, no_route_barge,
                                                                               no_route_truck, segment_in_dynamic, check_uncertainty_in_insertion_by_RL_or_not, new_row, finish_or_begin, uncertainty_index, store_routes_for_another_k, store_R_pool_for_another_k,duration, congestion_link,
                             congestion_node, index)
    else:
        h = int(np.random.choice(all_ok_position2, size=(1,)))
        all_obj_positions_list, new_try = one_position_insert_delivery(routes, R, has_end_depot, i, k, j, h, R_i,
                                                                       original_route_no_columns,
                                                                       all_obj_positions_list, Trans, Trans_Td,
                                                                       Trans_secondTp, Trans_secondTd, new_try,
                                                                       fixed_vehicles_percentage, Fixed, K,
                                                                       hash_table_1v_all_fail, hash_table_1v_all,
                                                                       no_route_barge, no_route_truck,segment_in_dynamic)
    # if Trans == 1 and Trans_Tp == 1:
    #     new_try = np.delete(new_try, list(new_try[4]).index(str(i) + 'Tp'), 1)
    # elif Trans == 1 and Trans_secondTp == 1:
    #     new_try = np.delete(new_try, list(new_try[4]).index(str(i) + 'secondTp'), 1)
    # else:
    #     new_try = np.delete(new_try, list(new_try[4]).index(str(i) + 'pickup'), 1)
    return all_obj_positions_list


# @profile()
# @time_me()
##@jit
def best_position_1_vehicle(R, no_route_barge, no_route_truck, hash_table_1v_all_fail, hash_table_1v_all, routes2,
                            fixed_vehicles_percentage, Fixed, K, hash_table_1v, hash_table_1v_fail, has_end_depot, R_i,
                            i, k, Trans, Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd, random_position=0,segment_in_dynamic=0, check_uncertainty_in_insertion_by_RL_or_not=0, new_row=-1, finish_or_begin=-1, uncertainty_index=-1, store_routes_for_another_k=-1, store_R_pool_for_another_k=-1,duration=-1, congestion_link=-1,
                                 congestion_node=-1,index=-1):
    global routes
    # print('routes2', routes2[83], 'routes', routes[83])
    routes = my_deepcopy(routes2)
    # print(routes[83])
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    obj_1_vehicle = []
    original_route = copy.copy(routes[k])
    #     original_route_no_columns = copy.copy(routes[k])
    # #    original_route_no_columns.columns = list(range(len(original_route_no_columns.columns)))
    #     original_route_no_columns = [tuple(x) for x in original_route_no_columns.to_records(index=False)]
    #     original_route_no_columns.append(tuple(routes[k][4]))
    #     original_route_no_columns = tuple(original_route_no_columns)
    #     original_route_no_columns = (original_route_no_columns)

    original_route_no_columns = route_no_columns(routes[k])
    key = get_key_1k(R_i, original_route_no_columns, k, fixed_vehicles_percentage, Fixed, K)
    if key in hash_table_1v_fail.keys():
        return obj_1_vehicle

    if key in hash_table_1v.keys():
        obj_1_vehicle.append([k, original_route, original_route_no_columns,
                              hash_table_1v[key][list(hash_table_1v[key])[0]]['cost_inserted_request'],
                              list(hash_table_1v[key])[0]])
        # caculate all positions' obj
    else:
        all_obj_positions_list = []
        # Pick up point
        if has_end_depot == 1:
            length = len(routes[k][4])
        else:
            length = len(routes[k][4]) + 1

        all_ok_position = ok_position(routes[k][0].tolist(), length)

        if random_position == 0:
            for j in all_ok_position:
                # if check_uncertainty_in_insertion_by_RL_or_not == 0:
                #     all_obj_positions_list = one_position_insert(R, has_end_depot, routes, j, random_position, i, k, Trans,
                #                                                  Trans_Tp, Trans_Td,
                #                                                  Trans_secondTp, Trans_secondTd, R_i,
                #                                                  original_route_no_columns, all_obj_positions_list,
                #                                                  all_ok_position, fixed_vehicles_percentage, Fixed, K,
                #                                                  hash_table_1v_all_fail, hash_table_1v_all, no_route_barge,
                #                                                  no_route_truck,segment_in_dynamic)
                # else:
                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                all_obj_positions_list = one_position_insert(R, has_end_depot, routes, j, random_position, i, k,
                                                             Trans,
                                                             Trans_Tp, Trans_Td,
                                                             Trans_secondTp, Trans_secondTd, R_i,
                                                             original_route_no_columns, all_obj_positions_list,
                                                             all_ok_position, fixed_vehicles_percentage, Fixed, K,
                                                             hash_table_1v_all_fail, hash_table_1v_all,
                                                             no_route_barge,
                                                             no_route_truck, segment_in_dynamic, check_uncertainty_in_insertion_by_RL_or_not, new_row, finish_or_begin, uncertainty_index, store_routes_for_another_k, store_R_pool_for_another_k,duration, congestion_link,
                             congestion_node,index)
        else:
            j = int(np.random.choice(all_ok_position, size=(1,)))
            all_obj_positions_list = one_position_insert(R, has_end_depot, routes, j, random_position, i, k, Trans,
                                                         Trans_Tp, Trans_Td,
                                                         Trans_secondTp, Trans_secondTd, R_i, original_route_no_columns,
                                                         all_obj_positions_list, all_ok_position,
                                                         fixed_vehicles_percentage, Fixed, K, hash_table_1v_all_fail,
                                                         hash_table_1v_all, no_route_barge, no_route_truck,segment_in_dynamic)
        if all_obj_positions_list:
            # all_position_obj = pd.DataFrame(all_obj_positions_list, columns=['position', 'cost_inserted_request'])
            all_position_obj = np.array(all_obj_positions_list)
            all_position_obj = all_position_obj[~np.isnan(list(all_position_obj[:,1]))]
            if all_position_obj.size != 0:
                dict_a_request_best_index = np.argmin(all_position_obj[:,1])
                dict_a_request_best_position,best_cost_inserted_request = all_position_obj[dict_a_request_best_index,:]
                # best_cost_inserted_request = all_position_obj['cost_inserted_request'][dict_a_request_best_index]
                if random_position == 0:
                    dict_a_request_best = hash_table_1v_all[key][dict_a_request_best_position]
                    if key not in hash_table_1v.keys():
                        hash_table_1v[key] = {}
                    hash_table_1v[key][dict_a_request_best_position] = copy.copy(dict_a_request_best)
                obj_1_vehicle.append([k, original_route, original_route_no_columns, best_cost_inserted_request,
                                      dict_a_request_best_position])
        else:
            hash_table_1v_fail[key] = {}
    return obj_1_vehicle


# @profile()
# @time_me()
def solve_relevant_try(relevant_try_copy,layer,aaa,check_preferences=0):
    global save_relevant_try_copy,relevant_try, check_start_position

    # the cross sysncronization will fall in dead loop, limit loop maximum number to avoid it
    #danger the 100 is depend on instance
    if aaa > 100:
        return 0
    # when len > 1, this function will be overwrited because the func solve_relevant_try called itself, so the second vehicle in relevant_try_copy will be wrong, I didn't find the solution
    # if len(relevant_try_copy) > 1:
    #     return 0
    # In fact, it only suitable for one relevant route, but no matter how deep it is, I mean the depth of relevan routes

    for v in relevant_try_copy.keys():
        if v == next(iter(relevant_try_copy)):
            if layer == 0:
                save_relevant_try_copy = {}
            save_relevant_try_copy[layer] = my_deepcopy(relevant_try_copy)
            layer = layer + 1

        # relevant_try_copy2 = my_deepcopy(relevant_try_copy)
        new_try = copy.copy(relevant_try_copy[v][0])
        check_start_position = relevant_try_copy[v][2]
        # this only do the check and doesn't change routes
        if check_preferences == 0:
            bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, v, new_try,relevant_try_copy[v][1])
        else:
            r_served_by_k = find_r_served_by_k(new_try)
            for r in r_served_by_k:
                #make sure v is the k1
                try:
                    k1, k2, k3 = find_used_k(r, v)
                except:
                    sys.exit(-8)
                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                if k2 == -1:
                    # r only served by the new_try, which is not be put into the routes, but r has not be checked
                    update_request_flow_t(new_try)
                    satisfy_preference = preference_constraints(r, k1, -1, -1, new_try, -1, -1)
                else:
                    if k3 == -1:
                        update_request_flow_t(new_try)
                        update_request_flow_t(routes[k2])
                        satisfy_preference = preference_constraints(r, k1, k2, -1, new_try, routes[k2], -1)
                    else:
                        update_request_flow_t(new_try)
                        update_request_flow_t(routes[k2])
                        update_request_flow_t(routes[k3])
                        satisfy_preference = preference_constraints(r, k1, k2, k3, new_try, routes[k2], routes[k3])
                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                if satisfy_preference == 0:
                    return 0
            bool_or_route = 1
        if isinstance(bool_or_route, bool):
            return 0
        else:
            #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
            #because when checking prefernces constraints, the uninserted new_try will be in relevant_try_copy, so I need to clarify that when preferences checking, the routes cannot be updated (and there are no changes when preferences checking)
            if check_preferences == 0:
                routes[v] = copy.copy(relevant_try_copy[v][0])
            #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
            if relevant_try:
                aaa = aaa + 1
                relevant_try_copy = my_deepcopy(relevant_try)
                result = solve_relevant_try(relevant_try_copy,layer,aaa,check_preferences)
                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                if result == 0:
                    return 0
                relevant_try_copy = my_deepcopy(save_relevant_try_copy[layer-1])
    return 1


# @profile()
# @time_me()
def reduce_ks(i, k1, k2, T_change, best_cost_inserted_request, all_ok_TK, delete_K_pair, not_consider_ks):
    # should be comment after know why sometimes best_cost_inserted_request < r_basic_cost
    index_r = list(R[:, 7]).index(i)
    for T_change_reduce in all_ok_TK[i].keys():
        all_ok_k_pair_reduce = np.array(delete_K_pair[T_change_reduce])
        if len(all_ok_k_pair_reduce) >= 1:
            for x_reduce in range(len(all_ok_k_pair_reduce)):
                if isinstance(T_change_reduce, int):
                    k1_reduce = all_ok_k_pair_reduce[x_reduce,0]
                    k2_reduce = all_ok_k_pair_reduce[x_reduce,1]
                    if [k1_reduce, k2_reduce, T_change_reduce] not in not_consider_ks:
                        r_basic_cost_reduce = get_r_basic_cost(R[index_r, 0], R[index_r, 1], i, k1_reduce, k2_reduce,
                                                               T_change_reduce)
                        if r_basic_cost_reduce > best_cost_inserted_request * 1.3:
                            not_consider_ks.append([k1_reduce, k2_reduce, T_change_reduce])
                else:
                    T_change_reduce1, T_change_reduce2 = T_change_reduce
                    k1_reduce, k2_reduce, k3_reduce = all_ok_k_pair_reduce[x_reduce]
                    if [k1_reduce, k2_reduce, k3_reduce, T_change_reduce1, T_change_reduce2] not in not_consider_ks:
                        r_basic_cost_reduce = get_r_basic_cost(R[index_r, 0], R[index_r, 1], i, k1_reduce, k2_reduce,
                                                               T_change_reduce1, k3_reduce, T_change_reduce2)
                        if r_basic_cost_reduce > best_cost_inserted_request * 1.3:
                            not_consider_ks.append([k1_reduce, k2_reduce, k3_reduce, T_change_reduce1, T_change_reduce2])
    return not_consider_ks

def check_a_k_in_dynamic(number_of_T, i, k1, k2 = -1, T1 = -1):
    if dynamic == 1 and dynamic_t > 0 and i in unchangeable_list.keys():
        # unchangeable_list[request_number] = [['k1'], T_k_record[number]] structure
        # when it is in unchangeable_list, it must be started by k1
        if k1 != unchangeable_list[i][1][2]:
            return -1
        if number_of_T == 1:
            if 'T1' in unchangeable_list[i][0] and T1 != unchangeable_list[i][1][0]:
                return -1
    #danger didn't consider two times of transshipments
    return 1

def get_request_list2(request_list2,i,Trans,Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd,dict_a_request_best_position):
    # if i == 3000000000001 or i == 3000000000006:
    #     print('wrong number in routes')

    if Trans == 1 and Trans_Tp == 1:
        request_list2.insert(list(dict_a_request_best_position)[0], str(i) + 'Tp')
    elif Trans == 1 and Trans_secondTp == 1:
        request_list2.insert(list(dict_a_request_best_position)[0], str(i) + 'secondTp')
    else:
        request_list2.insert(list(dict_a_request_best_position)[0], str(i) + 'pickup')
    if Trans == 1 and Trans_Td == 1:
        request_list2.insert(list(dict_a_request_best_position)[1], str(i) + 'Td')
    elif Trans == 1 and Trans_secondTd == 1:
        request_list2.insert(list(dict_a_request_best_position)[1], str(i) + 'secondTd')
    else:
        request_list2.insert(list(dict_a_request_best_position)[1], str(i) + 'delivery')
    return request_list2
# @profile()
# @time_me()
def insert1vehicle_base(R_i, i, K, Trans, Trans_Tp, Trans_Td, random_position=0, regret=0, segment_in_dynamic=0, check_uncertainty_in_insertion_by_RL_or_not=0, new_row=-1, finish_or_begin=-1, uncertainty_index=-1, store_routes_for_another_k=-1, store_R_pool_for_another_k=-1,duration=-1, congestion_link=-1,
                                 congestion_node=-1,index=-1):
    global routes
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    # one vehicle
    #lost_r()
    # if i == 2:
    #     print('sf')
    check_served_R()
    if segment_in_dynamic == 1:
        index_r = list(request_segment_in_dynamic[:, 7]).index(i)
    else:
        index_r = list(R[:, 7]).index(i)
    obj_list = []
    Trans_secondTp, Trans_secondTd = 0, 0
    if k_random_or == 1:
        random_k = int(np.random.choice([0, 1], size=(1,), p=[5. / 10, 5. / 10]))
    else:
        random_k = 0
    random_k = 0
    #lost_r()
    # #the changed r can't use
    # if dynamic == 0 or (dynamic == 1 and i not in R_change_dynamic[:, 7]):
    if random_k == 0 and random_position == 0:
        routes_tuple = get_routes_tuple(routes)
        # R_pool_tuple = tuple(.to_records)
        top_key = (i, routes_tuple, 'insert1vehicle')
        if top_key in hash_top.keys():
            print('top')
            obj_best_k, route_best_k, position, cost_inserted_request = copy.copy(
                hash_top[top_key]['obj_best_k']), copy.copy(hash_top[top_key]['route_best_k']), copy.copy(
                hash_top[top_key]['position']), copy.copy(hash_top[top_key]['cost_inserted_request'])
            # print('1430',routes['Barge3'],obj_best_k, route_best_k)
            return obj_best_k, route_best_k, position, cost_inserted_request
    #lost_r()
    if random_k == 0:
        no_train_truck = 0
        for k in K_R['1k'][i]:
            if check_a_k_in_dynamic(0, i, k) == -1:
                continue
            ########
            # danger forget why set this when cahnge df to array 20201124
            # if not isinstance(k, str):
            #     continue
            #######
            # I guess the above is for k == 0 (which means infeasible), so I change it to if k==-1 (infeasible in current version continue
            if not (isinstance(k, (int, np.integer)) and k != -1):
                continue
            if segment_in_dynamic == 1:
                #then this r's another segment is already in another k, so check the used_k of another segment and the k in used_k can not be used
                used_k_ = find_used_k(i - (i - i % 10000) % big_r)
                if k in used_k_:
                    continue
            if no_train_truck == 1:
                if k in train_truck:
                    continue

            if random_position == 0:
                #print(routes[83])
                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                # print(routes[83])
                obj_1_vehicle = best_position_1_vehicle(R, no_route_barge, no_route_truck, hash_table_1v_all_fail,
                                                            hash_table_1v_all, routes, fixed_vehicles_percentage, Fixed, K,
                                                            hash_table_1v, hash_table_1v_fail, has_end_depot, R_i, i, k,
                                                            Trans, Trans_Tp, Trans_Td, Trans_secondTp,
                                                            Trans_secondTd,segment_in_dynamic=segment_in_dynamic, check_uncertainty_in_insertion_by_RL_or_not=check_uncertainty_in_insertion_by_RL_or_not, new_row=new_row, finish_or_begin=finish_or_begin, uncertainty_index=uncertainty_index, store_routes_for_another_k=store_routes_for_another_k, store_R_pool_for_another_k=store_R_pool_for_another_k, duration=duration, congestion_link=congestion_link,
                congestion_node=congestion_node,index=index)


                #lost_r()
            else:
                obj_1_vehicle = best_position_1_vehicle(R, no_route_barge, no_route_truck, hash_table_1v_all_fail,
                                                        hash_table_1v_all, routes, fixed_vehicles_percentage, Fixed, K,
                                                        hash_table_1v, hash_table_1v_fail, has_end_depot, R_i, i, k,
                                                        Trans, Trans_Tp, Trans_Td, Trans_secondTp,
                                                        Trans_secondTd, random_position,segment_in_dynamic=segment_in_dynamic)
            if obj_1_vehicle:
                obj_list.append(obj_1_vehicle)
                # if minimize cost, then if I found a solution which use barge, then the solutiong which use train and truck are not be considereder
                # but i need to make sure if use barge there is no delay
                if multi_obj == 0:
                    if K[k, 5] == 1:
                        if segment_in_dynamic == 1:
                            #danger here maybe wrong
                            r_basic_cost = get_r_basic_cost(request_segment_in_dynamic[index_r, 0], request_segment_in_dynamic[index_r, 1], i, k)
                        else:
                            r_basic_cost = get_r_basic_cost(R[index_r, 0], R[index_r, 1], i, k)
                        if obj_1_vehicle[0][3] < r_basic_cost + R[index_r, 6] * 2 * c_storage - 0.1:
                            if i not in no_T_R:
                                no_T_R.append(i)
                            no_train_truck = 1
                            if regret == 0:
                                if obj_1_vehicle[0][3] < r_basic_cost + 0.1:
                                    break

    else:
        # k = 0
        # while not isinstance(k, str):
        k = random.choice(K_R['1k'][i])
        if check_a_k_in_dynamic(0, i, k) == -1:
            return -1, 0, 0, 0
        if random_position == 0:
            obj_1_vehicle = best_position_1_vehicle(R, no_route_barge, no_route_truck, hash_table_1v_all_fail,
                                                    hash_table_1v_all, routes, fixed_vehicles_percentage, Fixed, K,
                                                    hash_table_1v, hash_table_1v_fail, has_end_depot, R_i, i, k, Trans,
                                                    Trans_Tp, Trans_Td, Trans_secondTp,
                                                    Trans_secondTd,segment_in_dynamic=segment_in_dynamic, check_uncertainty_in_insertion_by_RL_or_not=check_uncertainty_in_insertion_by_RL_or_not, new_row=new_row, finish_or_begin=finish_or_begin, uncertainty_index=uncertainty_index, store_routes_for_another_k=store_routes_for_another_k, store_R_pool_for_another_k=store_R_pool_for_another_k, duration=duration, congestion_link=congestion_link,
                congestion_node=congestion_node,index=index)
        else:
            obj_1_vehicle = best_position_1_vehicle(R, no_route_barge, no_route_truck, hash_table_1v_all_fail,
                                                    hash_table_1v_all, routes, fixed_vehicles_percentage, Fixed, K,
                                                    hash_table_1v, hash_table_1v_fail, has_end_depot, R_i, i, k, Trans,
                                                    Trans_Tp, Trans_Td, Trans_secondTp,
                                                    Trans_secondTd, random_position,segment_in_dynamic=segment_in_dynamic)
        if obj_1_vehicle:
            obj_list.append(obj_1_vehicle)
    #lost_r()
    if obj_list:
        #lost_r()
        # if i in [31, 46, 50, 65]:
        #     print('wfw')
        obj_df_one_column = pd.DataFrame(obj_list, columns=['one_column'])
        obj_df = pd.DataFrame(obj_df_one_column['one_column'].values.tolist(),
                              columns=['k', 'original_route', 'original_route_no_columns', 'cost_inserted_request',
                                       'dict_a_request_best_position'])
        obj_df = obj_df.values
        obj_best = obj_df[np.argmin(obj_df[:,3],axis=0),:]
        # print(obj_best)

        best_k,original_route,original_route_no_columns,cost_inserted_request,dict_a_request_best_position = obj_best
        # print(best_k,routes[best_k])
        key = get_key_1k(R_i, original_route_no_columns, best_k, fixed_vehicles_percentage, Fixed, K)

        routes_copy = my_deepcopy(routes)
        routes_copy[best_k] = copy.copy(hash_table_1v_all[key][dict_a_request_best_position]['route'])
        request_list2 = list(original_route[4])
        request_list2 = get_request_list2(request_list2,i,Trans,Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd,dict_a_request_best_position)
        # print('1513', routes[best_k], routes_copy[best_k])
        routes_copy[best_k][4] = copy.copy(request_list2)
        # print('1515', routes[best_k], routes_copy[best_k])
        # if isinstance(dict_a_request_best_position,int):
        #     dict_a_request_best_position=[dict_a_request_best_position]
        if random_k == 0 and random_position == 0:
            hash_top[top_key] = {}
            hash_top[top_key]['obj_best_k'], hash_top[top_key]['route_best_k'], hash_top[top_key]['position'], \
            hash_top[top_key]['cost_inserted_request'] = copy.copy(best_k), copy.copy(
                routes_copy[best_k]), copy.copy(dict_a_request_best_position), cost_inserted_request

        #lost_r()
        # print('1522',routes[best_k],routes_copy[best_k])
        #check_capacity(routes)
        return best_k, routes_copy[best_k], dict_a_request_best_position, cost_inserted_request
    #lost_r()
    check_served_R()
    if random_k == 0 and random_position == 0:
        hash_top[top_key] = {}
        hash_top[top_key]['obj_best_k'], hash_top[top_key]['route_best_k'], hash_top[top_key]['position'], \
        hash_top[top_key]['cost_inserted_request'] = -1, 0, 0, 0
    return -1, 0, 0, 0


# @profile()
# @time_me()
def insert1vehicle(R_i, i, K, Trans, Trans_Tp, Trans_Td, regret=0, segment_in_dynamic = 0, check_uncertainty_in_insertion_by_RL_or_not=0, new_row=-1, finish_or_begin=-1, uncertainty_index=-1, store_routes_for_another_k=-1, store_R_pool_for_another_k=-1,duration=-1, congestion_link=-1,
                                 congestion_node=-1,index=-1):
    obj_best_k, route_best_k, position, cost_inserted_request = insert1vehicle_base(R_i, i, K, Trans, Trans_Tp,
                                                                                    Trans_Td, regret=0, segment_in_dynamic = segment_in_dynamic, check_uncertainty_in_insertion_by_RL_or_not=check_uncertainty_in_insertion_by_RL_or_not, new_row=new_row, finish_or_begin=finish_or_begin, uncertainty_index=uncertainty_index, store_routes_for_another_k=store_routes_for_another_k, store_R_pool_for_another_k=store_R_pool_for_another_k, duration=duration, congestion_link=congestion_link,
                congestion_node=congestion_node,index=index)
    # print('1540',obj_best_k, route_best_k)
    return obj_best_k, route_best_k, position, cost_inserted_request


# @profile()
# @time_me()
def random_insert1vehicle(R_i, i, K, Trans, Trans_Tp, Trans_Td, regret=0):
    random_position = 1
    obj_best_k, route_best_k, position, cost_inserted_request = insert1vehicle_base(R_i, i, K, Trans, Trans_Tp,
                                                                                    Trans_Td, random_position, regret=0)
    return obj_best_k, route_best_k, position, cost_inserted_request


# @profile()
# @time_me()
##@jit
# this function is used to get the best position for both k, and hash_table_2v_all contains k, but hash_table_2v doesn't contain k
def insert2vehicle_T(no_route_barge, no_route_truck, has_end_depot, routes, i, T_change, k1, k2, R_i,
                     original_route_no_columns1, original_route_no_columns2, key, Trans,
                     obj_list, random_position, hash_table_2v_all_fail, hash_table_2v_all, R_pool_2v, R, hash_table_1v,
                     hash_table_1v_fail, hash_table_1v_all, hash_table_1v_all_fail, request_flow_t,
                     fixed_vehicles_percentage, Fixed, K, check_uncertainty_in_insertion_by_RL_or_not=0, new_row=-1, finish_or_begin=-1, uncertainty_index=-1, store_routes_for_another_k=-1, store_R_pool_for_another_k=-1,duration=-1, congestion_link=-1,
                                 congestion_node=-1,index=-1):
    global relevant_request_position_number
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    Trans_secondTp, Trans_secondTd = 0, 0
    if key in hash_table_2v_all_fail.keys():
        if T_change in hash_table_2v_all_fail.keys():
            if parallel == 0 and parallel_thread == 0:
                return obj_list
            else:
                return obj_list, 'nothing', 0, 0, 0

    if key in hash_table_2v_all.keys():
        if T_change in hash_table_2v_all[key].keys():
            if list(hash_table_2v_all[key][T_change]):
                obj_list.append([T_change, list(hash_table_2v_all[key][T_change])[0],
                                 hash_table_2v_all[key][T_change][list(hash_table_2v_all[key][T_change])[0]][
                                     'cost_inserted_request']])
    else:
        if Demir == 1:
            if k1 in [0,1,2] or k2 in [0,1,2]:
                index_r = list(R[:, 7]).index(i)
                T_k_record[index_r,2] = k1
                T_k_record[index_r, 3] = k2
        seg_r_tuple1 = tuple(zip(R_pool_2v[R_i][T_change][0], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
        seg_r_tuple2 = tuple(zip(R_pool_2v[R_i][T_change][1], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))

        R_i_list = []
        for item in seg_r_tuple1:
            R_i_list.append(list(item))

        R_i_list[4][0] = 'no'
        R_i_list[5][0] = 'no'

        R_i_back = []
        for item in R_i_list:
            R_i_back.append(tuple(item))
        seg_r_tuple1 = tuple(R_i_back)

        R_i_list2 = []
        for item in seg_r_tuple2:
            R_i_list2.append(list(item))

        R_i_list2[2][0] = 'no'
        R_i_list2[3][0] = 'no'

        R_i_back2 = []
        for item in R_i_list2:
            R_i_back2.append(tuple(item))
        seg_r_tuple2 = tuple(R_i_back2)
        key1 = get_key_1k(seg_r_tuple1, original_route_no_columns1, k1, fixed_vehicles_percentage, Fixed, K)
        key2 = get_key_1k(seg_r_tuple2, original_route_no_columns2, k2, fixed_vehicles_percentage, Fixed, K)

        if key1 in hash_table_1v:
            new_k1 = k1
            new_try1 = copy.copy(hash_table_1v[key1][list(hash_table_1v[key1])[0]]['route'])
            position1 = copy.copy(list(hash_table_1v[key1])[0])
            insert_r_cost1 = hash_table_1v[key1][list(hash_table_1v[key1])[0]]['cost_inserted_request']
        else:
            Trans_Tp = 0
            Trans_Td = 1
            #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
            obj_1_vehicle = best_position_1_vehicle(R, no_route_barge, no_route_truck, hash_table_1v_all_fail,
                                                    hash_table_1v_all, routes, fixed_vehicles_percentage, Fixed, K,
                                                    hash_table_1v, hash_table_1v_fail, has_end_depot, seg_r_tuple1, i,
                                                    k1, Trans, Trans_Tp, Trans_Td, Trans_secondTp,
                                                    Trans_secondTd, random_position, check_uncertainty_in_insertion_by_RL_or_not=check_uncertainty_in_insertion_by_RL_or_not, new_row=new_row, finish_or_begin=finish_or_begin, uncertainty_index=uncertainty_index, store_routes_for_another_k=store_routes_for_another_k, store_R_pool_for_another_k=store_R_pool_for_another_k, duration=duration, congestion_link=congestion_link,
                congestion_node=congestion_node,index=index)
            #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
            if obj_1_vehicle:
                new_k1 = k1
                if random_position == 0:
                    new_try1 = copy.copy(hash_table_1v[key1][list(hash_table_1v[key1])[0]]['route'])
                    position1 = copy.copy(list(hash_table_1v[key1])[0])
                    insert_r_cost1 = hash_table_1v[key1][list(hash_table_1v[key1])[0]]['cost_inserted_request']
                else:
                    position1 = get_best_position(obj_1_vehicle)
                    new_try1 = copy.copy(hash_table_1v_all[key1][position1]['route'])
                    insert_r_cost1 = hash_table_1v_all[key1][position1]['cost_inserted_request']
        if 'new_k1' in locals():
            index_r = list(R[:, 7]).index(i)
            new_k1_copy = copy.copy(new_k1)
            del new_k1
            if str(i) + "delivery" in new_try1[4]:
                new_try1[4, list(new_try1[4]).index(str(i) + "delivery")] = str(i) + "Td"
                # new_try1.rename(columns={str(i) + "delivery": str(i) + "Td"})
                
                request_flow_t[index_r,1] = request_flow_t[index_r,5]
            #if there is a T, the second k will be influenced by the first k, for example, the storage time, waiting time. So the time constraints and objective need to be recalculated,
            if key2 in hash_table_1v:

                new_k2 = k2
                new_try2 = hash_table_1v[key2][list(hash_table_1v[key2])[0]]['route']
                position2 = copy.copy(list(hash_table_1v[key2])[0])
                # insert_r_cost2 = hash_table_1v[key2][list(hash_table_1v[key2])[0]]['cost_inserted_request']
                relevant_request_position_number = {}
                check_start_position = position2[0]
                bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, new_k2, new_try2, i)
                if isinstance(bool_or_route, bool):
                    if parallel == 0 and parallel_thread == 0:
                        return obj_list
                    else:
                        return obj_list, 'nothing', 0, 0, 0
                else:
                    new_try2 = bool_or_route

                insert_r_cost2 = objective_value_i(i, new_k2, new_try2)[0]
                if new_try2[2, position2[0]] >= request_flow_t[index_r,1]:
                    pass
                else:

                    Trans_Tp = 1
                    Trans_Td = 0
                    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                    obj_1_vehicle = best_position_1_vehicle(R, no_route_barge, no_route_truck, hash_table_1v_all_fail,
                                                            hash_table_1v_all, routes, fixed_vehicles_percentage, Fixed,
                                                            K, hash_table_1v, hash_table_1v_fail, has_end_depot,
                                                            seg_r_tuple2, i, k2, Trans, Trans_Tp, Trans_Td,
                                                            Trans_secondTp, Trans_secondTd, random_position, check_uncertainty_in_insertion_by_RL_or_not=check_uncertainty_in_insertion_by_RL_or_not, new_row=new_row, finish_or_begin=finish_or_begin, uncertainty_index=uncertainty_index, store_routes_for_another_k=store_routes_for_another_k, store_R_pool_for_another_k=store_R_pool_for_another_k, duration=duration, congestion_link=congestion_link,
                congestion_node=congestion_node,index=index)
                    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                    if obj_1_vehicle:

                        new_k2 = k2
                        if random_position == 0:
                            new_try2 = copy.copy(hash_table_1v[key2][list(hash_table_1v[key2])[0]]['route'])
                            position2 = copy.copy(list(hash_table_1v[key2])[0])
                            insert_r_cost2 = hash_table_1v[key2][list(hash_table_1v[key2])[0]]['cost_inserted_request']
                        else:
                            position2 = get_best_position(obj_1_vehicle)
                            new_try2 = copy.copy(hash_table_1v_all[key2][position2]['route'])
                            insert_r_cost2 = hash_table_1v_all[key2][position2]['cost_inserted_request']
            else:
                Trans_Tp = 1
                Trans_Td = 0
                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                obj_1_vehicle = best_position_1_vehicle(R, no_route_barge, no_route_truck, hash_table_1v_all_fail,
                                                        hash_table_1v_all, routes, fixed_vehicles_percentage, Fixed, K,
                                                        hash_table_1v, hash_table_1v_fail, has_end_depot, seg_r_tuple2,
                                                        i, k2, Trans, Trans_Tp, Trans_Td, Trans_secondTp,
                                                        Trans_secondTd, random_position, check_uncertainty_in_insertion_by_RL_or_not=check_uncertainty_in_insertion_by_RL_or_not, new_row=new_row, finish_or_begin=finish_or_begin, uncertainty_index=uncertainty_index, store_routes_for_another_k=store_routes_for_another_k, store_R_pool_for_another_k=store_R_pool_for_another_k, duration=duration, congestion_link=congestion_link,
                congestion_node=congestion_node,index=index)
                if obj_1_vehicle:
                    new_k2 = k2
                    if random_position == 0:
                        new_try2 = copy.copy(hash_table_1v[key2][list(hash_table_1v[key2])[0]]['route'])
                        position2 = copy.copy(list(hash_table_1v[key2])[0])
                        insert_r_cost2 = hash_table_1v[key2][list(hash_table_1v[key2])[0]]['cost_inserted_request']
                    else:
                        position2 = get_best_position(obj_1_vehicle)
                        new_try2 = copy.copy(hash_table_1v_all[key2][position2]['route'])
                        insert_r_cost2 = hash_table_1v_all[key2][position2]['cost_inserted_request']
            #check every possible solution's preference constraints in case it is ued in regret insertion
            if 'new_k2' in locals() and heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
                update_request_flow_t(new_try1)
                update_request_flow_t(new_try2)
                satisfy_preference = preference_constraints(i, new_k1_copy, new_k2, -1, new_try1, new_try2, -1)
                if satisfy_preference == 1:
                    if preference_relevant(new_k1_copy, new_try1, i) != 1:
                        satisfy_preference = 0
                    else:
                        if preference_relevant(new_k2, new_try2, i) != 1:
                            satisfy_preference = 0
                if satisfy_preference == 0:
                    del new_k2
            if 'new_k2' in locals():

                positions = [position1, position2]
                positions = tuple(positions)
                bottom = {'k1': new_k1_copy, 'k2': new_k2,
                          'route1': copy.copy(new_try1),
                          'route2': copy.copy(new_try2),
                          'cost_inserted_request1': insert_r_cost1,
                          'cost_inserted_request2': insert_r_cost2,
                          'cost_inserted_request': insert_r_cost1 + insert_r_cost2}
                if parallel == 0 and parallel_thread == 0:
                    if key not in hash_table_2v_all.keys():
                        hash_table_2v_all[key] = {}
                    hash_table_2v_all[key][T_change] = {}

                    hash_table_2v_all[key][T_change][positions] = bottom

                del new_k2
                obj_list.append([T_change, positions, insert_r_cost1 + insert_r_cost2])
                if parallel == 0 and parallel_thread == 0:
                    return obj_list
                else:
                    return obj_list, key, T_change, positions, bottom
            else:
                if parallel == 0 and parallel_thread == 0:
                    if key not in hash_table_2v_all_fail.keys():
                        hash_table_2v_all_fail[key] = {}
                    hash_table_2v_all_fail[key][T_change] = {}
                else:
                    return obj_list, key, T_change, 0, 0
        else:
            if parallel == 0 and parallel_thread == 0:
                if key not in hash_table_2v_all_fail.keys():
                    hash_table_2v_all_fail[key] = {}
                hash_table_2v_all_fail[key][T_change] = {}
            else:
                return obj_list, key, T_change, 0, 0
    if parallel == 0 and parallel_thread == 0:
        return obj_list
    else:
        return obj_list, 'nothing', 0, 0, 0

def convert_T(T_change):
    if isinstance(T_change,list):
        if len(T_change) == 1:
            T_change = T_change[0]
        else:
            if len(T_change) == 2:
                T_change = tuple(T_change)
    return T_change

# @profile()
# @time_me()
##@jit
# obj_list is the solution with best position
# obj_list_best_T contains all solutions with different T
def insert2vehicle_k(parallel, no_route_barge, no_route_truck, has_end_depot, i, R_i, T_change, k1, k2,
                     fixed_vehicles_percentage, K, Fixed, obj_list_best_T_local_local, Trans, random_position, routes,
                     hash_table_2v_fail, hash_table_2v, hash_table_2v_all_fail, hash_table_2v_all, R_pool_2v, R,
                     hash_table_1v, hash_table_1v_fail, hash_table_1v_all, hash_table_1v_all_fail, request_flow_t, check_uncertainty_in_insertion_by_RL_or_not=0, new_row=-1, finish_or_begin=-1, uncertainty_index=-1, store_routes_for_another_k=-1, store_R_pool_for_another_k=-1,duration=-1, congestion_link=-1,
                                 congestion_node=-1,index=-1):
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    original_route_no_columns1 = route_no_columns(routes[k1])
    original_route_no_columns2 = route_no_columns(routes[k2])
    fix_k1_ap, fix_k1_bp = get_fix_k_0_ap(k1, fixed_vehicles_percentage, Fixed)
    fix_k2_ap, fix_k2_bp = get_fix_k_0_ap(k2, fixed_vehicles_percentage, Fixed)
    T_change = convert_T(T_change)
    key = (T_change, R_i, original_route_no_columns1, K[k1, 0], K[k1, 1], fix_k1_ap, fix_k1_bp, original_route_no_columns2, K[k2, 0], K[k2, 1], fix_k2_ap, fix_k2_bp)
    hash_table_2v_all_fail_local_local = {}
    hash_table_2v_all_local_local = {}

    best_cost_inserted_request = 99999999999999

    if key in hash_table_2v_fail.keys():
        if parallel == 1 or parallel_thread == 1:
            return obj_list_best_T_local_local, best_cost_inserted_request, hash_table_2v_all_local_local, hash_table_2v_all_fail_local_local, 'nothing', 0, 0, 0
        else:
            return obj_list_best_T_local_local, best_cost_inserted_request

    #check_capacity(routes)
    if key in hash_table_2v.keys():
        linshi = hash_table_2v[key]
        best_T = list(linshi.keys())[0]
        #                best_k1 = linshi[best_T][list(linshi[best_T])[0]]['k1']
        #                best_k2 = linshi[best_T][list(linshi[best_T])[0]]['k2']
        #                best_route1 = linshi[best_T][list(linshi[best_T])[0]]['route1']
        #                best_route2 = linshi[best_T][list(linshi[best_T])[0]]['route2']
        best_postions = list(linshi[best_T])[0]
        best_cost_inserted_request = linshi[best_T][list(linshi[best_T])[0]]['cost_inserted_request']
        obj_list_best_T_local_local.append([k1, k2, best_T, best_postions, best_cost_inserted_request])

    else:
        obj_list = []
        if parallel == 0 and parallel_thread == 0:
            obj_list = insert2vehicle_T(no_route_barge, no_route_truck, has_end_depot, routes, i, T_change, k1, k2, R_i,
                                        original_route_no_columns1, original_route_no_columns2,
                                        key, Trans, obj_list, random_position, hash_table_2v_all_fail,
                                        hash_table_2v_all, R_pool_2v, R, hash_table_1v, hash_table_1v_fail,
                                        hash_table_1v_all, hash_table_1v_all_fail, request_flow_t,
                                        fixed_vehicles_percentage, Fixed, K, check_uncertainty_in_insertion_by_RL_or_not=check_uncertainty_in_insertion_by_RL_or_not, new_row=new_row, finish_or_begin=finish_or_begin, uncertainty_index=uncertainty_index, store_routes_for_another_k=store_routes_for_another_k, store_R_pool_for_another_k=store_R_pool_for_another_k, duration=duration, congestion_link=congestion_link,
                congestion_node=congestion_node,index=index)
        else:
            obj_list, key, T_change, positions, bottom = insert2vehicle_T(no_route_barge, no_route_truck, has_end_depot,
                                                                          routes, i, T_change, k1, k2, R_i,
                                                                          original_route_no_columns1,
                                                                          original_route_no_columns2,
                                                                          key, Trans, obj_list, random_position,
                                                                          hash_table_2v_all_fail, hash_table_2v_all,
                                                                          R_pool_2v, R, hash_table_1v,
                                                                          hash_table_1v_fail, hash_table_1v_all,
                                                                          hash_table_1v_all_fail, request_flow_t,
                                                                          fixed_vehicles_percentage, Fixed, K)
            if not isinstance(key, str):
                # key not 'nothing', which means do nothing
                if isinstance(bottom, int):
                    # fail
                    if key not in hash_table_2v_all_fail_local_local.keys():
                        hash_table_2v_all_fail_local_local[key] = {}
                    hash_table_2v_all_fail_local_local[key][T_change] = {}
                else:
                    if key not in hash_table_2v_all_local_local.keys():
                        hash_table_2v_all_local_local[key] = {}
                    hash_table_2v_all_local_local[key][T_change] = {}
                    hash_table_2v_all_local_local[key][T_change][positions] = bottom
        # because it only get the best position of both k, so there is at most only one feasible solution, and also store it to hash_table_2v
        if obj_list:
            #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
            obj_df = pd.DataFrame(obj_list, columns=['T', 'positions', 'cost_inserted_request'])
            obj_best = obj_df.loc[obj_df['cost_inserted_request'] == obj_df['cost_inserted_request'].min()]
            best_T, best_postions, best_cost_inserted_request = obj_best.iloc[0]

            if random_position == 0 and parallel == 0 and parallel_thread == 0:
                hash_table_2v[key] = {}
                hash_table_2v[key][best_T] = {}
                hash_table_2v[key][best_T][obj_best.iloc[0]['positions']] = copy.copy(
                    hash_table_2v_all[key][best_T][obj_best.iloc[0]['positions']])

            obj_list_best_T_local_local.append([k1, k2, best_T, best_postions, best_cost_inserted_request])
            obj_list = []
            if random_position == 0 and (parallel == 1 or parallel_thread == 1):
                return obj_list_best_T_local_local, best_cost_inserted_request, hash_table_2v_all_local_local, hash_table_2v_all_fail_local_local, key, best_T, \
                       obj_best.iloc[0]['positions'], hash_table_2v_all_local_local[key][best_T][
                           obj_best.iloc[0]['positions']]
        else:
            if random_position == 0 and parallel == 0 and parallel_thread == 0:
                hash_table_2v_fail[key] = {}
            if random_position == 0 and (parallel == 1 or parallel_thread == 1):
                return obj_list_best_T_local_local, best_cost_inserted_request, hash_table_2v_all_local_local, hash_table_2v_all_fail_local_local, key, T_change, 0, 0
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    if parallel == 0 and parallel_thread == 0:
        return obj_list_best_T_local_local, best_cost_inserted_request
    else:
        return obj_list_best_T_local_local, best_cost_inserted_request, hash_table_2v_all_local_local, hash_table_2v_all_fail_local_local, 'random_or_nothing', 0, 0, 0


# @profile()
# @time_me()
def insert2vehicle_best(obj_list_best_T, R_i, i):
    global routes, R_pool, check_start_position, relevant_request_position_number
    k1, k2 = -1, -1
    best_T, best_positions = -1, -1
    # if i in [31, 46, 50, 65]:
    #     print('wfw')
    if obj_list_best_T:
        # obj_list_best_T should add k1 and k2, and the name of k's columns should be changed
        # obj_df_best_T = pd.DataFrame(obj_list_best_T,
        #                              columns=['k1', 'k2', 'T', 'best_positions', 'cost_inserted_request'])
        obj_df_best_T = np.array(obj_list_best_T)
        obj_best_T = obj_df_best_T[np.argmin(obj_df_best_T[:,4],axis=0)]
        k1, k2, best_T, best_positions, cost_inserted_request = obj_best_T

        #        original_route1 = my_deepcopy(routes[k1])
        #        original_route_no_columns1 = my_deepcopy(routes[k1])
        # #       original_route_no_columns1.columns = list(range(len(original_route_no_columns1.columns)))
        #        original_route_no_columns1 = [tuple(x) for x in original_route_no_columns1.to_records(index=False)]
        #        original_route_no_columns1.append(tuple(routes[k1][4]))
        #        original_route_no_columns1 = tuple(original_route_no_columns1)
        original_route_no_columns1 = route_no_columns(routes[k1])
        original_route_no_columns2 = route_no_columns(routes[k2])
        #        original_route2 = my_deepcopy(routes[k2])
        #        original_route_no_columns2 = my_deepcopy(original_route2)
        # #       original_route_no_columns2.columns = list(range(len(original_route_no_columns2.columns)))
        #        original_route_no_columns2 = [tuple(x) for x in original_route_no_columns2.to_records(index=False)]
        #        original_route_no_columns2.append(tuple(original_route2.columns))
        #        original_route_no_columns2 = tuple(original_route_no_columns2)
        fix_k1_ap, fix_k1_bp = get_fix_k_0_ap(k1, fixed_vehicles_percentage, Fixed)
        fix_k2_ap, fix_k2_bp = get_fix_k_0_ap(k2, fixed_vehicles_percentage, Fixed)

        best_T = convert_T(best_T)
        key = (
            best_T, R_i, original_route_no_columns1, K[k1, 0], K[k1, 1], fix_k1_ap, fix_k1_bp, original_route_no_columns2, K[k2, 0], K[k2, 1], fix_k2_ap, fix_k2_bp)
        # print(key)
        try:
            linshi = hash_table_2v_all[key]
        except:
            sys.exit('check point 118')
        best_route1 = linshi[best_T][best_positions]['route1']
        best_route2 = linshi[best_T][best_positions]['route2']

        request_list_first = list(routes[k1][4])
        request_list_first.insert(best_positions[0][0], str(i) + 'pickup')
        request_list_first.insert(best_positions[0][1], str(i) + 'Td')
              
        request_list_second = list(routes[k2][4])
        request_list_second.insert(best_positions[1][0], str(i) + 'Tp')
        request_list_second.insert(best_positions[1][1], str(i) + 'delivery')

        best_route1[4] = request_list_first
        best_route2[4] = request_list_second
        #check_satisfy_constraints()
        routes_save = my_deepcopy(routes)

        # because the changes in each route may influent aother one, so need to do this relevant check
        # k: A B C_ D
        # l: A_ B_ C E F
        # k+: A B G C_ D -- G->C_ C_->C,E,F checked in first check. Similiar reason with the follow one, need check again
        # l+: G_ A_ B_ C E F -- G_->A_ B_ C E F, and C->C_, D checked in first check. But, in the first check, there is no G in k. so it needs to be checked again when G is inserted to k.
        check_start_position = best_positions[0][0]

        relevant_request_position_number = {}
        bool_or_route1, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k1, best_route1,i)
        if isinstance(bool_or_route1, bool):
            k1, k2 = -1, -1
            return k1, k2, routes_save, R_pool, [0]
        #check_relevant_try_not_in_routes()
        relevant_try_copy1 = my_deepcopy(relevant_try)

        layer,aaa = 0,0
        final_ok_or1 = solve_relevant_try(relevant_try_copy1,layer,aaa)

        check_start_position = best_positions[1][0]

        relevant_request_position_number = {}
        bool_or_route2, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k2, best_route2,i)
        if isinstance(bool_or_route2, bool):
            k1, k2 = -1, -1
            return k1, k2, routes_save, R_pool, [0]
        relevant_try_copy2 = my_deepcopy(relevant_try)
        layer,aaa = 0,0
        final_ok_or2 = solve_relevant_try(relevant_try_copy2,layer,aaa)
        satisfy_preference = 1
        if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
            update_request_flow_t(best_route1)
            update_request_flow_t(best_route2)
            satisfy_preference = preference_constraints(i, k1, k2, -1, best_route1, best_route2, -1)
            if satisfy_preference == 1:
                # when check preference, only the inserted r is checked, other r' in k has not be checked, and r may influce r', so the k itself should also be checked
                relevant_try_copy1[k1] = [copy.copy(best_route1), i, 0]
                layer, aaa = 0, 0
                preference_final_ok_or1 = solve_relevant_try(relevant_try_copy1, layer, aaa, 1)
                if preference_final_ok_or1 == 1:
                    # when check preference, only the inserted r is checked, other r' in k has not be checked, and r may influce r', so the k itself should also be checked
                    relevant_try_copy2[k2] = [copy.copy(best_route2), i, 0]
                    layer, aaa = 0, 0
                    preference_final_ok_or2 = solve_relevant_try(relevant_try_copy2, layer, aaa, 1)
                    if preference_final_ok_or2 != 1:
                        satisfy_preference = 0
                else:
                    satisfy_preference = 0
        if final_ok_or1 == 0 or final_ok_or2 == 0 or (heterogeneous_preferences == 1 and satisfy_preference == 0 and heterogeneous_preferences_no_constraints == 0):
            k1, k2 = -1, -1
            return k1, k2, routes_save, R_pool, [0]
        else:
            #check_satisfy_constraints()
            routes[k1] = copy.copy(best_route1)
            routes[k2] = copy.copy(best_route2)

            R_pool = R_pool[~(R_pool[:, 7] == i)]
            #check_satisfy_constraints()
            #lost_r()
    return k1, k2, routes, R_pool, [best_T]

def add_delay_unit_penalty(R):
    if Demir != 1:
        c_delay_list = []
        for request_number in R[:, 7]:
            index_r = list(R[:, 7]).index(request_number)
            if R[index_r, 5] - R[index_r, 2] < 30:
                c_delay_list.append(100)
            else:
                if R[index_r, 5] - R[index_r, 2] < 54:
                    c_delay_list.append(70)

                else:
                    c_delay_list.append(50)
        # R['c_delay'] = c_delay_list
        R = np.append(R, np.c_[c_delay_list], axis=1)
    return R
def read_R_K(request_number_in_R, what='all'):

    Data = pd.ExcelFile(data_path)
    if what == 'K' or what == 'revert_K':
        K = pd.read_excel(Data, 'K')
        K = K.set_index('K')
        if what == 'revert_K':
            revert_K = dict(zip(K.index, range(len(K))))
            return revert_K
        K = K.values
        return K
    if what == 'all' or 'noR_pool':
        R = pd.read_excel(Data, 'R_'+str(request_number_in_R))
        revert_r = R['p'][0]

        if isinstance(revert_r, str):
            names = revert_names('str')
        else:
            names = revert_names('int')
        R['p'] = R['p'].map(names).fillna(R['p'])
        R['d'] = R['d'].map(names).fillna(R['d'])
        R.insert(7, 'r', range(len(R)))
        R['p'] = R['p'].astype('int')
        R['d'] = R['d'].astype('int')
        R = R.values

        # change name of r to carrier00request_number
        if by_wenjing != 1:
            for index in range(len(R)):
                R[index, 7] = R[index, 7] + big_r * parallel_number

        R = add_delay_unit_penalty(R)
        if heterogeneous_preferences == 1:

            R_info = pd.read_excel(Data, 'R_'+str(request_number_in_R) + '_info')

            R_info = R_info.values
            #add R_info to R
            R = np.hstack((R,R_info))
            # if Demir != 1:
            #     revert_r = R['p'][0].item()
            # else:
            #danger here I didn't change values in R after I add R_info to R, so
            if fuzzy_probability == 1:
                R_info = R_info.astype('float64')
                for info in range(len(R_info)):
                    #level 1 is the highest level
                    #add 0.3 on cost/emission without storage/delay/waiting
                    #cost
                    if R_info[info, 0] == 1:
                        R_info[info, 0] = 0.28
                    else:
                        if R_info[info, 0] == 2:
                            R_info[info, 0] = 0.43
                        else:
                            if R_info[info, 0] == 3:
                                R_info[info, 0] = 0.75
                            else:
                                #no requirement on this attribute
                                R_info[info, 0] = 100000
                    #speed, minus 5 km/h for each mode
                    if R_info[info, 1] == 1:
                        R_info[info, 1] = 70
                    else:
                        if R_info[info, 1] == 2:
                            R_info[info, 1] = 40
                        else:
                            if R_info[info, 1] == 3:
                                R_info[info, 1] = 10
                            else:
                                # no requirement on this attribute
                                R_info[info, 1] = 0
                    #delay

                    #transshipment

                    #emissions
                    if R_info[info, 4] == 1:
                        R_info[info, 4] = 0.25
                    else:
                        if R_info[info, 4] == 2:
                            R_info[info, 4] = 0.33
                        else:
                            if R_info[info, 4] == 3:
                                R_info[info, 4] = 0.91
                            else:
                                # no requirement on this attribute
                                R_info[info, 4] = 100000
        else:
            R_info = -1
        if CP == 1:

            # add carrier label
            l = np.array(np.empty(shape=(len(R), 1)),dtype=int)
            l[:] = parallel_number
            R = np.append(R, l, axis=1)

            #danger here if time window is float, then this is wrong
            # R = R.astype(int)
        R_pool = R.copy()
        K = pd.read_excel(Data, 'K')
        K = K.set_index('K')
        K = K.values
        if what == 'noR_pool':
            return R, K
        if what == 'all':
            return R, K, R_pool


def read_D(what, K):
    if different_companies == 1:
        D_path = "C:\Intermodal\Case study\CP\EGS Contargo CTT\D_all_companies_full.xlsx"
        # if parallel_number == 1:
        #     D_path = "/data/yimeng/Case study/Preferences/D_EGS - 10r.xlsx"
        # elif parallel_number == 2:
        #     D_path = "C:\Intermodal\Case study\CP\EGS Contargo CTT\D_Contargo.xlsx"
        # else:
        #     D_path = "C:\Intermodal\Case study\CP\EGS Contargo CTT\D_CTT.xlsx"
    else:
        if Demir == 1:
            D_path = "/data/yimeng/Case study/Demir/D_Demir - 5r.xlsx"
        else:
            D_path = "A:/MYpython/34959_RL/D_EGS - 10r.xlsx"
    D_origin_barge = pd.read_excel(D_path, 'Barge')
    D_origin_train = pd.read_excel(D_path, 'Train')
    D_origin_truck = pd.read_excel(D_path, 'Truck')

    D_origin_barge = D_origin_barge.set_index('N')
    D_origin_train = D_origin_train.set_index('N')
    D_origin_truck = D_origin_truck.set_index('N')

    D_origin_barge = D_origin_barge.values
    D_origin_train = D_origin_train.values
    D_origin_truck = D_origin_truck.values
    # N_origin=N_origin.set_index('N')
    #    N = N.set_index('N')
    #    T = T.set_index('T')
    #    T_all = T_all.set_index('T_all')
    D = {}
    for k in range(len(K)):
        if K[k, 5] == 1:
            D[k] = D_origin_barge.copy()
        else:
            if K[k, 5] == 2:
                D[k] = D_origin_train.copy()
            else:
                D[k] = D_origin_truck.copy()
    delay_TRE = 0
    if dynamic == 1 and dynamic_t == 65 and delay_TRE == 1:
        D[1][1,7] = 100000000000

    if what == 'D':
        return D
    D_origin_All = pd.read_excel(D_path, 'All')
    # if different_companies == 0:
    D_origin_All = D_origin_All.set_index('N')
    D_origin_All = D_origin_All.values
    if what == 'D_All':
        return D, D_origin_All
    if what == 'all':
        return D, D_origin_All, D_origin_barge, D_origin_train, D_origin_truck


def read_no_route():
    if Demir == 1:
        Barge_no_land_path = "/data/yimeng/Case study/Demir/Barge_no_land_Demir.xlsx"
    else:
        Barge_no_land_path = "A:/MYpython/34959_RL/Barge_no_land.xlsx"
    no_route_barge = pd.read_excel(Barge_no_land_path, 'Barge')
    no_route_truck = pd.read_excel(Barge_no_land_path, 'Truck')
    names = revert_names()
    no_route_barge['O'] = no_route_barge['O'].map(names).fillna(no_route_barge['O'])
    no_route_barge['D'] = no_route_barge['D'].map(names).fillna(no_route_barge['D'])
    no_route_barge = no_route_barge.values
    no_route_truck = no_route_truck.values
    return no_route_barge, no_route_truck


def read_Fixed(request_number_in_R, percentage, Fixed=None):
    if Fixed == None:
        if different_companies == 1:
            if parallel_number == 1:
                fixed_data_path = 'A:/MYpython/34959_RL/Fixed_right_real.xlsx'
            elif parallel_number == 2:
                fixed_data_path = "C:\Intermodal\Case study\CP\EGS Contargo CTT\Fixed_services_Contargo.xlsx"
            elif parallel_number == 3:
                fixed_data_path = "C:\Intermodal\Case study\CP\EGS Contargo CTT\Fixed_services_HSL.xlsx"
            else:
                fixed_data_path = "C:\Intermodal\Case study\CP\EGS Contargo CTT\Fixed_services_Merge.xlsx"
        else:
            if Demir == 1:
                fixed_data_path = "/data/yimeng/Case study/Demir/Fixed_Demir.xlsx"
            else:
                fixed_data_path = 'A:/MYpython/34959_RL/Fixed_right_real.xlsx'
        Fixed_Data = pd.ExcelFile(fixed_data_path)
        Fixed = pd.read_excel(Fixed_Data, None)
        revert_Fixed = Fixed['FixedK']['FixedK'][0]

        if not isinstance(revert_Fixed, int):
            revert_K = read_R_K(request_number_in_R, what='revert_K')
            for k in Fixed['FixedK']['FixedK']:
                try:
                    Fixed[revert_K[k]] = Fixed.pop(k)
                except:
                    pass
            Fixed['FixedK']['FixedK'] = Fixed['FixedK']['FixedK'].map(revert_K).fillna(Fixed['FixedK']['FixedK'])
        names = revert_names()
        for k in list(Fixed.keys())[1:]:

            Fixed[k]['p'] = Fixed[k]['p'].map(names).fillna(Fixed[k]['p'])
            Fixed[k] = Fixed[k].values
        if Demir == 1 and Demir_barge_free == 1:
            for k in [0,1,2]:
                Fixed[k][0][1:3]=[0,10000]
                Fixed[k][1][1:3] = [0, 10000]
        return Fixed
    fixed_vehicles = Fixed['FixedK']['FixedK'].tolist()
    if isinstance(percentage, list):
        fixed_vehicles_percentage = fixed_vehicles[int(percentage[0] * len(fixed_vehicles)):int(percentage[1] * len(fixed_vehicles))]
    else:
        fixed_vehicles_percentage = fixed_vehicles[int(percentage * len(fixed_vehicles)):]
    return fixed_vehicles_percentage


def fixed_data(request_number_in_R, percentage):
    parallel, fuel_cost, c_storage, initial_solution_no_wait_cost, insert_multiple_r, b1, b2, b3, b4, b5, b6, b7, b8, b9, b10, alpha, belta, transshipment_time, service_time, truck_time_free, has_end_depot, Trans, random_position = [
        1, 0, 1, 0, 1, 0, 5, 7, 9, 13, 13, 17, 19, 21, 24, 2, 1.5, 1, 1, 1, 1, 1, 0,
    ]
    Fixed = read_Fixed(request_number_in_R, percentage)
    no_route_barge, no_route_truck = read_no_route()
    #danger here is wrong in CP, because R's name is not got from range(len(R)) after the initial optimization
    R, K = read_R_K(request_number_in_R, 'noR_pool')
    D = read_D('D', K)
    return parallel, fuel_cost, c_storage, initial_solution_no_wait_cost, insert_multiple_r, b1, b2, b3, b4, b5, b6, b7, b8, b9, b10, alpha, belta, transshipment_time, service_time, truck_time_free, has_end_depot, Trans, random_position, Fixed, no_route_barge, no_route_truck, R, D, K


# @profile()
def parallel_insert2vehicle_k_loop(args):
    global no_route_barge, no_route_truck, has_end_depot, i, R_i, T_change, k1, k2, fixed_vehicles_percentage, K,Fixed, obj_list_best_T_local_local, Trans, random_position, routes, hash_table_2v_fail, hash_table_2v,hash_table_2v_all_fail, hash_table_2v_all, R_pool_2v, R, hash_table_1v, hash_table_1v_fail, hash_table_1v_all,hash_table_1v_all_fail, request_flow_t
    times = timeit.default_timer()
    if parallel_thread != 1:
        global exp_number, parallel_number, hash_df_table, parallel, Fixed, fuel_cost, c_storage, initial_solution_no_wait_cost, T_k_record, insert_multiple_r, b1, b2, b3, b4, b5, b6, b7, b8, b9, b10, alpha, belta, R, transshipment_time, service_time, has_end_depot, fixed_vehicles_percentage, K, truck_time_free, request_flow_t, D

        x, exp_number, parallel_number, all_ok_k_pair, T_change = args
        request_number_in_R = load_obj(['request_number_in_R'])[0]
        percentage, hash_df_table, T_k_record, i, R_i, delete_K_pair, routes, request_flow_t, hash_table_2v_fail, hash_table_2v, hash_table_2v_all_fail, hash_table_2v_all, R_pool_2v, hash_table_1v, hash_table_1v_fail, hash_table_1v_all, hash_table_1v_all_fail = load_obj(
            ['percentage', 'hash_df_table', 'T_k_record', 'i', 'R_i',
             'delete_K_pair', 'routes',
             'request_flow_t', 'hash_table_2v_fail', 'hash_table_2v', 'hash_table_2v_all_fail', 'hash_table_2v_all',
             'R_pool_2v',
             'hash_table_1v', 'hash_table_1v_fail', 'hash_table_1v_all', 'hash_table_1v_all_fail'])

        parallel, fuel_cost, c_storage, initial_solution_no_wait_cost, insert_multiple_r, b1, b2, b3, b4, b5, b6, b7, b8, b9, b10, alpha, belta, transshipment_time, service_time, truck_time_free, has_end_depot, Trans, random_position, Fixed, no_route_barge, no_route_truck, R, D, K = fixed_data(
            request_number_in_R, percentage)

        fixed_vehicles_percentage = read_Fixed(request_number_in_R, percentage, Fixed)
    else:
        i, R_i, Trans, random_position = load_obj(['i', 'R_i', 'Trans', 'random_position'])
        x, exp_number, parallel_number, all_ok_k_pair, T_change = args

    obj_list_best_T_local_local = []
    # print(all_ok_k_pair,x)
    k1,k2 = all_ok_k_pair[x,:]
    # try:
    obj_list_best_T_local_local, best_cost_inserted_request, hash_table_2v_all_local_local, hash_table_2v_all_fail_local_local, key, best_T, best_position, bottom = insert2vehicle_k(
        parallel, no_route_barge, no_route_truck, has_end_depot, i, R_i, T_change, k1, k2, fixed_vehicles_percentage, K,
        Fixed, obj_list_best_T_local_local, Trans, random_position, routes, hash_table_2v_fail, hash_table_2v,
        hash_table_2v_all_fail, hash_table_2v_all, R_pool_2v, R, hash_table_1v, hash_table_1v_fail, hash_table_1v_all,
        hash_table_1v_all_fail, request_flow_t)
    # except:
    #     print('sf')
    #     sys.exit(-1)
    # (no_route_barge,no_route_truck,has_end_depot,i, R_i, T_change, k1, k2, obj_list_best_T, Trans, random_position,routes,K,hash_table_2v_fail,hash_table_2v,fixed_vehicles_percentage,Fixed,hash_table_2v_all_fail,hash_table_2v_all,R_pool_2v,R,hash_table_1v,hash_table_1v_fail,hash_table_1v_all,hash_table_1v_all_fail,request_flow_t)
    print('parallel_insert2vehicle_k_loop', timeit.default_timer() - times)
    # print(obj_list_best_T_local)
    return obj_list_best_T_local_local, key, best_T, best_position, bottom, hash_table_2v_all_local_local, hash_table_2v_all_fail_local_local


# @profile()
def parallel_insert2vehicle_T_loop(args):
    # global K,R,transshipment_time,service_time,max_processors,has_end_depot,fixed_vehicles_percentage,K,truck_time_free,request_flow_t,D
    global exp_number, parallel_number
    times = timeit.default_timer()
    T_change, exp_number, parallel_number = args

    # request_number_in_R = load_obj(['request_number_in_R'])[0]
    # parallel, fuel_cost, c_storage, initial_solution_no_wait_cost, insert_multiple_r, b1, b2, b3, b4, b5, b6, b7, b8, b9, b10, alpha, belta, transshipment_time, service_time, truck_time_free, has_end_depot, Trans, random_position, Fixed, no_route_barge, no_route_truck, R, D, K = fixed_data(request_number_in_R)
    delete_K_pair = load_obj(['delete_K_pair'])[0]
    # all_ok_k_pair = np.array(delete_K_pair[T_change])
    all_ok_k_pair = np.array(delete_K_pair[T_change])
    obj_list_best_T_local = []
    hash_table_2v_fail_local = {}
    hash_table_2v_local = {}
    hash_table_2v_all_fail_local = {}
    hash_table_2v_all_local = {}
    if len(all_ok_k_pair) >= 1:
        parallel_nested = 0
        if parallel_nested == 1:
            iterate_what = []

            for x in range(len(all_ok_k_pair)):
                iterate_what.append([x, exp_number, parallel_number, all_ok_k_pair, T_change])
            # save_obj([all_ok_k_pair, T_change],['all_ok_k_pair', 'T_change'])
            with ProcessPoolExecutor() as e:
                results = e.map(parallel_insert2vehicle_k_loop, iterate_what)

            for result in results:
                obj_list_best_T_local_local, key, best_T, best_position, bottom, hash_table_2v_all_local_local, hash_table_2v_all_fail_local_local = result
                obj_list_best_T_local.extend(obj_list_best_T_local_local)
                hash_table_2v_all_fail_local.update(hash_table_2v_all_fail_local_local)
                hash_table_2v_all_local.update(hash_table_2v_all_local_local)
                # print('locallocal',hash_table_2v_all_local_local.keys())
                # print('local', hash_table_2v_all_local.keys())
                if not isinstance(key, str):
                    # not 'random' or 'nothing'
                    if isinstance(bottom, int):
                        hash_table_2v_fail_local[key] = {}
                    else:
                        hash_table_2v_local[key] = {}
                        hash_table_2v_local[key][best_T] = {}
                        hash_table_2v_local[key][best_T][best_position] = copy.copy(bottom)
                        # print(obj_list_best_T_local)
        else:
            not_consider_ks = []
            request_number_in_R = load_obj(['request_number_in_R'])[0]
            parallel = 1
            percentage, hash_df_table, T_k_record, i, R_i, delete_K_pair, routes, request_flow_t, hash_table_2v_fail, hash_table_2v, hash_table_2v_all_fail, hash_table_2v_all, R_pool_2v, hash_table_1v, hash_table_1v_fail, hash_table_1v_all, hash_table_1v_all_fail = load_obj(
                ['percentage', 'hash_df_table', 'T_k_record', 'i', 'R_i',
                 'delete_K_pair', 'routes',
                 'request_flow_t', 'hash_table_2v_fail', 'hash_table_2v', 'hash_table_2v_all_fail', 'hash_table_2v_all',
                 'R_pool_2v',
                 'hash_table_1v', 'hash_table_1v_fail', 'hash_table_1v_all', 'hash_table_1v_all_fail'])

            parallel, fuel_cost, c_storage, initial_solution_no_wait_cost, insert_multiple_r, b1, b2, b3, b4, b5, b6, b7, b8, b9, b10, alpha, belta, transshipment_time, service_time, truck_time_free, has_end_depot, Trans, random_position, Fixed, no_route_barge, no_route_truck, R, D, K = fixed_data(
                request_number_in_R, percentage)

            fixed_vehicles_percentage = read_Fixed(request_number_in_R, percentage, Fixed)

            for x in range(len(all_ok_k_pair)):
                k1,k2 = all_ok_k_pair[x,:]

                if [k1, k2, T_change] not in not_consider_ks:

                    obj_list_best_T, best_cost_inserted_request = insert2vehicle_k(parallel, no_route_barge,
                                                                                   no_route_truck,
                                                                                   has_end_depot, i, R_i, T_change, k1,
                                                                                   k2, fixed_vehicles_percentage, K,
                                                                                   Fixed, obj_list_best_T, Trans,
                                                                                   random_position, routes,
                                                                                   hash_table_2v_fail, hash_table_2v,
                                                                                   hash_table_2v_all_fail,
                                                                                   hash_table_2v_all, R_pool_2v, R,
                                                                                   hash_table_1v, hash_table_1v_fail,
                                                                                   hash_table_1v_all,
                                                                                   hash_table_1v_all_fail,
                                                                                   request_flow_t)
                    #20201209 comment this because it may lose some solution which is good for current r but bad for all r, because it will be better to use the used k to serve other r
                    if best_cost_inserted_request != 99999999999999:
                        # find the not_consider_ks only when there is a better cost
                        if best_cost_inserted_request < best_cost:
                            best_cost = best_cost_inserted_request
                            not_consider_ks = reduce_ks(i, k1, k2, T_change, best_cost_inserted_request, all_ok_TK,
                                                        delete_K_pair,
                                                        not_consider_ks)
    print('parallel_insert2vehicle_T_loop', timeit.default_timer() - times)
    return obj_list_best_T_local, hash_table_2v_local, hash_table_2v_fail_local, hash_table_2v_all_fail_local, hash_table_2v_all_local


# @profile()
# @time_me()
##@jit
def insert2vehicle(i, K, random_k, random_position, check_uncertainty_by_RL = 0, check_uncertainty_in_insertion_by_RL_or_not=0, new_row=-1, finish_or_begin=-1, uncertainty_index=-1, store_routes_for_another_k=-1, store_R_pool_for_another_k=-1,duration=-1, congestion_link=-1,
                                 congestion_node=-1,index = -1):
    # two vehicles
    global routes, R_pool, max_processors,request_flow_t
    # if i in [31,46,50,65]:
    #     print('wfw')
    # if i in [46,50]:
    #     print('afs')
    # times = timeit.default_timer()
    #check_capacity(routes)
    check_served_R()
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    index_r = list(R[:, 7]).index(i)
    top_key = 0
    if random_k == 0 and random_position == 0:

        routes_tuple = get_routes_tuple(routes)
        # R_pool_tuple = tuple(.to_records)
        top_key = (i, routes_tuple, 'insert2vehicle')
        if top_key in hash_top.keys():
            print('top')
            routes = hash_top[top_key]['routes']
            R_pool = hash_top[top_key]['R_pool']
            best_T = hash_top[top_key]['best_T']
            request_flow_t = hash_top[top_key]['request_flow_t']
            return my_deepcopy(routes), copy.copy(R_pool), top_key, hash_top[top_key]['k'], best_T

    Trans = 1
    obj_list_best_T = []
    R_i = tuple(zip(R[index_r], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
    delete_K_pair = delete_k(i)
    not_consider_ks = []
    # sort_T = pd.DataFrame(columns=['sort_T_d'], index=all_ok_TK[i].keys())
    T_1 = []
    for T_change in all_ok_TK[i].keys():
        if isinstance(T_change, int):
            T_1.append(T_change)
    sort_T = np.empty(shape=(len(T_1),2),dtype='object')
    sort_T[:,1] = T_1

    for T_index in range(len(T_1)):
        T_change = list(all_ok_TK[i].keys())[T_index]
        sort_T[T_index,0] = D_origin_All[R[index_r, 0]][T_change] + D_origin_All[T_change][R[index_r, 1]]
        # sort_T = sort_T.sort_values(by=['sort_T_d'])

    sort_T = sort_T[np.argsort(sort_T[:, 0])]
    best_cost = 99999999999999

    # parallel_thread = 1
    if parallel == 1 or parallel_thread == 1:
        if parallel_thread == 1:
            save_obj([i,R_i,Trans,random_position],['i','R_i','Trans','random_position'])
        else:
            save_obj(
                [request_number_in_R, percentage, hash_df_table, T_k_record, i, R_i, delete_K_pair, routes, request_flow_t,
                 hash_table_2v_fail, hash_table_2v, hash_table_2v_all_fail, hash_table_2v_all, R_pool_2v, hash_table_1v,
                 hash_table_1v_fail, hash_table_1v_all, hash_table_1v_all_fail],
                ['request_number_in_R', 'percentage', 'hash_df_table', 'T_k_record', 'i', 'R_i', 'delete_K_pair', 'routes',
                 'request_flow_t', 'hash_table_2v_fail', 'hash_table_2v', 'hash_table_2v_all_fail', 'hash_table_2v_all',
                 'R_pool_2v', 'hash_table_1v', 'hash_table_1v_fail', 'hash_table_1v_all', 'hash_table_1v_all_fail'])
        parallel_nested = 0
        if parallel_nested == 1:
            iterate_what = []
            for T_change in sort_T[:,1]:
                iterate_what.append([T_change, exp_number, parallel_number])
            # time_s = timeit.default_timer()
            # import test_parallel_question
            # test_parallel_question.main()
            # print(timeit.default_timer()-time_s)
            # time_s = timeit.default_timer()
            with ProcessPoolExecutor() as e:
                results = e.map(parallel_insert2vehicle_T_loop, iterate_what)
            # print(timeit.default_timer() - time_s)
            for result in results:
                obj_list_best_T_local, hash_table_2v_local, hash_table_2v_fail_local, hash_table_2v_all_fail_local, hash_table_2v_all_local = result
                # only obj_list_best_T should be in a loop, obj_list is just used as temporily storage
                obj_list_best_T.extend(obj_list_best_T_local)
                hash_table_2v.update(hash_table_2v_local)
                hash_table_2v_fail.update(hash_table_2v_fail_local)
                hash_table_2v_all.update(hash_table_2v_all_local)

                # print('local',hash_table_2v_all_local.keys())
                # print(hash_table_2v_all.keys())
                hash_table_2v_all_fail.update(hash_table_2v_all_fail_local)
            # print(obj_list_best_T)
        else:

            for T_change in sort_T[:,1]:

                obj_list_best_T_local = []
                all_ok_k_pair = np.array(delete_K_pair[T_change])
                if len(all_ok_k_pair) >= 1:
                    if parallel_thread == 1:
                        iterate_what = []
                        for x in range(len(all_ok_k_pair)):
                            iterate_what.append([x, exp_number, parallel_number, all_ok_k_pair, T_change])
                        with ThreadPoolExecutor() as e:
                            results = e.map(parallel_insert2vehicle_k_loop, iterate_what)
                    else:
                        iterate_what = []
                        for x in range(len(all_ok_k_pair)):
                            iterate_what.append([x, exp_number, parallel_number, all_ok_k_pair, T_change])
                        # save_obj([all_ok_k_pair, T_change],['all_ok_k_pair', 'T_change'])
                        with ProcessPoolExecutor() as e:
                            results = e.map(parallel_insert2vehicle_k_loop, iterate_what)

                    for result in results:
                        obj_list_best_T_local_local, key, best_T, best_position, bottom, hash_table_2v_all_local_local, hash_table_2v_all_fail_local_local = result
                        obj_list_best_T_local.extend(obj_list_best_T_local_local)
                        hash_table_2v_all_fail.update(hash_table_2v_all_fail_local_local)
                        hash_table_2v_all.update(hash_table_2v_all_local_local)
                        # print('locallocal',hash_table_2v_all_local_local.keys())
                        # print('local', hash_table_2v_all_local.keys())
                        if not isinstance(key, str):
                            # not 'random' or 'nothing'
                            if isinstance(bottom, int):
                                hash_table_2v_fail[key] = {}
                            else:
                                hash_table_2v[key] = {}
                                hash_table_2v[key][best_T] = {}
                                hash_table_2v[key][best_T][best_position] = copy.copy(bottom)
    else:
        for T_change in sort_T[:,1]:
            all_ok_k_pair = np.array(delete_K_pair[T_change])
            if len(all_ok_k_pair) >= 1:

                for x in range(len(all_ok_k_pair)):

                    k1,k2 = all_ok_k_pair[x,:]

                    if [k1, k2, T_change] not in not_consider_ks:
                        #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                        if dynamic == 1:
                            if check_a_k_in_dynamic(1, i, k1, k2, T_change) == -1:
                                continue
                        obj_list_best_T, best_cost_inserted_request = insert2vehicle_k(parallel, no_route_barge,
                                                                                       no_route_truck, has_end_depot, i,
                                                                                       R_i, T_change, k1, k2,
                                                                                       fixed_vehicles_percentage, K,
                                                                                       Fixed, obj_list_best_T, Trans,
                                                                                       random_position, routes,
                                                                                       hash_table_2v_fail,
                                                                                       hash_table_2v,
                                                                                       hash_table_2v_all_fail,
                                                                                       hash_table_2v_all, R_pool_2v, R,
                                                                                       hash_table_1v,
                                                                                       hash_table_1v_fail,
                                                                                       hash_table_1v_all,
                                                                                       hash_table_1v_all_fail,
                                                                                       request_flow_t, check_uncertainty_in_insertion_by_RL_or_not=check_uncertainty_in_insertion_by_RL_or_not, new_row=new_row, finish_or_begin=finish_or_begin, uncertainty_index=uncertainty_index, store_routes_for_another_k=store_routes_for_another_k, store_R_pool_for_another_k=store_R_pool_for_another_k, duration=duration, congestion_link=congestion_link,
                congestion_node=congestion_node,index=index)
                        #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                        #20201209 comment this because it may lose some solution which is good for current r but bad for all r, because it will be better to use the used k to serve other r
                        if best_cost_inserted_request != 99999999999999:
                            # find the not_consider_ks only when there is a better cost
                            if best_cost_inserted_request < best_cost:
                                best_cost = best_cost_inserted_request
                                not_consider_ks = reduce_ks(i, k1, k2, T_change, best_cost_inserted_request, all_ok_TK,
                                                            delete_K_pair,
                                                            not_consider_ks)
    #check_capacity(routes)
    k1, k2, routes, R_pool, best_T = insert2vehicle_best(obj_list_best_T, R_i, i)
    #check_capacity(routes)
    if random_k == 0 and random_position == 0:
        hash_top[top_key] = {}
        hash_top[top_key]['routes'] = routes
        hash_top[top_key]['R_pool'] = R_pool
        hash_top[top_key]['k'] = [k1, k2]
        hash_top[top_key]['best_T'] = best_T
        hash_top[top_key]['request_flow_t'] = copy.copy(request_flow_t)
    # print('insert2vehicle', timeit.default_timer() - times)
    check_served_R()
    return my_deepcopy(routes), copy.copy(R_pool), top_key, [k1, k2], best_T


def random_k_insert2vehicle(i, K, random_k, random_position):
    # if I only have two vehicle, I will not do the compare
    global routes, R_pool
    index_r = list(R[:, 7]).index(i)
    # two vehicles
    v_has_r = [-1,-1,-1]
    used_T = [-1,-1]
    best_T = [-1]
    Trans = 1

    obj_list_best_T = []
    R_i = tuple(zip(R[index_r], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
    delete_K_pair = delete_k(i)
    for T_change in all_ok_TK[i].keys():
        if not isinstance(T_change, int):
            continue
        all_ok_k_pair = np.array(delete_K_pair[T_change])
        if len(all_ok_k_pair) >= 1:
            a = len(R_pool)
            x = random.choice(range(len(all_ok_k_pair)))
            k1,k2 = all_ok_k_pair[x,:]

            obj_list_best_T, best_cost_inserted_request = insert2vehicle_k(parallel, no_route_barge, no_route_truck,
                                                                           has_end_depot, i, R_i, T_change, k1, k2,
                                                                           fixed_vehicles_percentage, K, Fixed,
                                                                           obj_list_best_T, Trans, random_position,
                                                                           routes, hash_table_2v_fail, hash_table_2v,
                                                                           hash_table_2v_all_fail, hash_table_2v_all,
                                                                           R_pool_2v, R, hash_table_1v,
                                                                           hash_table_1v_fail, hash_table_1v_all,
                                                                           hash_table_1v_all_fail, request_flow_t)
            k1, k2, routes, R_pool, best_T = insert2vehicle_best(obj_list_best_T, R_i, i)
            b = len(R_pool)
            if a != b:
                v_has_r[0:2] = k1, k2
                used_T[0] = int(best_T[0])
                break

    return routes, R_pool, 0, v_has_r, used_T


def update_hash_top(random_k,random_position,top_key,routes, R_pool,k1, k2, k3,best_T1, best_T2):
    if random_k == 0 and random_position == 0:
        hash_top[top_key] = {}
        hash_top[top_key]['routes'] = routes
        hash_top[top_key]['R_pool'] = R_pool
        hash_top[top_key]['k'] = [k1, k2, k3]
        hash_top[top_key]['best_T'] = [best_T1, best_T2]
        hash_top[top_key]['request_flow_t'] = copy.copy(request_flow_t)
# @profile()
# @time_me()
##@jit
def insert3vehicle(i, K, random_k, random_position):
    global routes, R_pool, check_start_position, relevant_request_position_number, request_flow_t
    index_r = list(R[:, 7]).index(i)
    if two_T == 0:
        return routes, R_pool, 0, [-1,-1,-1], [-1,-1]
    else:
        pass
    # if i in [14]:
    #     print('wfsa')
    top_key = 0
    if random_k == 0 and random_position == 0:

        routes_tuple = get_routes_tuple(routes)
        # R_pool_tuple = tuple(.to_records)
        top_key = (i, routes_tuple, 'insert3vehicle')
        if top_key in hash_top.keys():
            routes = hash_top[top_key]['routes']
            R_pool = hash_top[top_key]['R_pool']
            best_T = hash_top[top_key]['best_T']
            ks = hash_top[top_key]['k']
            request_flow_t = hash_top[top_key]['request_flow_t']
            return my_deepcopy(routes), copy.copy(R_pool), top_key, ks, copy.copy(best_T)
    # three vehicles
    Trans = 1
    obj_list = []
    # obj_list_best_T1 = []
    # obj_list_best_T2 = []

    R_i = tuple(zip(R[index_r], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
    delete_K_pair = delete_k(i)
    # danger delete_K_pair[T_change] maybe empty
    if len(all_ok_TK[i].keys()) >= 2:
        # all_ok_k_pair = np.array(delete_K_pair[T_change])
        for T_change in all_ok_TK[i].keys():
            if isinstance(T_change, int):
                continue
            all_ok_k_pair = np.array(delete_K_pair[T_change])
            if len(all_ok_k_pair) >= 1:
                for x in range(len(all_ok_k_pair)):
                    T_change1, T_change2 = T_change
                    k1,k2,k3 = all_ok_k_pair[x,:]
                    
                    if Demir == 1:
                        if k1 in [0, 1, 2] or k2 in [0, 1, 2] or k3 in [0, 1, 2]:
                            T_k_record[index_r, 2] = k1
                            T_k_record[index_r, 3] = k2
                            T_k_record[index_r, 4] = k3
                    original_route_no_columns1 = route_no_columns(routes[k1])
                    original_route_no_columns2 = route_no_columns(routes[k2])
                    original_route_no_columns3 = route_no_columns(routes[k3])

                    fix_k1_ap, fix_1k_bp = get_fix_k_0_ap(k1,
                                                                                        fixed_vehicles_percentage,
                                                                                        Fixed)
                    fix_k2_ap, fix_k2_bp = get_fix_k_0_ap(k2,
                                                                                        fixed_vehicles_percentage,
                                                                                        Fixed)
                    fix_k3_ap, fix_k3_bp = get_fix_k_0_ap(k3,
                                                                                        fixed_vehicles_percentage,
                                                                                        Fixed)

                    # T_change = convert_T(T_change)
                    # T_change2 = convert_T(T_change2)
                    key = (
                        R_i, original_route_no_columns1, K[k1, 0], K[k1, 1], fix_k1_ap, fix_k1_bp, original_route_no_columns2, K[k2, 0], K[k2, 1],
                        fix_k2_ap, fix_k2_bp, original_route_no_columns3,
                        K[k3, 0], K[k3, 1], fix_k3_ap, fix_k3_bp,
                        T_change1, T_change2)

                    if key in hash_table_3v.keys():
                       routes[k1] = hash_table_3v[key][0]
                       routes[k2] = hash_table_3v[key][1]
                       routes[k3] = hash_table_3v[key][2]
                       R_pool=R_pool[~(R_pool[:,7]==i)]
                       for k in [k1,k2,k3]:
                           bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot,routes,K,k,routes[k],i)
                           if isinstance(bool_or_route, bool):
                               k1, k2, k3 = -1, -1, -1
                               v_has_r = [k1, k2, k3]
                               return routes, R_pool, top_key, v_has_r, [-1, -1]
                       update_hash_top(random_k, random_position, top_key, routes, R_pool, k1, k2, k3, T_change1, T_change2)
                       return routes, R_pool, top_key, [k1,k2,k3], [T_change1, T_change2]
                    seg_r_tuple1 = tuple(
                        zip(R_pool_3v[R_i][tuple([T_change1, T_change2])][0],
                            ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
                    Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd = 0, 1, 0, 0
                    obj_vehicle_1 = best_position_1_vehicle(R, no_route_barge, no_route_truck,
                                                            hash_table_1v_all_fail, hash_table_1v_all,
                                                            routes, fixed_vehicles_percentage, Fixed, K,
                                                            hash_table_1v, hash_table_1v_fail,
                                                            has_end_depot, seg_r_tuple1, i, k1, Trans,
                                                            Trans_Tp, Trans_Td,
                                                            Trans_secondTp, Trans_secondTd, random_position)
                    if obj_vehicle_1:
                        seg_r_tuple2 = tuple(
                            zip(R_pool_3v[R_i][tuple([T_change1, T_change2])][1],
                                ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
                        Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd = 1, 0, 0, 1
                        obj_vehicle_2 = best_position_1_vehicle(R, no_route_barge, no_route_truck,
                                                                hash_table_1v_all_fail, hash_table_1v_all,
                                                                routes, fixed_vehicles_percentage, Fixed, K,
                                                                hash_table_1v, hash_table_1v_fail,
                                                                has_end_depot, seg_r_tuple2, i, k2, Trans,
                                                                Trans_Tp,
                                                                Trans_Td, Trans_secondTp, Trans_secondTd,
                                                                random_position)
                        if obj_vehicle_2:
                            seg_r_tuple3 = tuple(
                                zip(R_pool_3v[R_i][tuple([T_change1, T_change2])][2],
                                    ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
                            Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd = 0, 0, 1, 0
                            obj_vehicle_3 = best_position_1_vehicle(R, no_route_barge, no_route_truck,
                                                                    hash_table_1v_all_fail,
                                                                    hash_table_1v_all, routes,
                                                                    fixed_vehicles_percentage, Fixed, K,
                                                                    hash_table_1v, hash_table_1v_fail,
                                                                    has_end_depot, seg_r_tuple3, i, k3,
                                                                    Trans, Trans_Tp,
                                                                    Trans_Td, Trans_secondTp,
                                                                    Trans_secondTd, random_position)
                            if obj_vehicle_3:
                                key1 = (
                                    seg_r_tuple1, obj_vehicle_1[0][2], K[k1, 0], K[k1, 1],
                                    fix_k1_0_ap,
                                    fix_k1_0_bp, fix_k1_1_ap, fix_k1_1_bp)
                                key2 = (
                                    seg_r_tuple2, obj_vehicle_2[0][2], K[k2, 0], K[k2, 1],
                                    fix_k2_0_ap,
                                    fix_k2_0_bp, fix_k2_1_ap, fix_k2_1_bp)
                                key3 = (
                                    seg_r_tuple3, obj_vehicle_3[0][2], K[k3, 0], K[k3, 1],
                                    fix_k3_0_ap,
                                    fix_k3_0_bp, fix_k3_1_ap, fix_k3_1_bp)

                                obj_3_vehilces = obj_vehicle_1[0][3] + obj_vehicle_2[0][3] + \
                                                 obj_vehicle_3[0][3]

                                obj_list.append([key, T_change1, T_change2, k1, k2, k3, key1, key2, key3,
                                                 obj_3_vehilces])
                                if parallel == 0 and parallel_thread == 0:
                                    position1, position2, position3 = list(hash_table_1v[key1])[0],list(hash_table_1v[key2])[0],list(hash_table_1v[key3])[0]

                                    bottom = {'k1': k1, 'k2': k2, 'k3': k3,

                                              'cost_inserted_request1': obj_vehicle_1[0][3],
                                              'cost_inserted_request2': obj_vehicle_2[0][3],
                                              'cost_inserted_request3': obj_vehicle_3[0][3],
                                              'cost_inserted_request': obj_vehicle_1[0][3] + obj_vehicle_2[0][3] + obj_vehicle_3[0][3]}

                                    positions = tuple([position1, position2, position3])

                                    if key not in hash_table_3v_all.keys():
                                        hash_table_3v_all[key] = {}
                                    hash_table_3v_all[key][T_change] = {}
                                    hash_table_3v_all[key][T_change][positions] = bottom


    best_T1, best_T2 = -1, -1
    v_has_r = [-1,-1,-1]
    if obj_list:
        obj_df = pd.DataFrame(obj_list, columns=['key', 'T1', 'T2', 'k1', 'k2', 'k3', 'key1', 'key2', 'key3',
                                                 'cost_inserted_request'])
        obj_best = obj_df.loc[obj_df['cost_inserted_request'] == obj_df['cost_inserted_request'].min()]
        key = obj_best.iloc[0]['key']
        key1 = obj_best.iloc[0]['key1']
        key2 = obj_best.iloc[0]['key2']
        key3 = obj_best.iloc[0]['key3']

        k1 = obj_best.iloc[0]['k1']
        k2 = obj_best.iloc[0]['k2']
        k3 = obj_best.iloc[0]['k3']

        best_T1 = obj_best.iloc[0]['T1']
        best_T2 = obj_best.iloc[0]['T2']

        new_try1 = copy.copy(hash_table_1v[key1][list(hash_table_1v[key1])[0]]['route'])
        new_try2 = copy.copy(hash_table_1v[key2][list(hash_table_1v[key2])[0]]['route'])
        new_try3 = copy.copy(hash_table_1v[key3][list(hash_table_1v[key3])[0]]['route'])

        positions1 = list(hash_table_1v[key1])[0]
        positions2 = list(hash_table_1v[key2])[0]
        positions3 = list(hash_table_1v[key3])[0]

        request_list_first = list(routes[k1][4])
        request_list_first.insert(positions1[0], str(i) + 'pickup')
        request_list_first.insert(positions1[1], str(i) + 'Td')

        request_list_second = list(routes[k2][4])
        request_list_second.insert(positions2[0], str(i) + 'Tp')
        request_list_second.insert(positions2[1], str(i) + 'secondTd')

        request_list_third = list(routes[k3][4])
        request_list_third.insert(positions3[0], str(i) + 'secondTp')
        request_list_third.insert(positions3[1], str(i) + 'delivery')

        new_try1[4] = request_list_first
        new_try2[4] = request_list_second
        new_try3[4] = request_list_third

        routes_save = my_deepcopy(routes)

        # relevant_request_position_number = {}

        check_start_position = positions1[0]

        relevant_request_position_number = {}
        bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k1, new_try1,i)
        if isinstance(bool_or_route, bool):
            k1, k2, k3 = -1, -1, -1
            v_has_r = [k1, k2, k3]
            return routes_save, R_pool, top_key, v_has_r, [best_T1, best_T2]
        relevant_try_copy = my_deepcopy(relevant_try)
        layer,aaa = 0,0
        final_ok_or1 = solve_relevant_try(relevant_try_copy,layer,aaa)

        check_start_position = positions2[0]

        relevant_request_position_number = {}
        bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k2, new_try2,i)
        if isinstance(bool_or_route, bool):
            k1, k2, k3 = -1, -1, -1
            v_has_r = [k1, k2, k3]
            return routes_save, R_pool, top_key, v_has_r, [best_T1, best_T2]
        relevant_try_copy = my_deepcopy(relevant_try)
        layer,aaa = 0,0
        final_ok_or2 = solve_relevant_try(relevant_try_copy,layer,aaa)

        check_start_position = positions3[0]

        relevant_request_position_number = {}
        bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k3, new_try3,i)
        if isinstance(bool_or_route, bool):
            k1, k2, k3 = -1, -1, -1
            v_has_r = [k1, k2, k3]
            return routes_save, R_pool, top_key, v_has_r, [best_T1, best_T2]
        relevant_try_copy = my_deepcopy(relevant_try)
        layer,aaa = 0,0
        final_ok_or3 = solve_relevant_try(relevant_try_copy,layer,aaa)

        if final_ok_or1 == 0 or final_ok_or2 == 0 or final_ok_or3 == 0:
            k1, k2, k3 = -1,-1,-1
            v_has_r = [k1, k2, k3]
            return routes_save, R_pool, top_key, v_has_r, [best_T1, best_T2]
        else:
            routes[k1] = copy.copy(new_try1)
            routes[k2] = copy.copy(new_try2)
            routes[k3] = copy.copy(new_try3)
            R_pool = R_pool[~(R_pool[:, 7] == i)]
            if random_position == 0:
                hash_table_3v[key] = [new_try1, new_try2, new_try3]

        update_hash_top(random_k,random_position,top_key,routes, R_pool,k1, k2, k3,best_T1, best_T2)
        v_has_r = [k1, k2, k3]
    return my_deepcopy(routes), copy.copy(R_pool), top_key, v_has_r, [best_T1, best_T2]


# if just insert to route, then  record_1_vehicle_new_try=0, regret = 0, position=position
# @profile()
# @time_me()
def change_route(k_order, k, ok_r, all_r_cost_copy, record_1_vehicle_new_try=0, regret=1, insert_terminals=0,
                 positions=0):
    capacity_full = 0
    if k_order == 1 and k not in fixed_vehicles_percentage and regret == 1:
        return record_1_vehicle_new_try[ok_r], capacity_full
        
    route = my_deepcopy(routes[k])

    insert_position1, insert_position2 = 100000, 100000
    if regret == 1:
        insert_position1 = 1
        insert_position2 = len(routes[k][4])
        insert_terminal1 = Fixed[k][0,0]
        insert_terminal2 = Fixed[k][1,0]
    else:
        insert_terminal1 = insert_terminals[0]
        insert_terminal2 = insert_terminals[1]
        for m in range(0, len(route[4])):
            if insert_terminal1 == route[0, m]:
                insert_position1 = m + 1
                break
        # if it's the bundle insert in regret insertion, the r may haven't be inserted to route before, so use the true position from hash_table

        if insert_position1 == 100000:
            if k_order == 1:
                insert_position1 = positions[0]
            if k_order == 21:
                insert_position1 = positions[0][0]
            if k_order == 22:
                insert_position1 = positions[1][0]
            if k_order == 31:
                insert_position1 = positions[0][0]
            if k_order == 32:
                insert_position1 = positions[1][0]
            if k_order == 33:
                insert_position1 = positions[2][0]

        for m in range(0, len(route[4])):
            if insert_terminal2 == route[0, m]:
                # this position is for duichen, not first load first unload
                insert_position2 = m + 1
                break
        if insert_position2 == 100000:
            if k_order == 1:
                insert_position2 = positions[1]
            if k_order == 21:
                insert_position2 = positions[0][1]

            if k_order == 22:
                insert_position2 = positions[1][1]
            if k_order == 31:
                insert_position2 = positions[0][1]
            if k_order == 32:
                insert_position2 = positions[1][1]
            if k_order == 33:
                insert_position2 = positions[2][1]

    if k_order == 1 and (k in fixed_vehicles_percentage or regret == 0):
        insert_str = ['pickup', 'delivery']
        # danger because I didn't add T as input, so this only suitable for fixed k with only two terminals(begin and end depot)
        route = np.insert(route, insert_position1,
                          [insert_terminal1, insert_terminal1, insert_terminal1, insert_terminal1,
                           str(ok_r) + insert_str[0]], axis=1)
        route = np.insert(route, insert_position2,
                          [insert_terminal2, insert_terminal2, insert_terminal2, insert_terminal2,
                           str(ok_r) + insert_str[1]], axis=1)
    if k_order != 1:
        if k not in fixed_vehicles_percentage and regret == 1:
            index_i = list(all_r_cost_copy[:, 3]).index(ok_r)
            return my_deepcopy(hash_top[all_r_cost_copy[index_i,0]]['routes'][k]), capacity_full
        else:
            if k_order == 21 or k_order == 31:
                insert_str = ['pickup', 'Td']
            if k_order == 22:
                insert_str = ['Tp', 'delivery']
            if k_order == 32:
                insert_str = ['Tp', 'secondTd']
            if k_order == 33:
                insert_str = ['secondTp', 'delivery']
            route = np.insert(route, insert_position1,
                              [insert_terminal1, insert_terminal1, insert_terminal1, insert_terminal1,
                               str(ok_r) + insert_str[0]], axis=1)
            route = np.insert(route, insert_position2,
                              [insert_terminal2, insert_terminal2, insert_terminal2, insert_terminal2,
                               str(ok_r) + insert_str[1]], axis=1)
    #check_capacity(routes)

    if capacity_constraints(has_end_depot, K, R, k, route) == False:
        if regret == 0:
            capacity_full = 1
        return False, capacity_full
    if new_subtour_constraints(route[0]) == False:
        return False, capacity_full
    # route = hash_top[all_r_cost_copy[index_i,0]]['routes'][k]
    bool_or_route = assign_time(k, route,ok_r,insert_position1)
    if isinstance(bool_or_route, bool):
        return False, capacity_full

    else:
        return bool_or_route, capacity_full

def preference_relevant(k1, route_k1, ok_r, remove = 0, original_route_k1 = -1):
    if len(route_k1[0]) == 2:
        return 1
    if remove == 1:
        relevant_try_copy1 = get_relevant_routes(0, k1, original_route_k1, ok_r)
    else:
        relevant_try_copy1 = get_relevant_routes(0, k1, route_k1, ok_r)
    # when check preference, only the inserted r is checked, other r' in k has not be checked, and r may influce r', so the k itself should also be checked
    relevant_try_copy1[k1] = [copy.copy(route_k1), ok_r, 0]
    layer, aaa = 0, 0
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    preference_final_ok_or1 = solve_relevant_try(relevant_try_copy1, layer, aaa, 1)
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    return preference_final_ok_or1

# @profile()
# @time_me()
def insert_a_r(all_k_r, ok_r, used_k, all_r_cost_copy, record_1_vehicle_new_try, regret_values_per_k='mark', regret=1,
               insert_terminals_original=[0, 0, 0, 0], positions=0, bundle=0):
    global routes, R_pool
    #find_unchecked_r_preference([6,45])
    # only when it is regret conflict, or bundle insert in regret insertion, the ok_r is the true index
    index_r = list(R[:, 7]).index(ok_r)
    if regret == 1 or isinstance(positions, tuple):
        k1, k2, k3 = used_k[list(used_k[:,3]).index(ok_r),0:3]
    else:
        k1, k2, k3 = used_k[0,0:3]
    # if infeasible then return
    if k1 == -1 and k2 == -1 and k3 == -1:
        return all_k_r, regret_values_per_k, 0
    # if k is used then return
    if regret == 1:
        if (isinstance(k1, (int, np.integer)) and k1 not in all_k_r.columns) or (
                isinstance(k2, (int, np.integer)) and k2 not in all_k_r.columns) or (
                isinstance(k3, (int, np.integer)) and k3 not in all_k_r.columns):
            return all_k_r, regret_values_per_k, 0
    if bundle == 1 and ((isinstance(k1, (int, np.integer)) and k1 != -1 and K[k1, 5] != 1) or (
            isinstance(k2, (int, np.integer)) and k2 != -1 and K[k2, 5] != 1) or (
                                isinstance(k3, (int, np.integer)) and k3 != -1 and K[k3, 5] != 1)):
        number = int(np.random.choice([1, 2], size=(1,), p=[0.7, 0.3]))
        if number == 1:
            return all_k_r, regret_values_per_k, 0
    route_k2 = 1
    route_k3 = 1
    violate_preference_constraints = 0
    #find_unchecked_r_preference([6,45])
    if isinstance(k3, (int, np.integer)) and k3 != -1:
        insert_terminals = [insert_terminals_original[0], insert_terminals_original[1]]
        route_k1, capacity_full = change_route(31, k1, ok_r, all_r_cost_copy, 0, regret, insert_terminals, positions)
        insert_terminals = [insert_terminals_original[1], insert_terminals_original[2]]
        route_k2, capacity_full = change_route(32, k2, ok_r, all_r_cost_copy, 0, regret, insert_terminals, positions)
        insert_terminals = [insert_terminals_original[2], insert_terminals_original[3]]
        route_k3, capacity_full = change_route(33, k3, ok_r, all_r_cost_copy, 0, regret, insert_terminals, positions)
        if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
            if not (isinstance(route_k1, bool) or isinstance(route_k2, bool) or isinstance(route_k3, bool)):
                if preference_constraints(ok_r, k1, k2, k3, route_k1, route_k2, route_k3) == 0:
                    violate_preference_constraints = 1
                else:
                    preference_final_ok_or1 = preference_relevant(k1, route_k1, ok_r)
                    if preference_final_ok_or1 == 0:
                        violate_preference_constraints = 1
                    else:
    
                        preference_final_ok_or2 = preference_relevant(k2, route_k2, ok_r)
                        if preference_final_ok_or2 == 0:
                            violate_preference_constraints = 1
                        else:
    
                            preference_final_ok_or3 = preference_relevant(k3, route_k3, ok_r)
                            if preference_final_ok_or3 == 0:
                                violate_preference_constraints = 1
    if (isinstance(k2, (int, np.integer)) and k2 != -1) and not (isinstance(k3, (int, np.integer)) and k3 != -1):
        insert_terminals = [insert_terminals_original[0], insert_terminals_original[1]]
        route_k1, capacity_full = change_route(21, k1, ok_r, all_r_cost_copy, 0, regret, insert_terminals, positions)
        insert_terminals = [insert_terminals_original[1], insert_terminals_original[2]]
        route_k2, capacity_full = change_route(22, k2, ok_r, all_r_cost_copy, 0, regret, insert_terminals, positions)
        r_basic_cost = get_r_basic_cost(R[index_r, 0], R[index_r, 1], ok_r, k1, k2, insert_terminals_original[1])
        if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
            if not (isinstance(route_k1, bool) or isinstance(route_k2, bool)):
                if preference_constraints(ok_r, k1, k2, -1, route_k1, route_k2, -1) == 0:
                    violate_preference_constraints = 1
                else:
    
                    preference_final_ok_or1 = preference_relevant(k1, route_k1, ok_r)
                    if preference_final_ok_or1 == 0:
                        violate_preference_constraints = 1
                    else:
    
                        preference_final_ok_or2 = preference_relevant(k2, route_k2, ok_r)
                        if preference_final_ok_or2 == 0:
                            violate_preference_constraints = 1
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    if not (isinstance(k2, (int, np.integer)) and k2 != -1) and not (isinstance(k3, (int, np.integer)) and k3 != -1):
        insert_terminals = [insert_terminals_original[0], insert_terminals_original[1]]
        route_k1, capacity_full = change_route(1, k1, ok_r, all_r_cost_copy, record_1_vehicle_new_try, regret,
                                               insert_terminals, positions)
        r_basic_cost = get_r_basic_cost(R[index_r, 0], R[index_r, 1], ok_r, k1)
        if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
            if not isinstance(route_k1, bool):
                if preference_constraints(ok_r, k1, -1, -1, route_k1, -1, -1) == 0:
                    violate_preference_constraints = 1
                else:
                    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                    preference_final_ok_or1 = preference_relevant(k1, route_k1, ok_r)
                    if preference_final_ok_or1 == 0:
                        violate_preference_constraints = 1
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    #find_unchecked_r_preference([6,45])
    if not (isinstance(route_k1, bool) or isinstance(route_k2, bool) or isinstance(route_k3, bool) or violate_preference_constraints == 1):
        v_has_r = [k1,-1,-1]
        used_T = [-1,-1]
        old_overall_cost = overall_obj(routes)[1]
        len1 = len(R_pool)
        # print(routes[k1],route_k1)
        routes[k1] = route_k1
        if not isinstance(route_k2, (int, np.integer)) and (isinstance(k2, (int, np.integer)) and k2 != -1):
            routes[k2] = route_k2
            v_has_r[1] = k2
            for col in route_k2[4]:
                request_number_col = ''.join(filter(str.isdigit, col))
                if str(ok_r) == request_number_col:
                    used_T[0] = route_k2[0][list(route_k2[4]).index(col)]
                    break
        if not isinstance(route_k3, (int, np.integer)) and (isinstance(k3, (int, np.integer)) and k3 != -1):
            routes[k3] = route_k3
            v_has_r[2] = k3
            for col in route_k3[4]:
                request_number_col = ''.join(filter(str.isdigit, col))
                if str(ok_r) == request_number_col:
                    used_T[1] = route_k3[0][list(route_k3[4]).index(col)]
                    break
        R_pool = R_pool[~(R_pool[:, 7] == ok_r)]
        #find_unchecked_r_preference([6,45])
        #lost_r()
        # don't let the r in bundle was inserted with too much delay cost/storage cost
        if bundle == 1:
            cost_inserted_request, r_hasbeen_caculated, routes_after_removed, R_pool_after_removed = get_r_cost_in_all_routes(
                ok_r)[0:4]
            if cost_inserted_request > r_basic_cost + R[index_r, 6] * 12* c_storage - 0.1:
                routes = routes_after_removed
                R_pool = R_pool_after_removed
        update_r_best_obj_in_insertion(ok_r, len1, old_overall_cost,v_has_r,used_T)
        if isinstance(k1, (int, np.integer)) and k1 != -1 and k1 not in fixed_vehicles_percentage and regret == 1:
            all_k_r.drop(k1, axis=1, inplace=True)

        if isinstance(k2, (int, np.integer)) and k2 != -1 and k2 not in fixed_vehicles_percentage and regret == 1:
            all_k_r.drop(k2, axis=1, inplace=True)

        if isinstance(k3, (int, np.integer)) and k3 != -1 and k3 not in fixed_vehicles_percentage and regret == 1:
            all_k_r.drop(k3, axis=1, inplace=True)
    # if ok_r using these k not violate time constraints, then it is inserted to routes and will be removed from regret_values_per_k,
    # if violate time constraints, then it's infeasible, also removed from regret_values_per_k
    if not isinstance(regret_values_per_k, str) and regret == 1:
        for k in regret_values_per_k.keys():
            if ok_r in regret_values_per_k[k][:,1]:
                regret_values_per_k[k] = np.delete(regret_values_per_k[k], list(regret_values_per_k[k][:,1]).index(ok_r), axis=0)
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    return all_k_r, regret_values_per_k, capacity_full


def check_bundle_in_k(i, used_k, possible_k, k_number, key, check_bundle_r_k):
    index_r = list(R[:, 7]).index(i)
    continue_or_not = 0
    break_or_not = 0
    if possible_k not in check_bundle_r_k.keys():
        check_bundle_r_k[possible_k] = []
    index = list(used_k[:, 3]).index(i)
    if len(check_bundle_r_k[possible_k]) > 0:
        # if the OD that k served is as same as i's, then k can serve this r
        if check_bundle_r_k[possible_k][0][0] == key:
            # check capacity
            k_load = 0
            for key_and_r in check_bundle_r_k[possible_k]:
                index_r = list(R[:,7]).index(key_and_r[1])
                k_load = k_load + R[index_r, 6]
            # I neglect the r in route, but I have capacity constraint so it's fine
            if k_load + R[index_r, 6] > K[possible_k, 0]:
                continue_or_not = 1
                return break_or_not, continue_or_not, check_bundle_r_k, used_k
            else:
                # I will insert as much as r which has the same OD to k
                # so I will not use all_r_cost, because I will insert this r to route directly
                # I only need k and insert terminals
                # all_r_cost[index_i,0] =
                if k_number == 1:
                    used_k[index,0] = possible_k
                if k_number == 2:
                    used_k[index,1] = possible_k
                if k_number == 3:
                    used_k[index,2] = possible_k
                check_bundle_r_k[possible_k].append([key, i])
                break_or_not = 1
                return break_or_not, continue_or_not, check_bundle_r_k, used_k
        else:
            continue_or_not = 1
            return break_or_not, continue_or_not, check_bundle_r_k, used_k
    else:
        check_bundle_r_k[possible_k].append([key, i])
        if k_number == 1:
            used_k[index,0] = possible_k
        if k_number == 2:
            used_k[index,1] = possible_k
        if k_number == 3:
            used_k[index,2] = possible_k
        break_or_not = 1
        return break_or_not, continue_or_not, check_bundle_r_k, used_k


# @profile()
# @time_me()
##@jit
def random_insert(i):
    global routes, R_pool
    #lost_r()
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    check_served_R()
    index_r = list(R[:, 7]).index(i)
    old_overall_cost = overall_obj(routes)[1]
    len1 = len(R_pool[:, 7])
    R_i = tuple(zip(R[index_r], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
    Trans = 0
    Trans_Tp = 0
    Trans_Td = 0
    number_T = -1
    best_T, top_key = [-1], -1
    v_has_r = [-1,-1,-1]
    used_T = [-1,-1]
    k, new_try, position, insert_r_cost = random_insert1vehicle(R_i, i, K, Trans, Trans_Tp, Trans_Td)
    if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
        satisfy_preference = preference_constraints(i, k, -1, -1, new_try, -1, -1)
        if satisfy_preference == 1:
            # when check preference, only the inserted r is checked, other r' in k has not be checked, and r may influce r', so the k itself should also be checked
            relevant_try_copy = {}
            relevant_try_copy[k] = [copy.copy(new_try), i, 0]
            layer, aaa = 0, 0
            preference_final_ok_or1 = solve_relevant_try(relevant_try_copy, layer, aaa, 1)
            if preference_final_ok_or1 == 0:
                k = -1
        else:
            k = -1
    v_has_r[0] = k

    if isinstance(k, (int, np.integer)) and k != -1:
        # print(routes[k],new_try)
        routes[k] = new_try

        R_pool = R_pool[~(R_pool[:, 7] == i)]
        #lost_r()
        number_T = 0
    len2 = len(R_pool[:, 7])
    if T_or == 1:
        if len1 == len2:
            number_T = -1
            if len(K) >= 2:
                # two vehicles
                random_k = int(np.random.choice([0, 1], size=(1,), p=[5. / 10, 5. / 10]))
                random_position = 1
                if random_k != 1:
                    routes, R_pool, top_key_2k, v_has_r, best_T_2k = insert2vehicle(i, K, random_k, random_position)
                else:
                    routes, R_pool, top_key_2k, v_has_r, best_T_2k = random_k_insert2vehicle(i, K, random_k, random_position)
                len4 = len(R_pool[:, 7])
                if len4<len1:
                    number_T = 1
                    used_T[0] = best_T
        len3 = len(R_pool[:, 7])
        if len3 == len1:
            number_T = -1
            if len(K) >= 3:
                random_k = int(np.random.choice([0, 1], size=(1,), p=[5. / 10, 5. / 10]))
                random_position = 1
                if random_k != 1:
                    #danger I haven't consider randomness in insert3vehicle
                    pass
                    # routes, R_pool, top_key_3k, v_has_r, best_T_3k = insert3vehicle(i, K, random_k, random_position)

                #                else:
                #                    routes, R_pool = random_k_insert3vehicle(i, K, random_k, random_position)
                len5 = len(R_pool[:, 7])
                if len5 < len1:
                    number_T = 2
                    used_T = best_T
    key = tuple([R[index_r, 0], R[index_r, 1]])
    update_r_best_obj_in_insertion(i, len1, old_overall_cost,v_has_r,best_T)
    if bundle_or_not == 1 and number_T != -1 and len(R_pool) > 0:
        if number_T == 0:
            insert_bundle_pre(i, key, number_T, best_T, top_key, k)
        else:
            if number_T == 1:
                insert_bundle_pre(i, key, number_T, best_T_2k, top_key_2k, k)
            else:
                insert_bundle_pre(i, key, number_T, best_T_3k, top_key_3k, k)
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    check_served_R()
    return routes, R_pool


# @profile()
# @time_me()
##@jit
def transshipment_insert(i):
    global routes, R_pool, transshipment_insert_number
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    index_r = list(R[:, 7]).index(i)
    if i in no_T_R:
        return routes, R_pool
    old_overall_cost = overall_obj(routes)[1]
    transshipment_insert_number = transshipment_insert_number + 1
    len1 = len(R_pool[:, 7])
    number_T = -1
    best_T, top_key = [-1], -1
    v_has_r = [-1, -1, -1]
    used_T = [-1, -1]
    if len(K) >= 2:
        # two vehicles
        random_k = int(np.random.choice([0, 1], size=(1,), p=[5. / 10, 5. / 10]))
        random_k = 0
        #        random_position = int(np.random.choice([0, 1], size=(1,), p=[5. / 10, 5. / 10]))
        random_position = 0
        if random_k != 1:
            routes, R_pool, top_key_2k, v_has_r, best_T = insert2vehicle(i, K, random_k, random_position)
        else:
            routes, R_pool, top_key_2k, v_has_r, best_T = random_k_insert2vehicle(i, K, random_k, random_position)
        number_T = 1
        used_T[0] = best_T
    len3 = len(R_pool[:, 7])
    if len3 == len1:
        number_T = -1
        if len(K) >= 3:
            random_k = 0
            random_position = 0
            if random_k != 1:
                routes, R_pool, top_key_3k, v_has_r, best_T = insert3vehicle(i, K, random_k, random_position)
                #                else:
                #                    routes, R_pool = random_k_insert3vehicle(i, K, random_k, random_position)
                number_T = 2
                used_T = best_T
    k = -1
    len2 = len(R_pool[:, 7])
    if len2 == len1:
        R_i = tuple(zip(R[index_r], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
        Trans = 0
        Trans_Tp = 0
        Trans_Td = 0
        k, new_try, position, insert_r_cost = insert1vehicle(R_i, i, K, Trans, Trans_Tp, Trans_Td)
        if isinstance(k, (int, np.integer)) and k != -1:
            routes[k] = new_try
            R_pool = R_pool[~(R_pool[:, 7] == i)]
            #lost_r()
            number_T = 0
            v_has_r[0] = k
    key = tuple([R[index_r, 0], R[index_r, 1]])
    update_r_best_obj_in_insertion(i, len1, old_overall_cost,v_has_r,used_T)
    if bundle_or_not == 1 and len(R_pool) > 0:
        if number_T == 1:
            insert_bundle_pre(i, key, number_T, best_T, top_key_2k, k)
        else:
            if number_T == 2:
                insert_bundle_pre(i, key, number_T, best_T, top_key_3k, k)
    # I want to keep the r if ater the insertion it's obj is better or equal to best history, but I found it's function is as same as history removal
    # if better_or_not == 1:
    #     #I should keep inserted position, k, and
    #     keep_r
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    return routes, R_pool


# @profile()
# @time_me()
##@jit
def greedy_insert(i, segment_in_dynamic = 0, check_uncertainty_in_insertion_by_RL_or_not=0, new_row=-1, finish_or_begin=-1, uncertainty_index=-1, store_routes_for_another_k=-1, store_R_pool_for_another_k=-1,duration=-1, congestion_link=-1,
                                 congestion_node=-1,index=-1):
    global request_segment_in_dynamic, routes, R_pool, check_start_position, relevant_request_position_number
    # if i in [275,147,168,205,107,233,181,395,61,239,3,240,166,153,133,173,25,257,199,177,420,383,101,277,212,289,150,550,11,110,271,573,286,119,210]:
    #find_unchecked_r_preference([6,45])
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    check_served_R()
    if segment_in_dynamic == 1:
        index_r = list(request_segment_in_dynamic[:, 7]).index(i)
        len1 = len(request_segment_in_dynamic[:, 7])
        R_i = tuple(zip(request_segment_in_dynamic[index_r], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r', 'delay_penalty', 'operation']))
    else:
        index_r = list(R[:, 7]).index(i)
        len1 = len(R_pool[:, 7])
        R_i = tuple(zip(R[index_r], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
    old_overall_cost = overall_obj(routes)[1]
    Trans = 0
    Trans_Tp = 0
    Trans_Td = 0
    number_T = -1
    #lost_r()
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    k, new_try, position, insert_r_cost = insert1vehicle(R_i, i, K, Trans, Trans_Tp, Trans_Td, segment_in_dynamic = segment_in_dynamic, check_uncertainty_in_insertion_by_RL_or_not=check_uncertainty_in_insertion_by_RL_or_not, new_row=new_row, finish_or_begin=finish_or_begin, uncertainty_index=uncertainty_index, store_routes_for_another_k=store_routes_for_another_k, store_R_pool_for_another_k=store_R_pool_for_another_k, duration=duration, congestion_link=congestion_link,
                congestion_node=congestion_node,index=index)
    
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    # print('2740',k,new_try)
    #find_unchecked_r_preference([6,45])
    best_T, top_key = [-1], -1
    v_has_r = [-1,-1,-1]
    used_T = [-1,-1]
    #find_unchecked_r_preference([6,45])
    if isinstance(k, (int, np.integer)) and k != -1:
        # print(routes[k])
        # routes_save = my_deepcopy(routes)
        
        check_start_position = position[0]

        relevant_request_position_number = {}
        bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k, new_try,i)
        if isinstance(bool_or_route, bool):

            return routes, R_pool
        # print('2750', k, new_try)
        relevant_try_copy = my_deepcopy(relevant_try)
        layer,aaa = 0,0
        final_ok_or = solve_relevant_try(relevant_try_copy,layer,aaa)
        if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
            satisfy_preference = preference_constraints(i, k, -1, -1, new_try, -1, -1)
            if satisfy_preference == 1:
                # when check preference, only the inserted r is checked, other r' in k has not be checked, and r may influce r', so the k itself should also be checked
                relevant_try_copy[k] = [copy.copy(new_try), i, 0]
                layer, aaa = 0, 0
                preference_final_ok_or1 = solve_relevant_try(relevant_try_copy, layer, aaa, 1)
                if preference_final_ok_or1 == 0:
                    final_ok_or = 0
            else:
                final_ok_or = 0
        if final_ok_or == 0:
            # print('final_ok_or',final_ok_or)
            #check_capacity(routes)
            #find_unchecked_r_preference([6,45])

            return routes, R_pool
        else:
            # if stochastic == 1 and add_RL == 0 and dynamic_t > 0:
            #     # only set the success k's insertion action = 1
            #     ALNS_insertion_implementation_store[(uncertainty_index, i)][k] = store_all(0, -1, 0)
            if segment_in_dynamic == 1:
                operation_previous = R_i[9][0]
                if operation_previous == 'pickup':
                    operation = 'Td'
                elif operation_previous == 'Tp':
                    operation = 'delivery'
                elif operation_previous == 'Tp2':
                    operation = 'secondTd'
                else:
                    operation = 'delivery'
                for m in range(1, len(new_try[0])-1):
                    if get_numbers(new_try[4,m]) == i:
                        if new_getLetters(new_try[4,m]) == 'pickup':
                            new_try[4, m] = str(i-10000) + operation_previous
                        else:
                            new_try[4, m] = str(i - 10000) + operation
                #caught_strange(new_try[4])
                request_segment_in_dynamic = request_segment_in_dynamic[~(request_segment_in_dynamic[:, 7] == i)]
            #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
            R_pool = R_pool[~(R_pool[:, 7] == i)]
            # print(routes[k],new_try)
            routes[k] = copy.copy(new_try)
            #lost_r()
            #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
            number_T = 0
            v_has_r[0] = k
    if not (isinstance(k, (int, np.integer)) and k != -1) and i in no_T_R:
        no_T_R.remove(i)
    if T_or == 1 and segment_in_dynamic == 0:
        len2 = len(R_pool[:, 7])
        if len2 == len1:
            number_T = -1
            if len(K) >= 2:
                # two vehicles
                #                random_k = int(np.random.choice([0, 1], size=(1,), p=[5. / 10, 5. / 10]))
                random_k = 0
                random_position = 0
                if random_k != 1:
                    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                    routes, R_pool, top_key_2k, v_has_r, best_T_2k = insert2vehicle(i, K, random_k, random_position, check_uncertainty_in_insertion_by_RL_or_not=check_uncertainty_in_insertion_by_RL_or_not, new_row=new_row, finish_or_begin=finish_or_begin, uncertainty_index=uncertainty_index, store_routes_for_another_k=store_routes_for_another_k, store_R_pool_for_another_k=store_R_pool_for_another_k, duration=duration, congestion_link=congestion_link,
                congestion_node=congestion_node,index=index)
                    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                else:
                    routes, R_pool, top_key_2k, v_has_r, best_T_2k = random_k_insert2vehicle(i, K, random_k, random_position)
                number_T = 1
                used_T[0] = best_T_2k
        len3 = len(R_pool[:, 7])
        if len3 == len1:
            number_T = -1
            if len(K) >= 3:
                random_k = 0
                random_position = 0
                if random_k != 1:
                    routes, R_pool, top_key_3k, v_has_r, best_T_3k = insert3vehicle(i, K, random_k, random_position)
                #                else:
                #                    routes, R_pool = random_k_insert3vehicle(i, K, random_k, random_position)
                number_T = 2
                used_T = best_T_3k
    key = tuple([R[index_r, 0], R[index_r, 1]])
    update_r_best_obj_in_insertion(i, len1, old_overall_cost,v_has_r,used_T)
    #find_unchecked_r_preference([6,45])
    if bundle_or_not == 1 and len(R_pool) > 0:
        if number_T == 0:
            insert_bundle_pre(i, key, number_T, best_T, top_key, k)
        else:
            if number_T == 1:
                insert_bundle_pre(i, key, number_T, best_T_2k, top_key_2k, k)
            else:
                if number_T == 2:
                    insert_bundle_pre(i, key, number_T, best_T_3k, top_key_3k, k)
    #check_capacity(routes)
    #find_unchecked_r_preference([6,45])
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    #check_served_R()
    return routes, R_pool

def return_routes_R_pool(k_number,routes_2k,R_pool_2k,i, key, best_T_2k, top_key_2k, k, final_ok_or, new_try,routes_3k, R_pool_3k, best_T_3k, top_key_3k):
    global routes, R_pool
    #lost_r()
    
    if k_number == 2:
        number_T = 1
        routes = my_deepcopy(routes_2k)
        R_pool = copy.copy(R_pool_2k)

        if bundle_or_not == 1 and len(R_pool) > 0:
            insert_bundle_pre(i, key, number_T, best_T_2k, top_key_2k, k)
        # check_capacity(routes)
        #lost_r()
        return routes, R_pool
        # routes = my_deepcopy(routes_2k)
        # if len2 < len1:
        #     R_pool=R_pool[~(R_pool[:,7]==i)]
    else:
        if k_number == 1:

            if final_ok_or == 1:

                R_pool = R_pool[~(R_pool[:, 7] == i)]
                # R_pool=R_pool[~(R_pool[:,7]==i)]

                routes[k] = copy.copy(new_try)
                #lost_r()
                # #lost_r()
                number_T = 0
                best_T, top_key = [-1], -1

                if bundle_or_not == 1 and len(R_pool) > 0:
                    insert_bundle_pre(i, key, number_T, best_T, top_key, k)
            # check_capacity(routes)
            #lost_r()
            return routes, R_pool
        else:
            number_T = 2
            routes = my_deepcopy(routes_3k)
            R_pool = copy.copy(R_pool_3k)
            if bundle_or_not == 1 and len(R_pool) > 0:
                insert_bundle_pre(i, key, number_T, best_T_3k, top_key_3k, k)
            # routes = my_deepcopy()
            # check_capacity(routes)
            #lost_r()
            return routes, R_pool

# @profile()
# @time_me()
##@jit
def real_greedy_insert(i):
    global routes, R_pool, check_start_position, relevant_request_position_number
    #find_unchecked_r_preference([6,45])
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    index_r = list(R[:, 7]).index(i)
    R_pool_copy = copy.copy(R_pool)
    routes_copy = my_deepcopy(routes)
    len_original = len(R_pool[:, 7])
    R_i = tuple(zip(R[index_r], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))

    old_overall_cost = overall_obj(routes_copy)[1]

    overall_cost = 999999999999999999999
    overall_cost_2k = 999999999999999999999
    overall_cost_3k = 999999999999999999999
    routes_2k, R_pool_2k,best_T_2k, top_key_2k, routes_3k, R_pool_3k, best_T_3k, top_key_3k = -1,-1,-1,-1,-1,-1,-1,-1
    Trans = 0
    Trans_Tp = 0
    Trans_Td = 0
    number_T = -1
    #find_unchecked_r_preference([6,45])
    k, new_try, position, insert_r_cost = insert1vehicle(R_i, i, K, Trans, Trans_Tp, Trans_Td)
    #find_unchecked_r_preference([6,45])
    final_ok_or = 0
    if isinstance(k, (int, np.integer)) and k != -1:

        check_start_position = position[0]


        relevant_request_position_number = {}
        bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k, new_try,i)
        if not isinstance(bool_or_route, bool):

            relevant_try_copy = my_deepcopy(relevant_try)
            layer,aaa = 0,0
            final_ok_or = solve_relevant_try(relevant_try_copy,layer,aaa)
            if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
                satisfy_preference = preference_constraints(i, k, -1, -1, new_try, -1, -1)
                if satisfy_preference == 1:
                    # when check preference, only the inserted r is checked, other r' in k has not be checked, and r may influce r', so the k itself should also be checked
                    relevant_try_copy[k] = [copy.copy(new_try), i, 0]
                    layer, aaa = 0, 0
                    preference_final_ok_or1 = solve_relevant_try(relevant_try_copy, layer, aaa, 1)
                    if preference_final_ok_or1 == 0:
                        final_ok_or = 0
                else:
                    final_ok_or = 0
            if final_ok_or == 0:
                pass
            else:

                R_pool = R_pool[~(R_pool[:, 7] == i)]
                routes[k] = copy.copy(new_try)
                #lost_r()
                overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost, overall_transshipment_cost, overall_un_load_cost, overall_emission_cost, overall_storage_cost, overall_delay_penalty, overall_number_transshipment, overall_average_speed, overall_average_time_ratio, overall_emission_transshipment = overall_obj(
                    routes)
                cost_inserted_request = overall_cost - old_overall_cost
                update_r_best_obj_record(i, cost_inserted_request,k,-1)
                if multi_obj == 0 and K[k, 5] == 1:
                    r_basic_cost = get_r_basic_cost(R[index_r, 0], R[index_r, 1], i, k)
                    if cost_inserted_request < r_basic_cost + R[index_r, 6] * 2 * c_storage - 0.1:
                        key = tuple([R[index_r, 0], R[index_r, 1]])
                        number_T = 0
                        best_T, top_key = [-1], -1
                        #find_unchecked_r_preference([6,45])
                        if bundle_or_not == 1 and len(R_pool) > 0:
                            insert_bundle_pre(i, key, number_T, best_T, top_key, k)
                        #lost_r()
                        #find_unchecked_r_preference([6,45])
                        return routes, R_pool
                R_pool = copy.copy(R_pool_copy)
                routes = my_deepcopy(routes_copy)
    
    if not (isinstance(k, (int, np.integer)) and k != -1) and i in no_T_R:
        no_T_R.remove(i)
    if T_or == 1 and i not in no_T_R:
        if len(K) >= 2:
            # two vehicles
            #                random_k = int(np.random.choice([0, 1], size=(1,), p=[5. / 10, 5. / 10]))
            random_k = 0
            random_position = 0
            if random_k == 0:
                routes_2k, R_pool_2k, top_key_2k,v_has_r_2k,  best_T_2k = insert2vehicle(i, K, random_k, random_position)
            else:
                routes_2k, R_pool_2k, top_key_2k, v_has_r_2k, best_T_2k = random_k_insert2vehicle(i, K, random_k, random_position)
            if len(R_pool_2k) == len_original:
                overall_cost_2k = 999999999999999999999
            else:
                overall_distance_2k, overall_cost_2k, overall_time_2k, overall_profit_2k, overall_emission_2k, served_requests_2k, overall_request_cost_2k, overall_vehicle_cost_2k, overall_wait_cost_2k, overall_transshipment_cost_2k, overall_un_load_cost_2k, overall_emission_cost_2k, overall_storage_cost_2k, overall_delay_penalty_2k, overall_number_transshipment_2k, overall_average_speed_2k, overall_average_time_ratio_2k, overall_emission_transshipment_2k = overall_obj(
                    routes_2k)
                cost_inserted_request = overall_cost_2k - old_overall_cost
                update_r_best_obj_record(i, cost_inserted_request,v_has_r_2k,best_T_2k)
            routes = my_deepcopy(routes_copy)
            R_pool = copy.copy(R_pool_copy)
    #lost_r()
    
    if len(K) >= 3:
        random_k = 0
        random_position = 0
        # if random_k != 1:
        # 20200901: this will change routes, and in the same time change routes_2k... so I need to cut the relationship between routes_2k and routes, so return copy in insert2vehicle
        routes_3k, R_pool_3k, top_key_3k, v_has_r_3k, best_T_3k = insert3vehicle(i, K, random_k, random_position)
        if len(R_pool_3k) == len_original:
            overall_cost_3k = 999999999999999999999
        else:
            overall_distance_3k, overall_cost_3k, overall_time_3k, overall_profit_3k, overall_emission_3k, served_requests_3k, overall_request_cost_3k, overall_vehicle_cost_3k, overall_wait_cost_3k, overall_transshipment_cost_3k, overall_un_load_cost_3k, overall_emission_cost_3k, overall_storage_cost_3k, overall_delay_penalty_3k, overall_number_transshipment_3k, overall_average_speed_3k, overall_average_time_ratio_3k, overall_emission_transshipment_3k = overall_obj(
                routes_3k)
            cost_inserted_request = overall_cost_3k - old_overall_cost
            update_r_best_obj_record(i, cost_inserted_request, v_has_r_3k, best_T_3k)
        routes = my_deepcopy(routes_copy)
        R_pool = copy.copy(R_pool_copy)
    #                else:
    #                    routes, R_pool = random_k_insert3vehicle(i, K, random_k, random_position)
    
    if overall_cost == 999999999999999999999 and overall_cost_2k == 999999999999999999999 and overall_cost_3k == 999999999999999999999:
        #lost_r()
        #find_unchecked_r_preference([6,45])
        return routes, R_pool
    key = tuple([R[index_r, 0], R[index_r, 1]])
    number_T = -1

    #if the cost is equal with each other, then compare the real cost, and return min one
    if Demir == 1:
        costs = [overall_cost, overall_cost_2k, overall_cost_3k]
        min_cost = min(costs)
        equal_number = 0
        equal_index = [0,0,0]
        for cost_index in range(len(costs)):
            if costs[cost_index] == min_cost:
                equal_number = equal_number + 1
                equal_index[cost_index] = 1
        if equal_number > 1:
            real_cost_1k = overall_request_cost + overall_vehicle_cost + overall_wait_cost + overall_transshipment_cost + overall_un_load_cost + overall_emission_cost + overall_storage_cost + overall_delay_penalty
            if overall_cost_2k < 999999999999999999999:
                real_cost_2k = overall_request_cost_2k + overall_vehicle_cost_2k + overall_wait_cost_2k + overall_transshipment_cost_2k + overall_un_load_cost_2k + overall_emission_cost_2k + overall_storage_cost_2k + overall_delay_penalty_2k
            else:
                real_cost_2k = 999999999999999999999
            if overall_cost_3k < 999999999999999999999:
                real_cost_3k = overall_request_cost_3k + overall_vehicle_cost_3k + overall_wait_cost_3k + overall_transshipment_cost_3k + overall_un_load_cost_3k + overall_emission_cost_3k + overall_storage_cost_3k + overall_delay_penalty_3k
            else:
                real_cost_3k = 999999999999999999999
            real_costs = [real_cost_1k,real_cost_2k,real_cost_3k]
            compare_cost = []
            compare_index = []
            for cost_index in range(len(costs)):
                if equal_index[cost_index] == 1:
                    compare_cost.append(real_costs[cost_index])
                    compare_index.append(cost_index)
            min_real_cost_index = compare_index[compare_cost.index(min(compare_cost))]
            #lost_r()
            #find_unchecked_r_preference([6,45])
            return return_routes_R_pool(min_real_cost_index+1,routes_2k,R_pool_2k,i, key, best_T_2k, top_key_2k, k, final_ok_or, new_try,routes_3k, R_pool_3k, best_T_3k, top_key_3k)

    if overall_cost_2k <= overall_cost_3k and overall_cost_2k <= overall_cost:
        k_number = 2
    else:
        if overall_cost <= overall_cost_3k and overall_cost <= overall_cost_2k:
            k_number = 1
        else:
            k_number = 3
    #lost_r()
    #find_unchecked_r_preference([6,45])
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    return return_routes_R_pool(k_number, routes_2k, R_pool_2k, i, key, best_T_2k,
                                          top_key_2k, k, final_ok_or,new_try, routes_3k,
                                          R_pool_3k, best_T_3k, top_key_3k)

# @profile()
# @time_me()
##@jit
def global_real_greedy_insert_base():
    global routes, R_pool, check_start_position, relevant_request_position_number
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    R_pool_copy = copy.copy(R_pool)
    routes_copy = my_deepcopy(routes)
    len_original = len(R_pool_copy)
    # all_r_cost = pd.DataFrame(columns=['record_top_key', 'overall_cost', 'insert_k_vehicles'], index=R_pool[:, 7])
    all_r_cost = np.array(np.empty(shape=(len(R_pool), 4)),dtype='object')
    all_r_cost[:] = np.NaN
    all_r_cost[:,3] = R_pool[:, 7]
    record_1_vehicle_new_try = {}

    old_overall_cost = overall_obj(routes)[1]
    # the i will be tried to insert to routes, and after each try, routes and R_pool  will be restored
    for i in R_pool[:, 7]:
        
        index_r = list(R[:, 7]).index(i)
        index_i = list(all_r_cost[:,3]).index(i)
        # if no_T == 1 means the barge can serve this r and not multi-obj
        no_T = 0
        R_i = tuple(zip(R[index_r], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
        Trans = 0
        Trans_Tp = 0
        Trans_Td = 0
        k, new_try, position, insert_r_cost = insert1vehicle(R_i, i, K, Trans, Trans_Tp, Trans_Td, 1)

        overall_cost = 999999999999999999999
        overall_cost_2k = 999999999999999999999
        overall_cost_3k = 999999999999999999999
        if isinstance(k, (int, np.integer)) and k != -1:
            check_start_position = position[0]

            relevant_request_position_number = {}
            bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k, new_try,i)
            if not isinstance(bool_or_route, bool):
                relevant_try_copy = my_deepcopy(relevant_try)
                layer,aaa = 0,0
                final_ok_or = solve_relevant_try(relevant_try_copy,layer,aaa)
                if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0:
                    satisfy_preference = preference_constraints(i, k, -1, -1, new_try, -1, -1)
                    if satisfy_preference == 1:
                        # when check preference, only the inserted r is checked, other r' in k has not be checked, and r may influce r', so the k itself should also be checked
                        relevant_try_copy[k] = [copy.copy(new_try), i, 0]
                        layer, aaa = 0, 0
                        preference_final_ok_or1 = solve_relevant_try(relevant_try_copy, layer, aaa, 1)
                        if preference_final_ok_or1 == 0:
                            final_ok_or = 0
                    else:
                        final_ok_or = 0
                if final_ok_or == 0:
                    pass
                else:

                    R_pool = R_pool[~(R_pool[:, 7] == i)]
                    routes_1k = my_deepcopy(routes_copy)
                    routes_1k[k] = copy.copy(new_try)
                    # #lost_r()
                    overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost, overall_transshipment_cost, overall_un_load_cost, overall_emission_cost, overall_storage_cost, overall_delay_penalty, overall_number_transshipment, overall_average_speed, overall_average_time_ratio, overall_emission_transshipment = overall_obj(
                        routes_1k)

                    if len(R_pool) < len_original and overall_cost < 100000000:
                        r_cost_in_all_routes = overall_cost - old_overall_cost
                        update_r_best_obj_record(i, r_cost_in_all_routes,k,-1)
                        if multi_obj == 0 and K[k, 5] == 1:
                            r_basic_cost = get_r_basic_cost(R[index_r, 0], R[index_r, 1], i, k)
                            if r_cost_in_all_routes < r_basic_cost + R[index_r, 6] * 2 * c_storage - 0.1:
                                no_T = 1
                                if i not in no_T_R:
                                    no_T_R.append(i)
                    R_pool = copy.copy(R_pool_copy)
        if not (isinstance(k, (int, np.integer)) and k != -1) and i in no_T_R:
            no_T_R.remove(i)
        if T_or == 1 and no_T == 0 and i not in no_T_R:
            if len(K) >= 2:
                # two vehicles
                #                random_k = int(np.random.choice([0, 1], size=(1,), p=[5. / 10, 5. / 10]))
                random_k = 0
                random_position = 0
                if random_k != 1:
                    routes_2k, R_pool_2k, top_key_2k, v_has_r_2k, best_T_2k = insert2vehicle(i, K, random_k, random_position)
                else:
                    routes_2k, R_pool_2k, top_key_2k, v_has_r_2k, best_T_2k = random_k_insert2vehicle(i, K, random_k, random_position)
                overall_distance_2k, overall_cost_2k, overall_time_2k, overall_profit_2k, overall_emission_2k, served_requests_2k, overall_request_cost_2k, overall_vehicle_cost_2k, overall_wait_cost_2k, overall_transshipment_cost_2k, overall_un_load_cost_2k, overall_emission_cost_2k, overall_storage_cost_2k, overall_delay_penalty_2k, overall_number_transshipment_2k, overall_average_speed_2k, overall_average_time_ratio_2k, overall_emission_transshipment_2k = overall_obj(
                    routes_2k)
                if len(R_pool_2k) == len_original:
                    overall_cost_2k = 999999999999999999999
                else:
                    if overall_cost_2k < 100000000:
                        r_cost_in_all_routes = overall_cost_2k - old_overall_cost
                        update_r_best_obj_record(i, r_cost_in_all_routes,v_has_r_2k,best_T_2k)

                routes = my_deepcopy(routes_copy)
                R_pool = copy.copy(R_pool_copy)

            if len(K) >= 3:
                random_k = 0
                random_position = 0
                # if random_k != 1:
                routes_3k, R_pool_3k, top_key_3k, v_has_r_3k, best_T_3k = insert3vehicle(i, K, random_k, random_position)
                overall_distance_3k, overall_cost_3k, overall_time_3k, overall_profit_3k, overall_emission_3k, served_requests_3k, overall_request_cost_3k, overall_vehicle_cost_3k, overall_wait_cost_3k, overall_transshipment_cost_3k, overall_un_load_cost_3k, overall_emission_cost_3k, overall_storage_cost_3k, overall_delay_penalty_3k, overall_number_transshipment_3k, overall_average_speed_3k, overall_average_time_ratio_3k, overall_emission_transshipment_3k = overall_obj(
                    routes_3k)
                if len(R_pool_3k) == len_original:
                    overall_cost_3k = 999999999999999999999
                else:
                    if overall_cost_3k < 100000000:
                        r_cost_in_all_routes = overall_cost_3k - old_overall_cost
                        update_r_best_obj_record(i, r_cost_in_all_routes,v_has_r_3k,best_T_3k)
                # in global_ these two statements are needed because routes need to be keep as the original one when next r is inserted in, but in real_greedy it is not needed because insert3vehicle is the end
                routes = my_deepcopy(routes_copy)
                R_pool = copy.copy(R_pool_copy)
        #                else:
        #                    routes, R_pool = random_k_insert3vehicle(i, K, random_k, random_position)
        if overall_cost == 999999999999999999999 and overall_cost_2k == 999999999999999999999 and overall_cost_3k == 999999999999999999999:
            all_r_cost[index_i,2] = 0
            continue
        if overall_cost_2k <= overall_cost_3k and overall_cost_2k <= overall_cost:
            all_r_cost[index_i,2] = 2
            record_top_key = top_key_2k
            all_r_cost[index_i,1] = overall_cost_2k
            # routes = my_deepcopy(routes_2k)
            # if len2 < len1:
            #     R_pool=R_pool[~(R_pool[:,7]==i)]
        else:
            if overall_cost <= overall_cost_3k and overall_cost <= overall_cost_2k and k != -1:
                all_r_cost[index_i,2] = 1
                record_top_key = tuple([k, i])
                all_r_cost[index_i,1] = overall_cost

                record_1_vehicle_new_try[i] = new_try
            else:
                # routes = my_deepcopy()
                all_r_cost[index_i,2] = 3
                record_top_key = top_key_3k
                all_r_cost[index_i,1] = overall_cost_3k

        all_r_cost[index_i,0] = record_top_key

    return all_r_cost, record_1_vehicle_new_try


# what I want is learn from experience, I believe in the past for every r, it has a high probility that it be inserted to the best position in the optimal solution, but how to find it?
# features: lowest obj, for all r, for each r, may lose something of one r to gain someting of another r, this is the key
# but, if there is no conflicts for r1, it should always be inserted to this position
# for wenjing's case, the only confict is capacity, if I can insert it to the lowest cost k.position, then no one has conflict with it
# if conflict, then choose the one with lower regret value, which means has lower impact on global
# therefore, it should be easy to find the optimal solution of wenjing
# for my case, the conflicts are more complex, because time is involved in, but it can be also done by compare regret
# problem is, we can only based on current routes, but maybe the best routes should use another insertion order
# how to escape it? learn from experience? -> maybe from other insertion oprators, there are other solution, with lower cost, then we can use it as current routes, this is what I do, but can I do it in a more direct way?
# first use conflict_regret, then use random/greedy/transshipment, then after find a better solution or worse solution, use conflict_regret, if it can find the better solution, then keep this loop; if worse,
# for all r, get all alternatives, and for each r, try to insert it to it's lowest-cost k and postion, if there are conflicts, choose the second,third,... one depending on regret value
# if r is inserted to a route with many terminals,
# for wenjing's case, only two terminals for one k, therefore it's no problem
# for PDP, if there is a r2 which have been inserted to route, but if it has conflict with the new r1, which has a lower regret value -> compared with what? r2 doesn't have regret value in fact,
# how to compare r2 and r1? r1 has insert cost and regret value table, r2?
# pull r2 out of route firstly, and treat it as same as r in R_pool, then try to insert it and get insert cost and regret value
# this is done in the previous iteration!
# if I go back to historical records, and find the lowest cost and the relevant routes, how could I destroy current routes and don't care about the inserted r in routes!
# but I can try, if there is no conflict, or 1 conflict r2, I can try to compare r1 and r2's regret -> but the routes are changed I can't get regret table of r2, so this idea is not work
# if there is conflict, try the next best insertion of r1 -> this is what greedy done!
# or just remove it to R_pool
# replace the replacable one when conflicts
# and insert as many as can be inserted r to routes after the regret greedy calculation, means all r without conflicts and all r with conflict and lowest regret value
# it seems if the best one in historical records will change the current routes, then it's hard to judge it's good or not. if it doesn't change, then it's same as the current operators

# for all r, get all alternatives, and choose to insert the r with lowest cost to the losest-cost k and position
# @profile()
# @time_me()
def global_real_greedy_insert():
    global routes, R_pool
    routes_tuple = get_routes_tuple(routes)
    R_pool_tuple = df_tuple(R_pool, 'R_pool')
    hash_top_R_pool_key = tuple([routes_tuple, R_pool_tuple])
    if hash_top_R_pool_key not in hash_top_R_pool.keys():
        hash_top_R_pool[hash_top_R_pool_key] = {}
    else:
        return my_deepcopy(hash_top_R_pool[hash_top_R_pool_key]['routes']), copy.copy(
            hash_top_R_pool[hash_top_R_pool_key]['R_pool'])
    all_r_cost, record_1_vehicle_new_try = global_real_greedy_insert_base()
    # if one r can't be inserted, then return original routes and R_pool
    # if 0 in all_r_cost.insert_k_vehicles:
    #     return routes, R_pool
    best_i_k = all_r_cost[np.argmin(all_r_cost[:,1], axis=0)]
    if best_i_k[1] == 999999999999999999999:
        hash_top_R_pool[hash_top_R_pool_key]['routes'] = my_deepcopy(routes)
        hash_top_R_pool[hash_top_R_pool_key]['R_pool'] = copy.copy(R_pool)
        return routes, R_pool
    else:
        if best_i_k[2] == 1:
            best_i = best_i_k[0][1]
            routes[best_i_k[0][0]] = copy.copy(record_1_vehicle_new_try[best_i])
            R_pool = R_pool[~(R_pool[:, 7] == best_i)]
            #lost_r()
            hash_top_R_pool[hash_top_R_pool_key]['routes'] = my_deepcopy(routes)
            hash_top_R_pool[hash_top_R_pool_key]['R_pool'] = copy.copy(R_pool)
            return routes, R_pool
        else:
            routes = my_deepcopy(hash_top[best_i_k[0]]['routes'])
            R_pool = copy.copy(hash_top[best_i_k[0]]['R_pool'])
            hash_top_R_pool[hash_top_R_pool_key]['routes'] = my_deepcopy(routes)
            hash_top_R_pool[hash_top_R_pool_key]['R_pool'] = copy.copy(R_pool)
            return routes, R_pool

def insert_1r_regret(record_1_vehicle_new_try,regret_values_all_r,all_r_cost,hash_top_R_pool_key):
    global routes, R_pool
    if np.size(regret_values_all_r) == 0:
        return routes, R_pool
    chose_r = regret_values_all_r[np.argmin(regret_values_all_r[:, 0], axis=0), 1]
    index_i = list(all_r_cost[:, 3]).index(chose_r)
    best_i_k = all_r_cost[index_i]
    print('current_i ', chose_r)
    if best_i_k[2] == 1:
        best_i = best_i_k[0][1]
        routes[best_i_k[0][0]] = copy.copy(record_1_vehicle_new_try[best_i])
        R_pool = R_pool[~(R_pool[:, 7] == best_i)]

        #lost_r()
        hash_top_R_pool[hash_top_R_pool_key]['routes'] = my_deepcopy(routes)
        hash_top_R_pool[hash_top_R_pool_key]['R_pool'] = copy.copy(R_pool)
        return routes, R_pool
    else:
        routes = my_deepcopy(hash_top[best_i_k[0]]['routes'])
        R_pool = copy.copy(hash_top[best_i_k[0]]['R_pool'])
        hash_top_R_pool[hash_top_R_pool_key]['routes'] = my_deepcopy(routes)
        hash_top_R_pool[hash_top_R_pool_key]['R_pool'] = copy.copy(R_pool)
        return routes, R_pool

# @profile()
# @time_me()
##@jit
def global_real_greedy_insert_regret():
    global routes, R_pool
    #find_unchecked_r_preference([6,45])
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    routes_tuple = get_routes_tuple(routes)
    R_pool_tuple = df_tuple(R_pool, 'R_pool')
    hash_top_R_pool_key = tuple([routes_tuple, R_pool_tuple])

    if hash_top_R_pool_key not in hash_top_R_pool.keys():
        hash_top_R_pool[hash_top_R_pool_key] = {}
    else:
        if len(hash_top_R_pool[hash_top_R_pool_key]) != 0:
            return my_deepcopy(hash_top_R_pool[hash_top_R_pool_key]['routes']), copy.copy(
                hash_top_R_pool[hash_top_R_pool_key]['R_pool'])
    all_r_cost, record_1_vehicle_new_try = global_real_greedy_insert_base()
    # if one r can't be inserted, then return original routes and R_pool
    # if 0 in list(all_r_cost.insert_k_vehicles):
    #     return routes, R_pool
    # calculate k_regret value
    # all_r_cost.dropna(inplace=True)
    try:
        all_r_cost = all_r_cost[~np.isnan(list(all_r_cost[:,1]))]
    except:
        print('caught')
    for i in all_r_cost[:,3]:
        index_i = list(all_r_cost[:, 3]).index(i)
        if all_r_cost[index_i,1] >= 9999999999999999 or all_r_cost[index_i,2] == 0:
            # all_r_cost.drop(i, axis=0, inplace=True)
            all_r_cost = np.delete(all_r_cost, index_i, axis=0)
            continue
        if all_r_cost.size == 0:
            hash_top_R_pool[hash_top_R_pool_key]['routes'] = my_deepcopy(routes)
            hash_top_R_pool[hash_top_R_pool_key]['R_pool'] = copy.copy(R_pool)
            return routes, R_pool
    # regret_values_all_r = pd.DataFrame(columns=['k_regret_value'], index=all_r_cost[:,3])
    regret_values_all_r = np.array(np.empty(shape=(len(all_r_cost), 2)))
    regret_values_all_r[:, 1] = all_r_cost[:, 3]
    #find_unchecked_r_preference([6,45])
    all_regret_values_df = {}
    for i in all_r_cost[:,3]:
        index_r = list(R[:, 7]).index(i)
        R_i = tuple(zip(R[index_r], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
        regret_values = []
        noT = 0
        for k1 in range(len(K)):

            original_route_no_columns = route_no_columns(routes[k1])
            key_1k = get_key_1k(R_i, original_route_no_columns, k1, fixed_vehicles_percentage, Fixed, K)
            if key_1k in hash_table_1v_all.keys():
                if k1 not in train_truck and len(hash_table_1v_all[key_1k].keys()) > 0:
                    noT = 1
                for position in hash_table_1v_all[key_1k].keys():
                    regret_values.append(
                        [k1, position, key_1k, hash_table_1v_all[key_1k][position]['cost_inserted_request'], '1k'])
        delete_K_pair = delete_k(i)
        if not (multi_obj == 0 and noT == 1):

            for T_change in all_ok_TK[i].keys():
                all_ok_k_pair = np.array(delete_K_pair[T_change])
                if len(all_ok_k_pair) >= 1:
                    if isinstance(T_change,int):
                        for x in range(len(all_ok_k_pair)):
                            
                            k1,k2 = all_ok_k_pair[x,:]
                            
                            original_route_no_columns1 = route_no_columns(routes[k1])
                            original_route_no_columns2 = route_no_columns(routes[k2])
                            fix_k1_ap, fix_k1_bp = get_fix_k_0_ap(k1,
                                                                                                fixed_vehicles_percentage,
                                                                                                Fixed)
                            fix_k2_ap, fix_k2_bp = get_fix_k_0_ap(k2,
                                                                                                fixed_vehicles_percentage,
                                                                                                Fixed)

                            key_2k = (
                                T_change, R_i, original_route_no_columns1, K[k1, 0], K[k1, 1], fix_k1_ap, fix_k1_bp, original_route_no_columns2, K[k2, 0], K[k2, 1], fix_k2_ap, fix_k2_bp)
                            if key_2k in hash_table_2v_all.keys():
                                if T_change in hash_table_2v_all[key_2k].keys():
                                    if list(hash_table_2v_all[key_2k][T_change]):
                                        regret_values.append(
                                            [[k1, k2], list(hash_table_2v_all[key_2k][T_change])[0], key_2k,
                                             hash_table_2v_all[key_2k][T_change][
                                                 list(hash_table_2v_all[key_2k][T_change])[0]]['cost_inserted_request'],
                                             T_change])
                    else:
                        T_change1, T_change2 = T_change
                        for x in range(len(all_ok_k_pair)):
                            k1, k2, k3 = all_ok_k_pair[x, :]

                            original_route_no_columns1 = route_no_columns(routes[k1])
                            original_route_no_columns2 = route_no_columns(routes[k2])
                            original_route_no_columns3 = route_no_columns(routes[k3])
                            fix_k1_ap, fix_k1_bp = get_fix_k_0_ap(k1,
                                                                                                fixed_vehicles_percentage,
                                                                                                Fixed)
                            fix_k2_ap, fix_k2_bp = get_fix_k_0_ap(k2,
                                                                                                fixed_vehicles_percentage,
                                                                                                Fixed)
                            fix_k3_ap, fix_k3_bp = get_fix_k_0_ap(k3,
                                                                                                fixed_vehicles_percentage,
                                                                                                Fixed)
                            key_3k = (
                                R_i, original_route_no_columns1, K[k1, 0], K[k1, 1], fix_k1_ap, fix_k1_bp, original_route_no_columns2, K[k2, 0], K[k2, 1],
                                fix_k2_ap, fix_k2_bp, original_route_no_columns3,
                                K[k3, 0], K[k3, 1], fix_k3_ap, fix_k3_bp,
                                T_change1, T_change2)

                            if key_3k in hash_table_3v_all.keys():
                                if T_change in hash_table_3v_all[key_3k].keys():
                                    if list(hash_table_3v_all[key_3k][T_change]):
                                        regret_values.append(
                                            [[k1, k2, k3], list(hash_table_3v_all[key_3k][T_change])[0], key_3k,
                                             hash_table_3v_all[key_3k][T_change][
                                                 list(hash_table_3v_all[key_3k][T_change])[0]]['cost_inserted_request'],
                                             T_change])
        # regret_values_df = pd.DataFrame(regret_values, columns=['k', 'position', 'key', 'cost', 'T'])
        regret_values_df = np.array(regret_values)
        # regret_values_df = regret_values_df.sort_values(by=['cost'])

        # regret_values.sort()
        if len(regret_values_df) > 0:
            regret_values_df = regret_values_df[np.argsort(regret_values_df[:, 3])]

            index_i = list(regret_values_all_r[:,1]).index(i)
            regret_values_all_r[index_i,0] = regret_values_df[min(len(regret_values_df) - 1, regret_k),3] - regret_values_df[0,3]
            all_regret_values_df[i] = copy.copy(regret_values_df)
    #find_unchecked_r_preference([6,45])
    old_length = len(R_pool)
    # regret_values_all_r['k_regret_value'] = pd.to_numeric(regret_values_all_r['k_regret_value'])
    if insert_multiple_r == 0:
        return insert_1r_regret(record_1_vehicle_new_try,regret_values_all_r, all_r_cost,hash_top_R_pool_key)
    else:
        # this means I insert as many as r to routes, if there are conflicts, compare r's regret value, and insert the lowest one,
        # but many r can be inserted to a same route, as long as capacity is not exceeded, the problem is not easy to find the time
        # time, maybe if fixed, then fixed time; if free, the main thing is the wait time, and the time windows of r, then the earlist time on
        # or, I can just insert r one by one (order is based on regret value), if time constraints fine, then fine, if not, then put into next regret_insert
        # and I must say, if the k is not conflict, the time constraints also must be checked due to the chain reaction
        # so, create a func which can insert multiple r
        # the advantage is it can save time; the disadvantage of this way is it will not find solutions which based on the routes after the insertion, and these solutions may better
        # and there is a better way, when conflict can't pass time constraints when insert multiple r to a same k, it can try the next k which can serve it, but it need to record all possibilities, and I
        # all_r_cost
        # used_k = pd.DataFrame(index=all_r_cost[:,3], columns=['k1', 'k2', 'k3'])
        used_k = np.array(np.empty(shape=(len(all_r_cost[:,3]), 4)), dtype='object')
        used_k[:]=-1
        used_k[:, 3] = all_r_cost[:,3]
        # get all used k for each r
        for inserted_r_index in all_r_cost[:,3]:
            index = list(used_k[:,3]).index(inserted_r_index)
            index_i = list(all_r_cost[:, 3]).index(inserted_r_index)
            if all_r_cost[index_i,2] == 1:
                used_k[index,0] = all_r_cost[index_i,0][0]
            else:
                if all_r_cost[index_i,2] == 2:
                    used_k[index,0], used_k[index,1] = \
                        hash_top[all_r_cost[index_i,0]]['k']
                else:
                    used_k[index,0], used_k[index,1], used_k[index,2] = \
                            hash_top[all_r_cost[index_i,0]]['k']

        # all k r pairs
        all_k_r = pd.DataFrame(columns=range(len(K)), index=all_r_cost[:,3])
        # all_k_r = np.array(np.empty(shape=(len(all_r_cost),len(K))))
        # all_k_r
        for r in used_k[:,3]:
            index = list(used_k[:,3]).index(r)
            try:
                all_k_r[used_k[index,0]][r] = 1
            except:
                #when the r is not be served, do nothing
                print(1)
            if isinstance(used_k[index,1], (int, np.integer)) and used_k[index,1] != -1:
                all_k_r[used_k[index,1]][r] = 1
                if isinstance(used_k[index,2], (int, np.integer)) and used_k[index,2] != -1:
                    all_k_r[used_k[index,2]][r] = 1
        #find_unchecked_r_preference([6,45])
        # no conflict r
        all_r_cost_copy = copy.copy(all_r_cost)
        for r in all_k_r.index:
            index_i = list(all_r_cost_copy[:, 3]).index(r)
            if all_k_r.loc[r].isnull().values.all():
                all_k_r.drop(r, axis=0, inplace=True)

                all_r_cost_copy = np.delete(all_r_cost_copy, index_i, axis=0)
                continue
            else:
                conflict = 0
                for k in all_k_r.loc[r].dropna().index:
                    if k in fixed_vehicles_percentage:
                        continue
                    else:
                        if len(all_k_r[k].dropna()) == 1:
                            continue
                        else:
                            # if
                            conflict = 1
                            break
                if conflict == 1:

                    all_r_cost_copy = np.delete(all_r_cost_copy, index_i, axis=0)
        # the left r in all_r_cost_copy can be inserted to routes directly
        for ok_r in all_r_cost_copy[:,3]:
            # if not inserted, inserted_or_not = 0, otherwise it is 'mark'
            all_k_r, inserted_or_not, capacity_full = insert_a_r(all_k_r, ok_r, used_k, all_r_cost_copy,
                                                                 record_1_vehicle_new_try)
            # if not isinstance(inserted_or_not, int):
            # conflict_r no matter ok_r is inserted or not, it's not conflict r. if it can't be inserted, then it's infeasible
            try:
                all_k_r.drop(ok_r, axis=0, inplace=True)
            except:
                sys.exit(-4)
            # key = tuple([R[index_r,0], R[index_r,1]])
            # #insert_bundle_pre(ok_r, key, number_T, best_T, top_key, k, 1, )
        # conflict_r
        # for r in used_k[:,3]:
        #     # delete the r without conflicts by only assign 1 to r which has conflict
        #     if r not in all_r_cost_copy[:,3]:
        #         all_k_r[used_k[index,0]][r] = 1
        #         if isinstance(used_k[index,1], str):
        #             all_k_r[used_k[index,1]][r] = 1
        #             if isinstance(used_k[index,2], str):
        #                 all_k_r[used_k[index,2]][r] = 1
        if not all_k_r.empty:
            for k in all_k_r.columns:
                if all_k_r[k].isnull().values.all():
                    all_k_r.drop(k, axis=1, inplace=True)
            # for r in all_k_r.index:
            #     if all_k_r.loc[r].isnull().values.all():
            #         all_k_r.drop(r, axis=0, inplace=True)

            regret_values_per_k = {}
            for k in all_k_r.columns:
                # for each k, compare regret value of all possible r
                # regret_values_per_k[k] = pd.DataFrame(columns=['k_regret_value'], index=all_r_cost[:,3])
                regret_values_per_k[k] = np.array(np.empty(shape=(len(all_r_cost[:,3]),2)))
                regret_values_per_k[k][:] = np.NaN
                regret_values_per_k[k][:,1] = all_r_cost[:,3]
                for r in all_k_r.index:
                    if isinstance(all_k_r[k][r], (int, np.integer)):
                        index_i = list(regret_values_all_r[:,1]).index(i)
                        regret_values_per_k[k][list(regret_values_per_k[k][:,1]).index(r),0] = regret_values_all_r[index_i,0]
                # regret_values_per_k[k] = regret_values_per_k[k].dropna()
                regret_values_per_k[k] = regret_values_per_k[k][~np.isnan(regret_values_per_k[k][:,0])]
                # regret_values_per_k[k][:,0] = pd.to_numeric(regret_values_per_k[k]['k_regret_value'])
            for k in all_k_r.columns:
                if k not in all_k_r.columns:
                    continue
                if np.size(regret_values_per_k[k]) == 0:
                    continue
                chose_r = int(regret_values_per_k[k][np.argmax(regret_values_per_k[k][:,0],axis=0),1])

                index_i = list(all_r_cost[:,3]).index(chose_r)
                index = list(used_k[:,3]).index(chose_r)
                if (isinstance(used_k[index,0], (int, np.integer)) and used_k[index,0] != -1 and
                    used_k[index,0] not in all_k_r.columns) or (
                        isinstance(used_k[index,1], (int, np.integer)) and used_k[index,1] != -1 and
                        used_k[index,1] not in all_k_r.columns) or \
                        (isinstance(used_k[index,2], (int, np.integer)) and used_k[index,2] != -1 and
                         used_k[index,2] not in all_k_r.columns):
                    continue
                chose_r_regret_value = regret_values_per_k[k][:,0].max()
                # if k can serve more than one r, then get the r which ranks second regret value
                if len(regret_values_per_k[k]) > 1:
                    this_k_regret_values = copy.copy(regret_values_per_k[k])
                    this_k_regret_values = np.delete(this_k_regret_values,list(this_k_regret_values[:,1]).index(chose_r),axis=0)
                    second_r_regret_value = max(this_k_regret_values[:,0])
                    second_r = int(this_k_regret_values[list(this_k_regret_values[:,0]).index(second_r_regret_value),1])
                    index_second_i = list(all_r_cost[:, 3]).index(second_r)
                else:
                    second_r_regret_value = 0
                    second_r = -1
                    index_second_i = -1
                # if the r with max regret value only use one k, then insert it direcly
                if all_r_cost[index_i,2] == 1:
                    # best_i_k = all_r_cost[index_i]
                    all_k_r, regret_values_per_k, capacity_full = insert_a_r(all_k_r, chose_r, used_k, all_r_cost,
                                                                             record_1_vehicle_new_try,
                                                                             regret_values_per_k)
                else:
                    if all_r_cost[index_i,2] == 2:
                        # then I need to compare the regret value between the chose_r and the sum of regret value of two other influented r
                        other_k_regret_value = copy.copy(regret_values_per_k[used_k[index,1]])
                        # if the other k is only serve chose_r, then doesn't matter
                        if len(other_k_regret_value) == 1:
                            # insert this r
                            all_k_r, regret_values_per_k, capacity_full = insert_a_r(all_k_r, chose_r, used_k,
                                                                                     all_r_cost,
                                                                                     record_1_vehicle_new_try,
                                                                                     regret_values_per_k)
                        # otherwise, get the max regret value of the other k except for chose_r's regret value, then sum it with the second one of this k
                        else:
                            other_k_regret_value = np.delete(other_k_regret_value, list(other_k_regret_value[:,1]).index(chose_r), axis=0)

                            second_r_regret_value_of_other_k = max(other_k_regret_value[:,0])
                            all_influenced_r_regret_value = second_r_regret_value + second_r_regret_value_of_other_k
                            if float(chose_r_regret_value) >= float(all_influenced_r_regret_value):
                                # insert
                                all_k_r, regret_values_per_k, capacity_full = insert_a_r(all_k_r, chose_r, used_k,
                                                                                         all_r_cost,
                                                                                         record_1_vehicle_new_try,
                                                                                         regret_values_per_k)
                            else:
                                # if second_r only use this k, then insert it
                                if second_r != 0 and all_r_cost[index_second_i,2] == 1:
                                    # insert second r
                                    all_k_r, regret_values_per_k, capacity_full = insert_a_r(all_k_r, second_r, used_k,
                                                                                             all_r_cost,
                                                                                             record_1_vehicle_new_try,
                                                                                             regret_values_per_k)
                                else:
                                    continue

                    else:
                        # then I need to compare the regret value between the chose_r and the sum of regret value of three other influented r
                        other_k_regret_value = copy.copy(regret_values_per_k[used_k[index, 1]])
                        third_k_regret_value = copy.copy(regret_values_per_k[used_k[index, 2]])
                        # if the other k and the third k is only serve chose_r, then doesn't matter
                        if len(other_k_regret_value) == 1 and len(third_k_regret_value) == 1:
                            # insert this r
                            all_k_r, regret_values_per_k, capacity_full = insert_a_r(all_k_r, chose_r, used_k,
                                                                                     all_r_cost,
                                                                                     record_1_vehicle_new_try,
                                                                                     regret_values_per_k)
                        # otherwise, get the max regret value of the other and the third k except for chose_r's regret value, then sum it with the second one of this k
                        else:
                            second_r_regret_value_of_other_k, second_r_regret_value_of_third_k = 0,0
                            if len(other_k_regret_value) != 1:
                                other_k_regret_value = np.delete(other_k_regret_value,
                                                             list(other_k_regret_value[:, 1]).index(chose_r), axis=0)
                                second_r_regret_value_of_other_k = max(other_k_regret_value[:, 0])
                            if len(third_k_regret_value) != 1:
                                third_k_regret_value = np.delete(third_k_regret_value,
                                                             list(third_k_regret_value[:, 1]).index(chose_r), axis=0)
                                second_r_regret_value_of_third_k = max(third_k_regret_value[:, 0])
                            all_influenced_r_regret_value = second_r_regret_value + second_r_regret_value_of_other_k + second_r_regret_value_of_third_k
                            if float(chose_r_regret_value) >= float(all_influenced_r_regret_value):
                                # insert
                                all_k_r, regret_values_per_k, capacity_full = insert_a_r(all_k_r, chose_r, used_k,
                                                                                         all_r_cost,
                                                                                         record_1_vehicle_new_try,
                                                                                         regret_values_per_k)
                            else:
                                # if second_r only use this k, then insert it
                                if second_r != -1 and all_r_cost[index_second_i, 2] == 1:
                                    # insert second r
                                    all_k_r, regret_values_per_k, capacity_full = insert_a_r(all_k_r, second_r, used_k,
                                                                                             all_r_cost,
                                                                                             record_1_vehicle_new_try,
                                                                                             regret_values_per_k)
                                else:
                                    continue

        # until now r with highest regret_value which conflicts with other r, is inserted if there it use only one k or (use two k and still the most regret). Danger 3k is not considered!
        # the left r are r which is second, third regret r for free k, it should be recalculated in the next round because the current k has been changed
        # the only difference between insert one r and multiple r is for free k, if it's inserting r one by one, maybe the second inserted r will find a better solution based on the first r's insertion
        # situations for example second_r use more than 1 k, the chose r use 3 k,

        # after insert the r with high regret value, try to insert as many as r which regret value == 0, because these r has high probability that this is the best position
        # insert r in a bundle way
        # and this also avoid uncessary rounds
        #find_unchecked_r_preference([6,45])
        rest_r = []
        for i in all_r_cost[:,3]:
            if i in all_regret_values_df.keys() and i in R_pool[:, 7]:
                rest_r.append(i)
        # used_k = pd.DataFrame(index=rest_r, columns=['k1', 'k2', 'k3'])
        used_k = np.array(np.empty(shape=(len(rest_r), 4)), dtype='object')
        used_k[:] = -1
        used_k[:, 3] = rest_r
        check_bundle_r_k = {}
        for i in rest_r:
            index_r = list(R[:, 7]).index(i)
            regret_values_df = all_regret_values_df[i]
            # if there are other alternatives that has the same cost, i.e., regret value == 0, then choose the k that serve other r in the same bundle
            possible_k = []
            min_cost = regret_values_df[:,3].min()
            for m in range(len(regret_values_df)):
                if regret_values_df[m,3] == min_cost:
                    possible_k.append(regret_values_df[m,0])

            regret_values_df_index = 0
            for n in possible_k:
                # the k which has been served r before the rest_r = [] can't be used again

                possible_k1, possible_k2, possible_k3 = 0, 0, 0
                if isinstance(n, (int, np.integer)) and n != -1:
                    possible_k1 = n
                    if possible_k1 not in all_k_r.columns:
                        regret_values_df_index = regret_values_df_index + 1
                        continue
                    key = tuple([R[index_r, 0], R[index_r, 1]])
                    break_or_not, continue_or_not, check_bundle_r_k, used_k = check_bundle_in_k(i, used_k, possible_k1,
                                                                                                1, key,
                                                                                                check_bundle_r_k)
                    if break_or_not == 1:
                        insert_terminals = [R[index_r, 0], R[index_r, 1]]
                        positions = regret_values_df[regret_values_df_index,1]
                        insert_a_r(0, i, used_k, 0, 0, 'mark', 0, insert_terminals, positions)
                        break
                    if continue_or_not == 1:
                        regret_values_df_index = regret_values_df_index + 1
                        continue
                else:
                    # danger 2T was not considered
                    if len(n) == 2:
                        possible_k1 = n[0]
                        possible_k2 = n[1]
                        if possible_k1 not in all_k_r.columns or possible_k2 not in all_k_r.columns:
                            regret_values_df_index = regret_values_df_index + 1
                            continue
                        # should get T from regret_values_df
                        # this should get the first T in all 'k'==n
                        T = regret_values_df[regret_values_df_index,4]
                        key1 = tuple([R[index_r, 0], T])
                        key2 = tuple([T, R[index_r, 1]])
                        # should test both key1 and key2 are satisfied I think
                        break_or_not1, continue_or_not1, check_bundle_r_k, used_k = check_bundle_in_k(i, used_k,
                                                                                                      possible_k1, 1, key1,
                                                                                                      check_bundle_r_k)
                        break_or_not2, continue_or_not2, check_bundle_r_k, used_k = check_bundle_in_k(i, used_k,
                                                                                                      possible_k2, 2, key2,
                                                                                                      check_bundle_r_k)
                        if break_or_not1 == 1 and break_or_not2 == 1:
                            insert_terminals = [R[index_r, 0], T, R[index_r, 1]]
                            positions = regret_values_df[regret_values_df_index,1]
                            insert_a_r(0, i, used_k, 0, 0, 'mark', 0, insert_terminals, positions)
                            break
                        if continue_or_not1 == 1 or continue_or_not2 == 1:
                            regret_values_df_index = regret_values_df_index + 1
                            continue
                    else:
                        #2T 3k
                        possible_k1,possible_k2,possible_k3 = n

                        if possible_k1 not in all_k_r.columns or possible_k2 not in all_k_r.columns or possible_k3 not in all_k_r.columns:
                            regret_values_df_index = regret_values_df_index + 1
                            continue
                        # should get T from regret_values_df
                        # this should get the first T in all 'k'==n
                        T = regret_values_df[regret_values_df_index, 4]
                        T1, T2 = T
                        key1 = tuple([R[index_r, 0], T1])
                        key2 = tuple([T1, T2])
                        key3 = tuple([T2, R[index_r, 1]])
                        # should test both key1 and key2 are satisfied I think
                        break_or_not1, continue_or_not1, check_bundle_r_k, used_k = check_bundle_in_k(i, used_k,
                                                                                                      possible_k1, 1,
                                                                                                      key1,
                                                                                                      check_bundle_r_k)
                        break_or_not2, continue_or_not2, check_bundle_r_k, used_k = check_bundle_in_k(i, used_k,
                                                                                                      possible_k2, 2,
                                                                                                      key2,
                                                                                                      check_bundle_r_k)
                        break_or_not3, continue_or_not3, check_bundle_r_k, used_k = check_bundle_in_k(i, used_k,
                                                                                                      possible_k3, 3,
                                                                                                      key3,
                                                                                                      check_bundle_r_k)
                        if break_or_not1 == 1 and break_or_not2 == 1 and break_or_not3 == 1:
                            insert_terminals = [R[index_r, 0], T1, T2, R[index_r, 1]]
                            positions = regret_values_df[regret_values_df_index, 1]
                            # positions1, positions2 = positions
                            insert_a_r(0, i, used_k, 0, 0, 'mark', 0, insert_terminals, positions)
                            break
                        if continue_or_not1 == 1 or continue_or_not2 == 1 or continue_or_not3 == 1:
                            regret_values_df_index = regret_values_df_index + 1
                            continue
                # regret_values_df is sorted so it equals to:
                regret_values_df_index = regret_values_df_index + 1
        if len(R_pool) == old_length:
            #find_unchecked_r_preference([6,45])
            routes, R_pool = insert_1r_regret(record_1_vehicle_new_try,regret_values_all_r, all_r_cost,hash_top_R_pool_key)
            #find_unchecked_r_preference([6,45])
        hash_top_R_pool[hash_top_R_pool_key]['routes'] = my_deepcopy(routes)
        hash_top_R_pool[hash_top_R_pool_key]['R_pool'] = copy.copy(R_pool)
        #find_unchecked_r_preference([6,45])
        #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
        return routes, R_pool


# @profile()
# @time_me()
def most_hard_first_insert():
    global routes, R_pool
    # hard_value_R_pool = pd.DataFrame(columns=['hard_value'], index=R_pool[:, 7])
    #find_unchecked_r_preference([6,45])
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    hard_value_R_pool = np.array(np.empty(shape=(len(R_pool), 2)))
    hard_value_R_pool[:] = np.NaN
    hard_value_R_pool[:,1] = R_pool[:, 7]
    for r in R_pool[:, 7]:
        try:
            hard_value_R_pool[list(hard_value_R_pool[:,1]).index(r),0] = hard_value['hard_value'][r]
        except:
            sys.exit(-7)
    hard_value_R_pool = hard_value_R_pool[np.argsort(-hard_value_R_pool[:,0])]
    for r in hard_value_R_pool[:,1]:
        if r in R_pool[:, 7]:
            #find_unchecked_r_preference([6,45])
            routes, R_pool = greedy_insert(int(r))
            #find_unchecked_r_preference([6,45])
    #find_unchecked_r_preference([6,45])
    check_served_R()
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    return routes, R_pool

def membership_function(request, attribute, actual_value, satisfactory_level_value):
    continue_or = 0
    if attribute != 'speed':
        if actual_value <= satisfactory_level_value:
            satisfactory_value = 100
        else:
            continue_or = 1
    else:
        if actual_value >= satisfactory_level_value:
            satisfactory_value = 100
        else:
            continue_or = 1
    if continue_or == 1:
        # if attribute == 'cost':
        gap = abs(satisfactory_level_value-actual_value)
        if gap >= satisfactory_level_value:
            satisfactory_value = 0
        else:
            satisfactory_value = (1 - gap / satisfactory_level_value) * 100
    return satisfactory_value

def probability_function(satisfactory_value):
    accept_probability = satisfactory_value / 100
    return accept_probability

def accept_function(request, attribute, actual_value, satisfactory_level_value,get_satisfactory_value=0):

    if fuzzy_probability == 1:
        satisfactory_value = membership_function(request, attribute, actual_value, satisfactory_level_value)

    else:
        satisfactory_value = fuzzy_HP.fuzzy_HP_one(attribute, actual_value, satisfactory_level_value)
    accept_probability = probability_function(satisfactory_value)
    #when actual value very close to the value of the level, then in fuzzy the satisfactory_value/accept_probability is like 4.999999994 not 0.5, so I change it to 0.49
    if accept_probability < 0.49:
        return 0
    else:
        if get_satisfactory_value == 0:
            return 1
        else:
            return satisfactory_value
def hard_preference_constraints(attribute, actual_value, preference, get_satisfactory_value=0):
    satisfactory_level_value = fuzzy_HP.get_value_from_preference(attribute, preference)
    if attribute != 'speed':
        if actual_value < satisfactory_level_value:
            if get_satisfactory_value == 0:
                return 1
            else:
                return 100
        else:
            return 0
    else:
        if actual_value > satisfactory_level_value:
            if get_satisfactory_value == 0:
                return 1
            else:
                return 100
        else:
            return 0


def fuzzy_interval(cost_per_container_per_km, time_ratio, reliability, trans, emission, cost_im, time_im, reliability_im, trans_im, emission_im):
    if cost_per_container_per_km > 0 and cost_per_container_per_km < 0.3:  # 改
        cost_sa = np.array([0.7, 0.9, 0.9, 1.0])
    if cost_per_container_per_km > 0.3 and cost_per_container_per_km < 0.6:
        cost_sa = np.array([0.5, 0.7, 0.7, 0.9])
    if cost_per_container_per_km > 0.6 and cost_per_container_per_km < 0.9:
        cost_sa = np.array([0.3, 0.5, 0.5, 0.7])
    if cost_per_container_per_km > 0.9 and cost_per_container_per_km < 1.2:
        cost_sa = np.array([0.1, 0.3, 0.3, 0.5])
    if cost_per_container_per_km > 1.2 and cost_per_container_per_km < 1.5:
        cost_sa = np.array([0, 0.1, 0.1, 0.3])

    if time_ratio > 0 and time_ratio < 0.05:  # 改
        time_sa = np.array([0.7, 0.9, 0.9, 1.0])
    if time_ratio > 0.05 and time_ratio < 0.1:
        time_sa = np.array([0.5, 0.7, 0.7, 0.9])
    if time_ratio > 0.1 and time_ratio < 0.2:
        time_sa = np.array([0.3, 0.5, 0.5, 0.7])
    if time_ratio > 0.2 and time_ratio < 0.3:
        time_sa = np.array([0.1, 0.3, 0.3, 0.5])
    if time_ratio > 0.3 and time_ratio < 0.4:
        time_sa = np.array([0, 0.1, 0.1, 0.3])

    if reliability > 0 and reliability < 0.05:  # 改
        reliability_sa = np.array([0.7, 0.9, 0.9, 1.0])
    if reliability > 0.05 and reliability < 0.1:
        reliability_sa = np.array([0.5, 0.7, 0.7, 0.9])
    if reliability > 0.1 and reliability < 0.2:
        reliability_sa = np.array([0.3, 0.5, 0.5, 0.7])
    if reliability > 0.2 and reliability < 0.3:
        reliability_sa = np.array([0.1, 0.3, 0.3, 0.5])
    if reliability > 0.3 and reliability < 0.4:
        reliability_sa = np.array([0, 0.1, 0.1, 0.3])

    if trans > 0 and trans < 0.05:  # 改
        trans_sa = np.array([0.7, 0.9, 0.9, 1.0])
    if trans > 0.05 and trans < 0.1:
        trans_sa = np.array([0.5, 0.7, 0.7, 0.9])
    if trans > 0.1 and trans < 0.2:
        trans_sa = np.array([0.3, 0.5, 0.5, 0.7])
    if trans > 0.2 and trans < 0.3:
        trans_sa = np.array([0.1, 0.3, 0.3, 0.5])
    if trans > 0.3 and trans < 0.4:
        trans_sa = np.array([0, 0.1, 0.1, 0.3])

    if emission > 0 and emission < 0.05:  # 改
        emission_sa = np.array([0.7, 0.9, 0.9, 1.0])
    if emission > 0.05 and emission < 0.1:
        emission_sa = np.array([0.5, 0.7, 0.7, 0.9])
    if emission > 0.1 and emission < 0.2:
        emission_sa = np.array([0.3, 0.5, 0.5, 0.7])
    if emission > 0.2 and emission < 0.3:
        emission_sa = np.array([0.1, 0.3, 0.3, 0.5])
    if emission > 0.3 and emission < 0.4:
        emission_sa = np.array([0, 0.1, 0.1, 0.3])

    index_r = list(R[:, 7]).index(r)
    if R[index_r, 7] == 1:
        cost_im = np.array([0.7, 0.9, 0.9, 1.0])
    if R[index_r, 7] == 2:
        cost_im = np.array([0.5, 0.7, 0.7, 0.9])
    if R[index_r, 7] == 3:
        cost_im = np.array([0.3, 0.5, 0.5, 0.7])
    if R[index_r, 7] == 4:
        cost_im = np.array([0, 0.1, 0.1, 0.3])
    if R[index_r, 7] == 5:
        cost_im = np.array([0, 0.1, 0.1, 0.3])

        # fuzzy rating vectors for cost
    if R[index_r, 7] == 1:
        time_im = np.array([0.7, 0.9, 0.9, 1.0])
    if R[index_r, 7] == 2:
        time_im = np.array([0.5, 0.7, 0.7, 0.9])
    if R[index_r, 7] == 3:
        time_im = np.array([0.3, 0.5, 0.5, 0.7])
    if R[index_r, 7] == 4:
        time_im = np.array([0, 0.1, 0.1, 0.3])
    if R[index_r, 7] == 5:
        time_im = np.array([0, 0.1, 0.1, 0.3])

    if R[index_r, 7] == 1:
        reliability_im = np.array([0.7, 0.9, 0.9, 1.0])
    if R[index_r, 7] == 2:
        reliability_im = np.array([0.5, 0.7, 0.7, 0.9])
    if R[index_r, 7] == 3:
        reliability_im = np.array([0.3, 0.5, 0.5, 0.7])
    if R[index_r, 7] == 4:
        reliability_im = np.array([0, 0.1, 0.1, 0.3])
    if R[index_r, 7] == 5:
        reliability_im = np.array([0, 0.1, 0.1, 0.3])

    if R[index_r, 7] == 1:
        trans_im = np.array([0.7, 0.9, 0.9, 1.0])
    if R[index_r, 7] == 2:
        trans_im = np.array([0.5, 0.7, 0.7, 0.9])
    if R[index_r, 7] == 3:
        trans_im = np.array([0.3, 0.5, 0.5, 0.7])
    if R[index_r, 7] == 4:
        trans_im = np.array([0, 0.1, 0.1, 0.3])
    if R[index_r, 7] == 5:
        trans_im = np.array([0, 0.1, 0.1, 0.3])

    if R[index_r, 7] == 1:
        emission_im = np.array([0.7, 0.9, 0.9, 1.0])
    if R[index_r, 7] == 2:
        emission_im = np.array([0.5, 0.7, 0.7, 0.9])
    if R[index_r, 7] == 3:
        emission_im = np.array([0.3, 0.5, 0.5, 0.7])
    if R[index_r, 7] == 4:
        emission_im = np.array([0, 0.1, 0.1, 0.3])
    if R[index_r, 7] == 5:
        emission_im = np.array([0, 0.1, 0.1, 0.3])


    request_sa_vector = (cost_sa * cost_im) + (time_sa * time_im) + (trans_sa * trans_im) + (reliability_sa * reliability_im) + (emission_sa * emission_im)
    satisfaction = (request_sa_vector[0] + request_sa_vector[1] + request_sa_vector[2] + request_sa_vector[3]) / 4

    return satisfaction



# this idea is not used because it hard to insert a r which is best in history but inserting it will conflict with other r
# can I insert r with lowest cost if it has no conflict with current r? then it will be greedy
# def learn_from_experience():

def preference_constraints(r,k1,k2,k3,best_route1,best_route2,best_route3,get_satisfactory_value=0,get_objectives=0):
    # if get_satisfactory_value == 0:
    #     #then all r in the route need to be checked
    #     if k
    # after_iteration = 0,
    # if emission_preference_constraints_after_iteration == 1 and after_iteration == 0:
    #     return 1

    index_r=list(R[:,7]).index(r)
    insert_r_cost_1, insert_r_emissions_1,insert_r_cost_2, insert_r_emissions_2,insert_r_cost_3, insert_r_emissions_3 = 0,0,0,0,0,0
    transshipment_times = 0


    k1_distance, k2_distance, k3_distance = 0,0,0
    if k1 != -1:
        pd = []
        for i in range(len(best_route1[0])):
            if str(r) in best_route1[4][i]:
                pd.append(best_route1[0][i])
            if len(pd) == 2:
                break
        k1_distance = D[k1][pd[0], pd[1]]
        pd = []
        if k2 != -1:
            for i in range(len(best_route2[0])):
                if str(r) in best_route2[4][i]:
                    pd.append(best_route2[0][i])
                if len(pd) == 2:
                    break
            k2_distance = D[k2][pd[0], pd[1]]
            pd = []
            if k3 != -1:
                for i in range(len(best_route3[0])):
                    if str(r) in best_route3[4][i]:
                        pd.append(best_route3[0][i])
                    if len(pd) == 2:
                        break
                k3_distance = D[k3][pd[0], pd[1]]
    # k1_distance = pd.append(best_route1[1][i])
    true_distance = k1_distance + k2_distance + k3_distance

    if k1 != -1:
        all_objs_1 = objective_value_i(r, k1, best_route1)
        insert_r_cost_1, insert_r_emissions_1 = all_objs_1[0], all_objs_1[2]
        # if only_eco_label == 0:
        #     all_objs_1 = objective_value_i(r, k1, best_route1)
        #     insert_r_cost_1, insert_r_emissions_1 = all_objs_1[0], all_objs_1[2]
        # else:
        #     if only_eco_label_add_cost == 1:
        #         all_objs_1 = objective_value_i(r, k1, best_route1)
        #         insert_r_cost_1, insert_r_emissions_1 = all_objs_1[0], all_objs_1[2]
        #     else:
        #         insert_r_emissions_1 = objective_value_i(r, k1, best_route1)[2]
    else:
        return 0
    if k2 != -1:
        all_objs_2 = objective_value_i(r, k2, best_route2)
        insert_r_cost_2, insert_r_emissions_2 = all_objs_2[0], all_objs_2[2]
        transshipment_times = transshipment_times + 1
        # if only_eco_label == 0:
        #     all_objs_2 = objective_value_i(r, k2, best_route2)
        #     insert_r_cost_2, insert_r_emissions_2 = all_objs_2[0], all_objs_2[2]
        #     transshipment_times = transshipment_times + 1
        # else:
        #     if only_eco_label_add_cost == 1:
        #         all_objs_2 = objective_value_i(r, k2, best_route2)
        #         insert_r_cost_2, insert_r_emissions_2 = all_objs_2[0], all_objs_2[2]
        #     else:
        #         insert_r_emissions_2 = objective_value_i(r, k2, best_route2)[2]
    if k3 != -1:
        all_objs_3 = objective_value_i(r, k3, best_route3)
        insert_r_cost_3, insert_r_emissions_3 = all_objs_3[0], all_objs_3[2]
        transshipment_times = transshipment_times + 1
        # if only_eco_label == 0:
        #     all_objs_3 = objective_value_i(r, k3, best_route3)
        #     insert_r_cost_3, insert_r_emissions_3 = all_objs_3[0], all_objs_3[2]
        #     transshipment_times = transshipment_times + 1
        # else:
        #     if only_eco_label_add_cost == 1:
        #         all_objs_3 = objective_value_i(r, k3, best_route3)
        #         insert_r_cost_3, insert_r_emissions_3 = all_objs_3[0], all_objs_3[2]
        #     else:
        #         insert_r_emissions_3 = objective_value_i(r, k3, best_route3)[2]
    transshipment_times = transshipment_times * R[index_r, 6]
    # if only_eco_label == 0:
    if insert_r_cost_1 == 10000000000000000000 or insert_r_cost_2 == 10000000000000000000 or insert_r_cost_3 == 10000000000000000000:
        return 0
    cost = insert_r_cost_1 + insert_r_cost_2 + insert_r_cost_3
    cost_per_container_per_km = cost/R[index_r,6]/true_distance
    # speed = D_origin_All[R[index_r, 0], R[index_r, 1]] / (request_flow_t[index_r, 5] - request_flow_t[index_r, 0])
    #the following time_ratio uses actual_time/time_window, but in practice time window usually is set longer than expected time, so it's not approriate
    # time_ratio = (request_flow_t[index_r, 5] - request_flow_t[index_r, 0]) / (R[index_r, 5] - R[index_r, 2])
    #so now I use actual time/(distance/average speed of all vehicles)
    time_ratio = (request_flow_t[index_r, 5] - request_flow_t[index_r, 0]) / (true_distance/25)
    # delay_time = request_flow_t[index_r, 5] - R[index_r, 5]
    delay_time_ratio = max(0, (request_flow_t[index_r, 5] - R[index_r, 5]) / (request_flow_t[index_r, 5] - request_flow_t[index_r, 0]))
    # else:
    #     cost_per_container_per_km, time_ratio, delay_time_ratio = 0, 0, 0
    emissions = insert_r_emissions_1 + insert_r_emissions_2 + insert_r_emissions_3
    #20211121 when in solve_relevant_try() function, if the initial request (who trigger the chain reaction) is served by more than 1 vehicle, it will check a request segment of the initial request, and the emission is the segment emimission but distance is the overall distance. But it doesn't matter, because preference constraints of this reqeust is checked before, and in this function, it will satisfied because emission is smaller than the overall emissions and distance is the same
        #but it still needs exist, because it can detect severe changes of this request which make it can't satisfy preference, especially when only use one vehicle to serve this request
    emissions_per_container_per_km = emissions / R[index_r, 6]/true_distance

    # time1_unit,time2_unit,time3_unit=0,0,0
    # # if value request_flow_t[index_r, 1] exists, then at least 1T is used
    # if not np.isnan(request_flow_t[index_r, 1]):
    #     time1 = request_flow_t[index_r, 1] - request_flow_t[index_r, 0]
    #     time1_unit = time1/K[k1,1]
    # else:
    #     time1 = request_flow_t[index_r, 5] - request_flow_t[index_r, 0]
    #     time1_unit = time1 / K[k1, 1]
    # #if value request_flow_t[index_r, 3] exists, then 2T is used
    # if not np.isnan(request_flow_t[index_r, 3]):
    #     time2 = request_flow_t[index_r, 3] - request_flow_t[index_r, 2]
    #     time2_unit = time1/K[k2,1]
    # else:
    #     # if value request_flow_t[index_r, 3] is nan, but request_flow_t[index_r, 2] exists, then 1T is used
    #     if not np.isnan(request_flow_t[index_r, 2]):
    #         time2 = request_flow_t[index_r, 5] - request_flow_t[index_r, 2]
    #         time2_unit = time1 / K[k2, 1]
    # if not np.isnan(request_flow_t[index_r, 4]):
    #     time3 = request_flow_t[index_r, 5] - request_flow_t[index_r, 4]
    #     time3_unit = time1/K[k3,1]
    # time = (time1_unit+time2_unit+time3_unit)/3
    if get_objectives == 1:
        return cost_per_container_per_km, time_ratio, emissions_per_container_per_km, delay_time_ratio, transshipment_times
    cost_preference, time_ratio_preference, delay_time_ratio_preference, transshipment_preference, emissions_preference = R[index_r,9:14]
    if get_satisfactory_value == 0:
        if fuzzy_constraints == 1:

            if only_eco_label == 0:

                if accept_function(r, 'Cost', cost_per_container_per_km,cost_preference) == 0 or accept_function(r, 'Time', time_ratio,time_ratio_preference) == 0 or\
                        accept_function(r, 'Delay', delay_time_ratio, delay_time_ratio_preference) == 0 or accept_function(r, 'Transshipment', transshipment_times,transshipment_preference) == 0 or\
                                accept_function(r, 'Emissions', emissions_per_container_per_km, emissions_preference) == 0:
                    return 0
            else:
                if only_eco_label_add_cost == 1:
                    if accept_function(r, 'Cost', cost_per_container_per_km, cost_preference) == 0 or accept_function(r, 'Emissions', emissions_per_container_per_km, emissions_preference) == 0:
                        return 0
                else:
                    if accept_function(r, 'Emissions', emissions_per_container_per_km, emissions_preference) == 0:
                        return 0
        else:
            if only_eco_label == 0:

                if hard_preference_constraints('Cost', cost_per_container_per_km, cost_preference) == 0 or \
                        hard_preference_constraints('Time', time_ratio, time_ratio_preference) == 0 or \
                        hard_preference_constraints('Delay', delay_time_ratio, delay_time_ratio_preference) == 0 or \
                        hard_preference_constraints('Transshipment', transshipment_times, transshipment_preference) == 0 or \
                        hard_preference_constraints('Emissions', emissions_per_container_per_km, emissions_preference) == 0:
                    return 0
            else:
                if only_eco_label_add_cost == 1:
                    if hard_preference_constraints('Cost', cost_per_container_per_km, cost_preference) == 0 or hard_preference_constraints('Emissions', emissions_per_container_per_km, emissions_preference) == 0:
                        return 0
                else:
                    if hard_preference_constraints('Emissions', emissions_per_container_per_km, emissions_preference) == 0:
                        return 0
        return 1
    else:

        if fuzzy_constraints == 1:
            if only_eco_label == 0:
                if get_satisfactory_value_one_by_one == 1:
                    satisfactory_value = accept_function(r, 'Cost', cost_per_container_per_km,cost_preference,1)+ accept_function(r, 'Time', time_ratio,time_ratio_preference,1)+\
                            accept_function(r, 'Delay', delay_time_ratio, delay_time_ratio_preference,1)+ accept_function(r, 'Transshipment', transshipment_times,transshipment_preference,1)+\
                                    accept_function(r, 'Emissions', emissions_per_container_per_km, emissions_preference,1)
                else:
                    satisfactory_value = fuzzy_HP.five_attributes_to_satisfactory(cost_preference, time_ratio_preference, delay_time_ratio_preference, transshipment_preference, emissions_preference, cost_per_container_per_km, time_ratio, delay_time_ratio, transshipment_times, emissions_per_container_per_km)
            else:
                if only_eco_label_add_cost == 1:
                    if get_satisfactory_value_one_by_one == 1:
                        satisfactory_value = accept_function(r, 'Cost', cost_per_container_per_km, cost_preference, 1) +  accept_function(r, 'Emissions', emissions_per_container_per_km,
                                                             emissions_preference, 1)
                    else:
                        satisfactory_value = fuzzy_HP.five_attributes_to_satisfactory(cost_preference,
                                                                                      time_ratio_preference,
                                                                                      delay_time_ratio_preference,
                                                                                      transshipment_preference,
                                                                                      emissions_preference,
                                                                                      cost_per_container_per_km,
                                                                                      0, 0,
                                                                                      0,
                                                                                      emissions_per_container_per_km,only_eco_label)
                else:
                    if get_satisfactory_value_one_by_one == 1:
                        satisfactory_value = accept_function(r, 'Emissions', emissions_per_container_per_km, emissions_preference,1)
                    else:
                        satisfactory_value = fuzzy_HP.five_attributes_to_satisfactory(cost_preference,
                                                                                      time_ratio_preference,
                                                                                      delay_time_ratio_preference,
                                                                                      transshipment_preference,
                                                                                      emissions_preference,
                                                                                      0, 0,
                                                                                      0, 0,
                                                                                      emissions_per_container_per_km,only_eco_label)
        else:
            if only_eco_label == 0:

                satisfactory_value = hard_preference_constraints('Cost',cost_per_container_per_km, cost_preference, 1)+ \
                        hard_preference_constraints('Time',time_ratio, time_ratio_preference, 1)+ \
                            hard_preference_constraints('Delay', delay_time_ratio, delay_time_ratio_preference, 1)+ \
                            hard_preference_constraints('Transshipment', transshipment_times, transshipment_preference, 1)+ \
                            hard_preference_constraints('Emissions', emissions_per_container_per_km, emissions_preference, 1)
                if satisfactory_value < 500:
                    # one attribute is not be satisfied
                    satisfactory_value = 0
                if get_satisfactory_value_one_by_one == 0:
                    satisfactory_value = satisfactory_value / 5

            else:
                if only_eco_label_add_cost == 1:
                    satisfactory_value = hard_preference_constraints('Cost', cost_per_container_per_km, cost_preference,
                                                                     1) + hard_preference_constraints('Emissions', emissions_per_container_per_km, emissions_preference, 1)
                    if satisfactory_value < 200:
                        # one attribute is not be satisfied
                        satisfactory_value = 0
                    if get_satisfactory_value_one_by_one == 0:
                        satisfactory_value = satisfactory_value / 2
                else:
                    satisfactory_value = hard_preference_constraints('Emissions', emissions_per_container_per_km, emissions_preference, 1)



        if only_eco_label == 0:
            if accept_function(r, 'Cost', cost_per_container_per_km, cost_preference) == 0 or accept_function(r, 'Time',time_ratio,time_ratio_preference) == 0 or \
                    accept_function(r, 'Delay', delay_time_ratio, delay_time_ratio_preference) == 0 or accept_function(r, 'Transshipment',transshipment_times,transshipment_preference) == 0 or \
                    accept_function(r, 'Emissions', emissions_per_container_per_km, emissions_preference) == 0:
                fuzzy_satisfy_or_not = 0
            else:
                fuzzy_satisfy_or_not = 1
            if hard_preference_constraints('Cost', cost_per_container_per_km, cost_preference) == 0 or \
                    hard_preference_constraints('Time', time_ratio, time_ratio_preference) == 0 or \
                    hard_preference_constraints('Delay', delay_time_ratio, delay_time_ratio_preference) == 0 or \
                    hard_preference_constraints('Transshipment', transshipment_times, transshipment_preference) == 0 or \
                    hard_preference_constraints('Emissions', emissions_per_container_per_km, emissions_preference) == 0:
                hard_satisfy_or_not = 0
            else:
                hard_satisfy_or_not = 1
        else:
            if only_eco_label_add_cost == 1:
                if accept_function(r, 'Cost', cost_per_container_per_km, cost_preference) == 0 or accept_function(r, 'Emissions', emissions_per_container_per_km, emissions_preference) == 0:
                    fuzzy_satisfy_or_not = 0
                else:
                    fuzzy_satisfy_or_not = 1
                if hard_preference_constraints('Cost', cost_per_container_per_km, cost_preference) == 0 or hard_preference_constraints('Emissions', emissions_per_container_per_km, emissions_preference) == 0:
                    hard_satisfy_or_not = 0
                else:
                    hard_satisfy_or_not = 1
            else:
                if accept_function(r, 'Emissions', emissions_per_container_per_km, emissions_preference) == 0:
                    fuzzy_satisfy_or_not = 0
                else:
                    fuzzy_satisfy_or_not = 1
                if hard_preference_constraints('Emissions', emissions_per_container_per_km, emissions_preference) == 0:
                    hard_satisfy_or_not = 0
                else:
                    hard_satisfy_or_not = 1
        return satisfactory_value, fuzzy_satisfy_or_not, hard_satisfy_or_not





    # @profile()
# @time_me()
##@jit
def satisfy_constraints(routes, has_end_depot, R, k, route, fixed_vehicles_percentage, K, no_route_barge,
                        no_route_truck,inserted_r,remove=0,only_time = 0):
    global check_start_position, relevant_request_position_number
    # 20200927 mute this because I only insert r/T which i,j in fixed route
    # if Fixed_route(k, route) == False:
    #     return False
    if len(route[0]) == 2:
        return route, '_'
    if only_time == 0:
        if remove == 0:
            if Barge_no_land(k, route, fixed_vehicles_percentage, K, no_route_barge, no_route_truck) == False:
                return False, '_'
            if new_subtour_constraints(route[0]) == False:
                return False, '_'
        # if capacity_constraints(has_end_depot, K, R, k, route) == False:
        #     return False
    #if remvoe and k is truck, then no need to check time_constraints
    if remove == 0 or (remove==1 and K[k, 5] != 3):
        relevant_request_position_number = {}
        check_start_position = 0
        bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k, route,inserted_r)
        if isinstance(bool_or_route, bool):
            return False, 'wrong_time'
        else:
            route = bool_or_route

    return route, '_'

def RL_insertion_constraints(index, route, inserted_r, new_row, finish_or_begin, uncertainty_index, k, store_routes_for_another_k, store_R_pool_for_another_k,duration, congestion_link,
                                 congestion_node):

    global RL_insertion_segment
    RL_insertion_segment = 0
    index_r = list(R[:, 7]).index(inserted_r)
    congestion_nodes_in_this_route, use_RL_in_insertion = get_congestion_nodes_in_this_route(route, duration)
    # the r's number without labeling of T
    inserted_r = inserted_r - (inserted_r - inserted_r % 10000) % big_r
    if dynamic_RL34959.implement == 1 and finish_or_begin == 'finish':
        return route
    break_flag4, feasibility = check_uncertainty_in_insertion_by_RL(implement_or_not, congestion_nodes_in_this_route, route, inserted_r, index_r, new_row, finish_or_begin, uncertainty_index, k, inserted_r, store_routes_for_another_k, store_R_pool_for_another_k,duration, congestion_link,
                                 congestion_node, index, routes_store, request_flow_t_store, k)
    if feasibility == 0:
        return False
    # elif feasibility == -1:
    #     #it's still in learning mode
    #     pass
    # else:
    #     pass
    RL_insertion_segment = 1
    return route

# @profile()
# @time_me()
def Fixed_route(k, route):
    if k in fixed_vehicles_percentage:
        if len(route[4]) >= 2:
            no_duplicate_route = unique(route[0])
            fixed_route = list(Fixed[k][:,0])
            for terminal in no_duplicate_route:
                if terminal not in fixed_route:
                    return False
            compare_order = [x for x in fixed_route if x in no_duplicate_route]
            if no_duplicate_route != compare_order:
                return False


# @profile()
# @time_me()
def Barge_no_land(k, route, fixed_vehicles_percentage, K, no_route_barge, no_route_truck):
    # all trains have fixed timetable, so only restrict barge and truck
    if k not in fixed_vehicles_percentage and (K[k, 5] == 1 or K[k, 5] == 2) and len(route[4]) > 2:
        no_duplicate_route = unique(route[0])
        for x, j in zip(no_duplicate_route[0::1], no_duplicate_route[1::1]):
            if K[k, 5] == 1:
                for m in range(0, len(no_route_barge)):
                    if x == no_route_barge[m,0] and j == no_route_barge[m,1]:
                        return False
            if K[k, 5] == 3:
                for m in range(0, len(no_route_truck)):
                    if x == no_route_truck[m,0] and j == no_route_truck[m,1]:
                        return False
def dedupe_adjacent(alist):
    for i in range(len(alist) - 1, 0, -1):
        if alist[i] == alist[i-1]:
            del alist[i]
    return alist

def new_subtour_constraints(terminals):
    alist = dedupe_adjacent(terminals.tolist())
    if len(alist) != len(set(alist)):
        return False

# @profile()
# @time_me()
def subtour_constraints(route):
    # subtour
    transposed_route = route[0].T
    res = [x[0] for x in groupby(transposed_route.tolist())]
    res = pd.DataFrame(res)

    # danger when begin depot and end depot are same, should unmute this, but this will casuse that the terminal in route may same as end depot
    # begin depot and end depot can be same
    # but it can have the r's terminal is as same as end terminal
    # if has_end_depot == 1:
    #     res.drop(res.tail(1).index, inplace=True)

    xx = res.duplicated()

    if xx.any():
        return False


# @profile()
# @time_me()
def capacity_constraints(has_end_depot, K, R, k, route, swap_r_load=0, calculate_load = 0):
    load = 0 + swap_r_load
    load_list = [0]
    if has_end_depot == 1:
        length = len(route[4])
    else:
        length = len(route[4]) + 1
    for m in range(1, length - 1):
        if hasNumbers(route[4, m]):
            request_number = int(''.join(filter(str.isdigit, route[4, m])))
            index_r = list(R[:, 7]).index(request_number)
            letters = new_getLetters(route[4, m])

            if letters == 'pickup' or letters == 'Tp' or letters == 'secondTp':
                load = load + R[index_r, 6]
                load_list.append(load)
            else:
                load = load - R[index_r, 6]
                load_list.append(load)
            if calculate_load == 0:
                if load > K[k, 0]:
                    return False
    if calculate_load == 1:
        load_max = max(load_list)
        left_capacity = K[k, 0] - load_max
        return load_max, left_capacity

def get_relevant(k,request_number,last_letter):
    relevant_try = {}
    l_list = list(range(len(K)))
    l_list.remove(k)
    for l in l_list:
        if has_end_depot == 1:
            length = len(routes[l][4])
        else:
            length = len(routes[l][4]) + 1
        for n in range(1, length - 1):
            name = routes[l][4, n]
            if hasNumbers(name):
                request_number_else = int(''.join(filter(str.isdigit, name)))
                if request_number == request_number_else:
                    #in time_constraints_relevant, I have limit that no relevant when last letter is 'delivery',
                    #when there is 2T and 'secondTd', because it will only influence the 'secondTp', so I mute the case of 'pickup'
                    if last_letter == 'secondTd':
                        if getLetters(name) == 'pickup':
                            break
                    relevant_request_position_number[l] = [n, request_number]
                    relevant_try[l] = [copy.copy(routes[l]), request_number, n]
                    #check_relevant_try_not_in_routes()
                    break
    return relevant_request_position_number, relevant_try

def get_relevant_routes(relevant_request_position_number,k,route,inserted_r):

    relevant_try = {}
    relevant_request_position_number_copy = copy.copy(relevant_request_position_number)
    relevant_request_position_number = {}
    # find all relevant requests. The found requests will repeat and add into relevant_try more than one time,
    # but it doesn't matter because the same route will be covered.

    if K[k, 5] == 1 or K[k, 5] == 2 or truck_fleet == 0:
        # if k is fixed, and not truck, it's time will not influence other k, and the if r use T, the second k will also be checked in the second k's constraints checking
        if k not in fixed_vehicles_percentage:
            if has_end_depot == 1:
                length = len(route[4])
            else:
                length = len(route[4]) + 1

            for m in range(check_start_position, length - 1):
                if hasNumbers(route[4, m]):

                    request_number = int(''.join(filter(str.isdigit, route[4, m])))
                    # letters = new_getLetters(route[4, m])
                    two_letters, two_m = remove_T_k_in_record(route, request_number)
                    #when check time (include_itself == 0), the all times of k itself has been checked, so if the r in k only transferred by k (two_letters[1] == 'delivery'), the k should be not checked

                    if two_letters[1] != 'delivery':

                        if relevant_request_position_number_copy:
                            try:
                                if request_number != relevant_request_position_number_copy[k][1]:
                                    relevant_request_position_number, relevant_try = get_relevant(k, request_number,
                                                                                                  two_letters[1])
                            except:
                                relevant_request_position_number, relevant_try = get_relevant(k, request_number,
                                                                                              two_letters[1])
                        else:
                            relevant_request_position_number, relevant_try = get_relevant(k, request_number,
                                                                                          two_letters[1])
    else:
        # if truck, then only if r is served by more than one k, the other k are relevant_try
        # if not math.isnan(T_k_record[inserted_r,0]):
        #     for l in T_k_record[inserted_r,2:5]:
        #         if not math.isnan(l) and l != k :
        #             relevant_try[l] = [copy.copy(routes[l]), inserted_r]
        # else:
        # if no record
        two_letters, two_m = remove_T_k_in_record(route, inserted_r)
        #when checking preference constraints in remove_a_request, the inserted_r has been removed, so two_letters is empty. I don't know what the inserted_r should be what in this case so I use the same one with insertion
        if two_letters:
            if two_letters[1] != 'delivery':
                relevant_request_position_number, relevant_try = get_relevant(k, inserted_r, two_letters[1])
    return relevant_try


# @profile()
# @time_me()
##@jit
def time_constraints_relevant(has_end_depot, routes, K, k, route, inserted_r):
    global waiting_times, only_check_this_influenced_r_in_dynamic_uncertainty, wait, relevant_try, check_start_position, relevant_request_position_number, fixed_wait


    fixed_wait = 0
    wait = 0
    # 20200927 this is muted before, I guess it's because the check_start position is defined uper the time_constraints_relevant function
    # check_start_position=0
    if dynamic == 1 and dynamic_t > 0 and stochastic == 1:
        only_check_this_influenced_r_in_dynamic_uncertainty = inserted_r
    waiting_times = {}
    bool_time, route, infeasible_request_terminal = time_constraints(k, route,inserted_r)
    #    stop=0
    while wait == 1:
        bool_time, route, infeasible_request_terminal = time_constraints(k, route,inserted_r)
    waiting_times = {}
    #        stop=stop+1
    #        if stop>len(route[4]):
    #            return False
    #            sys.exit("aa! errors!")
    #        if stop>20:
    #            return False
    # if isinstance(time_constraints(k, route), pd.DataFrame):
    #     route=time_constraints(k, route)
    #    bool_time, route = time_constraints(k, route)

    if bool_time == True:

        relevant_try = get_relevant_routes(relevant_request_position_number,k,route,inserted_r)
        #check_relevant_try_not_in_routes()
        relevant_try_copy = my_deepcopy(relevant_try)
        for l in relevant_try:
            # 20200927 mute this because I only insert r/T which i,j in fixed route
            # if Fixed_route(l, relevant_try[l]) == False:
            #     return False

            # the relevant_try can also add wait time if it not satisfy constraints, but it will be too complex,
            # so it not be considered

            #            try:
            # 20200927 add this because I afraid the check_start_position use the one for vehilce k
            check_start_position = relevant_try[l][2]
            waiting_times = {}
            bool_time_relevant, route_relevant, infeasible_request_terminal = time_constraints(l, relevant_try[l][0], relevant_try[l][1])
            while wait == 1:
                bool_time_relevant, route_relevant, infeasible_request_terminal = time_constraints(l, relevant_try[l][0], relevant_try[l][1])
            waiting_times = {}
            #            except:
            #                sys.exit('sda')
            if bool_time_relevant == False:
                relevant_try = my_deepcopy(relevant_try_copy)
                return False, infeasible_request_terminal
            else:
                relevant_try[l] = [copy.copy(route_relevant), relevant_try[l][1], relevant_try[l][2]]
        # # if all relevant_try are satisfied
        # for l in relevant_try:
        #     bool_time_relevant, relevant_try[l] = time_constraints(l, relevant_try[l])
        return route, -1
    else:
        return False, infeasible_request_terminal


# @profile()
# @time_me()
def get_travel_time(x1, x2, y1, y2, departure_time):
    return (y1 - y2) / (x1 - x2) * departure_time + y1 - (y1 - y2) / (x1 - x2) * x1


# @profile()
# @time_me()
def get_travel_time_pre(departure_time, original_travel_time):
    global time_dependent_truck_travel_time
    if time_dependent_truck_travel_time == 0:
        return original_travel_time
    if departure_time <= b2 or departure_time >= b9:
        truck_travel_time = original_travel_time
    if departure_time > b2 and departure_time <= b3:
        truck_travel_time = get_travel_time(b2, b3, original_travel_time, alpha * original_travel_time, departure_time)
    if (departure_time > b3 and departure_time <= b4) or (departure_time > b7 and departure_time <= b8):
        truck_travel_time = alpha * original_travel_time
    if departure_time > b4 and departure_time <= b5:
        truck_travel_time = get_travel_time(b4, b5, alpha * original_travel_time, belta * original_travel_time,
                                            departure_time)
    if departure_time > b5 and departure_time <= b6:
        truck_travel_time = belta * original_travel_time
    if departure_time > b6 and departure_time <= b7:
        truck_travel_time = get_travel_time(b6, b7, belta * original_travel_time, alpha * original_travel_time,
                                            departure_time)
    if departure_time > b8 and departure_time < b9:
        truck_travel_time = get_travel_time(b8, b9, alpha * original_travel_time, original_travel_time, departure_time)
    return truck_travel_time

def remove_T_k_in_record(route,inserted_r):
    # global T_k_record

    two_letters, two_m = [], []
    if has_end_depot == 1:
        length = len(route[4])
    else:
        length = len(route[4]) + 1
    for m in range(1,length-1):
        z = route[4,m]
        check_r_use_T_r = int(''.join(filter(str.isdigit, z)))
        if check_r_use_T_r == inserted_r:
            two_letters.append(getLetters(z))
            two_m.append(m)
            if len(two_m) == 2:
                break

    return two_letters, two_m

def update_T_k_record_request_flow_t(routes_local):
    for k, route in routes_local.items():
        length = len(routes_local[k][0])
        if K[k, 5] == 1 or K[k, 5] == 2 or Demir == 1:
            for m in range(0, length - 1):
                if hasNumbers(route[4, m]):
                    request_number = int(''.join(filter(str.isdigit, route[4, m])))
                    index_r = list(R[:, 7]).index(request_number)
                    letters = new_getLetters(route[4, m])
                    if letters == 'Td':
                        # transshipment time for Td
                        T_k_record[index_r, 0] = route[0, m]

                        request_flow_t[index_r, 1] = route[3, m]
                    if letters == 'secondTd':
                        # transshipment time for secondTd
                        T_k_record[index_r, 1] = route[0, m]
                        request_flow_t[index_r, 3] = route[3, m]

                    #            wait = 0
                    if letters == 'pickup':

                        request_flow_t[index_r, 0] = route[2, m]
                        T_k_record[index_r, 2] = k


                    if letters == 'Tp':
                        request_flow_t[index_r, 2] = route[2, m]
                        T_k_record[index_r, 0] = route[0, m]
                        T_k_record[index_r, 3] = k


                    if letters == 'secondTp':
                        request_flow_t[index_r, 4] = route[2, m]
                        T_k_record[index_r, 1] = route[0, m]
                        T_k_record[index_r, 4] = k


                    if letters == 'delivery':

                        request_flow_t[index_r, 5] = route[2, m]
        # truck fleet
        else:
            served_r = find_r_served_by_k(route)
            for inserted_r in served_r:
                two_letters, two_m = remove_T_k_in_record(route, inserted_r)
                request_number = inserted_r
                index_r = list(R[:, 7]).index(request_number)
                letters = two_letters[0]

                m = two_m[0]
                if letters == 'pickup':

                    T_k_record[index_r, 2] = k
                    request_flow_t[index_r, 0] = route[2, m]

                # danger if there are 2T, time of Tp may need same with time in secondTd
                # danger the second truck maybe can't arrive T on time
                if letters == 'Tp':

                    request_flow_t[index_r, 2] = route[2, m]
                    T_k_record[index_r, 0] = route[0, m]
                    T_k_record[index_r, 3] = k
                if letters == 'secondTp':

                    request_flow_t[index_r, 4] = route[2, m]
                    T_k_record[index_r, 1] = route[0, m]
                    T_k_record[index_r, 4] = k

                if letters == 'delivery':
                    # remove T in T_rcord if request_number not use T
                    if two_letters[0] == 'pickup':
                        T_k_record[index_r] = np.nan
                    else:
                        if two_letters[0] == 'Tp':
                            T_k_record[index_r, 1] = np.nan
                            T_k_record[index_r, 4] = np.nan

                    # T_k_record T1,T2,k1,k2,k3
                    # 2T


                    request_flow_t[index_r, 5] = route[2, m]
                if letters == 'Td':
                    request_flow_t[index_r, 1] = route[3, m]
                    T_k_record[index_r, 0] = route[0, m]
                if letters == 'secondTd':

                    T_k_record[index_r, 1] = route[0, m]
                    request_flow_t[index_r, 3] = route[3, m]


def time_constraints_only_check(k, route, inserted_r):
    global wait, wait_time, check_start_position, fixed_wait, request_flow_t, service_time, transshipment_time
    index_inserted_r = list(R[:, 7]).index(inserted_r)
    # when remove, it may empty
    if len(route[0]) < 3:
        return True, route


    # change wait to 0 when the vehicle is impossible to transport the request, even wait time was added; or the vehicle pass the constraint
    if has_end_depot == 1:
        length = len(route[4])
    else:
        length = len(route[4]) + 1

    fixed = 0


    request_flow_t_copy = copy.copy(request_flow_t)

    if K[k, 5] == 1 or K[k, 5] == 2 or Demir == 1:
        for m in range(check_start_position, length - 1):
            if hasNumbers(route[4, m]):
                request_number = int(''.join(filter(str.isdigit, route[4, m])))
                index_r = list(R[:, 7]).index(request_number)
                letters = new_getLetters(route[4, m])

                # when there is double wait, i.e., both fixed timetable and earlier arrival need add wait time, the time shouldn't be refresh, because the wait time of fixed timetable may makes the wait_time less than actual wait_time due to earlier arrival, then the earlier arrival will never be make up and stuck into dead loop

                # open window of fixed routes' terminals
                if fixed == 1:
                    if Demir == 1:
                        service_time = N[route[0, m], 1] * R[index_inserted_r, 6]
                        departure_time = Fixed[k][Fixed[k][:, 0] == route[0, m], 1][0]
                        real_departure_time = route[2, m] + service_time
                    else:
                        # try:
                        departure_time = Fixed[k][Fixed[k][:, 0] == route[0, m], 1][0] + service_time
                        # except:
                        #     print('caught_it')
                        if K[k, 5] == 1 or K[k, 5] == 2:
                            real_departure_time = route[2, m] + service_time
                        else:
                            real_departure_time = route[2, m]
                    # to make the departure time totally same with departure time of fixed k
                    if real_departure_time < departure_time:

                        fixed_wait = 1
                        wait = 1
                        # wait_time = departure_time - route[1, m] + 0.000001
                        if K[k, 5] == 1 or K[k, 5] == 2:
                            wait_time = departure_time - route[2, m] - service_time
                        else:
                            wait_time = departure_time - route[2, m]
                        check_start_position = m
                        return False, route
                    else:
                        # try:
                        departure_final_time = Fixed[k][Fixed[k][:, 0] == route[0, m], 2][0]
                        # except:
                        #     print('find error')
                        #     sys.exit(7)
                        wait = 0
                        # fixed:
                        # pickup a-1 a
                        # delivery b b+1
                        # route:
                        # pickup delivery
                        # a-1 b
                        # a b+1

                        # so real_departure_time can't bigger than departure_final_time
                        if Demir == 1 and (letters == 'delivery' or letters == 'Td' or letters == 'secondTd'):
                            if real_departure_time > departure_final_time + service_time:
                                request_flow_t = copy.copy(request_flow_t_copy)
                                return False, route
                        else:
                            if real_departure_time > departure_final_time:
                                request_flow_t = copy.copy(request_flow_t_copy)
                                return False, route

                if letters == 'Td':
                    # transshipment time for Td
                    if Demir == 1:

                        T_k_record[index_r, 0] = route[0, m]

                    request_flow_t[index_r, 1] = route[3, m]
                if letters == 'secondTd':
                    # transshipment time for secondTd
                    if Demir == 1:

                        T_k_record[index_r, 1] = route[0, m]

                    request_flow_t[index_r, 3] = route[3, m]

                #            wait = 0
                if letters == 'pickup':

                    request_flow_t[index_r, 0] = route[2, m]
                    if Demir == 1:
                        T_k_record[index_r, 2] = k

                    if request_flow_t[index_r, 0] < R[index_r, 2]:
                        wait = 1
                        wait_time = R[index_r, 2] - request_flow_t[index_r, 0] + 0.000001
                        check_start_position = m
                        return False, route
                    if Demir != 1 and request_flow_t[index_r, 0] + service_time > R[index_r, 3]:
                        # above last is R[index_r,3] because the containers can only be stored in the pickup time window, if exceed, then can't pickup

                        wait = 0
                        request_flow_t = copy.copy(request_flow_t_copy)
                        return False, route

                if letters == 'Tp':
                    request_flow_t[index_r, 2] = route[2, m]
                    if Demir == 1:
                        T_k_record[index_r, 0] = route[0, m]
                        T_k_record[index_r, 3] = k
                    # after the inserted request's route's has other request's Tp/secondTp was considered,
                    # it also has probability that in relevant routes, similiar situation may happen, which will not be considered
                    if pd.isnull(request_flow_t[index_r, 1]):
                        wait = 0
                        request_flow_t = copy.copy(request_flow_t_copy)
                        return False, route
                    if request_flow_t[index_r, 2] < request_flow_t[index_r, 1]:
                        wait = 1
                        wait_time = request_flow_t[index_r, 1] - request_flow_t[index_r, 2] + 0.000001
                        check_start_position = m
                        return False, route
                    else:

                        request_flow_t[index_r, 2] = route[2, m]

                if letters == 'secondTp':
                    request_flow_t[index_r, 4] = route[2, m]
                    if Demir == 1:
                        T_k_record[index_r, 1] = route[0, m]
                        T_k_record[index_r, 4] = k
                    # after the inserted request's route's has other request's Tp/secondTp was considered,
                    # it also has probability that in relevant routes, similiar situation may happen, which will not be considered
                    if pd.isnull(request_flow_t[index_r, 3]):
                        wait = 0
                        request_flow_t = copy.copy(request_flow_t_copy)
                        return False, route
                    if request_flow_t[index_r, 4] < request_flow_t[index_r, 3]:
                        wait = 1
                        wait_time = request_flow_t[index_r, 3] - request_flow_t[index_r, 4] + 0.000001
                        check_start_position = m
                        return False, route
                    else:

                        request_flow_t[index_r, 4] = route[2, m]

                if letters == 'delivery':

                    request_flow_t[index_r, 5] = route[2, m]

                    if fixed == 1:
                        # if the route has any signal that the time exceed the arrival final time + 1 of fixed route, then infeasible
                        arrival_final_time = Fixed[k][Fixed[k][:, 0] == route[0, m], 2][0]
                        if Demir == 1:
                            if route[2, m] > arrival_final_time:
                                wait = 0
                                request_flow_t = copy.copy(request_flow_t_copy)
                                return False, route
                        else:
                            if route[3, m] > arrival_final_time:
                                wait = 0
                                request_flow_t = copy.copy(request_flow_t_copy)
                                return False, route
                    if forbid_much_delay == 1:
                        if route[3, m] > R[index_r, 5] + 2:
                            wait = 0
                            request_flow_t = copy.copy(request_flow_t_copy)
                            return False, route
                    # allow delay
                    # else:
                    #     if Demir == 1:
                    #         service_time = N[route[0, m], 1] * R[index_inserted_r, 6]
                    #
                    #     if route[3, m] > R[index_r, 5] + service_time:
                    #         wait = 0
                    #         request_flow_t = copy.copy(request_flow_t_copy)
                    #         return False, route
                # else:
                #     # seems repeat with the begining fixed constraints check 20201106
                #     if fixed == 1:
                #         departure_final_time = Fixed[k][Fixed[k][:,0] == route[0, m],2][0]
                #
                #         if route[3, m] > departure_final_time:
                #             wait = 0
                #             request_flow_t = copy.copy(request_flow_t_copy)
                #             return False, route
                fixed_wait = 0

            #wait = 0
    # truck fleet
    else:
        two_letters, two_m = remove_T_k_in_record(route, inserted_r)
        request_number = inserted_r
        index_r = list(R[:, 7]).index(request_number)
        letters = two_letters[0]

        m = two_m[0]
        if letters == 'pickup':

            T_k_record[index_r, 2] = k
            request_flow_t[index_r, 0] = route[2, m]

        # danger if there are 2T, time of Tp may need same with time in secondTd
        # danger the second truck maybe can't arrive T on time
        if letters == 'Tp':
            if pd.isnull(request_flow_t[index_r, 1]):
                wait = 0
                request_flow_t = copy.copy(request_flow_t_copy)
                return False, route

            request_flow_t[index_r, 2] = route[2, m]
            T_k_record[index_r, 0] = route[0, m]
            T_k_record[index_r, 3] = k
        if letters == 'secondTp':
            if pd.isnull(request_flow_t[index_r, 3]):
                wait = 0
                request_flow_t = copy.copy(request_flow_t_copy)
                return False, route

            request_flow_t[index_r, 4] = route[2, m]
            T_k_record[index_r, 1] = route[0, m]
            T_k_record[index_r, 4] = k
        letters = two_letters[1]
        m = two_m[1]

        if letters == 'delivery':
            # remove T in T_rcord if request_number not use T
            if two_letters[0] == 'pickup':
                T_k_record[index_r] = np.nan
            else:
                if two_letters[0] == 'Tp':
                    T_k_record[index_r, 1] = np.nan
                    T_k_record[index_r, 4] = np.nan

            # T_k_record T1,T2,k1,k2,k3
            # 2T
            if not math.isnan(T_k_record[index_r, 1]):

                if pd.isnull(request_flow_t[index_r, 4]):
                    wait = 0
                    request_flow_t = copy.copy(request_flow_t_copy)
                    return False, route


            else:
                # 1T
                if not math.isnan(T_k_record[index_r, 0]):
                    if pd.isnull(request_flow_t[index_r, 2]):
                        wait = 0
                        request_flow_t = copy.copy(request_flow_t_copy)
                        return False, route




            # can't delay more than 1 hour. During the optimization, the r's delivery time may influenced by other newly inserted r, so I add this
            if forbid_much_delay == 1:
                if route[3, m] > R[index_r, 5] + 2:
                    wait = 0
                    request_flow_t = copy.copy(request_flow_t_copy)
                    return False, route
            request_flow_t[index_r, 5] = route[2, m]
        if letters == 'Td':

            request_flow_t[index_r, 1] = route[3, m]
            T_k_record[index_r, 0] = route[0, m]
        if letters == 'secondTd':
            if pd.isnull(request_flow_t[index_r, 2]):
                wait = 0
                request_flow_t = copy.copy(request_flow_t_copy)
                return False, route

            T_k_record[index_r, 1] = route[0, m]
            request_flow_t[index_r, 3] = route[3, m]

    return True, route

def check_delay_and_return_under_uncertainty(k, route, request_number, m, index_r, request_flow_t_copy):
    global wait, request_flow_t, dynamic_time_false, re_plan_when_event_finishes
    if RL_is_trained_or_evaluated_or_ALNS_is_evaluated == 1:
        for operation_index in range(len(route[4, 1:-1])):
            if request_number == get_numbers(route[4, operation_index + 1]):
                operation_type = new_getLetters(route[4, operation_index + 1])
                pickup_time_of_this_r_at_this_route = route[2, operation_index + 1]
                break
        # if not pd.isnull(request_flow_t[index_r, 4]):
        #     pickup_time_of_this_r_at_this_route = request_flow_t[index_r, 4]
        # elif not pd.isnull(request_flow_t[index_r, 2]):
        #     pickup_time_of_this_r_at_this_route = request_flow_t[index_r, 2]
        # else:
        #     pickup_time_of_this_r_at_this_route = request_flow_t[index_r, 0]

        if route[3, m] > R[index_r, 5]:
            if dynamic == 1 and dynamic_t > 0 and dynamic_t_begin > pickup_time_of_this_r_at_this_route:
                dynamic_time_false = 1
            else:
                if dynamic == 1 and dynamic_t > 0 and dynamic_t_begin <= pickup_time_of_this_r_at_this_route:
                    re_plan_when_event_finishes = 1
                    add_one_row_for_re_plan_when_event_finishes(operation_type, k, request_number, route, m,
                                                                operation_index, pickup_time_of_this_r_at_this_route)
                wait = 0
                request_flow_t = copy.copy(request_flow_t_copy)
                if only_check_this_influenced_r_in_dynamic_uncertainty == request_number: #only check the influenced r is delayed or not, ignore other r's delay, to give RL the correct reward
                    return False, route, request_number
    return -1, -1, -1
# @profile()
# @time_me()
def time_constraints(k, route, inserted_r):
    global waiting_times, re_plan_when_event_finishes, dynamic_time_false, delayed_time_table, wait, wait_time, check_start_position, fixed_wait, request_flow_t,service_time,transshipment_time

    index_inserted_r = list(R[:, 7]).index(inserted_r)
    add_one_row_to_T_k_record(index_inserted_r)
    add_one_row_to_request_flow(index_inserted_r)
    #when remove, it may empty
    if len(route[0])<3:
        return True, route, -1
    # if k == 119 and str(inserted_r)[-1] == '0' and str(inserted_r)[-2] == '0':
    #     print('sfafs')
    route[1:4, 0] = 0

    # change wait to 0 when the vehicle is impossible to transport the request, even wait time was added; or the vehicle pass the constraint
    if has_end_depot == 1:
        length = len(route[4])
    else:
        length = len(route[4]) + 1

    fixed = 0

    if k in fixed_vehicles_percentage and (Demir != 1 and (K[k, 5] == 1 or K[k, 5] == 2 or truck_time_free == 0)):
        fixed = 1
        if K[k, 5] == 1 or K[k, 5] == 2:
            route[1:4, 0] = Fixed[k][0, 1]
    else:
        if k in fixed_vehicles_percentage and Demir == 1 and K[k, 5] != 3:
            fixed = 1
            if Fixed[k][0,2] - Fixed[k][0,1] <= 1:
                service_times = []
                for m in range(len(route[0])):
                    if hasNumbers(route[4, m]):
                        if route[0,m] != o[k,0]:
                            break
                        request_number = int(''.join(filter(str.isdigit, route[4, m])))
                        index_r = list(R[:, 7]).index(request_number)
                        service_times.append(N[route[0,m],1] * R[index_r,6])
                max_service_time = max(service_times)
                route[1:4, 0] = Fixed[k][0, 2] - max_service_time
            else:
                pickup_times = []
                for m in range(len(route[0])):
                    if hasNumbers(route[4, m]):
                        if route[0, m] != o[k, 0]:
                            break
                        request_number = int(''.join(filter(str.isdigit, route[4, m])))
                        index_r = list(R[:, 7]).index(request_number)
                        pick_type = getLetters(route[4, m])
                        if pick_type == 'pickup':
                            pickuptime = R[index_r,2]
                        else:
                            if pick_type == 'Tp':
                                pickuptime = request_flow_t[index_r,1]
                            else:
                                pickuptime = request_flow_t[index_r, 3]
                            if np.isnan(pickuptime):
                                pickuptime = Fixed[k][0,1]
                        ##method 1## this will increase storage time, but Demir donesn't consider storage cost so it's fine
                        if pickuptime <= Fixed[k][0,1]:
                            #if pickup time is earlier than earliest departure time, then the pickup time is assumed as the earliest departure time
                            pickup_times.append(Fixed[k][0,1])
                            break
                        else:
                            #otherwise the pickup time is the earliest pickup time
                            pickup_times.append(pickuptime)
                        ####
                        ##method 2## just let the pickup time is the earliest pickup time, this will increase waiting time
                        # pickup_times.append(pickuptime)
                route[1:4, 0] = min(pickup_times)
        else:
            if route[0, 0] == route[0, 1]:
                request_number = int(''.join(filter(str.isdigit, route[4][1])))
                index_r = list(R[:, 7]).index(request_number)
                if new_getLetters(route[4][1]) == 'pickup':
                    route[1:4, 0] = R[index_r, 2]
                else:
                    if new_getLetters(route[4][1]) == 'Tp':
                        if pd.isnull(request_flow_t[index_r,1]):
                            wait = 0
                            return False, route, request_number
                        route[1:4, 0] = request_flow_t[index_r,1]
                    else:
                        if pd.isnull(request_flow_t[index_r,3]):
                            wait = 0
                            return False, route, request_number
                        route[1:4, 0] = request_flow_t[index_r,3]
    request_flow_t_copy = copy.copy(request_flow_t)
    
    if K[k, 5] == 1 or K[k, 5] == 2 or Demir == 1:
        if dynamic == 1 and unexpected_events == 1:
            delay_is_already_added_in_the_first_request = {}
        for m in range(0, length - 1):
            if hasNumbers(route[4, m]):
                request_number = int(''.join(filter(str.isdigit, route[4, m])))
                index_r = list(R[:, 7]).index(request_number)
                letters = new_getLetters(route[4, m])

                # when there is double wait, i.e., both fixed timetable and earlier arrival need add wait time, the time shouldn't be refresh, because the wait time of fixed timetable may makes the wait_time less than actual wait_time due to earlier arrival, then the earlier arrival will never be make up and stuck into dead loop
                if fixed_wait != 1:

                    if Demir == 1:
                        if route[0, m] == route[0, m - 1]:
                            #assume parallel loading/unloading in Demir's model
                            route[1, m] = route[2, m - 1]
                        else:
                            route[1, m] = route[3, m - 1] + D[k][int(route[0, m-1]), int(route[0, m])] / K[k, 1]
                    else:
                        # wenjing: multiple requests pickup/deliveried at the same terminal, only in one service time
                        if route[0, m] == route[0, m - 1]:
                            route[1, m] = route[2, m - 1]
                        else:
                            route[1, m] = route[3, m - 1] + D[k][int(route[0, m-1]), int(route[0, m])] / K[k, 1]

                    route[2, m] = route[1, m]
                    #dynamic under congestion
                    if dynamic == 1 and unexpected_events == 1:
                        if route[0, m] not in delay_is_already_added_in_the_first_request.keys():
                            delay_is_already_added_in_the_first_request[route[0, m]] = 1
                            route[2, m] = route[2, m] + delayed_time_table[k][route[0,m]]


                    route[3, m] = route[2, m]
                if m in waiting_times.keys():
                    wait_time = waiting_times[m]
                    if dynamic == 1 and unexpected_events == 1:
                        if wait_time <= delayed_time_table[k][route[0, m]]:
                            pass
                        else:
                            route[2, m] = route[2, m] + wait_time - delayed_time_table[k][route[0, m]]
                    else:
                        route[2, m] = route[2, m] + wait_time
                # if wait == 1:
                #     if check_start_position == m:
                #         if dynamic == 1 and unexpected_events == 1:
                #             if wait_time <= delayed_time_table[k][route[0,m]]:
                #                 pass
                #             else:
                #                 route[2, m] = route[2, m] + wait_time - delayed_time_table[k][route[0,m]]
                #         else:
                #             route[2, m] = route[2, m] + wait_time
                #         wait = 0
                # open window of fixed routes' terminals
                if fixed == 1:
                    if Demir == 1:
                        service_time = N[route[0,m],1] * R[index_inserted_r,6]
                        departure_time = Fixed[k][Fixed[k][:, 0] == route[0, m], 1][0]
                        real_departure_time = route[2, m] + service_time
                    else:
                        # try:
                        departure_time = Fixed[k][Fixed[k][:,0] == route[0, m],1][0] + service_time
                        # except:
                        #     print('caught_it')
                        if K[k, 5] == 1 or K[k, 5] == 2:
                            real_departure_time = route[2, m] + service_time
                        else:
                            real_departure_time = route[2, m]
                    # to make the departure time totally same with departure time of fixed k
                    if real_departure_time < departure_time:

                        fixed_wait = 1
                        wait = 1
                        # wait_time = departure_time - route[1, m] + 0.000001
                        if K[k, 5] == 1 or K[k, 5] == 2:
                            wait_time = departure_time - route[2, m] - service_time
                        else:
                            wait_time = departure_time - route[2, m]
                        check_start_position = m
                        waiting_times[m] = wait_time
                        terminal = route[0, m]
                        return False, route, str(terminal)

                    else:
                        # try:
                        departure_final_time = Fixed[k][Fixed[k][:,0] == route[0, m],2][0]
                        # except:
                        #     print('find error')
                        #     sys.exit(7)
                        #wait = 0
                        # fixed:
                        # pickup a-1 a
                        # delivery b b+1
                        # route:
                        # pickup delivery
                        # a-1 b
                        # a b+1

                        # so real_departure_time can't bigger than departure_final_time
                        if Demir == 1 and (letters == 'delivery' or letters == 'Td' or letters == 'secondTd'):
                            if real_departure_time > departure_final_time + service_time:
                                request_flow_t = copy.copy(request_flow_t_copy)
                                wait = 0
                                return False, route, request_number
                        else:
                            if stochastic == 1 and dynamic_t > 0:
                                #under uncertain travel time, the fixed schedules can be voilated
                                pass
                            else:
                                if real_departure_time > departure_final_time:
                                    request_flow_t = copy.copy(request_flow_t_copy); wait = 0
                                    return False, route, request_number
                
                if letters == 'Td':
                    # transshipment time for Td
                    if Demir == 1:
                        route[3, m] = route[2, m] + N[route[0, m], 1] * R[index_inserted_r, 6]
                        T_k_record[index_r, 0] = route[0, m]
                    else:
                        if K[k, 5] != 3:
                            route[3, m] = route[2, m] + transshipment_time
                        else:
                            route[3, m] = route[2, m]
                    request_flow_t[index_r,1] = route[3, m]
                if letters == 'secondTd':
                    # transshipment time for secondTd
                    if Demir == 1:
                        route[3, m] = route[2, m] + N[route[0, m], 1] * R[index_inserted_r, 6]
                        T_k_record[index_r, 1] = route[0, m]
                    else:
                        if K[k, 5] != 3:
                            route[3, m] = route[2, m] + transshipment_time
                        else:
                            route[3, m] = route[2, m]
                    request_flow_t[index_r,3] = route[3, m]

                #            wait = 0
                if letters == 'pickup':

                    request_flow_t[index_r,0] = route[2, m]
                    if Demir == 1:
                        T_k_record[index_r, 2] = k
                    if Demir == 1:
                        service_time = N[route[0, m], 1] * R[index_inserted_r, 6]
                        route[3, m] = route[2, m] + service_time
                    else:
                        if K[k, 5] == 1 or K[k, 5] == 2:
                            route[3, m] = route[2, m] + service_time
                        else:
                            route[3, m] = route[2, m]
                    if request_flow_t[index_r,0] < R[index_r, 2]:
                        wait = 1
                        wait_time = R[index_r, 2] - request_flow_t[index_r,0] + 0.000001
                        check_start_position = m
                        waiting_times[m] = wait_time
                        return False, route, request_number
                    if Demir != 1 and request_flow_t[index_r,0] + service_time > R[index_r, 3]:
                        # above last is R[index_r,3] because the containers can only be stored in the pickup time window, if exceed, then can't pickup
                        if dynamic == 1 and dynamic_t > 0 and dynamic_t_begin >= route[2, m]:
                            #dynamic_time_false = 1
                            pass
                        else:
                            if dynamic == 1 and dynamic_t > 0 and dynamic_t_begin < route[2, m]:
                                re_plan_when_event_finishes = 1

                                add_one_row_for_re_plan_when_event_finishes('pickup', k, request_number, route, m,
                                                                            -1, -1)
                            wait = 0
                            request_flow_t = copy.copy(request_flow_t_copy)

                            return False, route, request_number
                
                if letters == 'Tp':
                    request_flow_t[index_r,2] = route[2, m]
                    if Demir == 1:
                        T_k_record[index_r, 0] = route[0, m]
                        T_k_record[index_r, 3] = k
                    # after the inserted request's route's has other request's Tp/secondTp was considered,
                    # it also has probability that in relevant routes, similiar situation may happen, which will not be considered
                    if pd.isnull(request_flow_t[index_r,1]):
                        wait = 0
                        request_flow_t = copy.copy(request_flow_t_copy)
                        return False, route, request_number
                    if request_flow_t[index_r,2] < request_flow_t[index_r,1]:
                        wait = 1
                        if dynamic == 1 and unexpected_events == 1:
                            #the  the waiting time is calculated based on the time after adding delay, which makes the waiting time is less than what it requires, so add the delay time
                            wait_time = request_flow_t[index_r, 1] - request_flow_t[index_r, 2] + 0.000001 + delayed_time_table[k][route[0, m]]
                        else:
                            wait_time = request_flow_t[index_r,1] - request_flow_t[index_r,2] + 0.000001
                        check_start_position = m
                        waiting_times[m] = wait_time
                        return False, route, request_number
                    else:
                        if Demir == 1:

                            #Demir's cost not include storage cost. but I need to let there is not too much storage time, for example request 1 and 2 in case 3 can use train 7 and train 11 for lower emission but storage time is too long and Demir not use them.
                            # if request_flow_t[index_r,2] - request_flow_t[index_r,1] > 50:
                            #     wait = 0
                            #     request_flow_t = copy.copy(request_flow_t_copy)
                            #     return False, route
                            route[3, m] = route[2, m] + N[route[0, m], 1] * R[index_inserted_r, 6]
                        else:
                            if K[k, 5] != 3:
                                route[3, m] = route[2, m] + transshipment_time
                            else:
                                route[3, m] = route[2, m]
                        request_flow_t[index_r,2] = route[2, m]

                if letters == 'secondTp':
                    request_flow_t[index_r,4] = route[2, m]
                    if Demir == 1:
                        T_k_record[index_r, 1] = route[0, m]
                        T_k_record[index_r, 4] = k
                    # after the inserted request's route's has other request's Tp/secondTp was considered,
                    # it also has probability that in relevant routes, similiar situation may happen, which will not be considered
                    if pd.isnull(request_flow_t[index_r,3]):
                        wait = 0
                        request_flow_t = copy.copy(request_flow_t_copy)
                        return False, route, request_number
                    if request_flow_t[index_r,4] < request_flow_t[index_r,3]:
                        wait = 1
                        wait_time = request_flow_t[index_r,3] - request_flow_t[index_r,4] + 0.000001
                        check_start_position = m
                        waiting_times[m] = wait_time
                        return False, route, request_number
                    else:
                        if Demir == 1:
                            # Demir's cost not include storage cost. but I need to let there is not too much storage time, for example request 1 and 2 in case 3 can use train 7 and train 11 for lower emission but storage time is too long and Demir not use them.
                            # if request_flow_t[index_r, 4] - request_flow_t[index_r, 3] > 50:
                            #     wait = 0
                            #     request_flow_t = copy.copy(request_flow_t_copy)
                            #     return False, route
                            route[3, m] = route[2, m] + N[route[0, m], 1] * R[index_inserted_r, 6]
                        else:
                            if K[k, 5] != 3:
                                route[3, m] = route[2, m] + transshipment_time
                            else:
                                route[3, m] = route[2, m]
                        request_flow_t[index_r,4] = route[2, m]

                if letters == 'delivery':

                    request_flow_t[index_r,5] = route[2, m]
                    if Demir == 1:
                        service_time = N[route[0, m], 1] * R[index_inserted_r, 6]
                        route[3, m] = route[2, m] + service_time
                    else:
                        if K[k, 5] == 1 or K[k, 5] == 2:
                            route[3, m] = route[2, m] + service_time
                        else:
                            route[3, m] = route[2, m]
                    if fixed == 1:
                        # if the route has any signal that the time exceed the arrival final time + 1 of fixed route, then infeasible
                        arrival_final_time = Fixed[k][Fixed[k][:,0] == route[0, m],2][0]
                        if Demir == 1:
                            if route[2, m] > arrival_final_time:
                                wait = 0
                                request_flow_t = copy.copy(request_flow_t_copy)
                                terminal = route[0, m]
                                return False, route, str(terminal)
                        else:

                            if route[3, m] > arrival_final_time:
                                if stochastic == 1 and dynamic == 1 and dynamic_t > 0:
                                    # under uncertain travel time, the fixed schedules can be voilated. and it needs to be feasible, because RL does not consider this in the state, if infeasible, RL will wrong
                                    pass
                                else:
                                    wait = 0
                                    request_flow_t = copy.copy(request_flow_t_copy)
                                    terminal = route[0, m]
                                    return False, route, str(terminal)
                    if stochastic == 1 and dynamic_t == 0:
                        #forbid delay for the initial solution
                        if route[3, m] > R[index_r, 5]:
                            wait = 0
                            request_flow_t = copy.copy(request_flow_t_copy)
                            return False, route, request_number
                    if stochastic == 1:
                        False_or_not, route_, request_number_ = check_delay_and_return_under_uncertainty(k, route, request_number, m, index_r, request_flow_t_copy)
                        if isinstance(False_or_not, bool):
                            return False_or_not, route_, request_number
                    else:
                        if forbid_much_delay == 1:
                            if route[3, m] > R[index_r, 5] + 2:
                                wait = 0
                                request_flow_t = copy.copy(request_flow_t_copy)
                                return False, route, request_number
                    #allow delay
                    # else:
                    #     if Demir == 1:
                    #         service_time = N[route[0, m], 1] * R[index_inserted_r, 6]
                    #
                    #     if route[3, m] > R[index_r, 5] + service_time:
                    #         wait = 0
                    #         request_flow_t = copy.copy(request_flow_t_copy)
                    #         return False, route
                # else:
                #     # seems repeat with the begining fixed constraints check 20201106
                #     if fixed == 1:
                #         departure_final_time = Fixed[k][Fixed[k][:,0] == route[0, m],2][0]
                #
                #         if route[3, m] > departure_final_time:
                #             wait = 0
                #             request_flow_t = copy.copy(request_flow_t_copy)
                #             return False, route
                fixed_wait = 0

        wait = 0
    # truck fleet
    else:
        two_letters,two_m = remove_T_k_in_record(route, inserted_r)
        request_number = inserted_r
        if request_number % 100000 > 10000:
            print('dsf')
            request_number_check_in_delayed_table = request_number - (request_number - request_number % 10000) % big_r
        else:
            request_number_check_in_delayed_table = request_number
        index_r = list(R[:, 7]).index(request_number)
        letters = two_letters[0]


        m = two_m[0]
        if letters == 'pickup':
            # dynamic under congestion
            if dynamic == 1 and unexpected_events == 1:
                route[1, m] = R[index_r, 2]
                if request_number_check_in_delayed_table in delayed_time_table[k][route[0, m]].keys():
                    route[2, m] = route[1, m] + delayed_time_table[k][route[0, m]][request_number_check_in_delayed_table]
                else:
                    route[2, m] = route[1, m]
                route[3, m] = route[2, m]
            else:
                route[1:4, m] = R[index_r, 2]
            # request_flow_t.loc[request_number] = np.nan
            # T_k_record.loc[request_number] = np.nan

            T_k_record[index_r, 2] = k

            request_flow_t[index_r,0] = route[2, m]

        # danger if there are 2T, time of Tp may need same with time in secondTd
        # danger the second truck maybe can't arrive T on time
        if letters == 'Tp':
            if pd.isnull(request_flow_t[index_r,1]):
                wait = 0
                request_flow_t = copy.copy(request_flow_t_copy)
                return False, route, request_number
            if dynamic == 1 and unexpected_events == 1:
                route[1, m] = request_flow_t[index_r,1]
                if request_number_check_in_delayed_table in delayed_time_table[k][route[0, m]].keys():
                    route[2, m] = route[1, m] + delayed_time_table[k][route[0, m]][request_number_check_in_delayed_table]
                else:
                    route[2, m] = route[1, m]
                route[3, m] = route[2, m]
            else:
                route[1:4, m] = request_flow_t[index_r,1]
            request_flow_t[index_r,2] = route[2, m]
            T_k_record[index_r,0] = route[0, m]
            T_k_record[index_r, 3] = k
        if letters == 'secondTp':
            if pd.isnull(request_flow_t[index_r,3]):
                wait = 0
                request_flow_t = copy.copy(request_flow_t_copy)
                return False, route, request_number
            if dynamic == 1 and unexpected_events == 1:
                route[1, m] = request_flow_t[index_r,3]
                if request_number_check_in_delayed_table in delayed_time_table[k][route[0, m]].keys():
                    route[2, m] = route[1, m] + delayed_time_table[k][route[0, m]][request_number_check_in_delayed_table]
                else:
                    route[2, m] = route[1, m]
                route[3, m] = route[2, m]
            else:
                route[1:4, m] = request_flow_t[index_r,3]
            request_flow_t[index_r,4] = route[2, m]
            T_k_record[index_r,1] = route[0, m]
            T_k_record[index_r, 4] = k
        letters = two_letters[1]
        m = two_m[1]

        if letters == 'delivery':
            #remove T in T_rcord if request_number not use T
            if two_letters[0] == 'pickup':
                T_k_record[index_r] = np.nan
            else:
                if two_letters[0] == 'Tp':
                    T_k_record[index_r, 1] = np.nan
                    T_k_record[index_r, 4] = np.nan

            #T_k_record T1,T2,k1,k2,k3
            #2T
            if not math.isnan(T_k_record[index_r,1]):

                if pd.isnull(request_flow_t[index_r,4]):
                    wait = 0
                    request_flow_t = copy.copy(request_flow_t_copy)
                    return False, route, request_number
                departure_time = request_flow_t[index_r,4] % 24
                original_travel_time = D[k][R[index_r, 1],int(T_k_record[index_r,1])] / \
                                       K[k, 1]
                truck_travel_time = get_travel_time_pre(departure_time, original_travel_time)
                if dynamic == 1 and unexpected_events == 1:
                    route[1, m] = request_flow_t[index_r,4] + truck_travel_time
                    if request_number_check_in_delayed_table in delayed_time_table[k][route[0, m]].keys():
                        route[2, m] = route[1, m] + delayed_time_table[k][route[0, m]][request_number_check_in_delayed_table]
                    else:
                        route[2, m] = route[1, m]
                    route[3, m] = route[2, m]
                else:
                    route[1:4, m] = request_flow_t[index_r,4] + truck_travel_time

            else:
                #1T
                if not math.isnan(T_k_record[index_r,0]):
                    if pd.isnull(request_flow_t[index_r,2]):
                        wait = 0
                        request_flow_t = copy.copy(request_flow_t_copy)
                        return False, route
                    departure_time = request_flow_t[index_r,2] % 24
                    original_travel_time = D[k][R[index_r, 1],int(T_k_record[index_r,0])] / K[k, 1]

                    truck_travel_time = get_travel_time_pre(departure_time, original_travel_time)
                    if dynamic == 1 and unexpected_events == 1:
                        route[1, m] = request_flow_t[index_r,2] + truck_travel_time
                        if request_number_check_in_delayed_table in delayed_time_table[k][route[0, m]].keys():
                            route[2, m] = route[1, m] + delayed_time_table[k][route[0, m]][request_number_check_in_delayed_table]
                        else:
                            route[2, m] = route[1, m]
                        route[3, m] = route[2, m]
                    else:
                        route[1:4, m] = request_flow_t[index_r,2] + truck_travel_time

                else:
                    #0T
                    departure_time = request_flow_t[index_r,0] % 24
                    original_travel_time = D[k][R[index_r, 1],R[index_r, 0]] / \
                                               K[k, 1]


                    truck_travel_time = get_travel_time_pre(departure_time, original_travel_time)
                    if dynamic == 1 and unexpected_events == 1:
                        route[1, m] = request_flow_t[index_r,0] + truck_travel_time
                        if request_number_check_in_delayed_table in delayed_time_table[k][route[0, m]].keys():
                            route[2, m] = route[1, m] + delayed_time_table[k][route[0, m]][request_number_check_in_delayed_table]
                        else:
                            route[2, m] = route[1, m]
                        route[3, m] = route[2, m]
                    else:
                        route[1:4, m] = request_flow_t[index_r,0] + truck_travel_time

            if stochastic == 1:
                if RL_is_trained_or_evaluated_or_ALNS_is_evaluated == 1:

                    if route[3, m] > R[index_r, 5]:
                        for operation_index in range(len(route[4, 1:-1])):
                            if request_number == get_numbers(route[4, operation_index + 1]):
                                operation_type = new_getLetters(route[4, operation_index + 1])
                                pickup_time_of_this_r_at_this_route = route[2, operation_index + 1]
                                break
                        # if not pd.isnull(request_flow_t[index_r, 4]):
                        #     pickup_time_of_this_r_at_this_route = request_flow_t[index_r, 4]
                        # elif not pd.isnull(request_flow_t[index_r, 2]):
                        #     pickup_time_of_this_r_at_this_route = request_flow_t[index_r, 2]
                        # else:
                        #     pickup_time_of_this_r_at_this_route = request_flow_t[index_r, 0]

                        if dynamic == 1 and dynamic_t > 0 and dynamic_t_begin > pickup_time_of_this_r_at_this_route:
                            dynamic_time_false = 1
                        else:
                            #when re_plan_when_event_finishes = 1, it means it already in the replanning, then just accept this route and then compare cost
                            if re_plan_when_event_finishes != 1:
                                if dynamic == 1 and dynamic_t > 0 and dynamic_t_begin <= pickup_time_of_this_r_at_this_route:
                                    re_plan_when_event_finishes = 1
                                    add_one_row_for_re_plan_when_event_finishes(operation_type, k, request_number, route, m, operation_index, pickup_time_of_this_r_at_this_route)
                                wait = 0
                                request_flow_t = copy.copy(request_flow_t_copy)
                                return False, route, request_number
            else:
                #can't delay more than 1 hour. During the optimization, the r's delivery time may influenced by other newly inserted r, so I add this
                if forbid_much_delay == 1:
                    if route[3, m] > R[index_r, 5] + 2:
                        wait = 0
                        request_flow_t = copy.copy(request_flow_t_copy)
                        return False, route, request_number

            request_flow_t[index_r,5] = route[2, m]
        if letters == 'Td':
            departure_time = R[index_r, 2] % 24
            original_travel_time = D[k][int(route[0, m]), R[index_r, 0]] / K[k, 1]
            truck_travel_time = get_travel_time_pre(departure_time, original_travel_time)
            if dynamic == 1 and unexpected_events == 1:
                route[1, m] = request_flow_t[index_r, 0] + truck_travel_time
                if request_number_check_in_delayed_table in delayed_time_table[k][route[0, m]].keys():
                    route[2, m] = route[1, m] + delayed_time_table[k][route[0, m]][request_number_check_in_delayed_table]
                else:
                    route[2, m] = route[1, m]
                route[3, m] = route[2, m]
            else:
                route[1:4, m] = R[index_r, 2] + truck_travel_time
            request_flow_t[index_r,1] = route[3, m]
            T_k_record[index_r,0] = route[0, m]
        if letters == 'secondTd':
            if pd.isnull(request_flow_t[index_r,2]):
                wait = 0
                request_flow_t = copy.copy(request_flow_t_copy)
                return False, route, request_number
            departure_time = request_flow_t[index_r,2] % 24
            T_k_record[index_r, 1] = route[0, m]

            original_travel_time = D[k][int(T_k_record[index_r,1]),int(T_k_record[index_r,0])] / K[
                k, 1]

            truck_travel_time = get_travel_time_pre(departure_time, original_travel_time)
            if dynamic == 1 and unexpected_events == 1:
                route[1, m] = request_flow_t[index_r,2] + truck_travel_time
                if request_number_check_in_delayed_table in delayed_time_table[k][route[0, m]].keys():
                    route[2, m] = route[1, m] + delayed_time_table[k][route[0, m]][request_number_check_in_delayed_table]
                else:
                    route[2, m] = route[1, m]
                route[3, m] = route[2, m]
            else:
                route[1:4, m] = request_flow_t[index_r,2] + truck_travel_time

            request_flow_t[index_r,3] = route[3, m]
    # if dynamic_time_false == 1:
    #     return False, route, -1
    # else:
    return True, route, -1

def add_one_row_for_re_plan_when_event_finishes(operation_type, k, request_number, route, m, operation_index, pickup_time_of_this_r_at_this_route):
    global re_plan_when_event_finishes_information, already_add_once_for_re_plan_when_event_finishes
    if at_get_reward == 1 and already_add_once_for_re_plan_when_event_finishes == 0:

        if operation_type == "pickup":
            re_plan_when_event_finishes_information = re_plan_when_event_finishes_information.append(pd.Series({'k': int(k), 'request_number': int(request_number), 'vehicle_stop_time': route[2, m], 'T1': -1, 'Td_time': -1, 'r_number': 'wait_for_input'}), ignore_index=True)
        elif operation_type == "Tp":
            re_plan_when_event_finishes_information = re_plan_when_event_finishes_information.append(pd.Series({'k': int(k), 'request_number': int(request_number), 'vehicle_stop_time': pickup_time_of_this_r_at_this_route, 'T1': int(route[0, operation_index + 1]), 'Td_time': pickup_time_of_this_r_at_this_route, 'r_number': 'wait_for_input'}), ignore_index=True)
    already_add_once_for_re_plan_when_event_finishes = 1

def calculate_emissions(k, d, load, emission):
    global during_iteration
    if load == 0 or d == 0:
        return emission
    if wtw_emissions == 0:
        # unit kg
        emission = emission + d * K[k, 4] * load
    else:
        # in fact trip type is only a name, and it not influence the emissions, but the load influence
        if K[k, 5] == 3:
            # truck
            trip_type = 'trip of 0.5 FTL'
        elif K[k, 5] == 2:
            # barge
            trip_type = 'trip of 50% utilized barge'
        else:
            # train
            trip_type = 'trip of 30 full train wagons'
        if note == '_sustainable_1':
            load_dependent_average_emission_rates = 1
        else:
            load_dependent_average_emission_rates = 0
        #a transport unit is 13 ton
        # the following one is wrong because it changes the load directly, rather than only load factor
        # 20211106 using uncertain * random.uniform(0.9, 1.1) will cause v_has_r in remove_a_request wrong, because it will let "satisfied" in remove_a_request unstable, so not use it
        # trip = Trip(name=trip_type, distance_in_km=d, load_in_ton=int(K[k, 0] * 0.6 * random.uniform(0.9, 1.1))  * 13, slope_profile=1, empty_trip_factor=0.2)
        # if load < K[k, 0] * 0.6:
        # trip = Trip(name=trip_type, distance_in_km=d,
        #             load_in_ton=int(K[k, 0] * 0.6) * 13, slope_profile=1,
        #             empty_trip_factor=0.2)

        if load_dependent_average_emission_rates == 1 and K[k, 5] != 3:
            #this one is correct and uses future loads (assume loads will be 60% in the end)
            trip = Trip(name=trip_type, distance_in_km=d,
                        load_in_ton=load * 13, slope_profile=1,
                        empty_trip_factor=0.2, during_iteration=during_iteration)
        else:
            #here I use during_iteration=0 to let the load factor is real
            trip = Trip(name=trip_type, distance_in_km=d, load_in_ton=load * 13, slope_profile=1, empty_trip_factor=0.2, during_iteration=0)
        if K[k, 5] == 2:
            wtw_emissions_in_tCO2e = trip.compute_train_wtw_emissions_in_tCO2e(setting='default')
        elif K[k, 5] == 3:
            wtw_emissions_in_tCO2e = trip.compute_truck_wtw_emissions_in_tCO2e(setting='default')
        elif K[k, 5] == 1:
            wtw_emissions_in_tCO2e = trip.compute_barge_wtw_emissions_in_tCO2e(setting='default')
        else:
            wtw_emissions_in_tCO2e = 0
        # unit ton -> kg
        emission = emission + wtw_emissions_in_tCO2e * 1000
    return emission

# @profile()
# @time_me()
def objective_value_k(k, new_try):
    global initial_solution_no_wait_cost
    
    vehicle_cost = 0
    time_on_route = 0
    time_at_terminal = 0
    request_cost = 0
    wait_cost = 0
    wait_time = 0
    transshipment_cost = 0
    un_load_cost = 0
    distance = 0
    profit = 0
    emission = 0
    emission_transshipment = 0
    emission_cost = 0
    storage_cost = 0
    delay_penalty = 0
    load = 0
    number_transshipment = 0
    if heterogeneous_preferences == 1 and only_eco_label == 0:
        use_T_label = {}
        if use_speed == 1:
            speed_k = {}
        else:
            time_ratio_k = {}
    #    transposed_route = new_try.iloc[0].T
    #    res = [i[0] for i in groupby(transposed_route.values.tolist())]
    #    res = pd.DataFrame(res)
    if has_end_depot == 1:
        length = len(new_try[4])
    else:
        length = len(new_try[4]) + 1
    if length <= 2:
        vehicle_cost = 0
        time_on_route = 0
    truck_time_record = {}

    for x in range(1, len(new_try[4])):
        travel_time = 0
        if K[k, 5] != 3 or Demir == 1:
            p1,d1 = int(new_try[0, x - 1]), int(new_try[0, x])
            if p1 == d1:
                d = 0
            else:
                d = D[k][p1,d1]

            if d == 1000000000:
                return 100000000000, 100000000000, 100000000000, 100000000000, 100000000000, 100000000000, 100000000000, 100000000000, 100000000000, 100000000000, 100000000000, 100000000000, 100000000000, 100000000000, 100000000000, 100000000000, 100000000000

            distance = distance + d
            vehicle_cost = vehicle_cost + d * fuel_cost


            #        nearest_terminal_index=res.index[res[0] == new_try[new_try[4][i]][0]][0] - 1
            #        if nearest_terminal_index >= 0:
            #            d = D[k][res[0][nearest_terminal_index]][new_try[new_try[4][i]][0]]
            #        else:
            #            d = 0

            if hasNumbers(new_try[4, x]):
                if new_try[0, x] == new_try[0, x-1] or new_try[4, x - 1] == 'begin_depot':
                    travel_time = 0
                    d_copy = 0
                else:
                    d_copy = d
                    travel_time = new_try[1, x] - new_try[3, x-1]
                request_number = int(''.join(filter(str.isdigit, new_try[4, x])))
                index_r = list(R[:, 7]).index(request_number)
                letters = new_getLetters(new_try[4, x])

                request_cost = request_cost + (K[k, 3] * d_copy + K[k, 2] * travel_time) * load
                profit = profit + (K[k, 3] * d_copy + K[k, 2] * travel_time) * load * 1.3
                emission = calculate_emissions(k, d, load, emission)
                if letters == 'pickup' or letters == 'Tp' or letters == 'secondTp':
                    load = load + R[index_r, 6]
                else:
                    load = load - R[index_r, 6]


        else:
            # truck fleet

            if hasNumbers(new_try[4, x]):
                request_number = int(''.join(filter(str.isdigit, new_try[4, x])))
                index_r = list(R[:, 7]).index(request_number)
                letters = new_getLetters(new_try[4, x])
                if letters == 'pickup' or letters == 'Tp' or letters == 'secondTp':
                    travel_time = 0
                    truck_time_record[request_number] = [new_try[0, x],
                                                         new_try[3, x]]
                else:
                    try:
                        d = D[k][int(new_try[0, x]),int(truck_time_record[request_number][0])]
                    except:
                        print('truck_time_record error', 'k', k, 'x', x,  'new_try', new_try, 'request_number', request_number)
                    #wenjing only calculate travel time as time-dependent but cost is fixed
                    time_dependent_cost = 0
                    if time_dependent_cost == 1:
                        travel_time = new_try[1, x] - truck_time_record[request_number][1]
                        request_cost = request_cost + (K[k, 3] * d + K[k, 2] * travel_time) * R[index_r, 6]
                        profit = profit + (K[k, 3] * d + K[k, 2] * travel_time) * R[index_r, 6] * 1.3
                    else:
                        travel_time = d / K[k, 1]
                        request_cost = request_cost + (K[k, 3] * d + K[k, 2] * travel_time) * R[index_r, 6]
                        profit = profit + (K[k, 3] * d + K[k, 2] * travel_time) * R[index_r, 6] * 1.3
                    emission = calculate_emissions(k, d, R[index_r, 6], emission)
                    # if wtw_emissions == 0:
                    #     emission = emission + d * K[k, 4] * R[index_r, 6]

        time_on_route = time_on_route + travel_time
        wait_and_service = new_try[3, x] - new_try[1, x]
        time_at_terminal = time_at_terminal + wait_and_service
        if hasNumbers(new_try[4, x]):

            # wait cost
            wait_cost = wait_cost + new_try[2, x] - new_try[1, x]
            wait_time = wait_time + new_try[2, x] - new_try[1, x]
            # loading and unloading emssions
            # if only_eco_label == 1:
            #     if K[k, 5] == 1:
            #         emission = 6.8 * R[index_r, 6] + emission
            #     elif K[k, 5] == 2:
            #         emission = 14 * R[index_r, 6] + emission
            #     else:
            #         emission = 15 * R[index_r, 6] + emission
            if only_eco_label == 1:
                #only calculate emissions during transshipment, which cludes both loading and unloading
                #danger here if vertical collaboration, needs to check the mode used by another operator
                if letters == 'Tp' or letters == 'secondTp':
                    # maybe another k is in another new_try and hasn't update T_k_record. In this case, use the average emissions
                    if letters == 'Tp':
                        if np.isnan(T_k_record[index_r, 2]):
                            type_of_another_k = -1
                        else:
                            type_of_another_k = K[T_k_record[index_r, 2],5]

                    else:
                        if np.isnan(T_k_record[index_r, 3]):
                            type_of_another_k = -1
                        else:
                            type_of_another_k = K[T_k_record[index_r, 3],5]
                    if type_of_another_k == -1:
                        emission_rate = (19.6 + 6.3 + 11.2) / 3
                    else:
                        # check using barge or not
                        if letters == 'Tp' and (K[k, 5] == 1 or type_of_another_k == 1):

                            if 2 in [K[k, 5], type_of_another_k]:
                                #train
                                emission_rate = 19.6
                            elif 3 in [K[k, 5], type_of_another_k]:
                                #truck
                                emission_rate = 6.3
                            else:
                                #another one is also barge
                                emission_rate = 11.2
                        elif letters == 'secondTp' and (K[k, 5] == 1 or type_of_another_k == 1):
                            if 2 in [K[k, 5], type_of_another_k]:
                                # train
                                emission_rate = 19.6
                            elif 3 in [K[k, 5], type_of_another_k]:
                                # truck
                                emission_rate = 6.3
                            else:
                                # another one is also barge
                                emission_rate = 11.2
                        else:
                            #barge is not used
                            emission_rate = 2.6
                    emission = emission_rate * R[index_r, 6] + emission
                    emission_transshipment = emission_rate * R[index_r, 6] + emission_transshipment
            # transshipment cost
            if letters == 'Td' or letters == 'secondTd' or letters == 'Tp' or letters == 'secondTp':
                if Demir == 1:
                    no_transshipment = 0
                    if k in [0,1,2]:
                        if letters == 'Td':
                            if T_k_record[index_r, 3] in [0,1,2]:
                                no_transshipment = 1
                        if letters == 'secondTd':
                            if T_k_record[index_r, 4] in [0, 1, 2]:
                                no_transshipment = 1
                        if letters == 'Tp':
                            if T_k_record[index_r, 2] in [0, 1, 2]:
                                no_transshipment = 1
                        if letters == 'secondTp':
                            if T_k_record[index_r, 3] in [0, 1, 2]:
                                no_transshipment = 1
                    if no_transshipment != 1:
                        transshipment_cost = N[new_try[0, x],2] * R[index_r, 6] + transshipment_cost
                        emission = N[new_try[0, x],3] * R[index_r, 6] + emission
                else:
                    if K[k, 5] == 1 or K[k, 5] == 2:
                        transshipment_cost = 18 * R[index_r, 6] + transshipment_cost
                    else:
                        transshipment_cost = 3 * R[index_r, 6] + transshipment_cost

            if heterogeneous_preferences == 1 and only_eco_label == 0:
                if letters == 'pickup':
                    use_T_label[request_number] = [0, new_try[0, x]]
                elif letters == 'Tp':
                    number_transshipment = number_transshipment + R[index_r, 6]
                    use_T_label[request_number] = [1, new_try[0, x]]
                elif letters == 'secondTp':
                    number_transshipment = number_transshipment + R[index_r, 6]
                    use_T_label[request_number] = [2, new_try[0, x]]
                elif use_speed == 1:
                    if letters == 'Td':
                        speed_k[request_number] = D[k][use_T_label[request_number][1]][new_try[0, x]] / (request_flow_t[index_r, 1] - request_flow_t[index_r, 0])
                    elif letters == 'secondTd':
                        speed_k[request_number] = D[k][use_T_label[request_number][1]][new_try[0, x]] / (request_flow_t[index_r, 3] - request_flow_t[index_r, 2])
                    #then check before delivery, what type of pickup
                    elif use_T_label[request_number][0] == 0:
                        speed_k[request_number] = D[k][use_T_label[request_number][1]][new_try[0, x]] / (request_flow_t[index_r, 5] - request_flow_t[index_r, 0])
                    elif use_T_label[request_number][0] == 1:
                        speed_k[request_number] = D[k][use_T_label[request_number][1]][new_try[0, x]] / (request_flow_t[index_r, 5] - request_flow_t[index_r, 2])
                    else:
                        speed_k[request_number] = D[k][use_T_label[request_number][1]][new_try[0, x]] / (request_flow_t[index_r, 5] - request_flow_t[index_r, 4])
                elif letters == 'delivery':
                    time_ratio_k[request_number] = (request_flow_t[index_r, 5] - request_flow_t[index_r, 0]) / (R[index_r, 5] - R[index_r, 2])
            if letters == 'pickup' or letters == 'delivery':
                if Demir == 1:
                    un_load_cost = N[new_try[0, x],2]* R[index_r, 6] + un_load_cost
                    emission = N[new_try[0, x], 3] * R[index_r, 6] + emission
                else:
                    if K[k, 5] == 1 or K[k, 5] == 2:
                        un_load_cost = 18 * R[index_r, 6] + un_load_cost
                    else:
                        un_load_cost = 3 * R[index_r, 6] + un_load_cost

            if letters == 'pickup':
                if new_try[2, x] > R[index_r, 2]:
                    storage_cost = storage_cost + c_storage * R[index_r, 6] * (
                            new_try[2, x] - R[index_r, 2])
            # if letters == 'Tp':
            #     if k != 17 and k != 18 and request_number == 8:
            #         print('wfw')
            if letters == 'Tp':
                #danger here request_flow_t is not refreshed, so maybe there is difference with real_cost


                if new_try[1, x] > request_flow_t[index_r,1]:
                    # storage_a_request = [new_try.iloc[3, i], request_flow_t[index_r,2]]
                    # storage_a_request.name = request_number
                    # if new_try.loc[0][i] not in storage.keys():
                    #     storage[new_try.loc[0][i]] = pd.DataFrame(columns=['storage_begin_time','storage_end_time'])
                    # storage[new_try.loc[0][i]].append(storage_a_request)
                    storage_cost = storage_cost + c_storage * R[index_r, 6] * (new_try[1, x] -
                            request_flow_t[index_r,1])
            if letters == 'secondTp':
                if new_try[1, x] > request_flow_t[index_r,3]:
                    storage_cost = storage_cost + c_storage * R[index_r, 6] * (new_try[1, x] -
                            request_flow_t[index_r,3])

            if letters == 'delivery':

                if new_try[3, x] < R[index_r, 4]:
                    # storage_a_request = [new_try.iloc[3, i],R[index_r,4]]
                    # storage_a_request.name = request_number
                    # if new_try.loc[0][i] not in storage.keys():
                    #     storage[new_try.loc[0][i]] = pd.DataFrame(columns=['storage_begin_time','storage_end_time'])
                    # storage[new_try.loc[0][i]].append(storage_a_request)
                    storage_cost = storage_cost + c_storage * R[index_r, 6] * (
                            R[index_r, 4] - new_try[3, x])

                if Demir == 1:
                    if new_try[2, x] > R[index_r, 5]:
                        delay_time = new_try[2, x] - R[index_r, 5]
                        delay_penalty = delay_penalty + R[index_r, 8] * delay_time
                else:
                    if new_try[3, x] > R[index_r, 5]:
                        delay_time = new_try[3, x] - R[index_r, 5]
                        delay_penalty = delay_penalty + R[index_r, 8] * delay_time * R[index_r, 6]

    if Demir == 1:
        storage_cost = 0
    if Demir == 1:
        emission_cost = emission / 1000 * 70
    else:
        emission_cost = emission / 1000 * 8

    # if I add vehicle cost in the future, then I can add the Pickup/Delivery Cluster Removal Heuristic in An ALNS for the PDP with Transfers because it can reduce the # vehicles
    vehicle_cost = 0
    # danger to same with Wenjing's model, set wait cost as 0, but it is wrong
    if initial_solution_no_wait_cost == 1 or Demir == 1:
        wait_cost = 0

    if Demir == 1:
        # cost = 0.1 * (request_cost + wait_cost + transshipment_cost + un_load_cost) + 0.8 * delay_penalty + 0.1 * (emission_cost)
        cost = w1 * (request_cost + wait_cost + transshipment_cost + un_load_cost) + w2 * delay_penalty + w3 * (emission_cost)
    else:
        cost = vehicle_cost + request_cost + wait_cost + transshipment_cost + un_load_cost + emission_cost + storage_cost + delay_penalty
    time = time_on_route + time_at_terminal
    # time = 0
    profit = profit - cost
    # print(k, cost)

    average_speed = 0
    average_time_ratio = 0
    if heterogeneous_preferences == 1 and only_eco_label == 0:
        if use_speed == 1:
            number_serve_r = len(speed_k.keys())
            if number_serve_r > 0:
                all_speed = 0
                for key in speed_k.keys():
                    all_speed = all_speed + speed_k[key]
                average_speed = all_speed / number_serve_r
        else:
            number_serve_r = len(time_ratio_k.keys())
            if number_serve_r > 0:
                all_time_ratio = 0
                for key in time_ratio_k.keys():
                    all_time_ratio = all_time_ratio + time_ratio_k[key]
                average_time_ratio = all_time_ratio / number_serve_r

    return round(cost, 3), round(time, 3), round(vehicle_cost, 3), round(request_cost, 3), round(wait_cost, 3), round(
        transshipment_cost, 3), round(un_load_cost, 3), round(distance, 3), round(profit, 3), round(emission, 3), round(
        emission_cost, 3), round(storage_cost, 3), round(delay_penalty, 3), number_transshipment, average_speed, average_time_ratio, round(emission_transshipment,3)


# @profile()
# @time_me()
def objective_value_i(i, k, new_try):
    global check_start_position, wait
    all_objs = objective_value_k(k, new_try)
    cost_all_requests, emissions_all_requests = all_objs[0], all_objs[9]
    new_try_copy = copy.copy(new_try)
    for j in new_try[4]:
        if hasNumbers(j):
            request_number = int(''.join(filter(str.isdigit, j)))
            if i == request_number:
                new_try_copy = np.delete(new_try_copy, list(new_try_copy[4]).index(j), 1)
    #20210301 mute this because if I want the cost of r in current route, I shouldn't change the time of it, although the inseted r may influence (add) cost of other rs, because the added cost belong to other rs
    # recalculate time
    # fixed_wait = 0
    # wait = 0
    # # 20200927 add this because it should be recalculated from beginning
    # if K[k, 5] == 1 or K[k, 5] == 2:
    #     check_start_position = 0
    #     bool_time, new_try_copy = time_constraints(k, new_try_copy, i)
    #     while wait == 1:
    #         bool_time, new_try_copy = time_constraints(k, new_try_copy, i)
    #     if bool_time == False:
    #         return 10000000000000000000, 10000000000000000000, 10000000000000000000
    all_objs_without_inserted_request = objective_value_k(k, new_try_copy)
    cost_without_inserted_request, emissions_without_inserted_request = all_objs_without_inserted_request[0], all_objs_without_inserted_request[9]
    cost_inserted_request = cost_all_requests - cost_without_inserted_request
    emissions_inserted_request = emissions_all_requests - emissions_without_inserted_request

    return round(cost_inserted_request, 3), cost_all_requests, round(emissions_inserted_request, 3)

def parallel_obj_func(k):
    return objective_value_k(k, routes_local2[k])

def update_request_flow_t(route):
    for m in range(1,len(route[0])-1):
        r = int(''.join(filter(str.isdigit, route[4][m])))
        index_r = list(R[:, 7]).index(r)
        letter = getLetters(route[4, m])
        if letter == 'pickup':
            request_flow_t[index_r,0] = route[2,m]
        else:
            if letter == 'delivery':
                request_flow_t[index_r,5] = route[3,m]
            else:
                if letter == 'Tp':
                    request_flow_t[index_r,2] = route[2,m]
                else:
                    if letter == 'Td':
                        request_flow_t[index_r, 1] = route[3, m]
                    else:
                        if letter == 'secondTp':
                            request_flow_t[index_r, 4] = route[2, m]
                        else:
                            if letter == 'secondTd':
                                request_flow_t[index_r, 3] = route[3, m]


def overall_satisfactory_values(routes_local2, get_objectives = 0):
    global during_iteration
    # here set during_iteration = 0 to calculate real emissions under eco-label A when calculating overall emissions
    during_iteration = 0
    for k in routes_local2.keys():
        update_request_flow_t(routes_local2[k])
    if get_objectives == 0:
        # satisfactory_value, fuzzy_satisfy_or_not, hard_satisfy_or_not
        satisfactory_values = [0,0,0]
    else:
        satisfactory_values = [0, 0, 0, 0, 0]
    # all_used_k = []
    served_requests = check_served_R()
    for r in R[:,7]:
        k1,k2,k3 = find_used_k(r)
        if k1 == -1:
            if get_objectives == 0:
                satisfactory_values_one_r = [0,0,0]
            else:
                satisfactory_values_one_r = [0, 0, 0, 0, 0]
        else:
            if k2 == -1:
                satisfactory_values_one_r = preference_constraints(r, k1, k2, k3, routes_local2[k1], -1, -1, 1, get_objectives)
            else:
                if k3 == -1:
                    satisfactory_values_one_r = preference_constraints(r, k1, k2, k3, routes_local2[k1], routes_local2[k2], -1, 1, get_objectives)
                else:
                    satisfactory_values_one_r = preference_constraints(r, k1, k2, k3, routes_local2[k1], routes_local2[k2], routes_local2[k3], 1, get_objectives)

        # try:
        satisfactory_values = list(np.add(satisfactory_values, list(satisfactory_values_one_r)))
        # except:
        #     print(satisfactory_values)
    if get_objectives == 0:
        if served_requests == 0:
            satisfactory_values[0] = 0
        else:
            satisfactory_values[0] = satisfactory_values[0] / served_requests
    else:
        #every overall objective is divided by served_requests, except for transshipment_times
        #cost_per_container_per_km, time_ratio, emissions_per_container_per_km, delay_time_ratio, transshipment_times
        for index in range(0,4):
            if served_requests == 0:
                satisfactory_values[index] = 0
            else:
                satisfactory_values[index] = satisfactory_values[index] / served_requests
    # here re-set during_iteration
    during_iteration = 1
    return satisfactory_values
# @profile()
# @time_me()
def overall_obj(routes_local):
    # number_of_R_served=0
    global routes_local2, during_iteration
    # here set during_iteration = 0 to calculate real emissions under eco-label A when calculating overall emissions
    during_iteration = 0
    routes_local2 = routes_local
    # routes_tuple = get_routes_tuple(routes_local)
    # if routes_tuple in hash_overall_obj_table.keys():
    #     return hash_overall_obj_table[routes_tuple]
    update_T_k_record_request_flow_t(routes_local)
    overall_request_cost = 0
    overall_vehicle_cost = 0
    overall_wait_cost = 0
    overall_transshipment_cost = 0
    overall_un_load_cost = 0
    overall_emission_cost = 0
    overall_storage_cost = 0
    overall_delay_penalty = 0
    overall_number_transshipment = 0
    overall_average_speed = 0
    overall_average_time_ratio = 0
    overall_speed = 0
    overall_time_ratio = 0
    number_used_k = 0

    overall_distance = 0
    overall_cost = 0
    overall_time = 0
    overall_profit = 0
    overall_emission = 0
    overall_emission_transshipment = 0
    parallel_obj = 0
    if parallel_obj == 1:
        parallel_k = []
        for k in range(len(K)):
            if len(routes_local[k][4]) > 2:
                parallel_k.append(k)
        with ThreadPoolExecutor() as e:
            results = e.map(parallel_obj_func, parallel_k)
        for result in results:
            cost, time, vehicle_cost, request_cost, wait_cost, transshipment_cost, un_load_cost, distance, profit, emission, emission_cost, storage_cost, delay_penalty = result
            overall_request_cost = overall_request_cost + request_cost
            overall_vehicle_cost = overall_vehicle_cost + vehicle_cost
            overall_wait_cost = overall_wait_cost + wait_cost
            overall_transshipment_cost = overall_transshipment_cost + transshipment_cost
            overall_un_load_cost = overall_un_load_cost + un_load_cost
            overall_emission_cost = overall_emission_cost + emission_cost
            overall_storage_cost = overall_storage_cost + storage_cost
            overall_delay_penalty = overall_delay_penalty + delay_penalty

            overall_distance = overall_distance + distance
            overall_cost = overall_cost + cost
            overall_time = overall_time + time
            overall_profit = overall_profit + profit
            overall_emission = overall_emission + emission
    else:

        for k in range(len(K)):

            if len(routes_local[k][4]) > 2:
                cost, time, vehicle_cost, request_cost, wait_cost, transshipment_cost, un_load_cost, distance, profit, emission, emission_cost, storage_cost, delay_penalty, number_transshipment, average_speed, average_time_ratio, emission_transshipment = objective_value_k(
                    k, routes_local[k])

                overall_request_cost = overall_request_cost + request_cost
                overall_vehicle_cost = overall_vehicle_cost + vehicle_cost
                overall_wait_cost = overall_wait_cost + wait_cost
                overall_transshipment_cost = overall_transshipment_cost + transshipment_cost
                overall_un_load_cost = overall_un_load_cost + un_load_cost
                overall_emission_cost = overall_emission_cost + emission_cost
                overall_storage_cost = overall_storage_cost + storage_cost
                overall_delay_penalty = overall_delay_penalty + delay_penalty
                if heterogeneous_preferences == 1 and only_eco_label == 0:
                    overall_number_transshipment = overall_number_transshipment + number_transshipment
                    if use_speed == 1:
                        overall_speed = overall_speed + average_speed
                    else:
                        overall_time_ratio = overall_time_ratio + average_time_ratio
                    number_used_k = number_used_k + 1
                overall_distance = overall_distance + distance
                overall_cost = overall_cost + cost
                overall_time = overall_time + time
                overall_profit = overall_profit + profit
                overall_emission = overall_emission + emission
                overall_emission_transshipment = overall_emission_transshipment + emission_transshipment
    # transfer includes both transshipment and un_load
    overall_transfer_cost = overall_transshipment_cost + overall_un_load_cost
    served_requests = check_served_R()
    if heterogeneous_preferences == 1 and number_used_k > 0 and only_eco_label == 0:
        if use_speed == 1:
            overall_average_speed = overall_speed / number_used_k
        else:
            overall_average_time_ratio = overall_time_ratio / number_used_k
            # if regular == 1:
    #     regular_cost = normalization(overall_cost, 'overall_cost')
    #     regular_time = normalization(overall_time, 'overall_time')
    #     regular_emission = normalization(overall_emission, 'overall_emission')
    #     regular_obj = regular_cost + regular_time + regular_emission
    #     overall_cost = regular_obj
    # if routes_tuple not in hash_overall_obj_table.keys():
    #     hash_overall_obj_table[routes_tuple] = [overall_distance, round(overall_cost, 3), round(overall_time, 3),
    #                                             round(overall_profit, 3), round(overall_emission, 3), served_requests,
    #                                             overall_request_cost, overall_vehicle_cost, overall_wait_cost,
    #                                             overall_transshipment_cost, overall_un_load_cost, overall_emission_cost,
    #                                             overall_storage_cost, overall_delay_penalty]
    # print("overall_request_cost ", overall_request_cost, 'overall_vehicle_cost ', overall_vehicle_cost,
    #       'overall_wait_cost ', overall_wait_cost, 'overall_transshipment_cost ', overall_transshipment_cost,
    #       'overall_un_load_cost ', overall_un_load_cost, 'overall_transfer_cost ', overall_transfer_cost,
    #       'overall_emission_cost ', overall_emission_cost, 'overall_storage_cost ', overall_storage_cost,
    #       'overall_delay_penalty ', overall_delay_penalty)
    # here re-set during_iteration = 1 after the overall_obj calculation
    during_iteration = 1
    return overall_distance, round(overall_cost, 3), round(overall_time, 3), round(overall_profit, 3), round(
        overall_emission,
        3), served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost, overall_transshipment_cost, overall_un_load_cost, overall_emission_cost, overall_storage_cost, overall_delay_penalty, overall_number_transshipment, overall_average_speed, overall_average_time_ratio, overall_emission_transshipment


# Very slow for many datapoints.  Fastest for many costs, most readable

def is_pareto_efficient_dumb(costs):
    """
    Find the pareto-efficient points
    :param costs: An (n_points, n_costs) array
    :return: A (n_points, ) boolean array, indicating whether each point is Pareto efficient
    """
    is_efficient = np.ones(costs.shape[0], dtype=bool)
    for i, c in enumerate(costs):
        is_efficient[i] = np.all(np.any(costs[:i] > c, axis=1)) and np.all(np.any(costs[i + 1:] > c, axis=1))
    return is_efficient


# Fairly fast for many datapoints, less fast for many costs, somewhat readable

def is_pareto_efficient_simple(costs):
    """
    Find the pareto-efficient points
    :param costs: An (n_points, n_costs) array
    :return: A (n_points, ) boolean array, indicating whether each point is Pareto efficient
    """
    is_efficient = np.ones(costs.shape[0], dtype=bool)
    for i, c in enumerate(costs):
        if is_efficient[i]:
            is_efficient[is_efficient] = np.any(costs[is_efficient] < c, axis=1)  # Keep any point with a lower cost
            is_efficient[i] = True  # And keep self
    return is_efficient


# Faster than is_pareto_efficient_simple, but less readable.

def is_pareto_efficient(costs, return_mask=True):
    """
    Find the pareto-efficient points
    :param costs: An (n_points, n_costs) array
    :param return_mask: True to return a mask
    :return: An array of indices of pareto-efficient points.
        If return_mask is True, this will be an (n_points, ) boolean array
        Otherwise it will be a (n_efficient_points, ) integer array of indices.
    """
    is_efficient = np.arange(costs.shape[0])
    n_points = costs.shape[0]
    next_point_index = 0  # Next index in the is_efficient array to search for
    while next_point_index < len(costs):
        nondominated_point_mask = np.any(costs < costs[next_point_index], axis=1)
        nondominated_point_mask[next_point_index] = True
        is_efficient = is_efficient[nondominated_point_mask]  # Remove dominated points
        costs = costs[nondominated_point_mask]
        next_point_index = np.sum(nondominated_point_mask[:next_point_index]) + 1
    if return_mask:
        is_efficient_mask = np.zeros(n_points, dtype=bool)
        is_efficient_mask[is_efficient] = True
        return is_efficient_mask
    else:
        return is_efficient


def normalization(value, obj_name):
    if value > obj_record[obj_name].max():
        return 1
    if value > obj_record[obj_name].min():
        if obj_record[obj_name].max() - obj_record[obj_name].min() > 0:
            norm_value = (value - obj_record[obj_name].min()) / (
                    obj_record[obj_name].max() - obj_record[obj_name].min())
        else:
            norm_value = 1
    else:
        norm_value = 0
    return norm_value


def dominate_1(overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests,
               old_overall_distance, old_overall_cost, old_overall_time, old_overall_profit, old_overall_emission,
               old_served_requests):
    if bi_obj_cost_emission == 1 or bi_obj_cost_time == 1:
        if bi_obj_cost_emission == 1:
            if overall_cost < old_overall_cost and overall_emission < old_overall_emission:
                return True
        else:
            if overall_cost < old_overall_cost and overall_time < old_overall_time:
                return True
    else:
        if overall_cost < old_overall_cost and overall_time < old_overall_time and overall_emission < old_overall_emission:
            return True

    # normalization
    overall_cost = normalization(overall_cost, 'overall_cost')
    old_overall_cost = normalization(old_overall_cost, 'overall_cost')
    overall_time = normalization(overall_time, 'overall_time')
    old_overall_time = normalization(old_overall_time, 'overall_time')
    overall_emission = normalization(overall_emission, 'overall_emission')
    old_overall_emission = normalization(old_overall_emission, 'overall_emission')

    if weight_interval == 1:
        if old_overall_cost - overall_cost >= 0:
            obj_g_cost = weight_min_cost * (old_overall_cost - overall_cost)
        else:
            obj_g_cost = weight_max_cost * (old_overall_cost - overall_cost)

        if old_overall_emission - overall_emission >= 0:
            obj_g_emission = weight_min_emission * (old_overall_emission - overall_emission)
        else:
            obj_g_emission = weight_max_emission * (old_overall_emission - overall_emission)
        if old_overall_time - overall_time >= 0:
            obj_g_time = weight_min_time * (old_overall_time - overall_time)
        else:
            obj_g_time = weight_max_time * (old_overall_time - overall_time)
        if bi_obj_cost_emission == 1 or bi_obj_cost_time == 1:
            if bi_obj_cost_emission == 1:
                obj_g_time = 0
            else:
                obj_g_emission = 0

        overall_obj_g = obj_g_cost + obj_g_time + obj_g_emission
        if overall_obj_g > 0:
            return True
        else:
            return abs(overall_obj_g)
    else:
        if bi_obj_cost_emission == 1 or bi_obj_cost_time == 1:
            if bi_obj_cost_emission == 1:
                if overall_cost * weight_cost + overall_emission * weight_emission <= old_overall_cost * weight_cost + old_overall_emission * weight_emission:
                    return True
                else:
                    return overall_cost * weight_cost + overall_emission * weight_emission - (
                            old_overall_cost * weight_cost + old_overall_emission * weight_emission)
            else:
                if overall_cost * weight_cost + overall_time * weight_time <= old_overall_cost * weight_cost + old_overall_time * weight_time:
                    return True
                else:
                    return overall_cost * weight_cost + overall_time * weight_time - (
                            old_overall_cost * weight_cost + old_overall_time * weight_time)
        else:
            if overall_cost * weight_cost + overall_time * weight_time + overall_emission * weight_emission <= old_overall_cost * weight_cost + old_overall_time * weight_time + old_overall_emission * weight_emission:
                return True
            else:
                return overall_cost * weight_cost + overall_time * weight_time + overall_emission * weight_emission - (
                        old_overall_cost * weight_cost + old_overall_time * weight_time + old_overall_emission * weight_emission)

#this function is used to check the current solution is the most dominant solution or not
def dominate(overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests):
    if ((obj_record['overall_cost'] == overall_cost) & (obj_record['overall_time'] == overall_time) & (
            obj_record['overall_emission'] == overall_emission)).any():
        return 0
    current_overall_distance = obj_record['overall_distance'][repeat - 1]
    current_overall_cost = obj_record['overall_cost'][repeat - 1]
    current_overall_time = obj_record['overall_time'][repeat - 1]
    current_overall_profit = obj_record['overall_profit'][repeat - 1]
    current_overall_emission = obj_record['overall_emission'][repeat - 1]
    current_served_requests = obj_record['served_requests'][repeat - 1]
    for h in range(0, repeat):
        old_overall_distance = obj_record['overall_distance'][h]
        old_overall_cost = obj_record['overall_cost'][h]
        old_overall_time = obj_record['overall_time'][h]
        old_overall_profit = obj_record['overall_profit'][h]
        old_overall_emission = obj_record['overall_emission'][h]
        old_served_requests = obj_record['served_requests'][h]
        if dominate_1(overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests,
                      old_overall_distance, old_overall_cost, old_overall_time, old_overall_profit,
                      old_overall_emission, old_served_requests) is not True:
            if dominate_1(overall_distance, overall_cost, overall_time, overall_profit, overall_emission,
                          served_requests, current_overall_distance, current_overall_cost, current_overall_time,
                          current_overall_profit, current_overall_emission, current_served_requests) is True:
                return 3
            else:
                return 4
    return 2


def Pareto_preference(h):
    overall_distance = obj_record.iloc[h]['overall_distance']
    overall_cost = obj_record.iloc[h]['overall_cost']
    overall_time = obj_record.iloc[h]['overall_time']
    overall_profit = obj_record.iloc[h]['overall_profit']
    overall_emission = obj_record.iloc[h]['overall_emission']
    served_requests = obj_record.iloc[h]['served_requests']
    for l in range(0, len(obj_record)):
        if l == h:
            continue
        old_overall_distance = obj_record.iloc[l]['overall_distance']
        old_overall_cost = obj_record.iloc[l]['overall_cost']
        old_overall_time = obj_record.iloc[l]['overall_time']
        old_overall_profit = obj_record.iloc[l]['overall_profit']
        old_overall_emission = obj_record.iloc[l]['overall_emission']
        old_served_requests = obj_record.iloc[l]['served_requests']
        if dominate_1(old_overall_distance, old_overall_cost, old_overall_time, old_overall_profit,
                      old_overall_emission, old_served_requests, overall_distance, overall_cost, overall_time,
                      overall_profit, overall_emission, served_requests) is True:
            return 0
    return 1


def draw_figures(obj_record_better, path, current_save):
    global regular_non_dominated
    fig, ax1 = plt.subplots()

    color = 'tab:red'
    ax1.set_xlabel('Iteration number')
    ax1.set_ylabel('Cost', color=color)
    ax1.plot(obj_record_better.index, obj_record_better['overall_cost'], color=color)
    ax1.tick_params(axis='y', labelcolor=color)

    ax2 = ax1.twinx()  # instantiate a second axes that shares the same x-axis

    color = 'tab:blue'
    ax2.set_ylabel('Served Requests', color=color)  # we already handled the x-label with ax1
    ax2.plot(obj_record_better.index, obj_record_better['served_requests'], color=color)
    ax2.tick_params(axis='y', labelcolor=color)
    if T_or == 1:
        plt.title('Served requests and cost change (' + str(len(T)) + 'T, ' + str(request_number) + 'r, ' + str(
            vehicle_number) + 'v)')
    else:
        plt.title('Served requests and cost change (noT, ' + str(request_number) + 'r, ' + str(vehicle_number) + 'v)')

    fig.tight_layout()  # otherwise the right y-label is slightly clipped
    plt.ticklabel_format(useOffset=False, style='plain')
    plt.savefig(path + current_save + '/better_obj_record' + current_save + str(exp_number - 1) + '.pdf', format='pdf')
    plt.close()
    fig, ax1 = plt.subplots()

    color = 'tab:red'
    ax1.set_xlabel('Iteration number')
    ax1.set_ylabel('Cost', color=color)
    ax1.plot(obj_record.index, obj_record['overall_cost'], color=color)
    ax1.tick_params(axis='y', labelcolor=color)

    ax2 = ax1.twinx()  # instantiate a second axes that shares the same x-axis

    color = 'tab:blue'
    ax2.set_ylabel('Served Requests', color=color)  # we already handled the x-label with ax1
    ax2.plot(obj_record.index, obj_record['served_requests'], color=color)
    ax2.tick_params(axis='y', labelcolor=color)
    if T_or == 1:
        plt.title('Served requests and cost change (' + str(len(T)) + 'T, ' + str(request_number) + 'r, ' + str(
            vehicle_number) + 'v)')
    else:
        plt.title('Served requests and cost change (noT, ' + str(request_number) + 'r, ' + str(vehicle_number) + 'v)')

    fig.tight_layout()  # otherwise the right y-label is slightly clipped
    plt.ticklabel_format(useOffset=False, style='plain')
    plt.savefig(path + current_save + '/obj_record' + current_save + str(exp_number - 1) + '.pdf', format='pdf')
    # plt.show()
    plt.close()
    try:
        all_Tem_df = all_Tem_df.astype(float)
        all_Tem_df.plot()
        #        all_pro_df.plot()
        plt.ticklabel_format(useOffset=False, style='plain')
        plt.savefig(path + current_save + '/Temperature' + current_save + str(exp_number - 1) + '.pdf', format='pdf')
        # plt.show()
        plt.close()
        all_pro_df[all_pro_df['Acceptance probability'] < 1].plot()
        #        all_pro_df.plot()
        plt.ticklabel_format(useOffset=False, style='plain')
        plt.savefig(path + current_save + '/all_worse_pro_df' + current_save + str(exp_number - 1) + '.pdf',
                    format='pdf')
        # plt.show()
        plt.close()
    except:
        pass

    if combination == 1:
        #        weight.plot()
        ax = plt.subplot(111)
        for x in operations['operation']:
            ax.plot(weight.index, weight[x])

        ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
        #    handles, labels = ax.gca().get_legend_handles_labels()
        #    by_label = dict(zip(labels, handles))
        #    ax.legend(by_label.values(), by_label.keys())
        #
        ax.set_xlabel('Segment number')
        ax.set_ylabel('Weight')
        ax.ticklabel_format(useOffset=False, style='plain')
        plt.savefig(
            path + current_save + '/weight' + current_save + str(exp_number - 1) + '.pdf',
            format='pdf', bbox_inches='tight')
        # plt.show()
        plt.close()
    else:
        #        weight_insertion.plot()
        #        weight_removal.plot()
        ax1 = plt.subplot(111)
        for x in insert_heuristic['operator']:
            ax1.plot(weight_insertion.index, weight_insertion[x])

        ax1.legend(loc='center left', bbox_to_anchor=(1, 0.5))

        ax1.set_xlabel('Segment number')
        ax1.set_ylabel('Weights of insertion operators')
        ax1.ticklabel_format(useOffset=False, style='plain')
        plt.savefig(
            path + current_save + '/weight_insertion' + current_save + str(exp_number - 1) + '.pdf',
            format='pdf', bbox_inches='tight')
        # plt.show()
        # plt.close()
        ax2 = plt.subplot(111)
        for x in removal_heuristic['operator']:
            ax2.plot(weight_removal.index, weight_removal[x])

        ax2.legend(loc='center left', bbox_to_anchor=(1, 0.5))

        ax2.set_xlabel('Segment number')
        ax2.set_ylabel('Weights of removal operators')
        ax2.ticklabel_format(useOffset=False, style='plain')
        plt.savefig(
            path + current_save + '/weight_removal' + current_save + str(exp_number - 1) + '.pdf',
            format='pdf', bbox_inches='tight')
        # plt.show()
        plt.close()
    # multi_obj
    if real_multi_obj == 1:
        global regular_non_dominated
    # if not multi-obj, then comment all follows until next function
    #no weight

        if regular == 1:
            if bi_obj_cost_emission == 1 or bi_obj_cost_time == 1:
                if bi_obj_cost_emission == 1:
                    obj_record_use = obj_record[['overall_cost', 'overall_emission']]
                    obj_record_use_array = obj_record_use.values
                    #    data = [[1,2], [3,4], [5,5]]
                    non_dominated = is_pareto_efficient(np.array(obj_record_use_array))

                    non_dominated2 = pd.DataFrame(obj_record_use[non_dominated])

                    regular_non_dominated = my_deepcopy(non_dominated2)

                    obj_record_copy = my_deepcopy(obj_record)
                    with pd.ExcelWriter(
                            path + current_save + '/regular' + current_save + '.xlsx') as writer:  # doctest: +SKIP
                        regular_non_dominated.to_excel(writer, sheet_name='regular_non_dominated')
                        obj_record_copy.to_excel(writer, sheet_name='obj_record')

                    plt.scatter(obj_record_use['overall_cost'], obj_record_use['overall_emission'], label='Dominated solutions')

                    plt.scatter(non_dominated2['overall_cost'], non_dominated2['overall_emission'], label='Nondominated solutions')
                    plt.xlabel('Cost (euro)')
                    plt.ylabel('Emissions (kg)')
                    plt.title('Pareto frontier of bi-objective optimization')
                    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                    plt.ticklabel_format(useOffset=False, style='plain')
                    plt.savefig(
                        path + current_save + '/2d_objective_traditional' + current_save + '.pdf',
                        format='pdf', bbox_inches='tight')
                    #plt.show()
                    plt.close()
                else:
                    obj_record_use = obj_record[['overall_cost', 'overall_time']]
                    obj_record_use_array = obj_record_use.values
                    #    data = [[1,2], [3,4], [5,5]]
                    non_dominated = is_pareto_efficient(np.array(obj_record_use_array))

                    non_dominated2 = pd.DataFrame(obj_record_use[non_dominated])

                    regular_non_dominated = my_deepcopy(non_dominated2)

                    obj_record_copy = my_deepcopy(obj_record)
                    with pd.ExcelWriter(
                            path + current_save + '/regular' + current_save + '.xlsx') as writer:  # doctest: +SKIP
                        regular_non_dominated.to_excel(writer, sheet_name='regular_non_dominated')
                        obj_record_copy.to_excel(writer, sheet_name='obj_record')

                    plt.scatter(obj_record_use['overall_cost'], obj_record_use['overall_time'],
                                label='Dominated solutions')

                    plt.scatter(non_dominated2['overall_cost'], non_dominated2['overall_time'],
                                label='Nondominated solutions')
                    plt.xlabel('Cost (euro)')
                    plt.ylabel('Time (h)')
                    plt.title('Pareto frontier of bi-objective optimization')
                    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                    plt.ticklabel_format(useOffset=False, style='plain')
                    plt.savefig(
                        path + current_save + '/2d_objective_traditional' + current_save + '.pdf',
                        format='pdf', bbox_inches='tight')
                    # plt.show()
                    plt.close()
            else:
                obj_record_use = obj_record[['overall_cost', 'overall_time', 'overall_emission']]
                obj_record_use_array = obj_record_use.values
                #    data = [[1,2], [3,4], [5,5]]
                non_dominated = is_pareto_efficient(np.array(obj_record_use_array))
                # plt.plot(data)

                plt.scatter(obj_record_use['overall_cost'], obj_record_use['overall_time'], label='Dominated solutions')
                non_dominated2 = pd.DataFrame(obj_record_use[non_dominated])

                plt.scatter(non_dominated2['overall_cost'], non_dominated2['overall_time'], label='Nondominated solutions')
                plt.xlabel('Cost (euro)')
                plt.ylabel('Time (h)')
                plt.title('Pareto frontier of bi-objective optimization')
                plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                plt.ticklabel_format(useOffset=False, style='plain')
                plt.savefig(
                    path + current_save + '/2d_objective_traditional' + current_save + '.pdf',
                    format='pdf', bbox_inches='tight')
                #plt.show()
                plt.close()
                fig = plt.figure()
                ax = fig.add_subplot(111, projection='3d')

                # For each set of style and range settings, plot n random points in the box
                # defined by x in [23, 32], y in [0, 100], z in [zlow, zhigh].

                xs2 = foo(list(non_dominated2['overall_cost']))
                ys2 = foo(list(non_dominated2['overall_time']))
                zs2 = foo(list(non_dominated2['overall_emission']))
                ax.scatter(xs2, ys2, zs2, marker='o', s=60, color='orange', zorder=1)

                regular_non_dominated = my_deepcopy(non_dominated2)

                obj_record_copy = my_deepcopy(obj_record)
                with pd.ExcelWriter(
                        path + current_save + '/regular' + current_save + '.xlsx') as writer:  # doctest: +SKIP
                    regular_non_dominated.to_excel(writer, sheet_name='regular_non_dominated')
                    obj_record_copy.to_excel(writer, sheet_name='obj_record')

                # if not non_dominated2.empty:
                # #     obj_record_copy = pd.merge(obj_record_copy, non_dominated2, indicator=True, how='outer').query('_merge=="left_only"').drop('_merge', axis=1)
                #     obj_record_copy = obj_record_copy[obj_record_copy.index.isin(obj_record_copy.index.difference(non_dominated2.index))]
                xs1 = foo(list(obj_record_copy['overall_cost']))
                ys1 = foo(list(obj_record_copy['overall_time']))
                zs1 = foo(list(obj_record_copy['overall_emission']))
                ax.scatter(xs1, ys1, zs1, marker='o', color='blue', zorder=2)


                ax.set_xlabel('Cost (euro)')
                ax.set_ylabel('Time (h)')
                ax.set_zlabel('Emission (kg)')
                ax.ticklabel_format(useOffset=False, style='plain')
                plt.savefig(
                    path + current_save + '/3d_objective_traditional' + current_save + '.pdf',
                    format='pdf', bbox_inches='tight')
                #plt.show()
                plt.close()
            # Barge_number = 0
            # Train_number = 0
            # Truck_number = 0
            # for non_dominated_index in range(0, len(regular_non_dominated)):
            #     Graph(all_routes[regular_non_dominated.index[non_dominated_index]], 1, non_dominated_index)
            #     with pd.ExcelWriter(path + current_save + '/non_dominated_routes' + str(
            #             non_dominated_index) + current_save + '.xlsx') as writer:  # doctest: +SKIP
            #         for key, value in all_routes[regular_non_dominated.index[non_dominated_index]].items():
            #             value.to_excel(writer, key)
            #             if len(value[4]) > 2:
            #                 if 'Barge' in key:
            #                     Barge_number = Barge_number + 1
            #                 if 'Train' in key:
            #                     Train_number = Train_number + 1
            #                 if 'Truck' in key:
            #                     Truck_number = Truck_number + 1
            # sum_number = Barge_number + Train_number + Truck_number
            # Barge_portion = Barge_number / sum_number
            # Train_portion = Train_number / sum_number
            # Truck_portion = Truck_number / sum_number
            #
            # Barge_number = Barge_number / len(regular_non_dominated)
            # Train_number = Train_number / len(regular_non_dominated)
            # Truck_number = Truck_number / len(regular_non_dominated)
            #
            # used_vehicle_number = pd.DataFrame(
            #     [[Barge_number, Train_number, Truck_number], [Barge_portion, Train_portion, Truck_portion]],
            #     columns=['Barge_number', 'Train_number', 'Truck_number'])
            # with pd.ExcelWriter(
            #         path + current_save + '/used_vehicle_number' + current_save + '.xlsx') as writer:  # doctest: +SKIP
            #     used_vehicle_number.to_excel(writer, 'used_vehicle_number')

            #
            # Graph(all_routes[regular_non_dominated.index[0]],1)
            # with pd.ExcelWriter(
            #         path + current_save + '/non_dominated_routes' + current_save + '.xlsx') as writer:  # doctest: +SKIP
            #     for key, value in all_routes[regular_non_dominated.index[0]].items():
            #         value.to_excel(writer, key)

        if regular == 0:
            if bi_obj_cost_emission == 1 or bi_obj_cost_time == 1:
                if bi_obj_cost_emission == 1:
                    obj_record.drop_duplicates(subset=['overall_cost', 'overall_emission'], inplace=True)
                    non_dominated_preference = pd.DataFrame(
                        columns=['overall_distance', 'overall_cost', 'overall_time', 'overall_profit', 'overall_emission',
                                 'served_requests', 'iteration_time'])
                    for h in range(0, len(obj_record)):
                        if Pareto_preference(h) == 1:
                            non_dominated_preference = non_dominated_preference.append(obj_record.iloc[h])

                    with pd.ExcelWriter(
                            path + current_save + '/preference' + current_save + '.xlsx') as writer:  # doctest: +SKIP
                        non_dominated_preference.to_excel(writer, sheet_name='preference_non_dominated')
                        regular_non_dominated.to_excel(writer, sheet_name='regular_non_dominated')
                        obj_record.to_excel(writer, sheet_name='obj_record')

                    plt.scatter(obj_record['overall_cost'], obj_record['overall_emission'],color = 'blue',
                                label='Dominated solutions')

                    plt.scatter(non_dominated_preference['overall_cost'], non_dominated_preference['overall_emission'],marker='^', s=80, color = 'red',
                                label='Nondominated solutions')
                    plt.xlabel('Cost (euro)')
                    plt.ylabel('Emissions (kg)')
                    plt.title('Pareto frontier of preference-based bi-objective optimization')
                    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                    plt.ticklabel_format(useOffset=False, style='plain')
                    plt.savefig(
                        path + current_save + '/2d_objective_preference' + current_save + '.pdf',
                        format='pdf', bbox_inches='tight')

                    #plt.show()
                    plt.close()
                    plt.scatter(obj_record['overall_cost'], obj_record['overall_emission'], color='blue',
                                label='Dominated solutions')
                    plt.scatter(regular_non_dominated['overall_cost'], regular_non_dominated['overall_emission'],
                                marker='o', s=60, color='orange', label='Nondominated solutions without preference')

                    plt.scatter(non_dominated_preference['overall_cost'], non_dominated_preference['overall_emission'],
                                marker='^', s=80, color='red',
                                label='Nondominated solutions with preference')
                    plt.xlabel('Cost (euro)')
                    plt.ylabel('Emissions (kg)')
                    plt.title('Comparison on Pareto frontiers with and without preference')
                    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                    plt.ticklabel_format(useOffset=False, style='plain')
                    plt.savefig(
                        path + current_save + '/2d_objective_preference_compare' + current_save + '.pdf',
                        format='pdf', bbox_inches='tight')
                    #plt.show()
                    plt.close()
                    plt.scatter(regular_non_dominated['overall_cost'], regular_non_dominated['overall_emission'], color='blue',
                                label='Nondominated solutions without preference')

                    plt.scatter(non_dominated_preference['overall_cost'], non_dominated_preference['overall_emission'],
                                marker='^', s=80, color='red',
                                label='Nondominated solutions with preference')
                    plt.xlabel('Cost (euro)')
                    plt.ylabel('Emissions (kg)')
                    plt.title('Comparison on Pareto frontiers with and without preference')
                    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                    plt.ticklabel_format(useOffset=False, style='plain')
                    plt.savefig(
                        path + current_save + '/2d_objective_compare' + current_save + '.pdf',
                        format='pdf', bbox_inches='tight')
                    #plt.show()
                    plt.close()
                else:
                    obj_record.drop_duplicates(subset=['overall_cost', 'overall_time'], inplace=True)
                    obj_record_save_for_mode_share = copy.copy(obj_record)
                    non_dominated_preference = pd.DataFrame(
                        columns=['overall_distance', 'overall_cost', 'overall_time', 'overall_profit',
                                 'overall_emission',
                                 'served_requests', 'iteration_time'])
                    mode_shares = pd.DataFrame(columns = ['number_used_vehicles', 'all_number_served_r_by_k', 'barge_seved_r_portion', 'train_seved_r_portion', 'truck_seved_r_portion'])

                    for h in range(0, len(obj_record)):
                        if Pareto_preference(h) == 1:
                            non_dominated_preference = non_dominated_preference.append(obj_record.iloc[h])
                            for l in range(0, len(obj_record_save_for_mode_share)):
                                if obj_record_save_for_mode_share.iloc[l][1] == obj_record.iloc[h][1]:
                                    routes_non_dominated = all_routes[l]
                                    number_used_vehicles, all_number, barge_seved_r_portion, train_seved_r_portion, truck_seved_r_portion = get_mode_share(l)
                                    mode_shares = mode_shares.append(pd.Series([number_used_vehicles, all_number, barge_seved_r_portion, train_seved_r_portion, truck_seved_r_portion], index = mode_shares.columns), ignore_index=True)
                                    break
                    with pd.ExcelWriter(
                            path + current_save + '/preference' + current_save + '.xlsx') as writer:  # doctest: +SKIP
                        non_dominated_preference.to_excel(writer, sheet_name='preference_non_dominated')
                        mode_shares.to_excel(writer, sheet_name='mode_shares')
                        regular_non_dominated.to_excel(writer, sheet_name='regular_non_dominated')
                        obj_record.to_excel(writer, sheet_name='obj_record')

                    plt.scatter(obj_record['overall_cost'], obj_record['overall_time'], color='blue',
                                label='Dominated solutions')

                    plt.scatter(non_dominated_preference['overall_cost'], non_dominated_preference['overall_time'],
                                marker='^', s=80, color='red',
                                label='Nondominated solutions')
                    plt.xlabel('Cost (euro)')
                    plt.ylabel('Time (h)')
                    plt.title('Pareto frontier of preference-based bi-objective optimization')
                    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                    plt.ticklabel_format(useOffset=False, style='plain')
                    plt.savefig(
                        path + current_save + '/2d_objective_preference' + current_save + '.pdf',
                        format='pdf', bbox_inches='tight')

                    # plt.show()
                    plt.close()
                    plt.scatter(obj_record['overall_cost'], obj_record['overall_time'], color='blue',
                                label='Dominated solutions')
                    plt.scatter(regular_non_dominated['overall_cost'], regular_non_dominated['overall_time'],
                                marker='o', s=60, color='orange', label='Nondominated solutions without preference')

                    plt.scatter(non_dominated_preference['overall_cost'], non_dominated_preference['overall_time'],
                                marker='^', s=80, color='red',
                                label='Nondominated solutions with preference')
                    plt.xlabel('Cost (euro)')
                    plt.ylabel('Time (h)')
                    plt.title('Comparison on Pareto frontiers with and without preference')
                    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                    plt.ticklabel_format(useOffset=False, style='plain')
                    plt.savefig(
                        path + current_save + '/2d_objective_preference_compare' + current_save + '.pdf',
                        format='pdf', bbox_inches='tight')
                    # plt.show()
                    plt.close()
                    plt.scatter(regular_non_dominated['overall_cost'], regular_non_dominated['overall_time'],
                                color='blue',
                                label='Nondominated solutions without preference')

                    plt.scatter(non_dominated_preference['overall_cost'], non_dominated_preference['overall_time'],
                                marker='^', s=80, color='red',
                                label='Nondominated solutions with preference')
                    plt.xlabel('Cost (euro)')
                    plt.ylabel('Time (h)')
                    plt.title('Comparison on Pareto frontiers with and without preference')
                    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                    plt.ticklabel_format(useOffset=False, style='plain')
                    plt.savefig(
                        path + current_save + '/2d_objective_compare' + current_save + '.pdf',
                        format='pdf', bbox_inches='tight')
                    # plt.show()
                    plt.close()
            else:
                obj_record.drop_duplicates(subset=['overall_cost', 'overall_time', 'overall_emission'], inplace=True)
                non_dominated_preference = pd.DataFrame(
                    columns=['overall_distance', 'overall_cost', 'overall_time', 'overall_profit', 'overall_emission',
                             'served_requests', 'iteration_time'])
                for h in range(0, len(obj_record)):
                    if Pareto_preference(h) == 1:
                        non_dominated_preference = non_dominated_preference.append(obj_record.iloc[h])


                fig = plt.figure()
                ax = fig.add_subplot(111, projection='3d')

                # For each set of style and range settings, plot n random points in the box
                # defined by x in [23, 32], y in [0, 100], z in [zlow, zhigh].

                obj_record_copy2 = my_deepcopy(obj_record)
                with pd.ExcelWriter(
                        path + current_save + '/preference' + current_save + '.xlsx') as writer:  # doctest: +SKIP
                    non_dominated_preference.to_excel(writer, sheet_name='preference_non_dominated')
                    regular_non_dominated.to_excel(writer, sheet_name='regular_non_dominated')
                    obj_record_copy2.to_excel(writer, sheet_name='obj_record')

                # if not non_dominated_preference.empty:
                # #     obj_record_copy2 = pd.merge(obj_record_copy2, non_dominated_preference, indicator=True, how='outer').query('_merge=="left_only"').drop('_merge', axis=1)
                #     obj_record_copy2 = obj_record_copy2[obj_record_copy2.index.isin(obj_record_copy2.index.difference(non_dominated_preference.index))]


                xs1 = foo(list(obj_record_copy2['overall_cost']))
                ys1 = foo(list(obj_record_copy2['overall_time']))
                zs1 = foo(list(obj_record_copy2['overall_emission']))
                ax.scatter(xs1, ys1, zs1, marker='o', color = 'blue',zorder=2)

                xs = foo(list(non_dominated_preference['overall_cost']))
                ys = foo(list(non_dominated_preference['overall_time']))
                zs = foo(list(non_dominated_preference['overall_emission']))
                ax.scatter(xs, ys, zs, marker='^', s=80, color = 'red',zorder=1)

                #
                ax.set_xlabel('Cost (euro)')
                ax.set_ylabel('Time (h)')
                ax.set_zlabel('Emission (kg)')
                ax.ticklabel_format(useOffset=False, style='plain')
                plt.savefig(
                    path + current_save + '/3d_objective_preference' + current_save + '.pdf',
                    format='pdf', bbox_inches='tight')

                #plt.show()
                plt.close()
                #compare preference-based and regular
                # if not non_dominated_preference.empty:
                #     #     obj_record_copy2 = pd.merge(obj_record_copy2, non_dominated_preference, indicator=True, how='outer').query('_merge=="left_only"').drop('_merge', axis=1)
                #     regular_non_dominated = regular_non_dominated[
                #         regular_non_dominated.index.isin(regular_non_dominated.index.difference(non_dominated_preference.index))]

                fig = plt.figure()
                ax = fig.add_subplot(111, projection='3d')

                xs3 = foo(list(regular_non_dominated['overall_cost']))
                ys3 = foo(list(regular_non_dominated['overall_time']))
                zs3 = foo(list(regular_non_dominated['overall_emission']))
                ax.scatter(xs3, ys3, zs3, marker='o', color='blue', zorder=1)

                xs = foo(list(non_dominated_preference['overall_cost']))
                ys = foo(list(non_dominated_preference['overall_time']))
                zs = foo(list(non_dominated_preference['overall_emission']))
                ax.scatter(xs, ys, zs, marker='o', s=60, color='orange', zorder=1)
                #
                ax.set_xlabel('Cost (euro)')
                ax.set_ylabel('Time (h)')
                ax.set_zlabel('Emission (kg)')
                ax.ticklabel_format(useOffset=False, style='plain')
                plt.savefig(
                    path + current_save + '/3d_objective_compare' + current_save + '.pdf',
                    format='pdf', bbox_inches='tight')
                #plt.show()
                plt.close()
                fig = plt.figure()
                ax = fig.add_subplot(111, projection='3d')

                xs1 = foo(list(obj_record_copy2['overall_cost']))
                ys1 = foo(list(obj_record_copy2['overall_time']))
                zs1 = foo(list(obj_record_copy2['overall_emission']))
                ax.scatter(xs1, ys1, zs1, marker='o', color='blue')

                xs3 = foo(list(regular_non_dominated['overall_cost']))
                ys3 = foo(list(regular_non_dominated['overall_time']))
                zs3 = foo(list(regular_non_dominated['overall_emission']))
                ax.scatter(xs3, ys3, zs3, marker='o', s=60, color='orange')

                xs = foo(list(non_dominated_preference['overall_cost']))
                ys = foo(list(non_dominated_preference['overall_time']))
                zs = foo(list(non_dominated_preference['overall_emission']))
                ax.scatter(xs, ys, zs, marker='^', s=80, color='red')
                #
                ax.set_xlabel('Cost (euro)')
                ax.set_ylabel('Time (h)')
                ax.set_zlabel('Emission (kg)')
                ax.ticklabel_format(useOffset=False, style='plain')
                plt.savefig(
                    path + current_save + '/3d_objective_compare_add_dominated' + current_save + '.pdf',
                    format='pdf', bbox_inches='tight')
                #plt.show()
                plt.close()
            # Barge_number = 0
            # Train_number = 0
            # Truck_number = 0
            # for non_dominated_index in range(0,len(non_dominated_preference)):
            #     Graph(all_routes[non_dominated_preference.index[non_dominated_index]],1,non_dominated_index)
            #     with pd.ExcelWriter(path + current_save + '/non_dominated_routes' + str(non_dominated_index) + current_save + '.xlsx') as writer:  # doctest: +SKIP
            #         for key, value in all_routes[non_dominated_preference.index[non_dominated_index]].items():
            #             value.to_excel(writer, key)
            #             if len(value[4]) > 2:
            #                 if 'Barge' in key:
            #                     Barge_number = Barge_number + 1
            #                 if 'Train' in key:
            #                     Train_number = Train_number + 1
            #                 if 'Truck' in key:
            #                     Truck_number = Truck_number + 1
            # sum_number = Barge_number + Train_number + Truck_number
            # Barge_portion = Barge_number / sum_number
            # Train_portion = Train_number / sum_number
            # Truck_portion = Truck_number / sum_number
            #
            # Barge_number = Barge_number/len(non_dominated_preference)
            # Train_number = Train_number/len(non_dominated_preference)
            # Truck_number = Truck_number/len(non_dominated_preference)
            #
            # used_vehicle_number = pd.DataFrame([[Barge_number,Train_number,Truck_number],[Barge_portion,Train_portion, Truck_portion]],columns=['Barge_number', 'Train_number', 'Truck_number'])
            # with pd.ExcelWriter(path + current_save + '/used_vehicle_number' + current_save + '.xlsx') as writer:  # doctest: +SKIP
            #     used_vehicle_number.to_excel(writer, 'used_vehicle_number')

def convert(k):
    if isinstance(k, (int, np.integer)):
        return list(revert_K.keys())[list(revert_K.values()).index(k)]
    else:
        return revert_K[k]

def Graph(routes, draw_non_dominated, path_change = 'a', non_dominated_index=0):
    # output routes as wenjing

    routes_match = pd.DataFrame(columns=R[:,7], index=range(0, 3))


    for k in routes:
        if has_end_depot == 1:
            length = len(routes[k][4])
        else:
            length = len(routes[k][4]) + 1
        if length > 2:
            labeled_begin = 0
            if has_end_depot == 0:
                length = length - 1
            for i in range(1, length - 1):
                request_number = int(''.join(filter(str.isdigit, routes[k][4, i])))
                letters = new_getLetters(routes[k][4, i])
                if letters == 'pickup':
                    routes_match[request_number][0] = convert(k)
                else:
                    if letters == 'Tp':
                        routes_match[request_number][1] = convert(k)
                    else:
                        if letters == 'secondTp':
                            routes_match[request_number][2] = convert(k)
    if path_change == 'a':
        with pd.ExcelWriter(path + current_save + '/routes_match' + current_save + str(
                exp_number - 1) + '.xlsx') as writer:  # doctest: +SKIP
            routes_match.to_excel(writer, sheet_name='routes_match' + str(exp_number))
    else:
        with pd.ExcelWriter(path_change) as writer:  # doctest: +SKIP
            routes_match.to_excel(writer, sheet_name='routes_match' + str(exp_number))
    return
    G = nx.DiGraph()
    edg = []
    for k in routes:
        if has_end_depot == 1:
            length = len(routes[k][4])
        else:
            length = len(routes[k][4]) + 1
        if length > 2:
            if has_end_depot == 0:
                length = length - 1
            for i in range(0, length - 1):
                edg.append((routes[k][0][i], routes[k][0][i + 1]))

    G.add_edges_from(edg)
    # terminal:(longitude, latitude)
    pos = {'Basel': (7.592673, 47.592874), 'Weil am Rhein': (7.591401, 47.606865), \
           'Ottmarsheim': (7.524833, 47.789135), 'Strasbourg': (7.791095, 48.579390), \
           'Karlsruhe': (8.311610, 49.017073), 'Worth': (8.297794, 49.053906), \
           'Ludwigshafen': (8.438255, 49.459181), 'Mannheim': (8.451327, 49.489829), \
           'Gustavsburg': (8.309391, 49.998306), 'Koblenz': (7.589802, 50.394149), \
           'Neuss': (6.708740, 51.214719), 'Duisburg': (6.736402, 51.448688), \
           'Emmelsum': (6.600983, 51.632549), 'Emmerich': (6.253026, 51.830966), \
           'Rotterdam': (4.145499, 51.950074), \
           'Antwerp': (4.406867, 51.241200), 'Bruay-sur-lEscaut': (3.546536, 50.391956), \
           'Dortmund': (7.438884, 51.529255), 'Frankfurt-Ost': (8.718539, 50.112908), \
           'Frankfurt-West': (8.530754, 50.086362), 'Delta': (4.031017, 51.958639), \
           'Euromax': (4.044132, 51.981154), 'HOME': (4.150638, 51.942798), \
           'Moerdijk': (4.582350, 51.692446), 'Willebroek': (4.365209, 51.074444), \
           'Venlo': (6.153538, 51.389351), 'Nuremberg': (11.061368, 49.396494)}

    edge_labels = {}

    for k in routes:
        served_requests = []
        if has_end_depot == 1:
            length = len(routes[k][4])
        else:
            length = len(routes[k][4]) + 1
        if length > 2:
            labeled_begin = 0
            if has_end_depot == 0:
                length = length - 1
            for i in range(0, length - 1):
                if labeled_begin == 0:
                    if str(routes[k][0][0]) != str(routes[k][0][i + 1]):
                        labeled_begin = 1
                        if (str(routes[k][0][0]), str(routes[k][0][i + 1])) not in edge_labels:
                            edge_labels[(str(routes[k][0][0]), str(routes[k][0][i + 1]))] = {}
                        edge_labels[(str(routes[k][0][0]), str(routes[k][0][i + 1]))].update({k: []})
                if i > 0:
                    request_number = int(''.join(filter(str.isdigit, routes[k][4, i])))
                    letters = new_getLetters(routes[k][4, i])
                    if letters == 'pickup' or letters == 'Tp' or letters == 'secondTp':
                        served_requests.append(request_number)
                    else:
                        served_requests.remove(request_number)
                    served_requests_copy = served_requests.copy()
                    if str(routes[k][0][i]) != str(routes[k][0][i + 1]):
                        if (str(routes[k][0][i]), str(routes[k][0][i + 1])) not in edge_labels:
                            edge_labels[(str(routes[k][0][i]), str(routes[k][0][i + 1]))] = {}
                        edge_labels[(str(routes[k][0][i]), str(routes[k][0][i + 1]))].update(
                            {k: served_requests_copy})

    plt.figure(1, figsize=(8, 8))

    nx.draw(G, pos, edge_color='black', width=1, linewidths=1, \
            node_size=1900, node_color='blue', alpha=0.6, \
            labels={node: node for node in G.nodes()}, font_size=15)

    nx.draw_networkx_edge_labels(G, pos, edge_labels, font_color='red', font_size=20, label_pos=0.5)
    #    fig, ax = plt.subplots()

    # We change the fontsize of minor ticks label
    plt.tick_params(axis='both', which='major', labelsize=20)
    plt.tick_params(axis='both', which='minor', labelsize=18)

    plt.axis('on')
    plt.ticklabel_format(useOffset=False, style='plain')
    if draw_non_dominated == 0:
        if T_or == 1:
            plt.savefig(path + current_save + '/Graph_ALNS_T' + str(exp_number - 1) + '.pdf', format='pdf')
        else:
            plt.savefig(path + current_save + '/Graph_ALNS_noT' + str(exp_number - 1) + '.pdf', format='pdf')
    else:
        plt.savefig(
            path + current_save + '/Graph_ALNS_non_dominated' + str(non_dominated_index) + str(exp_number - 1) + '.pdf',
            format='pdf')
    # plt.show()
    plt.close()

def create_routes_R_pool_initial(xls):
    global routes, R_pool, request_flow_t
    routes = pd.read_excel(xls, None, index_col=0)
    routes_new = {}
    names = revert_names()
    # if VCP_coordination == 1:
    #     global revert_K
    #     revert_K = read_R_K(request_number_in_R, what='revert_K')
    for k in routes.keys():
        routes[k].iloc[0] = routes[k].iloc[0].map(names).fillna(routes[k].iloc[0])
        route_array = routes[k].values
        routes_new[convert(k)] = np.vstack([route_array, routes[k].columns])
    routes = routes_new
    R_pool = create_R_pool()
        # if Demir == 1:
        #     R_pool = np.array(np.empty(shape=(0, 14)), dtype='object')
        # else:
        #     R_pool = np.array(np.empty(shape=(0, 8)), dtype='object')
    # add request_flow_t
    for k in routes.keys():
        update_request_flow_t_and_T_k_record(k)

    return routes, R_pool

def update_request_flow_t_and_T_k_record(k):
    global routes, request_flow_t, T_k_record
    for h in range(len(routes[k][4])):
        col = routes[k][4, h]
        if hasNumbers(col):
            request_number = get_numbers(col)
            index_r = list(R[:, 7]).index(request_number)
            name = new_getLetters(col)
            if name == 'Td':
                request_flow_t[index_r, 1] = routes[k][3, h]
            if name == 'secondTd':
                request_flow_t[index_r, 3] = routes[k][3, h]
            if name == 'pickup':
                request_flow_t[index_r, 0] = routes[k][2, h]
                T_k_record[index_r, 2] = k
            if name == 'delivery':
                request_flow_t[index_r, 5] = routes[k][2, h]
            if name == 'Tp':
                request_flow_t[index_r, 2] = routes[k][2, h]
                T_k_record[index_r, 3] = k
            if name == 'secondTp':
                request_flow_t[index_r, 4] = routes[k][2, h]
                T_k_record[index_r, 4] = k
            if name == 'Td' or name == 'Tp':
                T_k_record[index_r, 0] = routes[k][0, h]
            if name == 'secondTd' or name == 'secondTp':
                T_k_record[index_r, 1] = routes[k][0, h]

def check_request_at_terminal_in_segment(k, infeasible_request_terminal, influenced_node_index):
    # check_request_at_influnced_terminal
    if get_numbers(routes[k][4, influenced_node_index]) != infeasible_request_terminal:
        get_influnced_terminal = routes[k][0, influenced_node_index]
        for column_index in range(1, len(routes[k][0]) - 1):
            if routes[k][0, column_index] == get_influnced_terminal and get_numbers(
                    routes[k][4, column_index]) == infeasible_request_terminal:
                influenced_node_index = column_index
                break
    return influenced_node_index
def remove_and_add_segment(insert_operation, k, infeasible_request_terminal, col_congestion, influenced_node_index, congestion_link, congestion_node, terminal, vehicle_stop_time, remove_previous=1, current_terminal = -1, pickup_terminal_for_replan_when_event_finish=-1, pickup_time_for_replan_when_event_finish=-1):
    global request_segment_in_dynamic, R, R_pool, T_k_record, request_flow_t
    influenced_node_index = check_request_at_terminal_in_segment(k, infeasible_request_terminal, influenced_node_index)
    col_congestion = check_request_at_terminal_in_segment(k, infeasible_request_terminal, col_congestion)
    #if remove_previous = 0, case 4; if remove_previous = 2, case 3
    first = 1
    for column_index in range(1, len(routes[k][0]) - 1):
        if get_numbers(routes[k][4, column_index]) == infeasible_request_terminal:
            if first == 1:
                first_terminal = int(routes[k][0, column_index])
                first = 0
            else:
                second_terminal = int(routes[k][0, column_index])
                break
    if congestion_node == -1:
        congestion_terminal = int(congestion_link[0])
    else:
        congestion_terminal = int(congestion_node)
    terminal = int(terminal)
    #here, if not use copy, the routes will be changed when changing new_column
    new_column = copy.copy(routes[k][:, col_congestion])
    if remove_previous == 1:
        if insert_operation == 'Tp':
            r_number = infeasible_request_terminal + 10000
            new_column[-1] = str(infeasible_request_terminal) + 'Td'
        elif insert_operation == 'secondTp':
            #case 5
            # removed (and will be inserted in the future) part will contains secondTd
            r_number = infeasible_request_terminal + 20000
            new_column[-1] = str(infeasible_request_terminal) + 'secondTd'

        routes[k] = np.insert(routes[k], col_congestion, new_column, axis=1)
        # remove the delivery operation
        routes[k] = np.delete(routes[k], influenced_node_index + 1, 1)
    elif remove_previous == 2:
        #case 3, future insertion part will contains secondTp
        r_number = infeasible_request_terminal + 30000
        # add the congestion node as Td
        new_column[-1] = str(infeasible_request_terminal) + 'Td'

        routes[k] = np.insert(routes[k], col_congestion, new_column,
                              axis=1)
        # this case need to remove Td operation
        routes[k] = np.delete(routes[k], influenced_node_index + 1, 1)

        # also need to change original Tp to secondTp
        # first need to find which vehicle is the next vehicle
        # how to find it? check all routes? is there any record?
        ks = find_used_k(infeasible_request_terminal)
        # check_relevant_try_not_in_routes()
        for l in ks:
            if l != k:
                if str(infeasible_request_terminal) + 'Tp' in routes[l][4]:
                    find_index = np.where(routes[l] == str(infeasible_request_terminal) + 'Tp')[1]
                    routes[l][4][find_index] = str(infeasible_request_terminal) + 'secondTp'
                    T_time = routes[l][1][find_index]
                    break
    elif remove_previous == 3:
        #case 6
        r_number = infeasible_request_terminal + 10000
        new_column[-1] = str(infeasible_request_terminal) + 'Td'
        routes[k] = np.insert(routes[k], col_congestion, new_column, axis=1)
        # remove the delivery operation
        routes[k] = np.delete(routes[k], influenced_node_index + 1, 1)
    elif remove_previous == 7:
        #case 7
        # future insertion part will contains Tp and delivery
        r_number = infeasible_request_terminal + 10000
        # this case need to remove Tp operation
        #here the influenced_node_index is the previous_operation_index, maybe different with other cases
        routes[k] = np.delete(routes[k], influenced_node_index, 1)
        # remove the delivery operation
        routes[k] = np.delete(routes[k], col_congestion - 1, 1)
    else:
        # future insertion part will contains Tp and delivery
        r_number = infeasible_request_terminal + 10000
        # this case need to remove Tp operation
        routes[k] = np.delete(routes[k], col_congestion, 1)
        # remove the delivery operation
        routes[k] = np.delete(routes[k], influenced_node_index - 1, 1)
    # recalculate schedules, and update request_flot_t
    # danger

    # new request segment
    # how to add it to R_pool or R_pool_2v, what I want is to only insert the request segment, which cannot be achiveved by the request R_pool, because it will be inserted as pickup, and it can also be cut. what i want is inert a request segment, with Tp and compatiable time with previous request segment. Need I set a new table and a new insertion operator?
    # can i use the current insertion operators? random? regret? swap? or only use greedy? How I can use current operators? why I cann't use? I can set a table where requests only insert as the Tp and not consider further be segmented. It's also in the greedy insertion? or all?
    # insert one reqeust,
    # or call the part of function insert2vehicles, i.e., the part that insert the later part only? then it's the same as in greedy, and will only use greedy
    # necessary thing: set a table where segments in
    # ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r', new_operation]
    for r_index in range(len(R[:, 7])):
        if R[r_index, 7] == infeasible_request_terminal:
            delivery_time = R[r_index, 5]
            load = R[r_index, 6]
            delay_penalty_in_segment = R[r_index, 8]
            break

    if remove_previous == 2:
        # case 3
        if 'T_time' not in locals():
            sys.exit('error')
        new_r_segment = [congestion_terminal, terminal,
                         vehicle_stop_time,
                         T_time, vehicle_stop_time,
                         T_time, load,
                         r_number, delay_penalty_in_segment,
                         insert_operation + '2']
        ap, bd = vehicle_stop_time, T_time
    elif remove_previous == 3:
        new_r_segment = [current_terminal, congestion_terminal, vehicle_stop_time,
                         delivery_time, vehicle_stop_time, delivery_time, load,
                         r_number, delay_penalty_in_segment, insert_operation]
        p, d, ap, bd = current_terminal, congestion_terminal, vehicle_stop_time, delivery_time
    elif remove_previous == 7:
        #case 7
        new_r_segment = [int(current_terminal), terminal, vehicle_stop_time,
                         delivery_time, vehicle_stop_time, delivery_time, load,
                         r_number, delay_penalty_in_segment, insert_operation]
        p, d, ap, bd = int(current_terminal), terminal, vehicle_stop_time, delivery_time
    else:
        if pickup_terminal_for_replan_when_event_finish != -1:
            congestion_terminal = pickup_terminal_for_replan_when_event_finish
        if pickup_time_for_replan_when_event_finish != -1:
            vehicle_stop_time = pickup_time_for_replan_when_event_finish
        new_r_segment = [congestion_terminal, terminal, vehicle_stop_time,
                         delivery_time, vehicle_stop_time, delivery_time, load,
                         r_number, delay_penalty_in_segment, insert_operation]
        ap, bd = vehicle_stop_time, delivery_time
    if 'p' not in locals():
        p, d = first_terminal, second_terminal
    new_r_segment = np.array(new_r_segment, dtype='O')
    if new_r_segment[0] == new_r_segment[1]:
        new_r_segment[0], new_r_segment[1] = first_terminal, second_terminal
    if r_number not in request_segment_in_dynamic[:, 7]:
        request_segment_in_dynamic = np.vstack(
            [request_segment_in_dynamic, new_r_segment])
    new_r_segment_add = [int(item) for item in new_r_segment[0:9]]
    add_r_segment(r_number,new_r_segment_add)
    #for z in routes.keys():
        #caught_strange(routes[z][4])
    return p, d, ap, bd, r_number
def remove_online(operation, k, infeasible_request_terminal, col_congestion, influenced_node_index, congestion_link, congestion_node, terminal, vehicle_stop_time,current_terminal=-1,check_it=-1, pickup_terminal_for_replan_when_event_finish=-1, pickup_time_for_replan_when_event_finish=-1):
    global routes, R_pool
    #danger the following only works for link congestion and node congestion need to consider again
    #it seems node and link congestion have the same procedure
    if operation == 'pickup':
        #case 1
        # this request can be removed fully
        routes, R_pool = remove_a_request(infeasible_request_terminal,
                                          routes,
                                          R_pool)[0:2]
        index = np.where(R == infeasible_request_terminal)[0][0]
        p, d, ap, bd, r_number = R[index, 0], R[index, 1], R[index, 2], R[index, 3], infeasible_request_terminal
    elif operation == 'Tp':
        # if previous_operation_index == col_congestion:
        # if the cogestion node is just the node of this request, then remove the part after the congestion node
        # not need to add new Td, because the previous vehicle already has it
        # case 4
        p, d, ap, bd, r_number = remove_and_add_segment('Tp', k, infeasible_request_terminal, col_congestion, influenced_node_index, congestion_link, congestion_node, terminal, vehicle_stop_time, 0, -1, pickup_terminal_for_replan_when_event_finish, pickup_time_for_replan_when_event_finish)
    elif operation == 'delivery':
        for request_operation_index in range(1, col_congestion + 1):
            if get_numbers(routes[k][4, request_operation_index]) == infeasible_request_terminal:
                previous_operation = new_getLetters(routes[k][4, request_operation_index])
                previous_operation_index = request_operation_index
                previous_terminal = routes[k][0, request_operation_index]
                previous_time = routes[k][1, request_operation_index]
                break

        #case 6 one situation is that the congested terminal is delivery terminal, What I can do is to use another k, set current terminal as T, and transfer
        if congestion_node == terminal:
            # here are two situations, one is after T's delivery or after pickup's delivery
            if previous_operation == 'pickup':
                #case 6
                p, d, ap, bd, r_number = remove_and_add_segment('Tp', k, infeasible_request_terminal, col_congestion, influenced_node_index, congestion_link, congestion_node, terminal, vehicle_stop_time, 3, current_terminal)
            elif previous_operation == 'Tp':
                #case 7
                p, d, ap, bd, r_number = remove_and_add_segment('Tp', k, infeasible_request_terminal, col_congestion,
                                                                previous_operation_index, congestion_link, congestion_node,
                                                                terminal, previous_time, 7, previous_terminal)
        else:
            # here are two situations, one is after T's delivery or after pickup's delivery

            if previous_operation == 'pickup':
                # case 2
                #set congestion node as T, and remove congestion node-delivery
                #insert Td
                p, d, ap, bd, r_number = remove_and_add_segment('Tp', k, infeasible_request_terminal, col_congestion, influenced_node_index, congestion_link, congestion_node, terminal, vehicle_stop_time)

                #then I need to consider how to insert these segments

            elif previous_operation == 'Tp':
                #case 5
                #this case need to add secondTp
                p, d, ap, bd, r_number = remove_and_add_segment('secondTp', k, infeasible_request_terminal, col_congestion, influenced_node_index, congestion_link, congestion_node, terminal, vehicle_stop_time)
            else:
            #if not above cases, it should be 'secondTp', I will not consider new segments of four vehicles
            #in this case, an emergency truck will be used with high cost.
                p, d, ap, bd, r_number = -1, -1, -1, -1, -1
    elif operation == 'Td':
        if terminal == routes[k][0,col_congestion]:
            #this case is that the congestion at T at the first k, because the congestion at next k will also be detected, no need to do anything
            p, d, ap, bd, r_number = -1, -1, -1, -1, -1
        else:
            #case 3
            #this case need to add congestion node as new Td, and change original Tp to secondTp
            #this case also needs to set the time window of Tp/secondTd as [Td, secondTp]
            p, d, ap, bd, r_number = remove_and_add_segment('Tp', k, infeasible_request_terminal, col_congestion, influenced_node_index, congestion_link, congestion_node, terminal, vehicle_stop_time, 2)
    elif operation == 'secondTp':
        print('error')
    #if r_number / big_r > 10:
        #print('caught wrong r number')
    try:
        return p, d, ap, bd, r_number
    except:
        print('error')
        return -1, -1, -1, -1, -1
def get_operation_terminal_index(k, start_check_index, influenced_r):
    delivery_operation = 0
    #if no operation of this request after the congestion terminal, then the congestion terminal is the delivery terminal, return it
    terminal = routes[k][0][start_check_index]
    influenced_node_index = start_check_index
    #find an error that col_column (start_check_index) is at later position when there is another r use the same congested terminal, and will make operation is not found, so set start_check_index = 1
    #fstart_check_index = 1
    for request_operation_index in range(start_check_index, len(routes[k][4]) - 1):
        if get_numbers(routes[k][4, request_operation_index]) == influenced_r:
            if delivery_operation == 0:
                # this operation is the pickup operation, or the delivery operation if the congested terminal is the delivery terminal or in the minddle of pickup and delivery terminal
                operation = new_getLetters(routes[k][4][request_operation_index])
            if request_operation_index > start_check_index:
                # this terminal is the operation terminal of this request after the cogestion terminal, could be pickup operation or delivery operation
                terminal = routes[k][0][request_operation_index]
                influenced_node_index = request_operation_index
            if delivery_operation == 1:
                break
            delivery_operation = delivery_operation + 1
    try:
        operation
    except:
        operation = 'operation of other relevant vehicles is infeasible, not consider to remove it currently'; print(operation)
    return operation, terminal, influenced_node_index

def add_duration(R_change_dynamic_travel_time, index, k, congestion_link, congestion_node, uncertainty_index, new_try, col_congestion, duration, vehicle_stop_time,request_number):
    # record the delayed time
    # terminal_number = new_try[0][col_congestion]
    # if terminal_number != old_terminal_number:
    if R_change_dynamic_travel_time['location_type'][index] == 'link':
        key_k_node = (k, congestion_link[1], request_number)
    else:
        key_k_node = (k, congestion_node, request_number)
    congestion_influenced_time = 0
    if duration[0] <= new_try[2][col_congestion] and duration[1] > new_try[1][col_congestion]:
        if key_k_node not in delayed_time_table_uncertainty_index.keys() or uncertainty_index not in \
                delayed_time_table_uncertainty_index[key_k_node]:
            # add duration of the congestion
            if R_change_dynamic_travel_time['location_type'][index] == 'link':
                if duration[0] <= new_try[3][col_congestion] and new_try[3][col_congestion] <= \
                        duration[1]:
                    congestion_influenced_time = duration[1] - new_try[3][col_congestion]
                else:
                    # congestions occurs in half way of the link
                    congestion_influenced_time = duration[1] - duration[0]
            else:
                congestion_influenced_time = duration[1] - vehicle_stop_time #vehicle_stop_time already consider the event begins before the arrival or not etc
            if congestion_influenced_time < 0:
                congestion_influenced_time = 0
            if K[k, 5] == influenced_mode_by_current_event:
                if K[k, 5] != 3:
                    #danger. in the current setting, I assume events happens one by one, so  delayed_time_table[key_k_node[0]][key_k_node[1]] should only equal to congestion_influenced_time, and i found a bug that there is already an value (should not be in this setting, see comment 1 in just tell in one notebook),  so i set it as it directly.
                    # delayed_time_table[key_k_node[0]][key_k_node[1]] = delayed_time_table[key_k_node[0]][
                    #                                                        key_k_node[1]] + congestion_influenced_time
                    delayed_time_table[key_k_node[0]][key_k_node[1]] = congestion_influenced_time
                else:
                    if request_number not in delayed_time_table[key_k_node[0]][key_k_node[1]].keys():
                        delayed_time_table[key_k_node[0]][key_k_node[1]][request_number] = 0
                    # delayed_time_table[key_k_node[0]][key_k_node[1]][request_number] = \
                    # delayed_time_table[key_k_node[0]][
                    #     key_k_node[1]][request_number] + congestion_influenced_time
                    delayed_time_table[key_k_node[0]][key_k_node[1]][request_number] = congestion_influenced_time
            if key_k_node not in delayed_time_table_uncertainty_index.keys():
                delayed_time_table_uncertainty_index[key_k_node] = [uncertainty_index]
            else:
                delayed_time_table_uncertainty_index[key_k_node].append(uncertainty_index)
    return key_k_node, congestion_influenced_time

def add_duration_and_check_feasibility(index, R_change_dynamic_travel_time, duration, new_try, col_congestion, vehicle_stop_time, k, inserted_r, congestion_link, congestion_node, get_reward = 0, assumed_duration = 0):
    global check_start_position, routes, delayed_time_table_uncertainty_index, dynamic_time_false
    uncertainty_index = R_change_dynamic_travel_time['uncertainty_index'][index]
    # old_terminal_number = -1
    key_k_node, congestion_influenced_time = add_duration(R_change_dynamic_travel_time, index, k, congestion_link, congestion_node, uncertainty_index, new_try, col_congestion, duration, vehicle_stop_time,inserted_r)
    # # old_terminal_number = terminal_number
    ######################
    # for col_influenced in range(col_congestion, len(new_try[0]) - 1):
    #     if R_change_dynamic_travel_time['location_type'][index] == 'link':
    #         if col_influenced == col_congestion:
    #             continue
    #         else:
    #             new_try[1][col_influenced] = new_try[1][
    #                                              col_influenced] + congestion_influenced_time
    #     else:
    #         if new_try[1][col_influenced] <= vehicle_stop_time:
    #             new_try[2][col_influenced] = new_try[2][
    #                                              col_influenced] + congestion_influenced_time
    #     # danger here if multiple requests in the same termimal, the same waiting time of a vehicle will be used multiple times to absorb delayed time
    #     if R_change_dynamic_travel_time['location_type'][index] == 'link' or (
    #             R_change_dynamic_travel_time['location_type'][index] == 'node' and ((
    #                                                                                         col_influenced == col_congestion and
    #                                                                                         new_try[2][
    #                                                                                             col_influenced] <= vehicle_stop_time) or col_influenced > col_congestion)):
    #         if col_influenced > col_congestion:
    #             new_try[1][col_influenced] = new_try[1][col_influenced] + congestion_influenced_time
    #         # check whether waiting, if waiting, waiting time can absorb delay time
    #         if new_try[2][col_influenced] > new_try[1][col_influenced]:
    #             waiting_time = new_try[2][col_influenced] - new_try[1][col_influenced]
    #
    #             if congestion_influenced_time - waiting_time <= 0:
    #                 not_change_anything = 1
    #
    #             else:
    #                 not_change_anything = 0
    #                 congestion_influenced_time = congestion_influenced_time - waiting_time
    #             if not_change_anything == 1:
    #                 break
    #             else:
    #
    #                 new_try[2][col_influenced] = new_try[2][col_influenced] + congestion_influenced_time
    #                 new_try[3][col_influenced] = new_try[3][
    #                                                  col_influenced] + congestion_influenced_time
    #         else:
    #             new_try[2][col_influenced] = new_try[2][
    #                                              col_influenced] + congestion_influenced_time
    #             new_try[3][col_influenced] = new_try[3][
    #                                              col_influenced] + congestion_influenced_time
    #     else:
    #         # node congestion and only new_try[3][col_influenced] is influenced
    #         new_try[3][col_influenced] = new_try[3][
    #                                          col_influenced] + congestion_influenced_time
    ##########################
    # if delayed_time_table[key_k_node[0]][key_k_node[1]] == 0:
    #     bool_or_route, infeasible_request_terminal =
    # else:
    #
    # check time constraints
    # only needs to check feasibility of influenced requests
    check_start_position = col_congestion
    # new_try = my_deepcopy(routes[k])
    # here need to find which request cause the infeasibility
    # if vehicles' routes and schedules are fixed, then only needs to check time constraints of requests,
    # if time windows at terminals have buffer times, then if vehicle's fixed time window is violated, then need to find removing which request at the lowest cost
    # how to find this? -> try one by one? random remove and once feasible then fine? remove how many requests?
    # it's not about requests, it's about terminals, so I can only remove this terminal and all related requests
    # if flexible vehicles, removing which request and how many of requests at the lowest cost will be more complex due to less restrictions
    # one method is just remove the first violated request according to time sequence, then remove one by one until all requests are feasible
    # but this can also be done by other vehicles when there are transshippments, i.e., adjust the time of another vehicle, which maybe has a lower cost
    # other vehicles' schedules are also unadjustable, or more specifically, the time can be shortented, because they already shortened. It's also unnecessary to remove requests in other vehicles to make this request feasible
    # and it can also be done by previous requests if they are adjustable, which may make this request feasible by adjusting previous schedules
    # are previous requests really adjustable? are they already optimal on shortening unnecessary time? Yes I think so, so I can't adjust it
    # in summary, two types of violation
    # request's time constraints, then remove the first request that violated constraints
    # vehicle's time constraints at the terminal with fixed time window, then remove first terminal
    # above only remove the request/terminal in the current vehicle. Only when a vehicle influence many other vehicles, it's worth to remove a terminal in that vehicle, which is not considered until now
    # so in the time_constraints function, I need to find which request/terminal's constraints are violated

    dynamic_time_false = 0
    bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k, new_try,
                                                                           inserted_r)

    if dynamic_time_false == 1:
        bool_or_route = False
    dynamic_time_false = 0
    if get_reward == 1 and get_reward_by_cost_gap == 0:
        return bool_or_route, infeasible_request_terminal
    else:
        if isinstance(bool_or_route, bool):
            if get_reward == 0 and add_RL == 0:
                # Dynamic_ALNS_RL34959.ALNS_removal_action_list_in_implementation.append(1)
                ALNS_removal_implementation_store[(uncertainty_index, inserted_r)] = store_all(0, -1, 1)
            if isinstance(infeasible_request_terminal, int):
                # request's time constraints are violated
                # if infeasible, then check the current operation at this terminal of this request
                if R_change_dynamic_travel_time['location_type'][index] == 'link':
                    start_check_index = col_congestion + 1
                else:
                    start_check_index = col_congestion
                operation, terminal, influenced_node_index = get_operation_terminal_index(k, start_check_index, infeasible_request_terminal)
                if operation != 'operation of other vehicles not consider to remove currently':
                    for check_it in range(len(routes[k][4]))[1:-1]:
                        if routes[k][1,check_it] >= duration[0]:
                            current_terminal_operation = routes[k][4,check_it]
                            current_terminal = routes[k][0,check_it]
                            break
                    if 'current_terminal_operation' in locals():
                        #if current operation is delivery at congested terminal, it can only be delayed
                        if not(current_terminal_operation == str(infeasible_request_terminal) + 'delivery' and operation == 'delivery' and key_k_node[1] == terminal):
                            #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                            p, d, ap, bd, r_number = remove_online(operation, k, infeasible_request_terminal, col_congestion, influenced_node_index, congestion_link, congestion_node, terminal, vehicle_stop_time,current_terminal,check_it)
                            #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
            else:
                # terminal's fixed time window is violated
                # remove node, but only remove segments after the current time
                node = int(infeasible_request_terminal)
                for col in new_try[4]:
                    if hasNumbers(col):
                        if new_try[0, list(new_try[4]).index(col)] == node:
                            request_number = get_numbers(col)
                            # routes_local = my_deepcopy(routes)
                            # R_pool_local = copy.copy
                            #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                            routes, R_pool = remove_a_request(request_number, routes,
                                                              R_pool)[0:2]
                            #check_repeat_r_in_R_pool(), check_T_k_record_and_R()

        else:
            if get_reward == 0 and add_RL == 0:
                ALNS_removal_implementation_store[(uncertainty_index, inserted_r)] = store_all(0, -1, 0)
            #     Dynamic_ALNS_RL34959.ALNS_removal_action_list_in_implementation.append(0)
            routes[k] = bool_or_route
    # if add_RL == 0 and "finish" not in R_change_dynamic_travel_time['type'][index]:
    #     #ALNS will use this function twice, one at event begins and another one at event finishes, the begining one is a assumed duration, so need to remove it to let the actual one is added when event finishes
    #     delayed_time_table_uncertainty_index[key_k_node].remove(uncertainty_index)

    if assumed_duration == 1:#ALNS's implementation during RL training/add_RL == 0 assumes duration, to avoid RL training/ALNS get rewards when event finishes and later other cases do not add the actual reward, remove related operations of assumed delay
        if K[k, 5] == influenced_mode_by_current_event:
            if K[k, 5] != 3:
                delayed_time_table[key_k_node[0]][key_k_node[1]] = \
                    delayed_time_table[key_k_node[0]][
                        key_k_node[1]] - congestion_influenced_time
            else:
                if inserted_r not in delayed_time_table[key_k_node[0]][key_k_node[1]].keys():
                    delayed_time_table[key_k_node[0]][key_k_node[1]][inserted_r] = 0
                delayed_time_table[key_k_node[0]][key_k_node[1]][inserted_r] = delayed_time_table[key_k_node[0]][
                                                                                       key_k_node[1]][
                                                                                       inserted_r] - congestion_influenced_time
        try:
            delayed_time_table_uncertainty_index[key_k_node].remove(
                uncertainty_index)
        except:
            pass

        # return key_k_node, congestion_influenced_time, uncertainty_index

def store_all(copy_state=0,new_row=0, action = -1,routes_copy=-1):
    if copy_state==0:
        if isinstance(routes_copy, int):
            return [my_deepcopy(routes), action]
        else:
            return [my_deepcopy(routes_copy), action]
    elif copy_state==1:
        return [my_deepcopy(routes), copy.copy(R_pool), copy.copy(request_flow_t), copy.copy(T_k_record), copy.copy(new_row)]
    else:
        return [my_deepcopy(routes), copy.copy(new_row), action]
def get_vehicle_stop_time(R_change_dynamic_travel_time, index, route, col_congestion, duration):
    if R_change_dynamic_travel_time['location_type'][index] == 'link':
        vehicle_stop_time = route[3][col_congestion]
    else:
        if route[1][col_congestion] >= duration[0]:
            vehicle_stop_time = route[1][col_congestion]
        elif route[2][col_congestion] >= duration[0]:
            vehicle_stop_time = duration[0]
        else:
            vehicle_stop_time = route[3][col_congestion]#when service already starts, the event cannot influence it
    return vehicle_stop_time

def take_action_to_remove(k, R_change_dynamic_travel_time, index, col_congestion, influenced_r, congestion_link, congestion_node, vehicle_stop_time, pickup_terminal_for_replan_when_event_finish = -1, pickup_time_for_replan_when_event_finish = -1):
    global routes
    if col_congestion == -1:
        for col_ in range(1,len(routes[k][0])-1):
            if routes[k][0, col_] == congestion_node:
                col_congestion = col_
                break
    if R_change_dynamic_travel_time['location_type'][
        index] == 'link':
        start_check_index = col_congestion + 1
    else:
        start_check_index = col_congestion
    operation, terminal, influenced_node_index = get_operation_terminal_index(k, start_check_index,
                                                                           influenced_r)

    # here needs to check remove which part
    p, d, ap, bd, r_number = remove_online(operation, k, influenced_r, col_congestion, influenced_node_index, congestion_link, congestion_node, terminal, vehicle_stop_time, -1, -1, pickup_terminal_for_replan_when_event_finish, pickup_time_for_replan_when_event_finish)
    if r_number / big_r > 10:
        print('caught wrong r number')
    return p, d, ap, bd, r_number

def add_r_segment(r_number,new_r_segment_add):
    global R, T_k_record, request_flow_t, R_pool
    if r_number not in R[:, 7]:
        R = np.vstack([R, new_r_segment_add])
        T_k_record = np.insert(T_k_record, len(T_k_record), np.nan, axis=0)
        request_flow_t = np.insert(request_flow_t, len(request_flow_t), np.nan, axis=0)
    else:
        # When add segment, it's better to check, if this segment is in R, remove it, and then insert the new one
        # Otherwise the segment's T and time will be different
        R = R[~(R[:, 7] == r_number)]
        R = np.vstack([R, new_r_segment_add])

    if r_number not in R_pool[:, 7]:
        R_pool = np.vstack([R_pool, new_r_segment_add])
    else:
        R_pool = R_pool[~(R_pool[:, 7] == r_number)]
        R_pool = np.vstack([R_pool, new_r_segment_add])

def save_action_reward_table(segment_length_in_RL_or_ALNS_implementation, reward_list_in_implementation):
    global ALNS_end_flag
    if add_RL == 0:
        if ALNS_greedy_under_unknown_duration_assume_duration != 3:
            stop_segments = number_of_implementation / segment_length_in_RL_or_ALNS_implementation
        else:
            stop_segments = (number_of_training + number_of_implementation) / segment_length_in_RL_or_ALNS_implementation
    else:
        stop_segments = 20
    if len(reward_list_in_implementation) / segment_length_in_RL_or_ALNS_implementation >= stop_segments:
        # write the results to a table
        if ALNS_greedy_under_unknown_duration_assume_duration != 3 or add_RL == 1:
            average_reward, std_reward = np.mean(reward_list_in_implementation), np.std(
                reward_list_in_implementation)
            print('reward_list', reward_list_in_implementation, 'average_reward, std_reward',
                  average_reward, std_reward)
        else:
            average_reward, std_reward = np.mean(reward_list_in_implementation[number_of_training:number_of_training+number_of_implementation]), np.std(
                reward_list_in_implementation[number_of_training:number_of_training+number_of_implementation])
            print('reward_list', reward_list_in_implementation[number_of_training:number_of_training+number_of_implementation], 'average_reward, std_reward',
                  average_reward, std_reward)
        if add_RL == 1:
            ALNS_ = 0
        else:
            ALNS_ = 1
        removal_operator_removal_action, removal_operator_removal_action_reward, removal_operator_waiting_action, removal_operator_waiting_action_reward, insertion_operator_insertion_action, insertion_operator_insertion_action_reward, insertion_operator_non_insertion_action, insertion_operator_non_insertion_action_reward = analyze_RL_or_ALNS_implementation_results(ALNS_)
        # so I need to establish a table and then write and save it
        action_reward_table = pd.DataFrame(
            columns=['removal_operator_removal_action', 'removal_operator_removal_action_reward',
                     'removal_operator_waiting_action', 'removal_operator_waiting_action_reward',
                     'insertion_operator_insertion_action', 'insertion_operator_insertion_action_reward',
                     'insertion_operator_non_insertion_action', 'insertion_operator_non_insertion_action_reward',
                     'average_reward', 'std_reward'], index=[0])
        action_reward_table.loc[0] = [removal_operator_removal_action, removal_operator_removal_action_reward,
                                      removal_operator_waiting_action, removal_operator_waiting_action_reward,
                                      insertion_operator_insertion_action, insertion_operator_insertion_action_reward,
                                      insertion_operator_non_insertion_action,
                                      insertion_operator_non_insertion_action_reward, average_reward, std_reward]
        path_action_reward_table = path + '/action_reward_table' + str(
            exp_number - 1) + 'R' + str(request_number_in_R) + '.xlsx'
        with pd.ExcelWriter(path_action_reward_table) as writer:  # doctest: +SKIP
            action_reward_table.to_excel(writer, sheet_name='RL' + str(add_RL) + 'type' + duration_type + 'strategy' + str(ALNS_greedy_under_unknown_duration_assume_duration))
        ALNS_end_flag = 1
        sys.exit('saved_table')

def stop_wait():
    if os.path.exists('34959.txt'):
        sys.exit(78)


def get_delay_tolerance(get_route, first_index, duration, index_r, influenced_r, k):
    storage_time_ = 0
    sevice_time_minus_one = 0
    used_k_ = update_time_caused_by_event([influenced_r], 1)
    if used_k_[1] != -1 and k == used_k_[0] and not np.isnan(request_flow_t[index_r, 2]) and not np.isnan(
            request_flow_t[index_r, 1]):
        storage_time_ = max(0, request_flow_t[index_r, 2] - request_flow_t[index_r, 1])
        # if K[used_k_[1], 5] == 1 or K[used_k_[1], 5] == 2:
        #     sevice_time_minus_one = 1
    if (used_k_[1] == -1 and K[k, 5] in [1, 2]) or (used_k_[1] != -1 and K[used_k_[1], 5] in [1, 2]):
        sevice_time_minus_one = 1 #principle is the last k is a train or barge, then request_flow_t[index_r, 5] is not the real delivery time, and need to consider service time: only use one k, and this k is barge or train: use two k, and the last k is barge or train
    if R[index_r, 5] - request_flow_t[index_r, 5] < 0:
        delay_tolerance = R[index_r, 5] - request_flow_t[index_r, 5]
    else:
        # waiting_time_ already be considered in get_route[2, index_r] - duration[0]
        if sevice_time_minus_one == 1:
            delay_tolerance = get_route[2, first_index] - duration[0] + (
                    R[index_r, 5] - request_flow_t[
                index_r, 5]) + storage_time_ - 1  # - 1 because the delay is calculated time is based the service start time, but delay is the final delivery time, for barge and trains, i need to consider 1h service time
        else:
            delay_tolerance = get_route[2, first_index] - duration[0] + (
                    R[index_r, 5] - request_flow_t[index_r, 5]) + storage_time_

    return delay_tolerance
#@profile()
def prepare_for_dynamic():
    global ALNS_implement_start_RL_can_move, number_of_training, delayed_time_table_uncertainty_index, ALNS_greedy_under_unknown_duration_assume_duration, dynamic_t, RL_is_trained_or_evaluated_or_ALNS_is_evaluated, routes_store, request_flow_t_store, RL_removal_implementation_store, ALNS_removal_implementation_store, state_reward_pairs, state_reward_pairs_insertion, dynamic_routes_store, request_segment_in_dynamic, old_current_save, delayed_time_table, R, routes, R_pool, R_change_dynamic, R_change_dynamic_travel_time, unchangeable_list, request_flow_t, T_k_record, r_best_obj_record

    dynamic_routes_store = {}
    # dynamic_routes_store_insertion = {}
    unchangeable_list = {}
    if VCP_coordination == 1:
        # open text file in read mode
        text_file = open(path + "old_current_save" + str(parallel_number) + ".txt", "r")
        # read whole file to a string
        old_current_save = text_file.read()
        # close file
        text_file.close()
        # get the routes at last time stamp
    xls = pd.ExcelFile(path + old_current_save + '/best_routes' + old_current_save + '_' + str(
        exp_number - 1) + '.xlsx')
    routes, R_pool = create_routes_R_pool_initial(xls)
    # get changed R
    #
    Data = pd.ExcelFile(data_path)
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()

    wb = load_workbook(data_path, read_only=True)  # open an Excel file and return a workbook

    if 'R_' + str(request_number_in_R) + '_' + str(dynamic_t) in wb.sheetnames:

        #below are demand uncertainty
        R_change_dynamic = pd.read_excel(Data, 'R_' + str(request_number_in_R) + '_' + str(dynamic_t))
        revert_r = R_change_dynamic['p'][0]

        if isinstance(revert_r, str):
            names = revert_names('str')
        else:
            names = revert_names('int')
        R_change_dynamic['p'] = R_change_dynamic['p'].map(names).fillna(R_change_dynamic['p'])
        R_change_dynamic['d'] = R_change_dynamic['d'].map(names).fillna(R_change_dynamic['d'])
        R_change_dynamic = np.array(R_change_dynamic)
        R_change_dynamic = add_delay_unit_penalty(R_change_dynamic)

        for index in range(len(R_change_dynamic)):
            number = R_change_dynamic[index, 7]
            R_change_dynamic[index, 7] = number + big_r * parallel_number
            request_number = R_change_dynamic[index, 7]
            # R_change_dynamic.loc[index][10] is the delay unit pealty, and R_change_dynamic.loc[index][9] is the indicator of it's new request or not
            if heterogeneous_preferences == 0:
                new_row = list(R_change_dynamic[index, 0:8]) + [R_change_dynamic[index, 9]]
            else:
                # add R_info about preferences to R directly
                new_row = list(R_change_dynamic[index, 0:8]) + [R_change_dynamic[index, 14]] + list(R_change_dynamic[index, 9:14])
            # in main():
            # request_flow_t = np.array(np.empty(shape=(len(R), 6)))
            # request_flow_t[:] = np.NaN
            # # T_k_record = pd.DataFrame(columns=['T1', 'T2', 'k1', 'k2', 'k3'], index=R[:,7])
            # T_k_record = np.array(np.empty(shape=(len(R), 5)), dtype='object')
            # T_k_record[:] = np.NaN
            if R_change_dynamic[index, 8] == 0:
                # replace changed R

                R[number] = new_row

                # remove unfinished part
                # check routes, whether this request has been transported by any vehicle
                # k1, k2, k3 = find_used_k(request_number)
                # check request_flow_t first
                # index_r = list(R[:, 7]).index(request_number)
                # 3 < np.NaN
                # Out[4]: False
                # np.NaN < 3
                # Out[5]: False
                # so only when it use T, there are values between request_flow_t[number, 0] and request_flow_t[number, 5] that can be compared with dynamic_t
                if request_flow_t[number, 0] <= dynamic_t:
                    # it has been transported by k1, then k1 and schedule can't be changed
                    # in this case, it can wait, transshipment, and detour
                    # unchangeable_list[request_number] = [['k1'], T_k_record[number]]
                    if request_flow_t[number, 1] <= dynamic_t:
                        # it has arrived at the first T, then the first vehicle and T cannot be changed,
                        # it can be transferred to another vehicle, or the predefined vehicle detour

                        if request_flow_t[number, 2] <= dynamic_t:
                            # it has been transported by k2, then the k1, T, k2 cannot be changed

                            if request_flow_t[number, 3] <= dynamic_t:
                                # it has arrived at the second T
                                if request_flow_t[number, 4] <= dynamic_t:
                                    # it has been transported by k3
                                    if request_flow_t[number, 5] <= dynamic_t:
                                        # it has been deliverred, nothing can be changed
                                        # danger I haven't finish the unchanged part of a r, and now the r can only changed before it be picked up
                                        pass
                # else:
                # it hasn't been transported, then remove the whole r
                # when it has been transported, also remove the whole r, but when insert, it can only continue the original transport by restricting the vehicle and routes
                routes, R_pool = remove_a_request(request_number, routes, R_pool)[0:2]
                # replace r's information in R_pool
                R_pool = R_pool[~(R_pool[:, 7] == request_number)]
                R_pool = np.vstack([R_pool, new_row])
            else:
                if R_change_dynamic[index, 8] == 1:
                    add_r_segment(request_number, new_row)
                    # empty_row = [np.NaN] * len(request_flow_t.columns)
                    request_flow_t = np.vstack([request_flow_t, [np.NaN] * request_flow_t.shape[1]])
                    T_k_record = np.vstack([T_k_record, [np.NaN] * T_k_record.shape[1]])
                    r_best_obj_record = np.vstack([r_best_obj_record, [np.NaN] * r_best_obj_record.shape[1]])
                else:
                    # -1, then the request is cancelled
                    # danger, here if a request is removed, then the order of r in R is changed, and somewhere maybe wrong if it use this order
                    #   for example, request_flow_t and T_k_record
                    R = R[~(R[:, 7] == request_number)]
                    R_pool = R_pool[~(R_pool[:, 7] == request_number)]
                    # danger, here should consider whether this request is cancelled partly or whole
        # insert new R or R_segment using the original one
    if 'R_' + str(request_number_in_R) + '_' + str(dynamic_t) + ' (2)' in wb.sheetnames:

        # below are travel time uncertainty, including delay and congestion at nodes and arcs
        R_change_dynamic_travel_time = pd.read_excel(Data, 'R_' + str(request_number_in_R) + '_' + str(dynamic_t) + ' (2)')
        influenced_requests_lists_different_uncertainties = {}
        for index in R_change_dynamic_travel_time.index:
            if add_RL == 1 and dynamic_RL34959.stop_everything_in_learning_and_go_to_implementation_phase == 1:
                break
            uncertainty_index = R_change_dynamic_travel_time['uncertainty_index'][index]
            influenced_requests_list = []
            # when it delays, just change one vehicle's D, congestion change all
            if R_change_dynamic_travel_time['type'][index] == 'congestion':
                # if dynamic_RL34959.implement == 1:
                #     #then send state to RL
                #
                # else:
                duration = eval(R_change_dynamic_travel_time['duration'][index])
                if R_change_dynamic_travel_time['location_type'][index] == 'link':
                    congestion_link = eval(R_change_dynamic_travel_time['location'][index])
                    congestion_links.append(congestion_link)

                    # for k in D.keys():
                    #     D[k][congestion_link[0],congestion_link[1]] = 100000000000
                    #check which vehicles are influenced by this congestion
                    congestion_node = -1

                else:
                    # congestion in node, remove current and future plans that use this node
                    congestion_node = R_change_dynamic_travel_time['location'][index]
                    congestion_nodes.append(congestion_node)
                    congestion_nodes_at_begining[congestion_node] = duration
                    congestion_link = -1
                remove_node_in_congestion_list = []
                for node in congestion_nodes_at_begining.keys():
                    if node != congestion_node:
                        if duration[0] >= congestion_nodes_at_begining[node][1]:
                            remove_node_in_congestion_list.append(node)
                for node in remove_node_in_congestion_list:
                    del congestion_nodes_at_begining[node]
                global influenced_mode_by_current_event
                influenced_mode_by_current_event = R_change_dynamic_travel_time['mode'][index]
                for k in routes.keys():
                    if len(routes[k][0]) <= 2:
                        continue
                    if K[k, 5] == influenced_mode_by_current_event:
                        passed_terminals = [x for x, y in groupby(routes[k][0])]
                        if (R_change_dynamic_travel_time['location_type'][index] == 'link' and congestion_link[0] in passed_terminals and congestion_link[1] in passed_terminals) or (R_change_dynamic_travel_time['location_type'][index] == 'node' and congestion_node in passed_terminals):
                            if (R_change_dynamic_travel_time['location_type'][index] == 'link' and passed_terminals.index(congestion_link[1]) - passed_terminals.index(
                                    congestion_link[0]) == 1) or R_change_dynamic_travel_time['location_type'][index] == 'node':
                                # if congestion happens after the vehicle use this link
                                if R_change_dynamic_travel_time['location_type'][index] == 'link':
                                    congestion_ = congestion_link[0]
                                else:
                                    congestion_ = congestion_node
                                col_congestions = [index3 for index3, element in enumerate(routes[k][0][0:-1]) if
                                                       element == congestion_]
                                # col_congestion = list(routes[k][0]).index(congestion_link[0])
                                influenced_operations = []
                                for col_congestion in col_congestions:
                                    if col_congestion == 0:
                                        # remove begin_depot
                                        continue

                                    #if the current operation is delivery, then need to check the current time is bigger than pickup's start time or not, because I consider the replanning from the pickup terminal when it is delivered
                                    if getLetters(routes[k][4, col_congestion]) in ['Td', 'delivery']:
                                        request_number_ = get_numbers(routes[k][4, col_congestion])
                                        for operation_index in range(len(routes[k][4, 1:-1])):
                                            if request_number_ == get_numbers(routes[k][4, operation_index + 1]):
                                                operation_type = new_getLetters(routes[k][4, operation_index + 1])
                                                pickup_time_of_this_r_at_this_route = routes[k][2, operation_index + 1]
                                                break
                                        if operation_type == 'Tp':
                                            if pickup_time_of_this_r_at_this_route >= duration[0]:
                                                influenced_operations.append(routes[k][4, col_congestion])
                                            else:
                                                #if the first route's pickup time is larger than congestion time, it is also replannable
                                                used_ks = find_used_k(request_number_)
                                                k1 = used_ks[0]
                                                for operation_index in range(len(routes[k1][4, 1:-1])):
                                                    if request_number_ == get_numbers(routes[k1][4, operation_index + 1]):
                                                        operation_type = new_getLetters(routes[k1][4, operation_index + 1])
                                                        pickup_time_of_this_r_at_first_route = routes[k1][2, operation_index + 1]
                                                        break
                                                if pickup_time_of_this_r_at_first_route >= duration[0]:
                                                    influenced_operations.append(routes[k][4, col_congestion])
                                        else:
                                            if pickup_time_of_this_r_at_this_route >= duration[0]:
                                                influenced_operations.append(routes[k][4, col_congestion])
                                    else:
                                        # only when the current time is bigger than congestion start time, it's influenced
                                        if routes[k][2, col_congestion] >= duration[0]:
                                            influenced_operations.append(routes[k][4, col_congestion])
                                if influenced_operations == []:
                                    continue
                                for influenced_operation in influenced_operations:
                                    #check_RL_ALNS_iteraction_bug()
                                    stop2 = 0
                                    for index2, element in enumerate(routes[k][4]):
                                        if element == influenced_operation:
                                            col_congestion = index2
                                            # just set the first r as the inserted_r
                                            inserted_r = get_numbers(routes[k][4, col_congestion])
                                            stop2 = 1
                                            break
                                    if stop2 == 0:
                                        continue
                                    # danger, after insertion, need to recalculate the congestion_influenced_time
                                    if R_change_dynamic_travel_time['location_type'][index] == 'link':
                                        check_start_time = routes[k][3][col_congestion]
                                    else:
                                        # previous may wait, now may not wait and transshipment directly
                                        check_start_time = routes[k][1][col_congestion]
                                    if 'depot' not in routes[k][4][col_congestion] and check_start_time >= duration[
                                        0] and ((check_start_time <= duration[1] and add_RL == 0) or add_RL == 1):
                                        vehicle_stop_time = get_vehicle_stop_time(R_change_dynamic_travel_time, index, routes[k], col_congestion, duration)
                                        # obtain the requests in this link, and all requests in the following links are influenced

                                        # ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r', new_operation]
                                        request_segment_in_dynamic = np.array(np.empty(shape=(0, len(R[0]) + 1)),
                                                                              dtype=np.object)
                                        # here copy routes[k] and use new_try to simulate the influences of unexpected events, and the delayed time will be added in time constraints function
                                        new_try = my_deepcopy(routes[k])
                                        # danger here should change the col after one iteration of the removing
                                        # for col in range(col_congestion, len(new_try[0])):
                                        influenced_r = get_numbers(influenced_operation)
                                        influenced_r_index = influenced_r % (big_r * parallel_number)
                                        influenced_requests_list.append(influenced_r)
                                        if stochastic == 1:
                                            # maybe other stochastic modelling approaches are used, so check add_RL seperately
                                            # and (after_action_review == 0 or implement_RL == 1)
                                            delay_tolerance = get_delay_tolerance(copy.deepcopy(routes[k]), col_congestion, duration, influenced_r_index, influenced_r, k)

                                            for influenced_operations_index in range(col_congestion,
                                                                                     len(routes[k][4])):
                                                if get_numbers(
                                                        routes[k][4][influenced_operations_index]) == influenced_r:
                                                    r_delivery_operation_index = influenced_operations_index
                                                    break
                                            influenced_passed_terminals = [x for x, y in groupby(
                                                routes[k][0][col_congestion : r_delivery_operation_index + 1])]
                                            # remove links/terminals that don't have congestion
                                            for influenced_passed_terminal in influenced_passed_terminals:
                                                if influenced_passed_terminal not in congestion_nodes:
                                                    influenced_passed_terminals.remove(influenced_passed_terminal)
                                            for passed_terminal in range(10 - len(influenced_passed_terminals)):
                                                influenced_passed_terminals.append(-1)
                                            if influenced_passed_terminals[1] != -1:
                                                print('more than one terminal is congested')
                                            new_row = pd.Series(data={'uncertainty_index': uncertainty_index, 'uncertainty_type': 'begin', 'request': influenced_r, 'vehicle': k, 'delay_tolerance': delay_tolerance, 'passed_terminals': influenced_passed_terminals, 'current_time': duration[0], 'action': -10000000, 'reward': -10000000})

                                            # if len(state_reward_pairs) > 1:
                                            #     print('len > 1')
                                            # save delay_tolerance, influenced_passed_terminals, and current time to an excel
                                            # parallel_save_excel(path + 'state_reward_pairs.xlsx',
                                            #                     state_reward_pairs, 'state_reward_pairs')
                                            # store original routes, and then check feasibility after the duration
                                            # store whole routes in case chain reaction
                                            # after each removal/insertion, the routes will change, so need to store routes of each t, k, r
                                            dynamic_routes_store[(uncertainty_index, influenced_r)] = store_all(1,new_row)
                                            if add_RL == 1 and dynamic_RL34959.implement == 1:
                                                #check_RL_ALNS_iteraction_bug()
                                                while True:
                                                    if dynamic_RL34959.stop_everything_in_learning_and_go_to_implementation_phase == 1:
                                                        action = 0
                                                        break
                                                    # print('prepare 1')
                                                    if dynamic_RL34959.clear_pairs_done == 1:
                                                        #it's begin and implementation, so add it in pairs
                                                        state_reward_pairs = pd.DataFrame(
                                                            columns=state_reward_pairs.columns)
                                                        state_reward_pairs = state_reward_pairs.append(new_row,
                                                                                                       ignore_index=True)
                                                        break
                                                if dynamic_RL34959.implement == 1: #only when implement it can be 1, otherwise, when learning phase, it is changed to 1 here, then in the insertion function, implement phase begins, then it is wrong, because RL is not end one iteration in implementation phase and not convert it to 0, and in the first implementation, RL move without the real 1 from ALNS and then wrong
                                                    ALNS_implement_start_RL_can_move = 1
                                                # if after_action_review == 0:
                                                # if len(state_reward_pairs) in [0,2] and dynamic_RL34959.implement == 1:
                                                #     print('here is wrong!!!')
                                                # for ALNS, random removal

                                                # for RL,

                                                # get action
                                                break_flag = 0
                                                while True:
                                                    # print('prepare 2')
                                                    if dynamic_RL34959.stop_everything_in_learning_and_go_to_implementation_phase == 1:
                                                        break
                                                    stop_wait()
                                                    # state_reward_pairs = parallel_read_excel(
                                                    #     path + 'state_reward_pairs.xlsx', 'state_reward_pairs')
                                                    for pair_index in state_reward_pairs.index:

                                                        if pair_index not in state_reward_pairs.index:
                                                            continue
                                                        try:
                                                            check = state_reward_pairs['uncertainty_index'][
                                                                pair_index] == uncertainty_index and \
                                                            state_reward_pairs['uncertainty_type'][
                                                                pair_index] == 'begin' and \
                                                            state_reward_pairs['vehicle'][pair_index] == k and \
                                                            state_reward_pairs['request'][
                                                                pair_index] == influenced_r and \
                                                            state_reward_pairs.loc[pair_index]['action'] != -10000000
                                                        except:
                                                            break
                                                        if check:
                                                            #check_RL_ALNS_iteraction_bug()
                                                            action = state_reward_pairs.loc[pair_index]['action']
                                                            dynamic_RL34959.ALNS_got_action_in_implementation = 1
                                                            # store the routes, new_row, action for evaluate the performance of RL during implementation
                                                            RL_removal_implementation_store[(uncertainty_index, influenced_r)] = store_all(2, new_row, action)
                                                            # action = 1
                                                            if action == 1:
                                                                # remove
                                                                #check_RL_ALNS_iteraction_bug()
                                                                p, d, ap, bd, r_number = take_action_to_remove(k, R_change_dynamic_travel_time, index, col_congestion, influenced_r, congestion_link, congestion_node, vehicle_stop_time)
                                                                if dynamic_RL34959.implement == 1:
                                                                    while True:
                                                                        # print('prepare 9')
                                                                        if dynamic_RL34959.clear_pairs_done == 1:
                                                                            break
                                                                # state_reward_pairs.loc[pair_index]['reward'] = 0
                                                            break_flag = 1
                                                            break
                                                    if break_flag == 1:
                                                        break

                                                # else:
                                                #removal is finished, then if removal action = 1, I need to insert the removed request by RL
                                                if action == 1:
                                                    if r_number != -1:
                                                        if R_change_dynamic_travel_time['location_type'][index] == 'link':
                                                            check_terminal = congestion_link[0]
                                                        else:
                                                            check_terminal = congestion_node
                                                        #check_RL_ALNS_iteraction_bug()
                                                        insert_r_in_learning_or_implementation(index, check_terminal, new_row, r_number, k,
                                                                                           uncertainty_index, influenced_r,
                                                                                           R_change_dynamic_travel_time,
                                                                                           duration, congestion_link,
                                                                                           congestion_node,dynamic_RL34959.implement)
                                                        #check_RL_ALNS_iteraction_bug()
                                                        if dynamic_RL34959.implement == 1:
                                                            while True:
                                                                # print('parepare 10')
                                                                if dynamic_RL34959.clear_pairs_done == 1:
                                                                    break
                                                
                                            else:
                                                duration_copy = copy.copy(duration)

                                                #check_RL_ALNS_iteraction_bug()
                                                if ALNS_greedy_under_unknown_duration_assume_duration == 0:
                                                    duration_copy[1] = duration_copy[0]
                                                elif ALNS_greedy_under_unknown_duration_assume_duration == 1:
                                                    duration_copy[1] = 99999
                                                elif ALNS_greedy_under_unknown_duration_assume_duration == 2:
                                                    duration_copy[1] = duration_copy[0] + random.randint(0, 40)
                                                elif ALNS_greedy_under_unknown_duration_assume_duration == 3:
                                                    if Dynamic_ALNS_RL34959.ALNS_calculates_average_duration_list == [] or (add_RL == 0 and len(Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation) <= number_of_training):
                                                        duration_copy[1] = duration_copy[0]
                                                    else:
                                                        duration_copy[1] = duration_copy[0] + int(np.mean(Dynamic_ALNS_RL34959.ALNS_calculates_average_duration_list))
                                                #check_RL_ALNS_iteraction_bug()
                                                #when RL is not implemented, still use ALNS
                                                add_duration_and_check_feasibility(index, R_change_dynamic_travel_time,
                                                                                   duration_copy, new_try, col_congestion,
                                                                                   vehicle_stop_time, k, influenced_r, congestion_link, congestion_node, 0, 1)
                                                #check_RL_ALNS_iteraction_bug()
                                                #then insert removed request/segment

                                                if influenced_r in R_pool[:, 7]:
                                                    #store duration table before adding assumed durations
                                                    store_delayed_time_table, store_delayed_time_table_uncertainty_index = my_deepcopy(delayed_time_table), my_deepcopy(delayed_time_table_uncertainty_index)
                                                    # len_1 = len(R_pool)
                                                    #check_RL_ALNS_iteraction_bug()
                                                    if influenced_r in request_segment_in_dynamic[:, 7]:
                                                        routes, R_pool = greedy_insert(influenced_r, 1, 0, -1, -1,
                                                                      uncertainty_index, -1,
                                                                      -1, duration_copy,
                                                                      congestion_link,
                                                                      congestion_node, index)
                                                    else:
                                                        routes, R_pool = greedy_insert(influenced_r, 0, 0, -1, -1,
                                                                      uncertainty_index, -1,
                                                                      -1, duration_copy,
                                                                      congestion_link,
                                                                      congestion_node, index)
                                                    #check_RL_ALNS_iteraction_bug()
                                                    #restore delay tables
                                                    delayed_time_table, delayed_time_table_uncertainty_index = my_deepcopy(
                                                        store_delayed_time_table), my_deepcopy(
                                                        store_delayed_time_table_uncertainty_index)
                                                    # len_2 = len(R_pool)
                                                    # if len_1 > len_2:
                                                    #     Dynamic_ALNS_RL34959.ALNS_insertion_action_list_in_implementation.append(1)
                                                    # else:
                                                    #     Dynamic_ALNS_RL34959.ALNS_insertion_action_list_in_implementation.append(0)
                                        else:
                                            add_duration_and_check_feasibility(index, R_change_dynamic_travel_time, duration, new_try, col_congestion, vehicle_stop_time, k, inserted_r, congestion_link, congestion_node)

                                        # if violated, then remove it

                                        # here, if one request has been started from previous terminal (not in this link), then only this part need to be reconsidered and previous one should be unchanged
                                        # in this case, the arrival time in this terminal should be the pickup start time of another vehicle
                                        # then, how to identify which request is this type?
                                        # and, after the influenced link, also all requests need to be canceled
                                        # in one word, all requests stop at the terminal of the begining of the link, and wait for transshipment or detour

                                        # influenced_requests_list.append(get_numbers(routes[k][4][col]))
                                        # remove this request, if this request is
                if 'influenced_requests_list' in locals():
                    influenced_requests_lists_different_uncertainties[uncertainty_index] = influenced_requests_list
                else:
                    influenced_requests_lists_different_uncertainties[uncertainty_index] = []
            elif R_change_dynamic_travel_time['type'][index] == 'delay':
                #delay
                D[R_change_dynamic_travel_time['vehicle'][index]][eval(R_change_dynamic_travel_time['location'][index])[0],
                     eval(R_change_dynamic_travel_time['location'][index])[1]] = 100000000000
            elif R_change_dynamic_travel_time['type'][index] == 'congestion_finish':

                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                if R_change_dynamic_travel_time['location_type'][index] == 'link':
                    congestion_link = eval(R_change_dynamic_travel_time['location'][index])
                    check_terminal = congestion_link[0]
                    congestion_links.remove(congestion_link)
                else:
                    congestion_node = R_change_dynamic_travel_time['location'][index]
                    check_terminal = congestion_node
                    congestion_nodes.remove(congestion_node)
                duration = eval(R_change_dynamic_travel_time['duration'][index])
                Dynamic_ALNS_RL34959.ALNS_calculates_average_duration_list.append(duration[1]-duration[0])
                #when event finishes, change current time to event finish time
                dynamic_t = duration[1]
                #check feasibility of influenced requests in the original routes
                # store routes, R_pool, etc, for each request
                # store_routes, store_R_pool, store_request_flow_t, store_T_k_record = store_all()
                # get the influenced requests
                influenced_requests = influenced_requests_lists_different_uncertainties[uncertainty_index]
                if add_RL == 1:
                    if dynamic_RL34959.implement == 0:
                        # store the routes, etc, without using RL, and after RL training over, restore them
                        [routes_without_RL, R_pool_without_RL, request_flow_t_without_RL,
                         T_k_record_without_RL] = my_deepcopy([routes, R_pool, request_flow_t, T_k_record])
    
                        if get_reward_by_cost_gap == 1:
                            all_new_cost_of_r_when_uncertainty_finishes = {}
                            for influenced_r in influenced_requests:
                                all_new_cost_of_r_when_uncertainty_finishes[influenced_r] = get_r_cost_in_all_routes(influenced_r)[0]
    
                        for influenced_r in influenced_requests:
                            if dynamic_RL34959.implement == 1:
                                break
                            if get_reward_by_cost_gap == 1:
                                #danger here all the checking should for the original routes, but when I call time_constraints function, the request_flow_t is changed
                                new_cost_of_r_when_uncertainty_finishes = all_new_cost_of_r_when_uncertainty_finishes[influenced_r]
                            else:
                                new_cost_of_r_when_uncertainty_finishes = -1
                            # no state here? But RL needs state and action,
                            # or the same sate and actions with previous one
                            # state can be same, how to let action also the same?
                            # here only get rewards, in fact no action is taken
    
                            # new_row = pd.Series(
                            #     data={'uncertainty_index': uncertainty_index, 'uncertainty_type': 'finish',
                            #           'request': influenced_r,
                            #           'vehicle': -1, 'delay_tolerance': -1,
                            #           'passed_terminals': [-1]*10, 'current_time': -1,
                            #           'action': -10000000, 'reward': -10000000})
                            # send_to_RL_and_train(0, influenced_r, check_terminal,
                            #                      R_change_dynamic_travel_time, index, duration, vehicle_stop_time,
                            #                      uncertainty_index, new_cost_of_r_when_uncertainty_finishes, congestion_link,
                            #                      congestion_node)
                            if after_action_review == 1:
                                RL_is_trained_or_evaluated_or_ALNS_is_evaluated = 1
                                #these are training for the removal operations first, and if removal, then also insertion operations
                                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                                [routes, R_pool, request_flow_t, T_k_record,
                                 state_when_uncertainty_begins] = my_deepcopy(
                                    dynamic_routes_store[
                                        (uncertainty_index,
                                         influenced_r)])
                                add_one_row_to_T_k_record(len(R) - 1)
                                add_one_row_to_request_flow(len(R) - 1)
                                routes_store = my_deepcopy(routes)
                                request_flow_t_store = copy.copy(request_flow_t)
                                k = state_when_uncertainty_begins['vehicle']
                                for index2, element in enumerate(routes[k][0]):
                                    if index2 == 0 or index2 == (len(routes[k][0]) - 1):
                                        continue
                                    if element == check_terminal and get_numbers(routes[k][4, index2]) == influenced_r:
                                        col_congestion = index2
                                        # not_pass_congested_terminal = 0
                                        break
                                vehicle_stop_time = get_vehicle_stop_time(R_change_dynamic_travel_time, index,
                                                                          routes[k], col_congestion, duration)
                                for repeat_scene in range(1):
                                    if dynamic_RL34959.implement == 1:
                                        break
                                    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                                    action = send_to_RL_and_train(repeat_scene, influenced_r, check_terminal,
                                                         R_change_dynamic_travel_time, index, duration, vehicle_stop_time,
                                                         uncertainty_index, new_cost_of_r_when_uncertainty_finishes,congestion_link,congestion_node,state_when_uncertainty_begins,routes_store,request_flow_t_store)
                                    if dynamic_RL34959.implement == 1:
                                        break
                                    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                                    # if r in routes is removed (action == 1), then need to reload routes etc for the next repeat_scene
                                    if action == 1 and repeat_scene < 10:
                                        [routes, R_pool, request_flow_t, T_k_record,
                                         state_when_uncertainty_begins] = my_deepcopy(
                                            dynamic_routes_store[
                                                (uncertainty_index,
                                                 influenced_r)])
                                        add_one_row_to_T_k_record(len(R) - 1)
                                        add_one_row_to_request_flow(len(R) - 1)
                                RL_is_trained_or_evaluated_or_ALNS_is_evaluated = 0
                        for k_tt in routes_without_RL.keys():
                            if not np.array_equal(routes_without_RL[k_tt], routes[k_tt], equal_nan=False):
                                print(routes_without_RL[k_tt], routes[k_tt])
                        [routes, R_pool, request_flow_t, T_k_record] = my_deepcopy([routes_without_RL, R_pool_without_RL, request_flow_t_without_RL,
                         T_k_record_without_RL])
                        #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                    else:

                        # evaluate RL's implementation performance
                        RL_is_trained_or_evaluated_or_ALNS_is_evaluated = 1
                        current_routes = my_deepcopy(routes)
                        for influenced_r in influenced_requests:

                            if (uncertainty_index,influenced_r) in RL_removal_implementation_store.keys():
                                routes, new_row_after_action, action = my_deepcopy(RL_removal_implementation_store[(uncertainty_index,influenced_r)])
                                if 'action' in locals():

                                    # get reward by comparing action and feasibility
                                    reward = get_reward(check_terminal, action, influenced_r, R_change_dynamic_travel_time,
                                                        index, duration,
                                                        congestion_link, congestion_node,
                                                        -1)

                                    print('evaluate implementation removal action', action, 'reward', reward, 'state', new_row_after_action)
                                    Dynamic_ALNS_RL34959.removal_state_list_in_implementation.append(new_row_after_action)
                                    Dynamic_ALNS_RL34959.removal_action_list_in_implementation.append(action)
                                    Dynamic_ALNS_RL34959.removal_reward_list_in_implementation.append(reward)
                                    Dynamic_ALNS_RL34959.reward_list_in_implementation.append(reward)
                                    print('implementation times', len(Dynamic_ALNS_RL34959.reward_list_in_implementation))
                                    segment_length_in_RL_or_ALNS_implementation = 10
                                    if len(Dynamic_ALNS_RL34959.reward_list_in_implementation) % segment_length_in_RL_or_ALNS_implementation == 0:
                                        last_segment = Dynamic_ALNS_RL34959.reward_list_in_implementation[len(
                                            Dynamic_ALNS_RL34959.reward_list_in_implementation) - segment_length_in_RL_or_ALNS_implementation: len(
                                            Dynamic_ALNS_RL34959.reward_list_in_implementation)]
                                        average_reward, std_reward = np.mean(last_segment), np.std(last_segment)
                                        print('last_segment',last_segment,'average_reward, std_reward',average_reward, std_reward)
                                        # if average_reward < 0 and std_reward >= 0:
                                        #     dynamic_RL34959.implement = 0
                                        analyze_RL_or_ALNS_implementation_results()
                                    save_action_reward_table(segment_length_in_RL_or_ALNS_implementation, Dynamic_ALNS_RL34959.reward_list_in_implementation)

                                    del action

                            if (uncertainty_index,influenced_r) in RL_insertion_implementation_store.keys():
                                for k in RL_insertion_implementation_store[(uncertainty_index,influenced_r)].keys():
                                    routes, new_row_after_action, action = my_deepcopy(RL_insertion_implementation_store[(uncertainty_index, influenced_r)][k])

                                    if 'action' in locals():
                                        # get reward by comparing ation and feasibility
                                        reward = get_reward(check_terminal, action, influenced_r, R_change_dynamic_travel_time, index, duration,
                                                            congestion_link, congestion_node,
                                                            -1)
                                        print('evaluate implementation insertion action', action, 'reward', reward, 'state', new_row_after_action)
                                        Dynamic_ALNS_RL34959.insertion_state_list_in_implementation.append(new_row_after_action)
                                        Dynamic_ALNS_RL34959.insertion_action_list_in_implementation.append(action)
                                        Dynamic_ALNS_RL34959.insertion_reward_list_in_implementation.append(reward)
                                        Dynamic_ALNS_RL34959.reward_list_in_implementation.append(reward)
                                        print('implementation times', len(Dynamic_ALNS_RL34959.reward_list_in_implementation))
                                        segment_length_in_RL_or_ALNS_implementation = 10
                                        if len(Dynamic_ALNS_RL34959.reward_list_in_implementation) % segment_length_in_RL_or_ALNS_implementation == 0:
                                            last_segment = Dynamic_ALNS_RL34959.reward_list_in_implementation[len(Dynamic_ALNS_RL34959.reward_list_in_implementation) - segment_length_in_RL_or_ALNS_implementation : len(Dynamic_ALNS_RL34959.reward_list_in_implementation)]
                                            average_reward, std_reward = np.mean(last_segment), np.std(last_segment)
                                            print('last_segment', last_segment, 'average_reward, std_reward', average_reward, std_reward)
                                            # if average_reward < 0 and std_reward >= 0:
                                            #     dynamic_RL34959.implement = 0
                                            analyze_RL_or_ALNS_implementation_results()
                                        save_action_reward_table(segment_length_in_RL_or_ALNS_implementation,Dynamic_ALNS_RL34959.reward_list_in_implementation)

                                        del action
                        routes = my_deepcopy(current_routes)
                        #also update the time caused by event
                        update_time_caused_by_event(influenced_requests)
                        RL_is_trained_or_evaluated_or_ALNS_is_evaluated = 0
                    # for pair_index in dynamic_RL34959.state_action_reward_collect.index:
                else:

                    #evaluate ALNS's performance
                    RL_is_trained_or_evaluated_or_ALNS_is_evaluated = 1
                    current_routes = my_deepcopy(routes)
                    for influenced_r in influenced_requests:

                        if (uncertainty_index, influenced_r) in ALNS_removal_implementation_store.keys():
                            routes, action = ALNS_removal_implementation_store[(uncertainty_index, influenced_r)]
                            if 'action' in locals():

                                # get reward by comparing action and feasibility
                                reward = get_reward(check_terminal, action, influenced_r, R_change_dynamic_travel_time,
                                                    index, duration,
                                                    congestion_link, congestion_node,
                                                    -1)
                                print('ALNS evaluate implementation removal action', action, 'reward', reward)
                                if ALNS_greedy_under_unknown_duration_assume_duration != 3 or (
                                        ALNS_greedy_under_unknown_duration_assume_duration == 3 and len(
                                    Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation) >= number_of_training):
                                    Dynamic_ALNS_RL34959.ALNS_removal_action_list_in_implementation.append(action)
                                    Dynamic_ALNS_RL34959.ALNS_removal_reward_list_in_implementation.append(reward)
                                Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation.append(reward)
                                print('implementation times', len(Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation))
                                segment_length_in_RL_or_ALNS_implementation = 10
                                if len(
                                        Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation) % segment_length_in_RL_or_ALNS_implementation == 0:
                                    last_segment = Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation[len(
                                        Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation) - segment_length_in_RL_or_ALNS_implementation: len(
                                        Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation)]
                                    average_reward, std_reward = np.mean(last_segment), np.std(last_segment)
                                    print('last_segment', last_segment, 'average_reward, std_reward', average_reward,
                                          std_reward)

                                    analyze_RL_or_ALNS_implementation_results(1)
                                save_action_reward_table(segment_length_in_RL_or_ALNS_implementation, Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation)

                                del action
                        if (uncertainty_index, influenced_r) in ALNS_insertion_implementation_store.keys():
                            for k in ALNS_insertion_implementation_store[(uncertainty_index, influenced_r)].keys():
                                routes, action = ALNS_insertion_implementation_store[(uncertainty_index, influenced_r)][k]

                                if 'action' in locals():
                                    # get reward by comparing action and feasibility
                                    reward = get_reward(check_terminal, action, influenced_r,
                                                        R_change_dynamic_travel_time, index, duration,
                                                        congestion_link, congestion_node,
                                                        -1)
                                    print('evaluate implementation insertion action', action, 'reward', reward)
                                    if ALNS_greedy_under_unknown_duration_assume_duration != 3 or (
                                            ALNS_greedy_under_unknown_duration_assume_duration == 3 and len(
                                        Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation) >= number_of_training):
                                        Dynamic_ALNS_RL34959.ALNS_insertion_action_list_in_implementation.append(action)
                                        Dynamic_ALNS_RL34959.ALNS_insertion_reward_list_in_implementation.append(reward)
                                    Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation.append(reward)
                                    print('implementation times', len(Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation))
                                    segment_length_in_RL_or_ALNS_implementation = 10
                                    if len(
                                            Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation) % segment_length_in_RL_or_ALNS_implementation == 0:
                                        last_segment = Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation[len(
                                            Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation) - segment_length_in_RL_or_ALNS_implementation: len(
                                            Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation)]
                                        average_reward, std_reward = np.mean(last_segment), np.std(last_segment)
                                        print('last_segment', last_segment, 'average_reward, std_reward',
                                              average_reward, std_reward)

                                        analyze_RL_or_ALNS_implementation_results(1)
                                    save_action_reward_table(segment_length_in_RL_or_ALNS_implementation, Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation)

                                    del action
                    routes = my_deepcopy(current_routes)
                    # also update the time caused by event
                    update_time_caused_by_event(influenced_requests)
                    RL_is_trained_or_evaluated_or_ALNS_is_evaluated = 0
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
        #get the influenced vehicles and requests
    return routes, R_pool, R

def update_time_caused_by_event(influenced_requests, in_get_delay_tolerance = 0):
    for influenced_r in influenced_requests:
        used_k = find_used_k(influenced_r)
        for k in used_k:
            if k == -1:
                break
            bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot,
                                                                                   routes, K, k,
                                                                                   routes[k],
                                                                                   influenced_r)
    if in_get_delay_tolerance == 1:
        return used_k
def analyze_RL_or_ALNS_implementation_results(ALNS=0):
    if ALNS == 1:
        add_the_prefix_ALNS = 'ALNS_'
    else:
        add_the_prefix_ALNS = ''
    insertion_action = 0
    insertion_reward = 0
    removal_action = 0
    removal_reward = 0
    #if ALNS_greedy_under_unknown_duration_assume_duration != 3:
    range_of_implementation = range(
        len(eval('Dynamic_ALNS_RL34959.' + add_the_prefix_ALNS + 'removal_action_list_in_implementation')))

    for action_index in range_of_implementation:
        if eval('Dynamic_ALNS_RL34959.' + add_the_prefix_ALNS + 'removal_action_list_in_implementation')[
            action_index] == 0:
            insertion_action += 1
            insertion_reward += \
                eval('Dynamic_ALNS_RL34959.' + add_the_prefix_ALNS + 'removal_reward_list_in_implementation')[action_index]
        else:
            removal_action += 1
            removal_reward += \
                eval('Dynamic_ALNS_RL34959.' + add_the_prefix_ALNS + 'removal_reward_list_in_implementation')[action_index]
    print('removal operator, wait action', insertion_action, insertion_reward, 'removal action', removal_action, removal_reward)
    removal_operator_removal_action, removal_operator_removal_action_reward, removal_operator_waiting_action, removal_operator_waiting_action_reward = removal_action, removal_reward, insertion_action, insertion_reward

    insertion_action = 0
    insertion_reward = 0
    removal_action = 0
    removal_reward = 0
    #if ALNS_greedy_under_unknown_duration_assume_duration != 3:
    range_of_implementation = range(
        len(eval('Dynamic_ALNS_RL34959.' + add_the_prefix_ALNS + 'insertion_action_list_in_implementation')))
    for action_index in range_of_implementation:
        if eval('Dynamic_ALNS_RL34959.' + add_the_prefix_ALNS + 'insertion_action_list_in_implementation')[
            action_index] == 0:
            insertion_action += 1
            insertion_reward += \
            eval('Dynamic_ALNS_RL34959.' + add_the_prefix_ALNS + 'insertion_reward_list_in_implementation')[action_index]
        else:
            removal_action += 1
            removal_reward += \
            eval('Dynamic_ALNS_RL34959.' + add_the_prefix_ALNS + 'insertion_reward_list_in_implementation')[action_index]
    print('insertion operator, insert action', insertion_action, insertion_reward, 'not insert action', removal_action, removal_reward)
    insertion_operator_insertion_action, insertion_operator_insertion_action_reward, insertion_operator_non_insertion_action, insertion_operator_non_insertion_action_reward = insertion_action, insertion_reward, removal_action, removal_reward
    return removal_operator_removal_action,removal_operator_removal_action_reward,removal_operator_waiting_action,removal_operator_waiting_action_reward,insertion_operator_insertion_action,insertion_operator_insertion_action_reward,insertion_operator_non_insertion_action,insertion_operator_non_insertion_action_reward

def add_one_row_to_T_k_record(index_r):
    global T_k_record
    while True:
        stop_wait()
        if index_r >= len(T_k_record):
            T_k_record = np.vstack([T_k_record, [np.NaN] * T_k_record.shape[1]])
        else:
            break
def add_one_row_to_request_flow(index_r):
    global request_flow_t
    while True:
        stop_wait()
        if index_r >= len(request_flow_t):
            request_flow_t = np.vstack([request_flow_t, [np.NaN] * request_flow_t.shape[1]])
        else:
            break


def get_congestion_nodes_in_this_route(get_route, duration):
    # if there is no congested terminal in route, then the current routes is fine, no need to use RL
    congestion_nodes_in_this_route = []
    if congestion_nodes_at_begining:
        for column_index in range(len(get_route[0])):
            terminal = get_route[0, column_index]
            if terminal in congestion_nodes_at_begining.keys():
                if get_route[1, column_index] >= duration[0]:
                    congestion_nodes_in_this_route.append(terminal)

    use_RL_in_insertion = 1
    if congestion_nodes_in_this_route == []:
        # in this case, RL is not needed
        use_RL_in_insertion = 0
    return congestion_nodes_in_this_route, use_RL_in_insertion

def insert_r_in_learning_or_implementation(index,check_terminal, new_row, r_number, vehicle, uncertainty_index, influenced_r, R_change_dynamic_travel_time2, duration, congestion_link,
                                         congestion_node, implement_or_not2, routes_store=-1, request_flow_t_store=-1, not_use_RL_in_replan_when_event_finishes = 0):
    global R_change_dynamic_travel_time, implement_or_not, ok_K_canpickr, request_flow_t, possible_K, state_reward_pairs, state_reward_pairs_insertion, routes, R_pool, request_flow_t, T_k_record, RL_insertion_implementation_store
    R_change_dynamic_travel_time = R_change_dynamic_travel_time2
    implement_or_not = implement_or_not2
    if add_RL == 1 and implement_or_not != dynamic_RL34959.implement:
        return
    r_number_deep_copy = copy.deepcopy(r_number) #do not know why sometimes r_number with 1 label (T) will be changed to without 1 in it and it will let it not change the column name, so deepcopy it
    # if implement_or_not == 1:
        #store the routes in case the insertion action is not insert and need to try another k
    store_routes_for_another_k = my_deepcopy(routes)
    store_R_pool_for_another_k = copy.copy(R_pool)
    store_request_flow_t_for_another_k = copy.copy(request_flow_t)
    # else:
    #     store_routes_for_another_k = -1
    #     store_R_pool_for_another_k = -1
    # for both implementation and training it is 'finish' because it doesn't matter and I do not know why I store it as 'begin' when implementing RL
    # but when implement it, it should be 'begin' right? because the unexpected envent has not finished
    if not_use_RL_in_replan_when_event_finishes == 0:
        if implement_or_not == 1:
            finish_or_begin = 'begin'
        else:
            finish_or_begin = 'finish'
            # if dynamic_RL34959.implement == 1:
            #     print('remove it when run it in server')
        new_row['uncertainty_type'] = finish_or_begin
    # new_row['action'] = -10000000
    # new_row['reward'] = -10000000
    # here I need to let the ALNS finds the good positions
    # and then ALNS inserts it directly, and then the RL evaluate it should be removed or not,
    # if should remove, then the insertion is wrong, so try next one; if not remove, then the insertion is good, stop here
    # greedy insertion, and possible for two vehicles? -> then RL needs to evaluate two routes, also seems fine
    # try one vehicle first

    # send state to RL insertion operator
    # add state in state_reward_pairs_insertion first
    # the r or segment's pickup and delivery terminal
    # new_row['passed_terminals'] = [p, d]

    # check which vehicles are suitable to transport this vehicle/segment
    ok_K_canpickr = func_ok_K_canpickr()
    # undate K_R for r. if r is a segment, r_number could be different with influenced_r
    get_K_R_unit2(r_number)
    possible_K = K_R['1k'][r_number]

    if r_number % big_r >= 10000:
        #then it has T, then only insert to one another vehicle
        # check capacity
        index_r = list(R[:, 7]).index(r_number)
        load = R[index_r, 6]
        # add the new segment to request_flow_t
        add_one_row_to_request_flow(index_r)
        # sometimes it cannot find a better solution, so I keep the original vehicle and comment the following codes
        # if vehicle in possible_K:
        #     possible_K.remove(vehicle)
        # here the checking assume the load is always added on the load, i.e., when actual max load | capacity, return false, don't consider the load is unloaded
        for k in possible_K:
            if capacity_constraints(has_end_depot, K, R, k, routes[k], load, 0) == False:
                possible_K.remove(k)
        if possible_K == []:
            return
        # rank k by unit distance cost
        for possible_k_index in range(len(possible_K)):
            possible_K[possible_k_index] = [possible_K[possible_k_index]]
        possible_k_with_unit_cost = np.array(possible_K)
        possible_k_with_unit_cost = np.append(possible_k_with_unit_cost, possible_k_with_unit_cost, axis=1)


        for k in possible_K:
            k = k[0]
            # just rank by the unit cost
            # if K[]
            # for k_in_K_index in range(len(K)):
            #     k_in_K = K[k_in_K_index,0]
            #     if k_in_K == k:
            possible_k_with_unit_cost[0, 1] = K[k, 2]
        possible_k_with_unit_cost = possible_k_with_unit_cost[np.argsort(possible_k_with_unit_cost[:, 1])]
        R_i = tuple(zip(R[index_r], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
        break_flag4 = 0
        for k in possible_k_with_unit_cost[:, 0]:
            if k != possible_k_with_unit_cost[0, 0]:
                routes = my_deepcopy(store_routes_for_another_k)
                R_pool = copy.copy(store_R_pool_for_another_k)
                request_flow_t = copy.copy(store_request_flow_t_for_another_k)
            r_number = r_number_deep_copy
            Trans, Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd = 0, 0, 0, 0, 0
            obj_1_vehicle = best_position_1_vehicle(R, no_route_barge, no_route_truck,
                                                    hash_table_1v_all_fail,
                                                    hash_table_1v_all, routes,
                                                    fixed_vehicles_percentage,
                                                    Fixed, K,
                                                    hash_table_1v, hash_table_1v_fail, has_end_depot,
                                                    R_i,
                                                    r_number, k,
                                                    Trans, Trans_Tp, Trans_Td, Trans_secondTp,
                                                    Trans_secondTd,
                                                    0)
            if obj_1_vehicle:
                best_k, original_route, original_route_no_columns, cost_inserted_request, dict_a_request_best_position = \
                    obj_1_vehicle[0]
                key = get_key_1k(R_i, original_route_no_columns, k, fixed_vehicles_percentage, Fixed, K)
                get_route = hash_table_1v_all[key][dict_a_request_best_position]['route']
            else:
                continue

            congestion_nodes_in_this_route, use_RL_in_insertion = get_congestion_nodes_in_this_route(get_route, duration)

            if use_RL_in_insertion == 0 and implement_or_not == 0 and not_use_RL_in_replan_when_event_finishes == 0:
                routes = my_deepcopy(routes_store)
                request_flow_t = copy.copy(request_flow_t_store)
                break
            if combine_insertion_and_removal_operators == 0:
                # new_cost_of_r_when_uncertainty_finishes = get_r_cost_in_all_routes(influenced_r)[0]
                # here I should insert it to the vehicle
                # r_number % 10000 is the r's real number
                check_the_case = (r_number - r_number % 10000) % big_r
                if check_the_case == 10000:
                    # then it's the normal case: the insert part is the Tp, delivery
                    Trans, Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd = 1, 1, 0, 0, 0
                elif check_the_case == 20000:
                    # then case 5: the insert part is the Tp+secondTd
                    Trans, Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd = 1, 0, 0, 0, 1
                elif check_the_case == 30000:
                    # then case 3: the insert part is the secondTp+delivery
                    Trans, Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd = 1, 0, 0, 1, 0
                else:
                    # check_the_case == 0, and no transshipment
                    Trans, Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd = 0, 0, 0, 0, 0

                routes[k] = copy.copy(get_route)
                request_list2 = list(original_route[4])
                request_list2 = get_request_list2(request_list2, r_number - check_the_case,
                                                  Trans, Trans_Tp, Trans_Td,
                                                  Trans_secondTp, Trans_secondTd,
                                                  dict_a_request_best_position)
                #caught_strange(request_list2)
                routes[k][4] = copy.copy(request_list2)
                R_pool = R_pool[~(R_pool[:, 7] == r_number)]
                if use_RL_in_insertion == 0 or not_use_RL_in_replan_when_event_finishes == 1:
                    break
            # the r's number without labeling of T
            r_number = r_number - (r_number - r_number % 10000) % big_r
            if not_use_RL_in_replan_when_event_finishes == 0:
                if use_RL_in_insertion == 1 and K[k, 5] == influenced_mode_by_current_event:
                    # if dynamic_RL34959.implement == 1 and finish_or_begin == 'finish':
                    #     print('wrong wrong double pairs')
                    break_flag4, feasibility = check_uncertainty_in_insertion_by_RL(implement_or_not, congestion_nodes_in_this_route, copy.deepcopy(routes[k]), r_number, index_r, new_row, finish_or_begin, uncertainty_index, vehicle, influenced_r, store_routes_for_another_k, store_R_pool_for_another_k,duration, congestion_link,
                                         congestion_node, index, routes_store, request_flow_t_store, k)
                    if break_flag4 == 1:
                        if implement_or_not == 0:  # when implement_or_not, the restore when break_flag4 == 1 is already cosidered in check_uncertainty_in_insertion_by_RL
                            routes = my_deepcopy(store_routes_for_another_k)
                            R_pool = copy.copy(store_R_pool_for_another_k)
                            request_flow_t = copy.copy(store_request_flow_t_for_another_k)
                        break
    else:
        #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
        #then it does not has T, so I need to consider both 1k and 2k
        if not_use_RL_in_replan_when_event_finishes == 0:
            routes, R_pool = greedy_insert(r_number, 0, 1, new_row, finish_or_begin, uncertainty_index, store_routes_for_another_k, store_R_pool_for_another_k, duration, congestion_link,
            congestion_node,index)
        else:
            routes, R_pool = greedy_insert(r_number)
        #check_repeat_r_in_R_pool(), check_T_k_record_and_R()

        #for 1k

        # possible_K_2k = K_R['2k'][r_number]


def caught_strange(request_list2):
    check_number_of_r_in_one_route = {}
    for r_label in request_list2[1:-1]:
        if '.' in r_label:
            print('caught dot in r label')
        r_num = get_numbers(r_label)
        if r_num in check_number_of_r_in_one_route.keys():
            check_number_of_r_in_one_route[r_num] += 1
        else:
            check_number_of_r_in_one_route[r_num] = 1
    for r_num in check_number_of_r_in_one_route.keys():
        if check_number_of_r_in_one_route[r_num] % 2 > 0:
            print('caught the strange insertion')

def check_RL_ALNS_iteraction_bug():
    if dynamic_RL34959.implement == 1 and ALNS_implement_start_RL_can_move == 1 and len(state_reward_pairs) == 0:
        print('gfsfsfagsgfd')
        print('gfsfsfagsgfd')
#@profile()
def check_uncertainty_in_insertion_by_RL(implement_or_not, congestion_nodes_in_this_route, get_route, r_number, index_r, new_row, finish_or_begin, uncertainty_index, vehicle, influenced_r, store_routes_for_another_k, store_R_pool_for_another_k,duration, congestion_link,
                                     congestion_node, index, routes_store, request_flow_t_store, k):
    global used_interrupt, ALNS_implement_start_RL_can_move, R, state_reward_pairs, request_flow_t, routes, R_pool, interrupt_by_implement_is_one_and_assign_action_once_only
    # if len(state_reward_pairs) == 1 and dynamic_RL34959.implement == 1:
    #     print('i should check this wrong')
    # if dynamic_RL34959.implement == 1 and finish_or_begin == 'finish':
    #     print('wrong wrong double pairs')
    # if dynamic_RL34959.implement == 1 and ALNS_implement_start_RL_can_move == 1:
    #     print('ALNS_implement_start_RL_can_move', ALNS_implement_start_RL_can_move)
    #     print(ALNS_implement_start_RL_can_move,'wrong...')
    #check_RL_ALNS_iteraction_bug()
    index_r = list(R[:, 7]).index(r_number)
    feasibility = -1
    break_flag4 = 0
    if congestion_nodes_in_this_route == []:
        #there is no congestion node, so no need to use RL and the route is feasible
        return 1, 1
    #in insertion, the check_terminal may not the congested terminal in the influnced route, because the route will be changed, so need to assign the new congested terminal in this route to check_terminal
    check_terminal = congestion_nodes_in_this_route[0]
    first_index = -1
    second_index = -1
    for coloumn_index in range(len(get_route[4]))[1:-1]:
        r_number_in_route = get_numbers(get_route[4][coloumn_index])
        real_r_number_in_route = r_number_in_route - (r_number_in_route - r_number_in_route % 10000) % big_r  # in case it has the T label in r_number_in_route
        if real_r_number_in_route == r_number:
            if first_index == -1:
                first_index = coloumn_index
            else:
                second_index = coloumn_index
                break
    new_row['passed_terminals'] = unique([int(x) for x in list(get_route[0][first_index:])])
    # in insertion operator, delay_tolerance is changed according to k, and may be different from the removal operator
    # new_row['delay_tolerance'] = R[index_r, 5] - get_route[2, second_index]
    # get delay tolerance of the influenced request = pickup_time (or Tp) - event_begin_time + (latest delivery time - delivery_time)
    add_one_row_to_request_flow(index_r)
    # if dynamic_RL34959.implement == 1 and ALNS_implement_start_RL_can_move == 1:
    #     print('wrong...')
    # if np.isnan(request_flow_t[index_r, 5]):
    for k_ in find_used_k(r_number):
        if k_ == -1:
            break
        time_constraints(k_, routes[k_], r_number)
    # if len(state_reward_pairs) == 1 and dynamic_RL34959.implement == 1:
    #     print('i should check this wrong')
    # if dynamic_RL34959.implement == 1 and ALNS_implement_start_RL_can_move == 1:
    #     print('wrong...')
    #check_RL_ALNS_iteraction_bug()
    new_row['delay_tolerance'] = get_delay_tolerance(copy.deepcopy(routes[k]), first_index, duration, index_r, r_number, k)
    new_row['action'] = -10000000
    new_row['reward'] = -10000000
    new_row['vehicle'] = vehicle
    # dynamic_routes_store_insertion[(uncertainty_index, influenced_r)] = store_all(1, new_row)
    # here need to check whether this k is suitable under the uncertainty
    # so here consider the state of insertion operator
    # insert_r_in_learning_or_implementation() function is used in both learning and implementation, so I have changed the new_row['uncertainty_type'] at the begining of this function
        # and it gurantee that the type under training is finish and under implementation is begin
    if dynamic_RL34959.implement == 1:
        while True:
            # if dynamic_RL34959.implement == 1 and ALNS_implement_start_RL_can_move == 1:
            #     print('wrong...')
            if implement_or_not != dynamic_RL34959.implement:
                return -1, -1
            # print('check_uncertainty_in_insertion_by_RL')
            # if dynamic_RL34959.ALNS_got_action_in_implementation == 0:
            #     #then it is the first action to take in the implementation phase, no need to wait the clear of state_reward_pairs
            #     break
            if dynamic_RL34959.clear_pairs_done == 1:
                # if len(state_reward_pairs) != 0:
                #     print('i doubt it should has clear pairs but not, and the done still is 1 so it is wrong')
                break
    # if len(state_reward_pairs) == 1 and dynamic_RL34959.implement == 1:
    #     print('i should check this wrong')
    #check_RL_ALNS_iteraction_bug()
    # if dynamic_RL34959.implement == 1 and dynamic_RL34959.clear_pairs_done == 0:
    #     print('1')
    # if dynamic_RL34959.implement == 1 and dynamic_RL34959.ALNS_got_action_in_implementation == 1:
    #     print('here must be wrong')
    # if dynamic_RL34959.implement == 1 and ALNS_implement_start_RL_can_move == 1:
    #     print('wrong...')
    # if dynamic_RL34959.time_s == 22:
    #     print('gf')
    while True:
        # if len(state_reward_pairs) == 1 and dynamic_RL34959.implement == 1:
        #     print('i should check this wrong')
        # if dynamic_RL34959.implement == 1 and dynamic_RL34959.ALNS_got_action_in_implementation == 1:
        #     print('wrong it should not')
        # if dynamic_RL34959.implement == 1 and ALNS_implement_start_RL_can_move == 1:
        #     print('wrong...')
        stop_wait()
        if implement_or_not != dynamic_RL34959.implement:
            return -1, -1
        # print('check_uncertainty_in_insertion_by_RL 2')
        if dynamic_RL34959.clear_pairs_done == 1 or dynamic_RL34959.implement == 0:
            while True:
                # if len(state_reward_pairs) == 1 and dynamic_RL34959.implement == 1:
                #     print('i should check this wrong')
                # if dynamic_RL34959.implement == 1 and dynamic_RL34959.ALNS_got_action_in_implementation == 1:
                #     print('wrong it should not')
                # if dynamic_RL34959.implement == 1 and ALNS_implement_start_RL_can_move == 1:
                #     print('wrong...')
                # if dynamic_RL34959.implement == 1 and dynamic_RL34959.ALNS_got_action_in_implementation == 1:
                #     print('here must be wrong')
                # if dynamic_RL34959.implement == 1 and dynamic_RL34959.clear_pairs_done == 0:
                #     print('1')
                # print('check_uncertainty_in_insertion_by_RL 3')
                # if dynamic_RL34959.implement == 1:
                #     print('sdf')
                stop_wait()
                try:
                    # if len(state_reward_pairs) == 1 and dynamic_RL34959.implement == 1:
                    #     print('i should check this wrong')
                    try:
                        state_reward_pairs = pd.DataFrame(columns=state_reward_pairs.columns)
                        state_reward_pairs = state_reward_pairs.append(new_row, ignore_index=True)

                    except:
                        if len(state_reward_pairs) == 0:
                            state_reward_pairs = pd.DataFrame(
                                columns=['uncertainty_index', 'uncertainty_type', 'request', 'vehicle',
                                         'delay_tolerance',
                                         'passed_terminals',
                                         'current_time', 'action', 'reward'])
                            state_reward_pairs = state_reward_pairs.append(new_row, ignore_index=True)

                        else:
                            print('state_reward_pairs = state_reward_pairs.append(new_row, ignore_index=True)')
                    if len(state_reward_pairs) > 0:
                        # print('print the state reward pairs here', state_reward_pairs)
                        # if dynamic_RL34959.implement == 1 and ALNS_implement_start_RL_can_move == 1:
                        #     print('wrong...')
                        # if used_interrupt == 1:
                        #     print('cao!!!')
                        # 
                        # if dynamic_RL34959.implement == 1 and ALNS_implement_start_RL_can_move == 1:
                        #     print('wrong...')
                        # if dynamic_RL34959.implement == 1 and dynamic_RL34959.ALNS_got_action_in_implementation == 1:
                        #     print('wrong it should not')
                        # if len(state_reward_pairs) in [0, 2] and dynamic_RL34959.implement == 1:
                        #     print('here is wrong!!!')
                        
                        break
                except:
                    print('state_reward_pairs = state_reward_pairs.append(new_row, ignore_index=True)')
            break
    # if dynamic_RL34959.implement == 1 and dynamic_RL34959.ALNS_got_action_in_implementation == 1:
    #     print('wrong it should not')
    # if len(state_reward_pairs) in [0,2] and dynamic_RL34959.implement == 1:
    #     print('here is wrong!!!')
    # if len(state_reward_pairs) > 1:
    #     print('len > 1')
    # when combine_insertion_and_removal_operators = 0, the state is inserted route, no matter what action it is
    #check_RL_ALNS_iteraction_bug()
    break_flag2 = 0
    while True:
        # if dynamic_RL34959.implement == 1 and dynamic_RL34959.ALNS_got_action_in_implementation == 1:
        #     print('wrong it should not')
        if implement_or_not != dynamic_RL34959.implement:
            return -1, -1
        #check_RL_ALNS_iteraction_bug()
        if dynamic_RL34959.implement == 1 and ALNS_implement_start_RL_can_move == 0:
            ALNS_implement_start_RL_can_move = 1
        # if len(state_reward_pairs) in [0,2] and dynamic_RL34959.implement == 1:
        #     print('here is wrong!!!')
        # print('check_uncertainty_in_insertion_by_RL 4')
        # if dynamic_RL34959.implement == 1:
        #     print('sdf')
        if interrupt_by_implement_is_one_and_assign_action_once_only == 0:
            if dynamic_RL34959.implement == 1:
                interrupt_by_implement_is_one_and_assign_action_once_only = 1
                # used_interrupt = 1
                return -1, -1
        stop_wait()
        for pair_index in state_reward_pairs.index:

            if pair_index not in state_reward_pairs.index:
                continue


            try:
                # state_reward_pairs['uncertainty_type'][pair_index] == finish_or_begin
                check = state_reward_pairs['uncertainty_type'][pair_index] == finish_or_begin and state_reward_pairs['uncertainty_index'][pair_index] == uncertainty_index and \
                state_reward_pairs['vehicle'][
                    pair_index] == vehicle and state_reward_pairs['request'][pair_index] == influenced_r and \
                state_reward_pairs.loc[pair_index]['action'] != -10000000 and state_reward_pairs.loc[pair_index]['reward'] == -10000000
            except:
                break
            if check:
                if implement_or_not != dynamic_RL34959.implement:
                    return -1, -1
                if implement_or_not == 1:
                    # if len(state_reward_pairs) in [0,2] and dynamic_RL34959.implement == 1:
                    #     print('here is wrong!!!')
                    action_insertion = state_reward_pairs.loc[pair_index]['action']
                    dynamic_RL34959.ALNS_got_action_in_implementation = 1
                    # store the routes, new_row, action for evaluate the performance of RL during implementation
                    if (uncertainty_index, influenced_r) not in RL_insertion_implementation_store.keys():
                        RL_insertion_implementation_store[(uncertainty_index, influenced_r)] = {}
                    try:
                        RL_insertion_implementation_store[(uncertainty_index, influenced_r)][k] = store_all(2, state_reward_pairs.loc[pair_index], action_insertion)
                    except:
                        print('RL_insertion_implementation_store[(uncertainty_index, influenced_r)][k] = store_all(2, state_reward_pairs.loc[pair_index], action_insertion)')
                    if action_insertion == 1:
                        #it means the insertion should be removed
                        #then restore the route to the one that the request is only removed
                        #if it's full request, then it is evaluated in greedy_insert as a constraints, so need to return a bool that feasible or not
                        if RL_insertion_segment == 1:
                            routes = my_deepcopy(store_routes_for_another_k)
                            R_pool = copy.copy(store_R_pool_for_another_k)
                        else:
                            #it's removal, so infeasible
                            feasibility = 0
                        # if len(state_reward_pairs) in [0,2] and dynamic_RL34959.implement == 1:
                        #     print('here is wrong!!!')
                    else:
                        #if insertion no problem, the loop for possible ks can stop at this k
                        feasibility = 1
                        break_flag4 = 1


# insert
                                    # if action_insertion == 0:
                                        #when combine_insertion_and_removal_operators = 0, the action_insertion == 0 means not removal, i.e., insert, in the application when RL is mature

                                        #then do nothing, and RL can't get reward from it?
                                         #no, Rl should learn from it that I don't insert it, is it good or not

                if implement_or_not != dynamic_RL34959.implement:
                    return -1, -1
                if implement_or_not == 0:
                    # if len(state_reward_pairs) in [0,2] and dynamic_RL34959.implement == 1:
                    #     print('here is wrong!!!')
                    get_and_send_rewards(uncertainty_index, R_change_dynamic_travel_time, check_terminal, influenced_r, vehicle, index, duration, congestion_link,
                                     congestion_node)
                    # if len(state_reward_pairs) in [0,2] and dynamic_RL34959.implement == 1:
                    #     print('here is wrong!!!')
                    # break_flag2 = 1
                    # break
                    # else:
                    #
                    #
                    #     get_and_send_rewards(uncertainty_index, R_change_dynamic_travel_time, check_terminal, influenced_r, vehicle, index, duration, congestion_link,
                    #                          congestion_node)
                        #then reset the routes, under this insertion and removal are evaluated by the same operator, the insertion will be repeat, so need to reset routes and also R, request flow etc
                        #but the original routes should before the removal, or not?
                        #should the the one after the removal
                if implement_or_not == 0:
                    routes = my_deepcopy(routes_store)
                    request_flow_t = copy.copy(request_flow_t_store)
                #here should be two flags, the break_flag2 is for the currrent k, which means the current k's action is found
                #another flag break_flag4 means action = 1, and no other k needs to be evaluated
                break_flag2 = 1
                # if combine_insertion_and_removal_operators == 1:


                break
        #     if len(state_reward_pairs) in [0,2] and dynamic_RL34959.implement == 1:
        #         print('here is wrong!!!')
        # if len(state_reward_pairs) in [0,2] and dynamic_RL34959.implement == 1:
        #     print('here is wrong!!!')
        if break_flag2 == 1:
            break


    return break_flag4, feasibility
#@profile()
def send_to_RL_and_train(repeat_scene, influenced_r, check_terminal, R_change_dynamic_travel_time, index, duration, vehicle_stop_time, uncertainty_index, new_cost_of_r_when_uncertainty_finishes,congestion_link,congestion_node,state_when_uncertainty_begins,routes_store,request_flow_t_store):
    global used_interrupt, ALNS_implement_start_RL_can_move, interrupt_by_implement_is_one_and_assign_action_once_only, request_flow_t, ok_K_canpickr, possible_K, state_reward_pairs, state_reward_pairs_insertion, routes, R_pool, request_flow_t, T_k_record
    # if repeat_scene > 0:
    #     #when RL finish the drop, begins a new repeat
    #     while True:
    #         if dynamic_RL34959.RL_drop_finish == 1:
    #             dynamic_RL34959.RL_drop_finish = 0
    #             break
    # get original routes of that time that r

    find_one_in_collect = 0
    #############################################
    # if np.size(dynamic_RL34959.state_action_reward_collect) > 0:
    #     if 'index_collect' in locals():
    #         index_collect += 1
    #     else:
    #         index_collect = 0
    #     length_collect = len(dynamic_RL34959.state_action_reward_collect)
    #     if index_collect == length_collect:
    #         index_collect = 0
    #     for find_suitable_index in range(length_collect):
    #         if dynamic_RL34959.state_action_reward_collect[index_collect,1] != 'begin':
    #             new_row = pd.Series(list(dynamic_RL34959.state_action_reward_collect[index_collect]), index = state_reward_pairs.columns)
    #             new_row[7] = -10000000
    #             new_row[8] = -10000000
    #             find_one_in_collect = 1
    #             break
    #         else:
    #             index_collect += 1
#############################################################

            # new_row['uncertainty_type'] = 'finish'
            #
    if find_one_in_collect == 0:

        original_cost_of_r_when_uncertainty_finishes = get_r_cost_in_all_routes(influenced_r)[0]
        new_row = my_deepcopy(state_when_uncertainty_begins)
        new_row['uncertainty_type'] = 'finish'
        # if dynamic_RL34959.implement == 1:
        #     print('remove it when run it in server')
        vehicle = new_row['vehicle']
        if after_action_review == 1:
            new_row['action'] = -10000000

    while True:
        if dynamic_RL34959.stop_everything_in_learning_and_go_to_implementation_phase == 1:
            break
        # print('send_to_RL_and_train 1')
        if dynamic_RL34959.implement == 1:
            while True:
                # print('send_to_RL_and_train 2')
                if dynamic_RL34959.clear_pairs_done == 1:
                    break
        try:
            state_reward_pairs = pd.DataFrame(columns=state_reward_pairs.columns)
            state_reward_pairs = state_reward_pairs.append(new_row, ignore_index=True)
            if len(state_reward_pairs) > 0:
                break
        except:
            if 'state_reward_pairs2' not in locals():
                state_reward_pairs2 = copy.deepcopy(state_reward_pairs)
            state_reward_pairs = pd.DataFrame(columns=state_reward_pairs2.columns)
            for ind_ in state_reward_pairs2.index:
                state_reward_pairs = state_reward_pairs.append(state_reward_pairs2.loc[ind_], ignore_index=True)
    if dynamic_RL34959.implement == 1:
        ALNS_implement_start_RL_can_move = 1
    # if len(state_reward_pairs) in [0,2] and dynamic_RL34959.implement == 1:
    #     print('here is wrong!!!')
    # if len(state_reward_pairs) > 1:
    #     print('len > 1')
    if after_action_review == 1:
        #get removal and insertion action from RL
        break_flag3 = 0
        while True:
            # print('send_to_RL_and_train 3')
            if interrupt_by_implement_is_one_and_assign_action_once_only == 0:
                if dynamic_RL34959.implement == 1:
                    interrupt_by_implement_is_one_and_assign_action_once_only = 1
                    # used_interrupt = 1
                    action = 0
                    # get_and_send_rewards(uncertainty_index, R_change_dynamic_travel_time, check_terminal, influenced_r, vehicle, index, duration, congestion_link, congestion_node)
                    break
            stop_wait()
            for pair_index in state_reward_pairs.index:

                if pair_index not in state_reward_pairs.index:
                    continue

                # if pair_index not in state_reward_pairs.index:
                #     continue
                # try:
                #     state_reward_pairs['uncertainty_index'][pair_index] == uncertainty_index and \
                #     state_reward_pairs['uncertainty_type'][pair_index] == 'finish' and state_reward_pairs['vehicle'][
                #         pair_index] == vehicle and state_reward_pairs['request'][pair_index] == influenced_r and \
                #     state_reward_pairs.loc[pair_index]['action'] != -10000000 and state_reward_pairs['reward'][
                #         pair_index] == -10000000
                # except:
                #     print('sfa')
                # try:
                # print('ALNS', state_reward_pairs)
                try:
                    check = state_reward_pairs['uncertainty_index'][pair_index] == uncertainty_index and \
                    state_reward_pairs['uncertainty_type'][pair_index] == 'finish' and state_reward_pairs['vehicle'][
                        pair_index] == vehicle and state_reward_pairs['request'][pair_index] == influenced_r and \
                    state_reward_pairs.loc[pair_index]['action'] != -10000000 and state_reward_pairs.loc[pair_index]['reward'] == -10000000
                except:
                    break
                if check:
                    action = state_reward_pairs.loc[pair_index]['action']
                    dynamic_RL34959.ALNS_got_action_in_implementation = 1
                    break_flag3 = 1
                    #take the new action and get new_cost_of_r_when_uncertainty_finishes
                    if action == 1:


                        if combine_insertion_and_removal_operators == 0:
                            get_and_send_rewards(uncertainty_index, R_change_dynamic_travel_time, check_terminal, influenced_r, vehicle, index, duration, congestion_link, congestion_node)
                        # remove
                        # apply_RL_to_make_decisions = 0
                        # if apply_RL_to_make_decisions == 1:

                        #remove for the insertion operator
                        p, d, ap, bd, r_number = take_action_to_remove(vehicle, R_change_dynamic_travel_time, index, 1,
                                               influenced_r, congestion_link, congestion_node,
                                               vehicle_stop_time)
                        if r_number != -1:
                            insert_r_in_learning_or_implementation(index, check_terminal, new_row, r_number, vehicle, uncertainty_index, influenced_r, R_change_dynamic_travel_time, duration, congestion_link,
                                             congestion_node,dynamic_RL34959.implement,routes_store,request_flow_t_store)


                    else:
                        # get_and_send_rewards(uncertainty_index, R_change_dynamic_travel_time, check_terminal, influenced_r, vehicle, index, duration, congestion_link, congestion_node)
                        if combine_insertion_and_removal_operators == 0:
                            get_and_send_rewards(uncertainty_index, R_change_dynamic_travel_time, check_terminal, influenced_r, vehicle, index, duration, congestion_link, congestion_node)
                if break_flag3 == 1:
                    break
            if break_flag3 == 1:
                break

    if combine_insertion_and_removal_operators == 1:
        get_and_send_rewards(uncertainty_index, R_change_dynamic_travel_time, check_terminal, influenced_r, k, index, duration, congestion_link, congestion_node)
    return action
#@profile()
def get_reward(check_terminal, action, influenced_r, R_change_dynamic_travel_time, index, duration, congestion_link, congestion_node, original_cost_of_r_when_uncertainty_finishes):
    global already_add_once_for_re_plan_when_event_finishes, at_get_reward, re_plan_when_event_finishes, dynamic_time_false, routes, re_plan_when_event_finishes_information
    if add_RL == 1 and dynamic_RL34959.implement == 0:
        routes_not_change = my_deepcopy(routes)
    used_k = find_used_k(influenced_r)
    # here I should check the influenced k of r, if congestion happens at T, then two vehicles are influenced, so both vehicle need to be updated
    # so I need to check all used_k, and update all times of these k, then calculate the cost
    reward = 0
    if used_k == [-1, -1, -1]:
        print('error')
    dynamic_time_false = 0
    at_get_reward = 1
    re_plan_when_event_finishes = 0
    for k in used_k:
        if k == -1:
            break
        not_pass_congested_terminal = 1
        for index2, element in enumerate(routes[k][0]):
            if index2 == 0 or index2 == (len(routes[k][0]) - 1):
                continue
            if element == check_terminal and get_numbers(routes[k][4,index2]) == influenced_r:
                col_congestion = index2
                not_pass_congested_terminal = 0
                break
        already_add_once_for_re_plan_when_event_finishes = 0
        if not_pass_congested_terminal == 1:
            bool_or_route, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k, my_deepcopy(
                                                                                            routes[k]), influenced_r)
        else:
            # change time according to the real duration
            # here I should obtain the cost of orinal cost under the delayed time

            vehicle_stop_time = get_vehicle_stop_time(R_change_dynamic_travel_time, index, routes[k],
                                                      col_congestion, duration)
            bool_or_route, infeasible_request_terminal = add_duration_and_check_feasibility(index,
                                                                                            R_change_dynamic_travel_time,
                                                                                            duration,
                                                                                            my_deepcopy(
                                                                                                routes[k]),
                                                                                            col_congestion,
                                                                                            vehicle_stop_time,
                                                                                            k,
                                                                                            influenced_r,
                                                                                            congestion_link,
                                                                                            congestion_node, 1)

        if not isinstance(bool_or_route, bool):

            routes[k] = my_deepcopy(bool_or_route)
            relevant_try_copy = my_deepcopy(relevant_try)
            layer, aaa = 0, 0
            final_ok_or = solve_relevant_try(relevant_try_copy, layer, aaa)
        else:
            #then I need to check whether I need to do the replan immediately
            #first check
            if re_plan_when_event_finishes == 1:
                #then remove this r/segment
                #first update the request_flow_t_and_T_k_record
                re_plan_index = 0
                while re_plan_when_event_finishes == 1:
                    stop_wait()
                    re_plan_when_event_finishes = 0
                    k, influenced_r, vehicle_stop_time, T1, Td_time, _ = re_plan_when_event_finishes_information.loc[re_plan_index]
                    k, influenced_r, T1 = int(k), int(influenced_r), int(T1)
                    used_k = find_used_k(influenced_r)
                    for k_ in used_k:
                        if k_ == -1:
                            break
                        update_request_flow_t_and_T_k_record(k_)
                    index_inluenced_r = list(R[:, 7]).index(influenced_r)
                    if np.isnan(T_k_record[index_inluenced_r, 0]):
                        T1 = -1
                    else:
                        T1 = T_k_record[index_inluenced_r, 0]
                    if np.isnan(request_flow_t[index_inluenced_r,1]):
                        Td_time = -1
                    else:
                        Td_time = request_flow_t[index_inluenced_r,1]
                    for index2, element in enumerate(routes[k][0]):
                        if index2 == 0 or index2 == (len(routes[k][0]) - 1):
                            continue
                        if element == congestion_node and get_numbers(routes[k][4, index2]) == influenced_r:
                            col_congestion = index2
                            # not_pass_congested_terminal = 0
                            break
                    vehicle_stop_time = get_vehicle_stop_time(R_change_dynamic_travel_time, index,
                                                              routes[k], col_congestion, duration)
                    p, d, ap, bd, r_number = take_action_to_remove(k, R_change_dynamic_travel_time, index, 1,
                                                                   influenced_r, congestion_link, congestion_node,
                                                                   vehicle_stop_time, T1, Td_time)
                    if r_number != -1:
                        re_plan_when_event_finishes_information['r_number'][re_plan_index] = r_number
                        already_add_once_for_re_plan_when_event_finishes = 0
                        if K[k, 5] != 3 and len(routes[k][0]) > 2:
                            bool_or_route2, infeasible_request_terminal = time_constraints_relevant(has_end_depot, routes, K, k,
                                                                                                   my_deepcopy(
                                                                                                       routes[k]), influenced_r)
                        re_plan_index += 1
                    # routes_store, request_flow_t_store = my_deepcopy(routes), copy.copy(request_flow_t)
                re_plan_when_event_finishes = 1
                if add_RL == 1:
                    implement_RL = dynamic_RL34959.implement
                else:
                    implement_RL = 0
                for re_plan_index in re_plan_when_event_finishes_information.index:
                    k, influenced_r, vehicle_stop_time, T1, Td_time, r_number = re_plan_when_event_finishes_information.loc[re_plan_index]
                    if r_number == 'wait_for_input' or r_number == -1:
                        continue
                    if r_number / big_r > 10:
                        print('caught wrong r number')
                    k, influenced_r, T1, r_number = int(k), int(influenced_r), int(T1), int(r_number)
                    insert_r_in_learning_or_implementation(index, check_terminal, -1, r_number, k,
                                                           -1, influenced_r, R_change_dynamic_travel_time,
                                                           duration, congestion_link,
                                                           congestion_node, implement_RL, -1, -1, 1)
                re_plan_when_event_finishes_information = pd.DataFrame(columns= ['k', 'request_number', 'vehicle_stop_time', 'T1', 'Td_time', 'r_number'])
                break

        # if used_k[1] != -1 and (routes[used_k[0]][2,1] - routes[used_k[0]][1,1]) > 2:
        #     print('sdfsafg')
    #here no matter what happens, just let the routes finishing updating, even though there are infeasible things, just let it be because I have not add replaning for r that still is influenced but can be changeable
    if dynamic_time_false == 1:
        bool_or_route = False
    if 'bool_or_route' not in locals():
        print('error')
    if isinstance(bool_or_route, bool):
        if get_reward_by_cost_gap == 1:
            reward = -999999

    if get_reward_by_cost_gap == 1:
        if reward != -999999:
            if ALNS_guides_RL == 0:
                new_cost_of_r_when_uncertainty_finishes = get_r_cost_in_all_routes(influenced_r)[0]
                reward = original_cost_of_r_when_uncertainty_finishes - new_cost_of_r_when_uncertainty_finishes
            else:
                pass
    else:


        if (action == 1 and isinstance(bool_or_route, bool)) or (
                action == 0 and not isinstance(bool_or_route, bool)):
            reward = 1
        else:
            reward =0 
        if add_RL == 1 and dynamic_RL34959.implement == 1 and R_change_dynamic_travel_time['type'][index] == 'congestion_finish':
            pass
        else:
            if add_RL == 1 and ((dynamic_RL34959.severity_level == action and reward == 0) or (dynamic_RL34959.severity_level != action and reward == 1)):
                print('wrong label', 'table number', Dynamic_ALNS_RL34959.table_number)
            else:
                print('correct label')
        print(bool_or_route, 'action', action, 'reward', reward)
    if add_RL == 1 and dynamic_RL34959.implement == 0:
        routes = my_deepcopy(routes_not_change)
    re_plan_when_event_finishes = 0
    dynamic_time_false = 0
    at_get_reward = 0

    return reward
#@profile()
def get_and_send_rewards(uncertainty_index, R_change_dynamic_travel_time, check_terminal, influenced_r, vehicle, index, duration, congestion_link, congestion_node):
    #note: if I use the combination of insertion and removal operators, here the insertion operator is not rewarded, because I haven't add these codes
    # and the state_reward_pairs_insertion should be added if I use the combination of insertion and removal operators,

   #don't know why here sometimes (after a random number of time steps, usually several hundreds, sometimes serveral thousands) can't get the reward and then going on, just like this loop is be skipped, so set while, and then only thy one more time, it will be fine
    break_flag = 0
    while True:
        if dynamic_RL34959.stop_everything_in_learning_and_go_to_implementation_phase == 1:
            break
        # print('get_and_send_rewards 1')
        stop_wait()
        for pair_index in state_reward_pairs.index:

            if pair_index not in state_reward_pairs.index:
                continue
            try:
                check = state_reward_pairs['uncertainty_index'][pair_index] == uncertainty_index and \
                    state_reward_pairs['uncertainty_type'][pair_index] == 'finish' and state_reward_pairs['vehicle'][
                pair_index] == vehicle and state_reward_pairs['request'][pair_index] == influenced_r and \
                    state_reward_pairs.loc[pair_index]['action'] != -10000000
            except:
                break
            if check:
                if get_reward_by_cost_gap == 0:
                    # here has a problem, this action is for this uncertainty, although this uncertainty is uncertainty_finishes, but the feasibility is for previous action, so it's wrong
                    # can I get previous action? or just get reward = 1 when it's feasible and 0 otherwise?
                    # according to
                    # here is when no action is taken, or taking action?
                    # it's taking action, and then rewards later, so I need to get reward = 1 when it's feasible and 0 otherwise
                    # so the state_reward_pairs['uncertainty_type'] == 'begin'
                    action = state_reward_pairs.loc[pair_index]['action']
                    dynamic_RL34959.ALNS_got_action_in_implementation = 1
                    new_experience = 1
                    # if np.size(dynamic_RL34959.state_action_reward_collect) > 0:
                    #     while True:
                    #         # print('get_and_send_rewards 3')
                    #         stop_wait()
                    #         try:
                    #             try_row1 = my_deepcopy(state_reward_pairs.loc[pair_index])
                    #             try_row2 = my_deepcopy(state_reward_pairs.loc[pair_index])
                    #             break
                    #         except:
                    #             print('try_row1 = state_reward_pairs.loc[pair_index]')
                    #
                    #     try_row1['reward'] = 1
                    #     try_row1 = list(try_row1)
                    #     try_row2['reward'] = 0
                    #     try_row2 = list(try_row2)
                    #
                    #     # try_row1_index = np.where(np.all(dynamic_RL34959.state_action_reward_collect == try_row1, axis=1))
                    #     repeat_find_experience_exist = 0
                    #     while True:
                    #         if dynamic_RL34959.stop_everything_in_learning_and_go_to_implementation_phase == 1:
                    #             break
                    #         # print('get_and_send_rewards 4')
                    #         stop_wait()
                    #         repeat_find_experience_exist += 1
                    #
                    #         if repeat_find_experience_exist > 100:
                    #             break
                    #         try_row1_index = np.where(np.all(dynamic_RL34959.state_action_reward_collect == try_row1, axis=1))
                    #
                    #         if np.size(try_row1_index) > 0:
                    #             if repeat_find_experience_exist > 2:
                    #                 print("repeat_find_experience_exist_row1", repeat_find_experience_exist)
                    #             reward = dynamic_RL34959.state_action_reward_collect[try_row1_index, -1]
                    #             try:
                    #                 state_reward_pairs.loc[pair_index]['reward'] = reward[0][0]
                    #             except:
                    #                 break
                    #             new_experience = 0
                    #             break
                    #         try_row2_index = np.where(np.all(dynamic_RL34959.state_action_reward_collect == try_row2, axis=1))
                    #         if np.size(try_row2_index) > 0:
                    #             if repeat_find_experience_exist > 2:
                    #                 print("repeat_find_experience_exist_row2", repeat_find_experience_exist)
                    #             reward = dynamic_RL34959.state_action_reward_collect[try_row2_index, -1]
                    #             try:
                    #                 state_reward_pairs.loc[pair_index]['reward'] = reward[0][0]
                    #             except:
                    if new_experience == 1:

                        reward = get_reward(check_terminal, action, influenced_r, R_change_dynamic_travel_time, index, duration, congestion_link, congestion_node, -1)
                        state_reward_pairs.loc[pair_index]['reward'] = reward

                else:
                    state_reward_pairs.loc[pair_index]['reward'] = reward
                break_flag = 1
                break
        if break_flag == 1:
            while True:
                # print('get_and_send_rewards 5')
                stop_wait()
                if dynamic_RL34959.RL_drop_finish == 1 or dynamic_RL34959.implement == 1 or state_reward_pairs['reward'][pair_index] != -10000000:
                    break
                else:
                    print("state_reward_pairs['reward'][pair_index] != -10000000")
                    state_reward_pairs['reward'][pair_index] = reward
            break
    while True:
        if dynamic_RL34959.stop_everything_in_learning_and_go_to_implementation_phase == 1:
            break
        stop_wait()
        if dynamic_RL34959.RL_drop_finish == 1 or dynamic_RL34959.implement == 1:
            dynamic_RL34959.RL_drop_finish = 0
            break
        else:
            try:
                state_reward_pairs['reward'][pair_index] = reward
            except:
                print('here state_reward_pairs is broken so wait for qiji')
    #wait for the final/evaluation iteration of RL, because there are some bugs between implementation = 0 and 1
    while dynamic_RL34959.wait_training_finish_last_iteration == 1:
        if dynamic_RL34959.stop_everything_in_learning_and_go_to_implementation_phase == 1:
            break
        stop_wait()
        continue
# @profile()
# @time_me()
def initial_solution():
    global routes, R_pool, request_flow_t, R
    # R_pool = pd.DataFrame(R_pool,columns=['p','d','ap','bp','ad','bd','qr','r'])
    if dynamic == 1 and dynamic_t > 0:
        global R_change_dynamic


    if get_initial_bymyself == 1:
        if weight_interval == 1:
            R_pool = copy.copy(R)
        # only in this way, all requests will be tried to insert into the routes. Otherwise only one request will
        initial_regret = 1
        if initial_regret == 1:
            if allow_infeasibility == 0:
                not_regret = 0
                while R_pool.size != 0:
                    left_r = len(R_pool)
                    print('left_r', left_r)
                    if dynamic == 1 and dynamic_t > 0:
                        while R_pool.size != 0:
                            for i in R_pool[:, 7]:
                                if i in R_pool[:, 7]:
                                    if i in request_segment_in_dynamic[:, 7]:
                                        routes, R_pool = greedy_insert(i,1)
                                    else:
                                        routes, R_pool = greedy_insert(i)
                    else:
                        if not_regret == 0:
                            routes, R_pool = global_real_greedy_insert_regret()


                    after_regret_len = len(R_pool)
                    if left_r == after_regret_len:
                        # not_regret = 1
                        # while R_pool.size != 0:
                        #     for i in R_pool[:, 7]:
                        #         if i in R_pool[:, 7]:
                        #             routes, R_pool = real_greedy_insert(i)
                        #     if R_pool.size != 0:
                        #         print('left_r',len(R_pool))
                        #         routes, R_pool = random_removal()
                        routes, R_pool = random_removal()
            else:
                repeat_times = 0
                while R_pool.size != 0:

                    routes, R_pool = global_real_greedy_insert_regret()
                    repeat_times = repeat_times + 1
                    print('repeat_times',repeat_times,'left_r',len(R_pool))
                    if repeat_times > min(20, max(5, len(R)/5)):
                        break
        else:
            for i in R_pool[:, 7]:
                if i in R_pool[:, 7]:

                    routes, R_pool = real_greedy_insert(i)
            for i in R_pool[:, 7]:
                if i in R_pool[:, 7]:

                    routes, R_pool = real_greedy_insert(i)
        ##the global_real_greedy_insert is used to get the initial solution
        # not_serve_all_times = 0
        # while R_pool.size != 0:
        #     routes, R_pool = global_real_greedy_insert()
        #
        #     not_serve_all_times = not_serve_all_times + 1
        #     if not_serve_all_times >= int(1.3 * len(R[:,7])) + 1:
        #         break
        if allow_infeasibility == 0:
            while R_pool.size != 0:
                p_insertion = []
                p_removal = []
                for j in range(len(insert_heuristic)):
                    p_insertion.append(1 / len(insert_heuristic))
                for j in range(len(removal_heuristic)):
                    p_removal.append(1 / len(removal_heuristic))
                number_insertion = int(np.random.choice(range(len(insert_heuristic)), size=(1,), p=p_insertion))
                number_removal = int(np.random.choice(range(len(removal_heuristic)), size=(1,), p=p_removal))
                print(removal_heuristic['operator'][number_removal])
                routes, R_pool = eval(removal_heuristic['operator'][number_removal] + '()')
                # routes, R_pool = random_removal()

                # to avoid repeat stores in hash tables
                for k_name in routes.keys():
                    if len(routes[k_name][4]) <= 2:
                        routes[k_name][1:4, 0] = routes[k_name][0, 0]
                print(insert_heuristic['operator'][number_insertion])
                if insert_heuristic['operator'][number_insertion] == 'global_real_greedy_insert' or \
                        insert_heuristic['operator'][number_insertion] == 'global_real_greedy_insert_regret' or \
                        insert_heuristic['operator'][number_insertion] == 'most_hard_first_insert':
                    # global_real_greedy_insert will get all potential alternatives, so if it can't get feasible solution for all r, then the current routes may hard be updated, so it needs to be destroyed again, if destroyed for 3 times it still infeasible, then give up

                    routes, R_pool = eval(insert_heuristic['operator'][number_insertion] + '()')

                else:
                    for h in range(int(len(R_pool[:, 7])*1.3)):
                        if R_pool.size != 0:
                            i = random.choice(R_pool[:, 7])
                            # routes, R_pool=greedy_insert(i)
                            routes, R_pool = eval(insert_heuristic['operator'][number_insertion] + '(i)')
                        else:
                            break
            #         greedy_insert(i)
            #        routes, R_pool = transshipment_insert(i)
    else:
        if by_wenjing == 1:
            xls_path = "/data/yimeng/Case study/Preferences/Wenjing matching/routes/"
            xls = pd.ExcelFile(xls_path + str(request_number_in_R) + "r_result_correct_right.xlsx")

        else:
            xls = pd.ExcelFile(path + old_current_save + '/best_routes' + old_current_save + str(
                exp_number - 1) + '.xlsx')
        routes, R_pool = create_routes_R_pool_initial(xls)
    #check_satisfy_constraints()
    return routes, R_pool



def possible_remove_r(k,route,load,v_has_r,old_obj,r,R_i):
    r_in_k = []
    if K[k, 5] == 3:
        return r_in_k
    capacity_full = 0
    may_violate_other_constraints = 0
    more_than_one_k = v_has_r[1]
    if capacity_constraints(has_end_depot, K, R, k, route, load) == False:
        capacity_full = 1
    else:
        if more_than_one_k == -1:
            if K[k, 5] != 3 and k not in fixed_vehicles_percentage:
                #try to insert, check other constraints
                Trans, Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd = 0, 0, 0, 0, 0
                # not use insert_a_r because it must know positions or the other r in a bundle has been inserted to the route before
                obj_1_vehicle = best_position_1_vehicle(R, no_route_barge, no_route_truck, hash_table_1v_all_fail,
                                                        hash_table_1v_all, routes, fixed_vehicles_percentage, Fixed, K,
                                                        hash_table_1v, hash_table_1v_fail, has_end_depot, R_i, r,
                                                        v_has_r[0],
                                                        Trans, Trans_Tp, Trans_Td, Trans_secondTp,
                                                        Trans_secondTd)
                routes_local = my_deepcopy(routes)
                R_pool_local = copy.copy(R_pool)
                routes_local, R_pool_local = insert_r_in_swap(r, R_i, routes_local, R_pool_local, obj_1_vehicle)
                new_obj = overall_obj(routes_local)[1]
                if new_obj > old_obj or len(R_pool_local) != 0:
                    may_violate_other_constraints = 1
        else:
            #if k is fixed, no other constraints can be violated
            if k not in fixed_vehicles_percentage:
                may_violate_other_constraints = 1
    if  capacity_full==1 or may_violate_other_constraints == 1:
        # remove r in route one by one, and try to insert r and compare overall obj
        for col in route[4][1:-1]:
            request_number_col = ''.join(filter(str.isdigit, col))
            if request_number_col not in r_in_k:
                request_number_col = int(request_number_col)
                if request_number_col not in r_in_k:
                    r_in_k.append(request_number_col)

    return r_in_k

def insert_r_in_swap(r,R_i,routes_local,R_pool_local,obj_1_vehicle):
    obj_list=[]
    if obj_1_vehicle:
        obj_list.append(obj_1_vehicle)

    if obj_list:
        obj_df_one_column = pd.DataFrame(obj_list, columns=['one_column'])
        obj_df = pd.DataFrame(obj_df_one_column['one_column'].values.tolist(),
                              columns=['k', 'original_route', 'original_route_no_columns',
                                       'cost_inserted_request',
                                       'dict_a_request_best_position'])
        obj_df = obj_df.values
        obj_best = obj_df[np.argmin(obj_df[:, 3], axis=0), :]
        best_k, original_route, original_route_no_columns, cost_inserted_request, dict_a_request_best_position = obj_best
        key = get_key_1k(R_i, original_route_no_columns, best_k, fixed_vehicles_percentage, Fixed, K)
        routes_local[best_k] = copy.copy(hash_table_1v_all[key][dict_a_request_best_position]['route'])
        request_list2 = list(original_route[4])
        request_list2.insert(list(dict_a_request_best_position)[0], str(r) + 'pickup')
        request_list2.insert(list(dict_a_request_best_position)[1], str(r) + 'delivery')
        routes_local[best_k][4] = copy.copy(request_list2)
        R_pool_local = R_pool_local[~(R_pool_local[:, 7] == r)]

    return routes_local, R_pool_local

def create_R_pool():
    # danger not sure here are all right, I mean the shape
    if CP == 1:
        if heterogeneous_preferences == 1:
            R_pool = np.array(np.empty(shape=(0, 15)), dtype='object')
        else:
            R_pool = np.array(np.empty(shape=(0, 10)), dtype='object')
    else:
        if heterogeneous_preferences == 1:
            R_pool = np.array(np.empty(shape=(0, 14)), dtype='object')
        else:
            R_pool = np.array(np.empty(shape=(0, 9)), dtype='object')
    R_pool[:] = np.NaN
    served, all_served_r = check_served_R(0, -1, 1)
    if served < len(R):
        for r in R[:,7]:
            if r not in all_served_r:
                R_pool = np.vstack([R_pool, R[list(R[:, 7]).index(r)]])

    return R_pool

def swap_it(compare_remove_r,compare_save_routes):
    global routes, R_pool,request_flow_t
    if compare_remove_r:
        print('swap_success')
        compare_remove_r_array = np.array(compare_remove_r)
        best_r = compare_remove_r_array[np.argmin(compare_remove_r_array[:, 1], axis=0)][0]
        routes,request_flow_t = compare_save_routes[best_r]

        # R_pool = np.array(np.empty(shape=(0,9)),dtype='int')
        R_pool = create_R_pool()
        print('obj_swap_it',overall_obj(routes)[1])
    check_served_R()
    return routes,R_pool

def format_v_has_r(v_has_r_local):
    break_or_not = 0
    if isinstance(v_has_r_local, (int, float)):
        if math.isnan(v_has_r_local):
            break_or_not = 1
        else:
            v_has_r_local = int(v_has_r_local)
            v_has_r_local = [v_has_r_local, -1, -1]
    else:
        if len(v_has_r_local) == 2:
            v_has_r_local = [v_has_r_local[0], v_has_r_local[1], -1]
        if v_has_r_local[0] == -1:

            break_or_not = 1
    return break_or_not, v_has_r_local
def special_swap():
    global routes,R_pool,request_flow_t
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
    #check_satisfy_constraints()
    r_cost_gap = history_removal(1)
    random_position = 1
    if len(r_cost_gap) == 0:
        return routes,R_pool
    # print(r_best_obj_record)
    current_v_has_r_v_has_r=0
    for r_index in range(len(r_cost_gap)):
        r = int(r_cost_gap[r_index,1])
        print('want_to_swap_request ',r)
        print('cost_gap', r_cost_gap[r_index,0])
        # to avoid if the swap not success, r still in R_pool, change the routes back to the initial one
        routes_initial = my_deepcopy(routes)
        R_pool_initial = copy.copy(R_pool)
        request_flow_t_initial = copy.copy(request_flow_t)
        index_r = list(R[:,7]).index(r)
        load = R[index_r,6]
        old_obj = overall_obj(routes)[1]
        # print('initial_obj',old_obj)
        # print('dont know', overall_obj(routes_initial)[1])
        R_i = tuple(zip(R[index_r], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))

        routes, R_pool, current_v_has_r, current_used_T = remove_a_request(r, routes, R_pool)
        #lost_r()
        routes_save, R_pool_save,request_flow_t_save = my_deepcopy(routes), copy.copy(R_pool), copy.copy(request_flow_t)
        v_has_r, used_T = r_best_obj_record[index_r,1:3]
        # break_or_not, current_v_has_r = format_v_has_r(current_v_has_r)
        # print('r',r)
        # if break_or_not == 1:
        #     print('current_v_has_r_break')
        #     routes = my_deepcopy(routes_initial)
        #     R_pool = copy.copy(R_pool_initial)
        #     request_flow_t = copy.copy(request_flow_t_initial)
        #     continue
        break_or_not, v_has_r = format_v_has_r(v_has_r)
        #check_satisfy_constraints()
        if break_or_not == 1:
            print('v_has_r_break')
            routes = my_deepcopy(routes_initial)
            R_pool = copy.copy(R_pool_initial)
            request_flow_t = copy.copy(request_flow_t_initial)
            continue
        # if current_v_has_r == v_has_r:
        #     current_v_has_r = current_v_has_r + 1
        #     print('current_v_has_r == v_has_r',current_v_has_r)
            # routes = my_deepcopy(routes_initial)
            # R_pool = copy.copy(R_pool_initial)
            # request_flow_t = copy.copy(request_flow_t_initial)
            # continue
        k1 = v_has_r[0]
        r_in_k1 = possible_remove_r(k1, routes[k1], load, v_has_r,old_obj,r,R_i)
        if v_has_r[1] != -1:
            k2 = v_has_r[1]
            T1 = used_T[0]
            r_in_k2 = possible_remove_r(k2, routes[k2], load, v_has_r,old_obj,r,R_i)
            if v_has_r[2] != -1:
                k3 = v_has_r[2]
                T2 = used_T[1]
                r_in_k3 = possible_remove_r(k3, routes[k3], load, v_has_r,old_obj,r,R_i)
        #if only use one k, then check capacity full or not, if full, remove r in k1 one by one, and compare overall obj
        #                                                       not full, then insert r to route directly
        #if use 2 k, then check two k's capacity, and remove r if any k's capacity is full
        if v_has_r[1] == -1:
            compare_remove_r = []
            compare_save_routes = {}
            if r_in_k1:
                # print('dont know4', overall_obj(routes_initial)[1])
                for r_remove in r_in_k1:

                    routes, R_pool = remove_a_request(r_remove, routes, R_pool)[0:2]
                    #lost_r()
                    Trans, Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd = 0, 0, 0, 0, 0
                    #not use insert_a_r because it must know positions or the other r in a bundle has been inserted to the route before
                    obj_1_vehicle = best_position_1_vehicle(R, no_route_barge, no_route_truck, hash_table_1v_all_fail,
                                                            hash_table_1v_all, routes, fixed_vehicles_percentage, Fixed, K,
                                                            hash_table_1v, hash_table_1v_fail, has_end_depot, R_i, r, v_has_r[0],
                                                            Trans, Trans_Tp, Trans_Td, Trans_secondTp,
                                                            Trans_secondTd)
                    if obj_1_vehicle:
                        routes, R_pool = insert_r_in_swap(r, R_i, routes, R_pool, obj_1_vehicle)
                        #insert the removed r
                        routes, R_pool = real_greedy_insert(r_remove)
                        if len(R_pool) == 0:
                            new_obj = overall_obj(routes)[1]
                            print('new_obj',new_obj,'old_obj',old_obj)
                            if new_obj < old_obj:
                                
                                compare_remove_r.append([r_remove,new_obj])
                                compare_save_routes[r_remove] = [my_deepcopy(routes),copy.copy(request_flow_t)]

                    routes = my_deepcopy(routes_save)
                    R_pool = copy.copy(R_pool_save)
                    request_flow_t = copy.copy(request_flow_t_save)
                routes,R_pool = swap_it(compare_remove_r,compare_save_routes)
                #check_satisfy_constraints()
            else:
                #check_satisfy_constraints()
                Trans, Trans_Tp, Trans_Td, Trans_secondTp, Trans_secondTd = 0, 0, 0, 0, 0
                # not use insert_a_r because it must know positions or the other r in a bundle has been inserted to the route before
                obj_1_vehicle = best_position_1_vehicle(R, no_route_barge, no_route_truck, hash_table_1v_all_fail,
                                                        hash_table_1v_all, routes, fixed_vehicles_percentage, Fixed, K,
                                                        hash_table_1v, hash_table_1v_fail, has_end_depot, R_i, r, v_has_r[0],
                                                        Trans, Trans_Tp, Trans_Td, Trans_secondTp,
                                                            Trans_secondTd)
                routes, R_pool = insert_r_in_swap(r, R_i,routes,R_pool, obj_1_vehicle)
                new_obj = overall_obj(routes)[1]
                if new_obj > old_obj or len(R_pool) != 0:
                    routes = my_deepcopy(routes_initial)
                    R_pool = copy.copy(R_pool_initial)
                    print('swap_fail when there is no capacity limitation')
                else:
                    print('swap_success')
            #check_satisfy_constraints()
        # print('dont know2', overall_obj(routes_initial)[1])
        if v_has_r[1] != -1 and v_has_r[2] == -1:
            #I don't know where makes this case happens
            #check_satisfy_constraints()
            if used_T[0] == -1:
                routes = my_deepcopy(routes_initial)
                R_pool = copy.copy(R_pool_initial)
                request_flow_t = copy.copy(request_flow_t_initial)
                continue
            compare_remove_r = []
            compare_save_routes = {}
            Trans = 1
            
            if r_in_k1 or r_in_k2:
                
                #both k1 and k2 are full
                if r_in_k1 and r_in_k2:
                    # print('dont know3', overall_obj(routes_initial)[1])
                    for r_remove1 in r_in_k1:
                        for r_remove2 in r_in_k2:
                            obj_list_best_T = []
                            routes, R_pool = remove_a_request(r_remove1, routes, R_pool)[0:2]
                            routes, R_pool = remove_a_request(r_remove2, routes, R_pool)[0:2]
                            #lost_r()
                            a = len(R_pool)
                            obj_list_best_T, best_cost_inserted_request = insert2vehicle_k(parallel, no_route_barge,
                                                                                           no_route_truck,
                                                                                           has_end_depot, r,
                                                                                           R_i, T1, v_has_r[0], v_has_r[1],
                                                                                           fixed_vehicles_percentage, K,
                                                                                           Fixed, obj_list_best_T,
                                                                                           Trans,
                                                                                           random_position, routes,
                                                                                           hash_table_2v_fail,
                                                                                           hash_table_2v,
                                                                                           hash_table_2v_all_fail,
                                                                                           hash_table_2v_all, R_pool_2v,
                                                                                           R,
                                                                                           hash_table_1v,
                                                                                           hash_table_1v_fail,
                                                                                           hash_table_1v_all,
                                                                                           hash_table_1v_all_fail,
                                                                                           request_flow_t)
                            k1, k2, routes, R_pool, best_T = insert2vehicle_best(obj_list_best_T, R_i, r)
                            #check_satisfy_constraints()
                            b = len(R_pool)
                            if a != b:
                                routes, R_pool = real_greedy_insert(r_remove1)
                                routes, R_pool = real_greedy_insert(r_remove2)
                                if len(R_pool) == 0:
                                    new_obj = overall_obj(routes)[1]
                                    if new_obj < old_obj:
                                        r_key = tuple([r_remove1, r_remove2])
                                        compare_remove_r.append([r_key, new_obj])
                                        compare_save_routes[r_key] = [my_deepcopy(routes),copy.copy(request_flow_t)]
                            #check_satisfy_constraints()
                            routes = my_deepcopy(routes_save)
                            R_pool = copy.copy(R_pool_save)
                            request_flow_t = copy.copy(request_flow_t_save)
                    routes,R_pool = swap_it(compare_remove_r,compare_save_routes)
                    #check_satisfy_constraints()
                else:
                    #only k1 is full
                    if r_in_k1:
                        # print('dont know5', overall_obj(routes_initial)[1])
                        for r_remove1 in r_in_k1:
                            obj_list_best_T = []
                            routes, R_pool = remove_a_request(r_remove1, routes, R_pool)[0:2]
                            #lost_r()
                            a = len(R_pool)
                            obj_list_best_T, best_cost_inserted_request = insert2vehicle_k(parallel, no_route_barge,
                                                                                           no_route_truck,
                                                                                           has_end_depot, r,
                                                                                           R_i, T1, v_has_r[0], v_has_r[1],
                                                                                           fixed_vehicles_percentage, K,
                                                                                           Fixed, obj_list_best_T,
                                                                                           Trans,
                                                                                           random_position, routes,
                                                                                           hash_table_2v_fail,
                                                                                           hash_table_2v,
                                                                                           hash_table_2v_all_fail,
                                                                                           hash_table_2v_all, R_pool_2v,
                                                                                           R,
                                                                                           hash_table_1v,
                                                                                           hash_table_1v_fail,
                                                                                           hash_table_1v_all,
                                                                                           hash_table_1v_all_fail,
                                                                                           request_flow_t)
                            k1, k2, routes, R_pool, best_T = insert2vehicle_best(obj_list_best_T, R_i, r)
                            #check_satisfy_constraints()
                            b = len(R_pool)
                            if a != b:
                                routes, R_pool = real_greedy_insert(r_remove1)

                                if len(R_pool) == 0:
                                    new_obj = overall_obj(routes)[1]
                                    if new_obj < old_obj:

                                        compare_remove_r.append([r_remove1, new_obj])
                                        compare_save_routes[r_remove1] = [my_deepcopy(routes),copy.copy(request_flow_t)]

                            routes = my_deepcopy(routes_save)
                            R_pool = copy.copy(R_pool_save)
                            request_flow_t = copy.copy(request_flow_t_save)
                        routes,R_pool = swap_it(compare_remove_r,compare_save_routes)
                        #check_satisfy_constraints()
                        # print('dont know6', overall_obj(routes_initial)[1])
                    # only k2 is full
                    else:
                        # print('dont know6', overall_obj(routes_initial)[1])
                        for r_remove2 in r_in_k2:
                            obj_list_best_T = []
                            routes, R_pool = remove_a_request(r_remove2, routes, R_pool)[0:2]
                            #lost_r()
                            a = len(R_pool)
                            obj_list_best_T, best_cost_inserted_request = insert2vehicle_k(parallel, no_route_barge,
                                                                                           no_route_truck,
                                                                                           has_end_depot, r,
                                                                                           R_i, T1, v_has_r[0], v_has_r[1],
                                                                                           fixed_vehicles_percentage, K,
                                                                                           Fixed, obj_list_best_T,
                                                                                           Trans,
                                                                                           random_position, routes,
                                                                                           hash_table_2v_fail,
                                                                                           hash_table_2v,
                                                                                           hash_table_2v_all_fail,
                                                                                           hash_table_2v_all, R_pool_2v,
                                                                                           R,
                                                                                           hash_table_1v,
                                                                                           hash_table_1v_fail,
                                                                                           hash_table_1v_all,
                                                                                           hash_table_1v_all_fail,
                                                                                           request_flow_t)
                            k1, k2, routes, R_pool, best_T = insert2vehicle_best(obj_list_best_T, R_i, r)
                            b = len(R_pool)
                            if a != b:

                                routes, R_pool = real_greedy_insert(r_remove2)
                                if len(R_pool) == 0:
                                    new_obj = overall_obj(routes)[1]
                                    if new_obj < old_obj:

                                        compare_remove_r.append([r_remove2, new_obj])
                                        compare_save_routes[r_remove2] = [my_deepcopy(routes),copy.copy(request_flow_t)]

                            routes = my_deepcopy(routes_save)
                            R_pool = copy.copy(R_pool_save)
                            request_flow_t = copy.copy(request_flow_t_save)
                        routes,R_pool = swap_it(compare_remove_r,compare_save_routes)
                        #check_satisfy_constraints()
                #check_satisfy_constraints()
            else:
                #check_satisfy_constraints()
                # print('dont know7', overall_obj(routes_initial)[1])
                obj_list_best_T = []
                obj_list_best_T, best_cost_inserted_request = insert2vehicle_k(parallel, no_route_barge,
                                                                               no_route_truck, has_end_depot, r,
                                                                               R_i, T1, v_has_r[0], v_has_r[1],
                                                                               fixed_vehicles_percentage, K,
                                                                               Fixed, obj_list_best_T, Trans,
                                                                               random_position, routes,
                                                                               hash_table_2v_fail,
                                                                               hash_table_2v,
                                                                               hash_table_2v_all_fail,
                                                                               hash_table_2v_all, R_pool_2v, R,
                                                                               hash_table_1v,
                                                                               hash_table_1v_fail,
                                                                               hash_table_1v_all,
                                                                               hash_table_1v_all_fail,
                                                                               request_flow_t)
                k1, k2, routes, R_pool, best_T = insert2vehicle_best(obj_list_best_T, R_i, r)
                new_obj = overall_obj(routes)[1]
                if new_obj > old_obj or len(R_pool) != 0:
                    routes = my_deepcopy(routes_initial)
                    R_pool = copy.copy(R_pool_initial)
                    print('swap_fail when there is no capacity limitation')
                else:
                    print('swap_success')
        # print('dont know8', overall_obj(routes_initial)[1])
        if len(R_pool) != 0:
            #check_satisfy_constraints()
            print('finally r not be inserted')
            routes = my_deepcopy(routes_initial)
            R_pool = copy.copy(R_pool_initial)
            request_flow_t = copy.copy(request_flow_t_initial)
            # a = (request_flow_t[:,0:3] == request_flow_t_initial[:,0:3])
            # b = (request_flow_t[:,5:6] == request_flow_t_initial[:,5:6])
            # def check(a):
            #     br = 0
            #     for u in a:
            #         for z in u:
            #             if z == False:
            #                 br = 1
            #                 break
            #         if br == 1:
            #             print('False')
            #             if overall_obj(routes)[1] != old_obj:
            #                 print('obj_not_equal')
            #             break
            # check(a)
            # check(b)

            # print('dont know9', overall_obj(routes_initial)[1])
    check_served_R()
    #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
        #check_satisfy_constraints()
    #check_satisfy_constraints()
    return routes,R_pool
# #get the best k for each r when all k are available -> as same as regret as initial solution
# def pure_best_historical_solution():



# @profile()
# @time_me()
##@jit
def Adaptive():
    global lowest_cost, routes_lowest_cost, segment, operations, theta, pai, R_pool, routes, removal_heuristic, insert_heuristic, theta_insert, theta_removal, weight_insertion, weight_removal
    if combination == 1:
        initial_weight = 1 / len(operations)
        if repeat == 1:
            weight.iloc[0] = initial_weight
            #        weight.iloc[0]['transshipment_insert_clear_a_route'] = initial_weight * 5
            weight['transshipment_insert_delete_node'][0] = initial_weight * 5
            segment = 0
        else:
            if repeat % segment_number == 0:
                #lost_r()
                segment = int(repeat / segment_number)
                for th in range(0, len(theta)):
                    if theta['theta'][th] == 0:
                        weight[operations['operation'][th]][segment] = weight[operations['operation'][th]][segment - 1]
                    else:
                        weight[operations['operation'][th]][segment] = weight[operations['operation'][th]][
                                                                           segment - 1] * (1 - r) + r * \
                                                                       pai[operations['operation'][th]][repeat - 1] / \
                                                                       theta['theta'][th]

                theta['theta'] = 0
                #lost_r()
        # Greedy Insert greedy Random (vehicle number) Insert random insert
        sum_weight = weight.iloc[segment].sum()
        p = []
        for j in range(len(operations)):
            p.append(weight[operations['operation'][j]][segment] / sum_weight)
            # all operators chose by same probability
        #        p.append(1/len(operations))
        number = int(np.random.choice(range(len(operations)), size=(1,), p=p))
    else:
        initial_weight_insertion = 1 / len(insert_heuristic)
        initial_weight_removal = 1 / len(removal_heuristic)
        if repeat == 1:
            weight_insertion.iloc[0] = initial_weight_insertion
            weight_removal.iloc[0] = initial_weight_removal
            #        weight.iloc[0]['transshipment_insert_clear_a_route'] = initial_weight * 5
            #         weight_removal['delete_node'][0] = initial_weight_removal * 2
            segment = 0


        else:
            if repeat % segment_number == 0:
                segment = int(repeat / segment_number)

                #                theta_insert['theta']=0
                #                theta_removal['theta']=0
                #
                for th in range(0, len(theta_insert)):
                    if theta_insert['theta'][th] == 0:
                        weight_insertion[insert_heuristic['operator'][th]][segment] = \
                            weight_insertion[insert_heuristic['operator'][th]][segment - 1]
                    else:
                        weight_insertion[insert_heuristic['operator'][th]][segment] = \
                            weight_insertion[insert_heuristic['operator'][th]][segment - 1] * (1 - r) + r * \
                            pai[insert_heuristic['operator'][th]][repeat - 1] / theta_insert['theta'][th]
                for th in range(0, len(theta_removal)):
                    if theta_removal['theta'][th] == 0:
                        weight_removal[removal_heuristic['operator'][th]][segment] = \
                            weight_removal[removal_heuristic['operator'][th]][segment - 1]
                    else:
                        weight_removal[removal_heuristic['operator'][th]][segment] = \
                            weight_removal[removal_heuristic['operator'][th]][segment - 1] * (1 - r) + r * \
                            pai[removal_heuristic['operator'][th]][repeat - 1] / theta_removal['theta'][th]

                theta_insert['theta'] = 0
                theta_removal['theta'] = 0
                if start_from_best_at_begin_of_segement == 1:
                    current_cost = overall_obj(routes)[1]
                    if current_cost < lowest_cost:
                        lowest_cost = current_cost
                        routes_lowest_cost = my_deepcopy(routes)
                    if parallel_ALNS == 1:
                        if not os.path.isdir(path + current_save):
                            Path(path + current_save).mkdir(parents=True, exist_ok=True)
                        parallel_best_cost_path = path + 'parallel_best_cost.xlsx'
                        best_routes_path = path + current_save + '/best_routes' + current_save + '_' + str(
                            exp_number - 1) + '.xlsx'
                        if not os.path.isfile(parallel_best_cost_path):
                            parallel_best_cost = pd.DataFrame(index=range(0, 1000), columns=['best_cost', 'not_skip'])
                            parallel_best_cost['not_skip'] = 0
                            parallel_best_cost['best_cost'][parallel_number] = lowest_cost
                            with pd.ExcelWriter(parallel_best_cost_path) as writer:  # doctest: +SKIP
                                parallel_best_cost.to_excel(writer, sheet_name='best_cost', index=False)
                            with pd.ExcelWriter(best_routes_path) as writer:  # doctest: +SKIP
                                for key, value in routes_lowest_cost.items():
                                    route_df = pd.DataFrame(value[0:4, :], columns=value[4])
                                    route_df.to_excel(writer, str(key))
                        else:
                            parallel_best_cost = pd.read_excel(parallel_best_cost_path, 'best_cost')
                            parallel_lowest_cost = parallel_best_cost['best_cost'].dropna().min()
                            if lowest_cost <= parallel_lowest_cost:
                                if lowest_cost != parallel_lowest_cost:

                                    routes = my_deepcopy(routes_lowest_cost)
                                    parallel_best_cost['best_cost'][parallel_number] = lowest_cost

                                    # write the current route to file
                                    with pd.ExcelWriter(best_routes_path) as writer:  # doctest: +SKIP
                                        for key, value in routes_lowest_cost.items():
                                            value.to_excel(writer, key)

                                    with pd.ExcelWriter(parallel_best_cost_path) as writer:  # doctest: +SKIP
                                        parallel_best_cost.to_excel(writer, sheet_name='best_cost', index=False)
                            else:
                                parallel_lowest_cost_index = parallel_best_cost.index[
                                    parallel_best_cost['best_cost'] == parallel_lowest_cost].tolist()[0]
                                current_save_parallel = 'percentage' + str(percentage) + 'parallel_number' + str(
                                    parallel_lowest_cost_index)
                                parallel_lowest_cost_routes_path = path + current_save_parallel + '/best_routes' + current_save_parallel + '_' + str(
                                    exp_number - 1) + '.xlsx'
                                #lost_r()
                                routes = pd.read_excel(parallel_lowest_cost_routes_path, None, index_col=0)
                                #lost_r()
                    else:
                        if current_cost > lowest_cost:
                            #lost_r()
                            routes = my_deepcopy(routes_lowest_cost)
                            #lost_r()
        sum_weight_insertion = weight_insertion.values[segment].sum()
        sum_weight_removal = weight_removal.values[segment].sum()
        p_insertion = []
        p_removal = []
        for j in range(len(insert_heuristic)):
            p_insertion.append(weight_insertion[insert_heuristic['operator'][j]][segment] / sum_weight_insertion)
        for j in range(len(removal_heuristic)):
            p_removal.append(weight_removal[removal_heuristic['operator'][j]][segment] / sum_weight_removal)

        number_insertion = int(np.random.choice(range(len(insert_heuristic)), size=(1,), p=p_insertion))
        number_removal = int(np.random.choice(range(len(removal_heuristic)), size=(1,), p=p_removal))

    not_serve_all_times = 0
    a = 0
    original_routes = my_deepcopy(routes)
    original_R_pool = copy.copy(R_pool)

    if dynamic == 1 and dynamic_t > 0:
        pass
    else:
        if combination == 1:
            routes, R_pool = eval(operations['removal'][number] + '()')
        else:
            #lost_r()
            print(removal_heuristic['operator'][number_removal])
            routes, R_pool = eval(removal_heuristic['operator'][number_removal] + '()')
        #lost_r()

    while R_pool.size != 0 or request_segment_in_dynamic.size != 0:
        if not_serve_all_times != 0:
            if dynamic == 1 and dynamic_t > 0:
                pass
            else:
                if combination == 1:
                    routes, R_pool = eval(operations['removal'][number] + '()')
                else:
                    #lost_r()
                    print(removal_heuristic['operator'][number_removal])
                    routes, R_pool = eval(removal_heuristic['operator'][number_removal] + '()')
                #lost_r()
        # to avoid repeat stores in hash tables
        for k_name in routes.keys():
            if len(routes[k_name][4]) <= 2:
                routes[k_name][1:4, 0] = routes[k_name][0, 0]

        if insert_heuristic['operator'][number_insertion] == 'global_real_greedy_insert' or \
                insert_heuristic['operator'][number_insertion] == 'global_real_greedy_insert_regret' or \
                insert_heuristic['operator'][number_insertion] == 'most_hard_first_insert':
            # global_real_greedy_insert will get all potential alternatives, so if it can't get feasible solution for all r, then the current routes may hard be updated, so it needs to be destroyed again, if destroyed for 3 times it still infeasible, then give up

            print(insert_heuristic['operator'][number_insertion])

            #lost_r()
            not_serve_all_times = not_serve_all_times + 1
            print('not_serve_all_times',not_serve_all_times)
            # lost_r()
            routes, R_pool = eval(insert_heuristic['operator'][number_insertion] + '()')
            if not_serve_all_times > 3:
                if allow_infeasibility == 0:
                    routes = my_deepcopy(original_routes)
                    R_pool = copy.copy(original_R_pool)
                    print(len(R_pool))
                else:
                    break

        else:
            not_serve_all_times = not_serve_all_times + 1
            if dynamic == 1 and dynamic_t > 0:
                for x in request_segment_in_dynamic[:,7]:
                    if request_segment_in_dynamic.size == 0:
                        break
                    x = int(x)
                    segment_in_dynamic = 1
                    if combination == 1:
                        routes, R_pool = eval(operations['insertion'][number] + '(x, segment_in_dynamic)')
                    else:
                        print(insert_heuristic['operator'][number_insertion])
                        # lost_r()
                        routes, R_pool = eval(insert_heuristic['operator'][number_insertion] + '(x, segment_in_dynamic)')
            # print('notserve_all_times', not_serve_all_times)
            for h in range(int(len(R_pool[:, 7])*1.3)):
                if R_pool.size == 0:
                    break

                i = random.choice(R_pool[:, 7])
                if combination == 1:
                    routes, R_pool = eval(operations['insertion'][number] + '(i)')
                else:
                    print(insert_heuristic['operator'][number_insertion])
                    #lost_r()
                    routes, R_pool = eval(insert_heuristic['operator'][number_insertion] + '(i)')
                    #lost_r()

            if not_serve_all_times > 3:
                if allow_infeasibility == 0:
                    routes = my_deepcopy(original_routes)
                    R_pool = copy.copy(original_R_pool)
                    break
                else:

                    break



    #                start2=timeit.default_timer()
    #
    #                while R_pool.size != 0:
    #                    routes, R_pool = random_removal()
    #                    for h in R_pool[:,7]:
    #                        if R_pool.size == 0:
    #                            break
    #                        i = random.choice(R_pool[:,7])
    #                        routes, R_pool=greedy_insert(i)
    #                    Running_Time2 = timeit.default_timer() - start2
    #                    if Running_Time2>=10*len(R[:,7]):
    #                        if a%len(R[:,7])==0:
    #                            routes, R_pool = remove_all()
    #                        a=a+1
    #                        print('bs')
    #lost_r()
    check_served_R()
    # for k in routes.keys():
    #     if isinstance(capacity_constraints(has_end_depot, K, R, k, routes[k]),bool):
    #         print('wfwf')
    #         sys.exit(0)
    overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost, overall_transshipment_cost, overall_un_load_cost, overall_emission_cost, overall_storage_cost, overall_delay_penalty, overall_number_transshipment, overall_average_speed, overall_average_time_ratio, overall_emission_transshipment = overall_obj(
        routes)
    if combination == 1:
        pai_operation = operations['operation'][number]
        theta['theta'][number] = theta['theta'][number] + 1
    else:
        pai_operation_insert = insert_heuristic['operator'][number_insertion]
        pai_operation_removal = removal_heuristic['operator'][number_removal]
        theta_insert['theta'][number_insertion] = theta_insert['theta'][number_insertion] + 1
        theta_removal['theta'][number_removal] = theta_removal['theta'][number_removal] + 1

    if repeat > 0:
        if repeat % segment_number != 0:
            for j in pai.columns:
                if combination == 1:
                    if j != pai_operation:
                        pai[j][repeat] = pai[j][repeat - 1]
                else:
                    if j != pai_operation_insert and j != pai_operation_removal:
                        pai[j][repeat] = pai[j][repeat - 1]
            if multi_obj == 0:
                has_found = 0
                if overall_cost in obj_record['overall_cost'].values:
                    index = obj_record[obj_record['overall_cost'] == overall_cost].index.values
                    has_found = 1
                if has_found == 1 and served_requests == obj_record['served_requests'][index[0]]:
                    if combination == 1:
                        pai[pai_operation][repeat] = pai[pai_operation][repeat - 1]
                    else:
                        pai[pai_operation_insert][repeat] = pai[pai_operation_insert][repeat - 1]
                        pai[pai_operation_removal][repeat] = pai[pai_operation_removal][repeat - 1]
                else:
                    if served_requests >= obj_record['served_requests'].max() and overall_cost < obj_record['overall_cost'].min():
                        if repeat / segment_number != 0:
                            if combination == 1:
                                pai[pai_operation][repeat] = pai[pai_operation][repeat - 1] + miu1
                            else:
                                pai[pai_operation_insert][repeat] = pai[pai_operation_insert][repeat - 1] + miu1
                                pai[pai_operation_removal][repeat] = pai[pai_operation_removal][repeat - 1] + miu1
                    else:
                        if served_requests >= obj_record['served_requests'][repeat - 1] and overall_cost < obj_record['overall_cost'][repeat - 1]:
                            if combination == 1:
                                pai[pai_operation][repeat] = pai[pai_operation][repeat - 1] + miu2
                            else:
                                pai[pai_operation_insert][repeat] = pai[pai_operation_insert][repeat - 1] + miu2
                                pai[pai_operation_removal][repeat] = pai[pai_operation_removal][repeat - 1] + miu2
                        else:
                            if combination == 1:
                                pai[pai_operation][repeat] = pai[pai_operation][repeat - 1] + miu3
                            else:
                                pai[pai_operation_insert][repeat] = pai[pai_operation_insert][repeat - 1] + miu3
                                pai[pai_operation_removal][repeat] = pai[pai_operation_removal][repeat - 1] + miu3
            else:
                if dominate(overall_distance, overall_cost, overall_time, overall_profit, overall_emission,
                            served_requests) == 1:
                    if combination == 1:
                        pai[pai_operation][repeat] = pai[pai_operation][repeat - 1]
                    else:
                        pai[pai_operation_insert][repeat] = pai[pai_operation_insert][repeat - 1]
                        pai[pai_operation_removal][repeat] = pai[pai_operation_removal][repeat - 1]
                else:
                    if dominate(overall_distance, overall_cost, overall_time, overall_profit, overall_emission,
                                served_requests) == 2:
                        if repeat / segment_number != 0:
                            if combination == 1:
                                pai[pai_operation][repeat] = pai[pai_operation][repeat - 1] + miu1
                            else:
                                pai[pai_operation_insert][repeat] = pai[pai_operation_insert][repeat - 1] + miu1
                                pai[pai_operation_removal][repeat] = pai[pai_operation_removal][repeat - 1] + miu1
                    else:
                        if dominate(overall_distance, overall_cost, overall_time, overall_profit, overall_emission,
                                    served_requests) == 3:
                            if combination == 1:
                                pai[pai_operation][repeat] = pai[pai_operation][repeat - 1] + miu2
                            else:
                                pai[pai_operation_insert][repeat] = pai[pai_operation_insert][repeat - 1] + miu2
                                pai[pai_operation_removal][repeat] = pai[pai_operation_removal][repeat - 1] + miu2
                        else:
                            if combination == 1:
                                pai[pai_operation][repeat] = pai[pai_operation][repeat - 1] + miu3
                            else:
                                pai[pai_operation_insert][repeat] = pai[pai_operation_insert][repeat - 1] + miu3
                                pai[pai_operation_removal][repeat] = pai[pai_operation_removal][repeat - 1] + miu3
    return routes, R_pool

def remove_request_due_to_load_dependent():
    global during_iteration, routes, R_pool
# during_iteration = 1
    if heterogeneous_preferences == 1 and heterogeneous_preferences_no_constraints == 0 and emission_preference_constraints_after_iteration == 1:

        during_iteration = 0
        for request_number in R[:,7]:
            if request_number not in R_pool[:,7]:

                k1, k2, k3 = find_used_k(request_number)
                #20211216 here should check k1 = -1 or not, because under CP, request in R maybe exchanged to other requests and not in R_pool
                if k1 == -1:
                    continue
                else:
                    if k2 == -1:
                        if preference_constraints(request_number, k1, k2, k3, routes[k1], -1, -1) == 0:
                            routes, R_pool = remove_a_request(request_number, routes, R_pool)[0:2]
                    else:
                        if k3 == -1:
                            if preference_constraints(request_number, k1, k2, k3, routes[k1], routes[k2], -1) == 0:
                                routes, R_pool = remove_a_request(request_number, routes, R_pool)[0:2]
                        else:
                            if preference_constraints(request_number, k1, k2, k3, routes[k1], routes[k2], routes[k3]) == 0:
                                routes, R_pool = remove_a_request(request_number, routes, R_pool)[0:2]
    #danger here I need to check constraints of all requests again, because after deleting some requests, load are changed, and some requests may satisfy emission constraints before but not now
    during_iteration = 1


# @profile()
# @time_me()
def main(R_pool2, parallel_number2, SA2, combination2, only_T2, has_end_depot2, T_or_not, path2, N2, T_change, K_change,
         o_change, R_change,
         iteration_number, current_save2, i, j, transshipment_time2, service_time2, transshipment_cost_p, fuel_cost2,
         segment_number2, r2, miu1_1, miu2_1, miu3_1, pro, Fixed2, percentage2, k_random_or2):
    global during_iteration, Best_Running_Time_as_initial, revert_K,parallel_number, routes_lowest_cost, lowest_cost, no_T_R, train_truck, bundle_R, initial_solution_no_wait_cost, T_k_record, hash_df_table, hash_top_R_pool, r_best_obj_record, hard_value, fixed_vehicles_percentage, K_R, hash_table_route_no_columns, hash_table_route_no_columns_top, hash_top, D_origin_All, ok_K_canpickr, no_route_barge, no_route_truck, SA, combination, all_Tem_df, all_pro, all_pro_df, hash_table_1v, hash_table_1v_all, hash_table_2v, hash_table_2v_all, hash_table_1v_fail, hash_table_1v_all_fail, hash_table_2v_fail, hash_table_2v_all_fail, hash_table_3v, hash_table_3v_all, hash_table_3v_fail, hash_table_3v_all_fail, only_T, has_end_depot, T_or, T, K, o, R, N, D, routes, R_pool, request_flow_t, check_start_position, Tem, weight, r, miu1, miu2, miu3, operations, pai, theta, obj_record, all_routes, repeat, request_number, vehicle_number, transshipment_time, service_time, transshipment_cost_per, path, current_save, first_time_random, transshipment_insert_number, R_pool_2v, R_pool_3v, fuel_cost, segment_number, r, miu1, miu1, miu1, Fixed, percentage, all_ok_TK, k_random_or, removal_heuristic, insert_heuristic, theta_insert, theta_removal, weight_insertion, weight_removal, storage, fixed_vehicles
    # I set the r which can be served by barge, not consider T, but if there is not enough barges, r can only be served by T, then maybe infeasible and solution is expensive because uses train or truck
    # but in greedy, it can use other mode or T
    # if CP_try_r_of_other_carriers == 0:

    R_pool = R_pool2
    parallel_number = parallel_number2
    no_T_R = []
    hash_df_table = {}
    hash_top = {}
    hash_top_R_pool = {}
    # hash_overall_obj_table = {}
    # r_best_obj_record = pd.DataFrame(columns=['cost','k','T','r'], index=range(len(R_change)))

    hash_table_route_no_columns = {}
    hash_table_route_no_columns_top = {}
    SA = SA2
    combination = combination2

    k_random_or = k_random_or2
    transshipment_insert_number = 0

    all_pro = []

    hash_table_1v = {}
    hash_table_1v_all = {}
    hash_table_2v = {}
    hash_table_2v_all = {}
    hash_table_3v = {}
    hash_table_3v_all = {}

    hash_table_1v_fail = {}
    hash_table_1v_all_fail = {}
    hash_table_2v_fail = {}
    hash_table_2v_all_fail = {}
    hash_table_3v_fail = {}
    hash_table_3v_all_fail = {}

    storage = {}

    percentage = percentage2

    Fixed = Fixed2
    fixed_vehicles_percentage = read_Fixed(request_number_in_R, percentage, Fixed)
    N = N2

    only_T = only_T2

    fuel_cost = fuel_cost2

    has_end_depot = has_end_depot2
    path = path2
    current_save = current_save2
    transshipment_cost_per = transshipment_cost_p
    T_or = T_or_not
    transshipment_time = transshipment_time2
    service_time = service_time2
    request_number = j
    vehicle_number = i
    T = copy.copy(T_change)
    K = copy.copy(K_change)
    o = copy.copy(o_change)
    R = copy.copy(R_change)
    if dynamic == 1 and dynamic_t == 0:
        global dynamic_start_time
        dynamic_start_time = timeit.default_timer()
    if dynamic == 1 and dynamic_t > 0:
        routes, R_pool, R = prepare_for_dynamic()




        if stochastic == 1:
            #save routes and R_pool and objectives (obj_record), and also routes_match in the file for this time step
            #until now, I didn't save R_pool because all r can be served
            # 计算运行时间和目标值
            running_time = timeit.default_timer() - dynamic_start_time
            overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost, overall_transshipment_cost, overall_un_load_cost, overall_emission_cost, overall_storage_cost, overall_delay_penalty, overall_number_transshipment, overall_average_speed, overall_average_time_ratio, overall_emission_transshipment = overall_obj(
                routes)
            obj_record = pd.DataFrame(index=range(0, 1),
                                      columns=['overall_distance', 'overall_cost', 'overall_time', 'overall_profit',
                                               'overall_emission', 'served_requests', 'overall_request_cost',
                                               'overall_vehicle_cost', 'overall_wait_cost',
                                               'overall_transshipment_cost',
                                               'overall_un_load_cost', 'overall_emission_cost', 'overall_storage_cost',
                                               'overall_delay_penalty', 'running_time'])
            obj_record.loc[0] = [overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost, overall_transshipment_cost, overall_un_load_cost, overall_emission_cost, overall_storage_cost, overall_delay_penalty, running_time]

            # 保存结果，确保总是包含obj_record
            if add_RL == 1 and dynamic_RL34959.implement == 0:
                save_results(-1, routes, obj_record)
            else:
                # it has the last round in training will be saved's risk, but it does not matter
            #     if skip_the_last_training_round_and_implement_is_one == 0:
            #         save_results(-1, routes, obj_record)
            #         skip_the_last_training_round_and_implement_is_one = 1
            #     else:
                save_results(-1, routes, obj_record)
            if Dynamic_ALNS_RL34959.dynamic_end == 1:
                # restore routes and R_pool when all uncertainty finisheds to test RL and let the optimization never end

                D, routes, R_pool_2v, R_pool_3v, no_route_barge, no_route_truck, D_origin_All = read_data()
                R_pool = copy.copy(R)
            return -1, -1, -1, -1
    r_best_obj_record = np.array(np.empty(shape=(len(R), 4)), dtype='object')
    r_best_obj_record[:] = np.NaN
    r_best_obj_record[:, 3] = range(len(R))
    request_flow_t = np.array(np.empty(shape=(len(R), 6)))
    request_flow_t[:] = np.NaN
    # T_k_record = pd.DataFrame(columns=['T1', 'T2', 'k1', 'k2', 'k3'], index=R[:,7])
    T_k_record = np.array(np.empty(shape=(len(R), 5)), dtype='object')
    T_k_record[:] = np.NaN

    D, routes, R_pool_2v, R_pool_3v, no_route_barge, no_route_truck, D_origin_All = read_data()
    revert_K = read_R_K(request_number_in_R, what='revert_K')
    bundle_R = bundle()
    # request_flow_t = pd.DataFrame(index=R[:,7],
    #                               columns=['pickup', 'Td', 'Tp', 'secondTd', 'secondTp', 'delivery'])

    first_time_random = 0
    # danger this should be changed depend on instance
    train_truck = range(len(K))[49:]

    # sort r depending on r is hard to insert or not
    hard_value = pd.DataFrame(columns=['hard_value'], index=R[:,7])
    real_hard_value = pd.DataFrame(columns=['distance', 'time', 'load'], index=R[:,7])
    gamma_distance, gamma_time, gamma_load = 0.5, 0.2, 0.3

    for r in R[:,7]:
        index_r = list(R[:,7]).index(r)

        real_hard_value['distance'][r] = D_origin_All[int(R[index_r, 0])][int(R[index_r, 1])]
        real_hard_value['time'][r] = abs(R[index_r, 3] - R[index_r, 2]) + abs(R[index_r, 5] - R[index_r, 4])
        real_hard_value['load'][r] = R[index_r, 6]

    max_distance = max(real_hard_value['distance'])
    max_time = max(real_hard_value['time'])
    max_load = max(real_hard_value['load'])

    for r in R[:,7]:
        hard_value['hard_value'][r] = gamma_distance * real_hard_value['distance'][r] / max_distance + \
                                      gamma_time / real_hard_value['time'][r] / max_time + \
                                      gamma_load * real_hard_value['load'][r] / max_load

    if check_obj == 0:
        ok_K_canpickr = func_ok_K_canpickr()

        K_R = get_K_R()

        all_ok_TK = {}
        for r in R[:,7]:
            all_ok_TK[r] = ok_TK(r)
    else:
        ok_K_canpickr, K_R, all_ok_TK = 0, 0, 0

    # when a heuristic was added, just add its name in the following statement, then all work will be done automatically
    if T_or == 1:
        if only_T == 1:
            insert_heuristic = pd.DataFrame(['transshipment_insert'], columns=['operator'])
        else:
            if get_initial_bymyself == 1:
                if request_number_in_R >= 200:
                    # insert_heuristic = pd.DataFrame(
                    #     ['most_hard_first_insert', 'global_real_greedy_insert_regret', 'real_greedy_insert',
                    #      'greedy_insert', 'random_insert'], columns=['operator'])
                    # insert_heuristic = pd.DataFrame(
                    #     ['most_hard_first_insert', 'global_real_greedy_insert_regret', 'real_greedy_insert',
                    #      'greedy_insert'], columns=['operator'])
                    insert_heuristic = pd.DataFrame(['greedy_insert'], columns=['operator'])
                else:
                    # insert_heuristic = pd.DataFrame(
                    #     ['most_hard_first_insert', 'global_real_greedy_insert_regret', 'real_greedy_insert',
                    #      'greedy_insert', 'random_insert'], columns=['operator'])
                    insert_heuristic = pd.DataFrame(['greedy_insert'], columns=['operator'])
            else:
                if request_number_in_R >= 200:
                    insert_heuristic = pd.DataFrame(
                        ['most_hard_first_insert', 'global_real_greedy_insert_regret', 'real_greedy_insert',
                         'greedy_insert', 'random_insert'], columns=['operator'])
                else:
                    insert_heuristic = pd.DataFrame(
                        ['most_hard_first_insert', 'global_real_greedy_insert_regret', 'real_greedy_insert',
                         'greedy_insert', 'transshipment_insert', 'random_insert'], columns=['operator'])
        #                insert_heuristic = pd.DataFrame(['real_greedy_insert'], columns=['operator'])
    # 'global_real_greedy_insert',
    # 'real_greedy_insert',
    #            insert_heuristic=pd.DataFrame(['greedy_insert', 'transshipment_insert'], columns=['operator'])
    else:
        insert_heuristic = pd.DataFrame(['greedy_insert', 'random_insert'], columns=['operator'])

    # insert_heuristic=pd.DataFrame(['real_greedy_insert'], columns=['operator'])

    #    insert_heuristic=pd.DataFrame(['greedy_insert', 'transshipment_insert'], columns=['operator'])
    if get_initial_bymyself == 1:
        # removal_heuristic = pd.DataFrame(
        #     ['history_removal', 'related_removal', 'clear_a_route', 'random_removal', 'worst_removal',
        #      'delete_node'], columns=['operator'])
        removal_heuristic=pd.DataFrame(['random_removal'], columns=['operator'])
    else:
        removal_heuristic = pd.DataFrame(
            ['history_removal', 'related_removal', 'clear_a_route', 'random_removal', 'worst_removal', 'delete_node'],
            columns=['operator'])
    #        removal_heuristic = pd.DataFrame(['delete_node'], columns=['operator'])
    if by_wenjing == 1:
        wenjing_best_time = {5:0.28,10:0.80,20:0.65,30:0.94,50:2.83,100:4.09,200:9.07,400:29.06,700:38.43,1000:78.94,1300:158.57,1600:302.41}

    initial_solution_no_wait_cost = 0
    start_initial = timeit.default_timer()

    routes, R_pool = initial_solution()
    if dynamic == 1 and dynamic_t > 0:
        pass
    else:
        if get_initial_bymyself == 1:
            if swap_or_not == 1:
                routes,R_pool = special_swap()
                print('swap_cost', overall_obj(routes)[1])
        # routes,R_pool = special_swap()
        # print('swap_cost', overall_obj(routes)[1])
    remove_request_due_to_load_dependent()
    if get_initial_bymyself == 1:
        running_time_initial = timeit.default_timer() - start_initial
    else:
        if by_wenjing == 0:
            running_time_initial = Best_Running_Time_as_initial
        else:
            running_time_initial = wenjing_best_time[request_number_in_R]

    print('running time of initial solution: ', running_time_initial)

    initial_distance, initial_cost, initial_time, initial_profit, initial_emission, initial_requests, overall_request_cost, initial_vehicle_cost, initial_wait_cost, initial_transshipment_cost, initial_un_load_cost, initial_emission_cost, initial_storage_cost, initial_delay_penalty, initial_overall_number_transshipment, initial_overall_average_speed, initial_overall_average_time_ratio, initial_overall_emission_transshipment = overall_obj(
        routes)
    print(initial_cost, initial_time, initial_profit, initial_requests)
    initial_solution_no_wait_cost = 0

    all_Tem = []
    if SA == 1:
        Tem2 = Symbol('Tem2')
        w = 1.4
        # if multi_obj == 0:
        # Tem = solve(exp(-int(w * initial_cost-initial_cost)/Tem2)-0.5)[0]
        if multi_obj == 0:
            Tem = initial_cost / 10
        else:
            if bi_obj_cost_emission == 1:
                Tem = (initial_cost + initial_emission) / 10
            elif  bi_obj_cost_time == 1:
                Tem = (initial_cost + initial_time) / 10
            else:
                Tem = (initial_cost + initial_time + initial_emission) / 10
        # else:
        #
        #     Tem = solve(exp(-int(w -1)/Tem2)-0.5)[0]
        #
        #         Tem = initial_cost/10
        all_Tem.append(Tem)
        # iteration_number = 100
        # Tem = 10

    segment_number = segment_number2
    if combination == 1:
        operations = pd.DataFrame(index=range(len(insert_heuristic) * len(removal_heuristic)),
                                  columns=['operation', 'insertion', 'removal'])
        operation_number = 0
        for x in insert_heuristic['operator']:
            for j in removal_heuristic['operator']:
                operations.iloc[operation_number] = [x + '_' + j, x, j]
                operation_number = operation_number + 1
        weight = pd.DataFrame(index=range(0, int(iteration_number / segment_number)), columns=operations['operation'])
        # score
        pai = pd.DataFrame(index=range(0, iteration_number), columns=operations['operation'])
        theta = pd.DataFrame(0, index=range(0, len(insert_heuristic) * len(removal_heuristic)), columns=['theta'])
        theta['theta'] = 0
    else:
        weight_insertion = pd.DataFrame(index=range(0, int(iteration_number / segment_number)),
                                        columns=insert_heuristic['operator'])
        weight_removal = pd.DataFrame(index=range(0, int(iteration_number / segment_number)),
                                      columns=removal_heuristic['operator'])
        # score
        pai = pd.DataFrame(index=range(0, iteration_number),
                           columns=insert_heuristic['operator'].append(removal_heuristic['operator']))
        theta_insert = pd.DataFrame(0, index=range(0, len(insert_heuristic)), columns=['theta'])
        theta_removal = pd.DataFrame(0, index=range(0, len(removal_heuristic)), columns=['theta'])
        theta_insert['theta'] = 0
        theta_removal['theta'] = 0
    #        insert_operators = pd.DataFrame(index=range(len(insert_heuristic)), columns=['insert'])
    #        removal_operators = pd.DataFrame(index=range(len(removal_heuristic)), columns=['removal'])
    #        insert_number = 0
    #        removal_number = 0
    #        for i in insert_heuristic['insert_heuristic']:
    #            insert_operators.iloc[insert_number] = [i]
    #            insert_number = insert_number + 1
    #        for j in removal_heuristic['removal_heuristic']:
    #            removal_operators.iloc[removal_number] = [j]
    #            removal_number = removal_number + 1

    r = r2

    miu1 = miu1_1
    miu2 = miu2_1
    miu3 = miu3_1

    for j in range(0, iteration_number):
        if j % segment_number == 0:
            pai.iloc[j] = 0
    if CP == 1:
        if heterogeneous_preferences == 1:
            obj_record = pd.DataFrame(index=range(0, iteration_number),
                                      columns=['overall_distance', 'overall_cost', 'overall_time', 'overall_profit',
                                               'overall_emission', 'served_requests', 'overall_request_cost',
                                               'overall_vehicle_cost', 'overall_wait_cost',
                                               'overall_transshipment_cost',
                                               'overall_un_load_cost', 'overall_emission_cost', 'overall_storage_cost',
                                               'overall_delay_penalty', 'iteration_time', 'barge_served_requests',
                                               'train_served_requests', 'truck_served_requests', 'satisfactory_value',
                                                   'fuzzy_satisfy_or_not', 'hard_satisfy_or_not', 'overall_number_transshipment', 'overall_average_time_ratio', 'cost_per_container_per_km', 'time_ratio', 'emissions_per_container_per_km', 'delay_time_ratio', 'transshipment_times', 'overall_emission_transshipment'])
        else:
            obj_record = pd.DataFrame(index=range(0, iteration_number),
                     columns=['overall_distance', 'overall_cost', 'overall_time', 'overall_profit',
                              'overall_emission', 'served_requests', 'overall_request_cost',
                              'overall_vehicle_cost', 'overall_wait_cost',
                              'overall_transshipment_cost',
                              'overall_un_load_cost', 'overall_emission_cost', 'overall_storage_cost',
                              'overall_delay_penalty', 'iteration_time', 'barge_served_requests',
                              'train_served_requests', 'truck_served_requests'])
    else:
        if heterogeneous_preferences == 1:
            if use_speed == 1:
                obj_record = pd.DataFrame(index=range(0, iteration_number),
                                          columns=['overall_distance', 'overall_cost', 'overall_time', 'overall_profit',
                                                   'overall_emission', 'served_requests', 'overall_request_cost',
                                                   'overall_vehicle_cost', 'overall_wait_cost', 'overall_transshipment_cost',
                                                   'overall_un_load_cost', 'overall_emission_cost', 'overall_storage_cost',
                                                   'overall_delay_penalty', 'iteration_time', 'satisfactory_value',
                                                   'fuzzy_satisfy_or_not', 'hard_satisfy_or_not', 'overall_number_transshipment', 'overall_average_speed'])
            else:
                obj_record = pd.DataFrame(index=range(0, iteration_number),
                                          columns=['overall_distance', 'overall_cost', 'overall_time', 'overall_profit',
                                                   'overall_emission', 'served_requests', 'overall_request_cost',
                                                   'overall_vehicle_cost', 'overall_wait_cost',
                                                   'overall_transshipment_cost',
                                                   'overall_un_load_cost', 'overall_emission_cost', 'overall_storage_cost',
                                                   'overall_delay_penalty', 'iteration_time', 'satisfactory_value',
                                                   'fuzzy_satisfy_or_not', 'hard_satisfy_or_not',
                                                   'overall_number_transshipment', 'overall_average_time_ratio', 'cost_per_container_per_km', 'time_ratio', 'emissions_per_container_per_km', 'delay_time_ratio', 'transshipment_times', 'overall_emission_transshipment'])
        else:
            obj_record = pd.DataFrame(index=range(0, iteration_number),
                                      columns=['overall_distance', 'overall_cost', 'overall_time', 'overall_profit',
                                               'overall_emission', 'served_requests', 'overall_request_cost',
                                               'overall_vehicle_cost', 'overall_wait_cost', 'overall_transshipment_cost',
                                               'overall_un_load_cost', 'overall_emission_cost', 'overall_storage_cost',
                                               'overall_delay_penalty', 'iteration_time'])

    if heterogeneous_preferences == 1:
        if use_speed == 1:
            obj_record_better = pd.DataFrame(index=range(0, iteration_number),
                                             columns=['overall_cost', 'overall_time', 'overall_profit', 'served_requests',
                                                      'iteration_time', 'satisfactory_value', 'fuzzy_satisfy_or_not',
                                                      'hard_satisfy_or_not', 'overall_number_transshipment', 'overall_average_speed'])
        else:
            obj_record_better = pd.DataFrame(index=range(0, iteration_number),
                                             columns=['overall_cost', 'overall_time', 'overall_profit',
                                                      'served_requests',
                                                      'iteration_time', 'satisfactory_value', 'fuzzy_satisfy_or_not',
                                                      'hard_satisfy_or_not', 'overall_number_transshipment',
                                                      'overall_average_time_ratio'])
    else:
        obj_record_better = pd.DataFrame(index=range(0, iteration_number),
                                         columns=['overall_cost', 'overall_time', 'overall_profit', 'served_requests',
                                                  'iteration_time'])

    all_routes = {}
    all_routes[0] = my_deepcopy(routes)
    # initial solution's obj
    if CP == 1:

        barge_served_requests, train_served_requests, truck_served_requests = CP_served_requests_mode()
        if heterogeneous_preferences == 1:
            satisfactory_value, fuzzy_satisfy_or_not, hard_satisfy_or_not = overall_satisfactory_values(routes)
            print(satisfactory_value, fuzzy_satisfy_or_not, hard_satisfy_or_not)
            cost_per_container_per_km, time_ratio, emissions_per_container_per_km, delay_time_ratio, transshipment_times = overall_satisfactory_values(
                routes, 1)
            obj_record.iloc[0] = [initial_distance, initial_cost, initial_time, initial_profit, initial_emission,
                                  initial_requests, overall_request_cost, initial_vehicle_cost, initial_wait_cost,
                                  initial_transshipment_cost, initial_un_load_cost, initial_emission_cost,
                                  initial_storage_cost,
                                  initial_delay_penalty, running_time_initial, barge_served_requests,
                                  train_served_requests, truck_served_requests, satisfactory_value, fuzzy_satisfy_or_not,
                                      hard_satisfy_or_not, initial_overall_number_transshipment,
                                      initial_overall_average_time_ratio, cost_per_container_per_km, time_ratio, emissions_per_container_per_km, delay_time_ratio, transshipment_times, initial_overall_emission_transshipment]
        else:
            obj_record.iloc[0] = [initial_distance, initial_cost, initial_time, initial_profit, initial_emission,
                          initial_requests, overall_request_cost, initial_vehicle_cost, initial_wait_cost,
                          initial_transshipment_cost, initial_un_load_cost, initial_emission_cost,
                          initial_storage_cost,
                          initial_delay_penalty, running_time_initial, barge_served_requests, train_served_requests, truck_served_requests]
    else:
        if heterogeneous_preferences == 1:
            satisfactory_value, fuzzy_satisfy_or_not, hard_satisfy_or_not = overall_satisfactory_values(routes)
            print(satisfactory_value, fuzzy_satisfy_or_not, hard_satisfy_or_not)
            cost_per_container_per_km, time_ratio, emissions_per_container_per_km, delay_time_ratio, transshipment_times = overall_satisfactory_values(
                routes, 1)
            if use_speed == 1:
                obj_record.iloc[0] = [initial_distance, initial_cost, initial_time, initial_profit, initial_emission,
                                      initial_requests, overall_request_cost, initial_vehicle_cost, initial_wait_cost,
                                      initial_transshipment_cost, initial_un_load_cost, initial_emission_cost,
                                      initial_storage_cost,
                                      initial_delay_penalty, running_time_initial, satisfactory_value, fuzzy_satisfy_or_not,
                                      hard_satisfy_or_not, initial_overall_number_transshipment, initial_overall_average_speed]
            else:
                obj_record.iloc[0] = [initial_distance, initial_cost, initial_time, initial_profit, initial_emission,
                                      initial_requests, overall_request_cost, initial_vehicle_cost, initial_wait_cost,
                                      initial_transshipment_cost, initial_un_load_cost, initial_emission_cost,
                                      initial_storage_cost,
                                      initial_delay_penalty, running_time_initial, satisfactory_value, fuzzy_satisfy_or_not,
                                      hard_satisfy_or_not, initial_overall_number_transshipment,
                                      initial_overall_average_time_ratio, cost_per_container_per_km, time_ratio, emissions_per_container_per_km, delay_time_ratio, transshipment_times, initial_overall_emission_transshipment]
        else:
            # overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests, overall_request_cost, overall_vehicle_cost,overall_wait_cost,overall_transshipment_cost,overall_un_load_cost, overall_emission_cost,overall_storage_cost,overall_delay_penalty = overall_obj(routes)
            obj_record.iloc[0] = [initial_distance, initial_cost, initial_time, initial_profit, initial_emission,
                                  initial_requests, overall_request_cost, initial_vehicle_cost, initial_wait_cost,
                                  initial_transshipment_cost, initial_un_load_cost, initial_emission_cost,
                                  initial_storage_cost,
                                  initial_delay_penalty, running_time_initial]

    lowest_cost = initial_cost
    routes_lowest_cost = my_deepcopy(routes)
    times_of_compare_cost_obj = 0
    start_time = process_time()
    start = timeit.default_timer()
    for repeat in range(1, iteration_number):
        print('iteration', repeat)
        old_routes = my_deepcopy(routes)
        old_R_pool = R_pool.copy()

        old_overall_distance, old_overall_cost, old_overall_time, old_overall_profit, old_overall_emission, old_served_requests = \
            obj_record.iloc[repeat - 1][0:6]

        #        routes, R_pool = random_removal()

        routes, R_pool = Adaptive()
        if dynamic == 1 and dynamic_t > 0:
            pass
        else:
            if swap_or_not == 1:
                routes,R_pool = special_swap()
        remove_request_due_to_load_dependent()
        # iteration_time = timeit.default_timer() - start + running_time_initial
        iteration_time = timeit.default_timer() - start
        if swap_or_not == 1:
            print('swap_cost', overall_obj(routes)[1])
        overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost, overall_transshipment_cost, overall_un_load_cost, overall_emission_cost, overall_storage_cost, overall_delay_penalty, overall_number_transshipment, overall_average_speed, overall_average_time_ratio, overall_emission_transshipment = overall_obj(
            routes)
        only_minimize_cost = 0
        if only_minimize_cost == 1:
            if overall_cost < lowest_cost:
                routes_lowest_cost = my_deepcopy(routes)
                lowest_cost = overall_cost
        else:
            if (served_requests == old_served_requests and overall_cost < lowest_cost) or served_requests > old_served_requests:
                routes_lowest_cost = my_deepcopy(routes)
                lowest_cost = overall_cost
        # accept the solution depend on prbability given by simulated annealing
        if SA == 1:
            Tem = Tem * c
            all_Tem.append(Tem)
            if multi_obj == 0:

                if only_minimize_cost == 1:
                    if overall_cost - old_overall_cost > 0:
                        pro = np.exp(float(-(overall_cost - old_overall_cost) / devide_value / Tem))
                    else:
                        pro = 1
                else:
                    if served_requests > old_served_requests:
                        pro = 1
                    else:
                        if served_requests < old_served_requests:
                            # pro = 0.1
                            pro = 0
                        else:
                            times_of_compare_cost_obj = times_of_compare_cost_obj + 1
                            if overall_cost - old_overall_cost > 0:
                                pro = np.exp(float(-(overall_cost - old_overall_cost) / devide_value / Tem))
                            else:
                                pro = 1
            else:
                sum_g_or_delta_obj = dominate_1(overall_distance, overall_cost, overall_time, overall_profit,
                                                overall_emission,
                                                served_requests, old_overall_distance, old_overall_cost,
                                                old_overall_time,
                                                old_overall_profit, old_overall_emission, old_served_requests)
                if sum_g_or_delta_obj is not True:
                    if weight_interval == 1:
                        pro = np.exp(float(- sum_g_or_delta_obj * (
                                overall_cost + overall_time + overall_emission) / devide_value / Tem))
                    else:
                        pro = np.exp(float(- sum_g_or_delta_obj / devide_value / Tem))
                else:
                    pro = 1
                #
                # overall_cost_norm = normalization(overall_cost, 'overall_cost')
                # old_overall_cost_norm = normalization(old_overall_cost, 'overall_cost')
                # overall_time_norm = normalization(overall_time, 'overall_time')
                # old_overall_time_norm = normalization(old_overall_time, 'overall_time')
                # overall_emission_norm = normalization(overall_emission, 'overall_emission')
                # old_overall_emission_norm = normalization(old_overall_emission, 'overall_emission')
                # if weight_interval == 1:
                #     weight_cost = (weight_max_cost + weight_min_cost) / 2
                #     weight_time = (weight_max_time + weight_min_time)/2
                #     weight_emission = (weight_max_emission + weight_min_emission) / 2
                # if bi_obj_cost_emission == 1:
                #     weight_sum_obj = weight_cost * overall_cost_norm + weight_emission * overall_emission_norm
                #     old_weight_sum_obj = weight_cost * old_overall_cost_norm + weight_emission  * old_overall_emission_norm
                #     pro = np.exp(float(-(weight_sum_obj - old_weight_sum_obj) * (initial_cost + initial_emission) /10 / Tem))
                # else:
                #     weight_sum_obj = weight_cost * overall_cost_norm + weight_time * overall_time_norm + weight_emission * overall_emission_norm
                #     old_weight_sum_obj = weight_cost * old_overall_cost_norm + weight_time * old_overall_time_norm + weight_emission * old_overall_emission_norm
                #     pro = np.exp(float(-(weight_sum_obj - old_weight_sum_obj) * (initial_cost + initial_time + initial_emission) /10 / Tem))
                #
        else:
            pro = 0.5
        if pro > 1:
            pro = 1
        all_pro.append(pro)
        if SA == 1:
            print('Acceptance Probability:' + str(pro))
            print('Temperature:' + str(Tem))

        if pro != 1:
            number = int(np.random.choice([1, 2], size=(1,), p=[pro, 1 - pro]))
            if number == 2:
                # if overall_cost > old_overall_cost or served_requests < old_served_requests:
                routes = my_deepcopy(old_routes)
                R_pool = old_R_pool.copy()
                overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost, overall_transshipment_cost, overall_un_load_cost, overall_emission_cost, overall_storage_cost, overall_delay_penalty = \
                    obj_record.iloc[repeat - 1][0:14]
        all_routes[repeat] = my_deepcopy(routes)

        print(overall_cost, overall_time, overall_profit, served_requests)
        # if overall_cost < 44197:
        #     print('asdf')
        if CP == 1:
            barge_served_requests, train_served_requests, truck_served_requests = CP_served_requests_mode()
            if heterogeneous_preferences == 0:

                obj_record.iloc[repeat] = [overall_distance, overall_cost, overall_time, overall_profit, overall_emission,
                                           served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost,
                                           overall_transshipment_cost, overall_un_load_cost, overall_emission_cost,
                                           overall_storage_cost, overall_delay_penalty, iteration_time, barge_served_requests, train_served_requests, truck_served_requests]
            else:
                satisfactory_value, fuzzy_satisfy_or_not, hard_satisfy_or_not = overall_satisfactory_values(routes)
                cost_per_container_per_km, time_ratio, emissions_per_container_per_km, delay_time_ratio, transshipment_times = overall_satisfactory_values(
                    routes, 1)
                print(satisfactory_value, fuzzy_satisfy_or_not, hard_satisfy_or_not)
                obj_record.iloc[repeat] = [overall_distance, overall_cost, overall_time, overall_profit,
                                           overall_emission,
                                           served_requests, overall_request_cost, overall_vehicle_cost,
                                           overall_wait_cost,
                                           overall_transshipment_cost, overall_un_load_cost, overall_emission_cost,
                                           overall_storage_cost, overall_delay_penalty, iteration_time, barge_served_requests, train_served_requests, truck_served_requests,
                                           satisfactory_value,
                                           fuzzy_satisfy_or_not, hard_satisfy_or_not, overall_number_transshipment,
                                           overall_average_time_ratio, cost_per_container_per_km, time_ratio,
                                           emissions_per_container_per_km, delay_time_ratio, transshipment_times, overall_emission_transshipment]
        else:
            if heterogeneous_preferences == 1:
                satisfactory_value, fuzzy_satisfy_or_not, hard_satisfy_or_not = overall_satisfactory_values(routes)
                cost_per_container_per_km, time_ratio, emissions_per_container_per_km, delay_time_ratio, transshipment_times = overall_satisfactory_values(routes, 1)
                print(satisfactory_value, fuzzy_satisfy_or_not, hard_satisfy_or_not)
                if use_speed == 1:
                    obj_record.iloc[repeat] = [overall_distance, overall_cost, overall_time, overall_profit, overall_emission,
                                               served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost,
                                               overall_transshipment_cost, overall_un_load_cost, overall_emission_cost,
                                               overall_storage_cost, overall_delay_penalty, iteration_time, satisfactory_value,
                                               fuzzy_satisfy_or_not, hard_satisfy_or_not, overall_number_transshipment, overall_average_speed]
                else:
                    obj_record.iloc[repeat] = [overall_distance, overall_cost, overall_time, overall_profit,
                                               overall_emission,
                                               served_requests, overall_request_cost, overall_vehicle_cost,
                                               overall_wait_cost,
                                               overall_transshipment_cost, overall_un_load_cost, overall_emission_cost,
                                               overall_storage_cost, overall_delay_penalty, iteration_time,
                                               satisfactory_value,
                                               fuzzy_satisfy_or_not, hard_satisfy_or_not, overall_number_transshipment,
                                               overall_average_time_ratio, cost_per_container_per_km, time_ratio, emissions_per_container_per_km, delay_time_ratio, transshipment_times, overall_emission_transshipment]
            else:
                obj_record.iloc[repeat] = [overall_distance, overall_cost, overall_time, overall_profit, overall_emission,
                                           served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost,
                                           overall_transshipment_cost, overall_un_load_cost, overall_emission_cost,
                                           overall_storage_cost, overall_delay_penalty, iteration_time]

        if timeit.default_timer() - start > stop_time * 3600:
            break
    Running_Time = timeit.default_timer() - start
    CPU_Time = process_time() - start_time
    print('Running_Time ', Running_Time + running_time_initial, 'CPU_Time ', CPU_Time + running_time_initial)
    print('times_of_compare_cost_obj', times_of_compare_cost_obj)
    for x in range(1, repeat + 2):
        obj_record_tem = obj_record.iloc[0:x]
        obj_record_better_find = obj_record_tem.loc[
            obj_record_tem['served_requests'] == obj_record_tem['served_requests'].max()]
        obj_record_better_find = obj_record_better_find.loc[
            obj_record_better_find['overall_cost'] == obj_record_better_find['overall_cost'].min()]
        obj_record_better.iloc[x - 1] = obj_record_better_find.iloc[0][obj_record_better.columns]

    # if multi_obj == 0:
    obj_record_best = obj_record.loc[obj_record['served_requests'] == obj_record['served_requests'].max()]
    if Demir == 1:
        obj_record_best['real_cost'] = obj_record_best['overall_delay_penalty'] + obj_record_best['overall_request_cost'] + obj_record_best['overall_emission_cost'] + obj_record_best['overall_un_load_cost'] + obj_record_best['overall_transshipment_cost']
        obj_record_best = obj_record_best.loc[obj_record_best['overall_cost'] == obj_record_best['overall_cost'].min()]
        obj_record_best.sort_values('real_cost', inplace=True)
    else:
        obj_record_best = obj_record_best.loc[obj_record_best['overall_cost'] == obj_record_best['overall_cost'].min()]
    obj_record_best['iteration_time'] = obj_record_best['iteration_time']
    if obj_record_best.iloc[0][1] <= initial_cost + 0.1 and obj_record_best.iloc[0][1] >= initial_cost - 0.1:
        Best_Running_Time = 0
    else:
        Best_Running_Time = obj_record_best.iloc[0]['iteration_time']
    print(obj_record_best)
    if not os.path.isdir(path + current_save):
        Path(path + current_save).mkdir(parents=True, exist_ok=True)

    number_used_vehicles, all_number, barge_seved_r_portion, train_seved_r_portion, truck_seved_r_portion = get_mode_share(obj_record_best.index[0])
    if not os.path.isfile(exps_record_path):
        if CP == 1:
            if heterogeneous_preferences == 1:
                exps_record = pd.DataFrame(
                        columns=['exp_number', 'parallel_number', 'obj_number', 'r', 'served_r', 'k', 'T', 'iterations',
                                 'stop_iterations',
                                 'segement', 'c', 'percentage', 'bundle_or_not', 'best_iteration_number', 'best_cost',
                                 'best_request_cost', 'best_vehicle_cost', 'best_wait_cost', 'best_transshipment_cost',
                                 'best_unload_cost', 'best_emission_cost', 'best_storage_cost', 'best_delay_penalty',
                                 'best_time',
                                 'total_time', 'initial_time', 'initial_cost', 'add_initial_best_time',
                                 'add_initial_total_time', 'nodes', 'processors', 'get_initial_bymyself', 'by_wenjing',
                                 'number_used_vehicles',
                                 'barge_seved_r_portion', 'train_seved_r_portion', 'truck_seved_r_portion', 'note', 'barge_served_requests', 'train_served_requests', 'truck_served_requests',
                                 'satisfactory_value', 'fuzzy_satisfy_or_not', 'hard_satisfy_or_not',
                                 'overall_number_transshipment', 'overall_average_time_ratio', 'cost_per_container_per_km', 'time_ratio', 'emissions_per_container_per_km', 'delay_time_ratio', 'transshipment_times', 'heterogeneous_preferences_no_constraints', 'heterogeneous_preferences', 'fuzzy_constraints', 'CP', 'overall_emission_transshipment'])
            else:
                exps_record = pd.DataFrame(
                    columns=['exp_number', 'parallel_number', 'obj_number', 'r', 'k', 'T', 'iterations', 'stop_iterations',
                             'segement', 'c', 'percentage', 'bundle_or_not', 'best_iteration_number', 'best_cost',
                             'best_request_cost', 'best_vehicle_cost', 'best_wait_cost', 'best_transshipment_cost',
                             'best_unload_cost', 'best_emission_cost', 'best_storage_cost', 'best_delay_penalty',
                             'best_time',
                             'total_time', 'initial_time', 'initial_cost', 'add_initial_best_time',
                             'add_initial_total_time', 'nodes', 'processors', 'get_initial_bymyself', 'by_wenjing',
                             'number_used_vehicles',
                             'barge_seved_r_portion', 'train_seved_r_portion', 'truck_seved_r_portion', 'note', 'barge_served_requests', 'train_served_requests', 'truck_served_requests'])
        else:
            if heterogeneous_preferences == 1:
                if use_speed == 1:
                    exps_record = pd.DataFrame(
                        columns=['exp_number', 'parallel_number', 'obj_number', 'r', 'k', 'T', 'iterations', 'stop_iterations',
                                 'segement', 'c', 'percentage', 'bundle_or_not', 'best_iteration_number', 'best_cost',
                                 'best_request_cost', 'best_vehicle_cost', 'best_wait_cost', 'best_transshipment_cost',
                                 'best_unload_cost', 'best_emission_cost', 'best_storage_cost', 'best_delay_penalty', 'best_time',
                                 'total_time', 'initial_time', 'initial_cost', 'add_initial_best_time', 'add_initial_total_time', 'nodes', 'processors', 'get_initial_bymyself', 'by_wenjing', 'number_used_vehicles',
                                 'barge_seved_r_portion', 'train_seved_r_portion', 'truck_seved_r_portion', 'note', 'satisfactory_value', 'fuzzy_satisfy_or_not', 'hard_satisfy_or_not', 'overall_number_transshipment','overall_average_speed'])
                else:
                    exps_record = pd.DataFrame(
                        columns=['exp_number', 'parallel_number', 'obj_number', 'r', 'served_r', 'k', 'T', 'iterations',
                                 'stop_iterations',
                                 'segement', 'c', 'percentage', 'bundle_or_not', 'best_iteration_number', 'best_cost',
                                 'best_request_cost', 'best_vehicle_cost', 'best_wait_cost', 'best_transshipment_cost',
                                 'best_unload_cost', 'best_emission_cost', 'best_storage_cost', 'best_delay_penalty',
                                 'best_time',
                                 'total_time', 'initial_time', 'initial_cost', 'add_initial_best_time',
                                 'add_initial_total_time', 'nodes', 'processors', 'get_initial_bymyself', 'by_wenjing',
                                 'number_used_vehicles',
                                 'barge_seved_r_portion', 'train_seved_r_portion', 'truck_seved_r_portion', 'note',
                                 'satisfactory_value', 'fuzzy_satisfy_or_not', 'hard_satisfy_or_not',
                                 'overall_number_transshipment', 'overall_average_time_ratio', 'cost_per_container_per_km', 'time_ratio', 'emissions_per_container_per_km', 'delay_time_ratio', 'transshipment_times', 'heterogeneous_preferences_no_constraints', 'heterogeneous_preferences', 'fuzzy_constraints', 'CP', 'overall_emission_transshipment'])
            else:
                exps_record = pd.DataFrame(
                    columns=['exp_number', 'parallel_number', 'obj_number', 'r', 'k', 'T', 'iterations', 'stop_iterations',
                             'segement', 'c', 'percentage', 'bundle_or_not', 'best_iteration_number', 'best_cost',
                             'best_request_cost', 'best_vehicle_cost', 'best_wait_cost', 'best_transshipment_cost',
                             'best_unload_cost', 'best_emission_cost', 'best_storage_cost', 'best_delay_penalty',
                             'best_time',
                             'total_time', 'initial_time', 'initial_cost', 'add_initial_best_time',
                             'add_initial_total_time', 'nodes', 'processors', 'get_initial_bymyself', 'by_wenjing',
                             'number_used_vehicles',
                             'barge_seved_r_portion', 'train_seved_r_portion', 'truck_seved_r_portion', 'note'])
    else:
        exps_record = pd.read_excel(exps_record_path, 'exps_record')
    # if obj_record_best.iloc[0][1] <= initial_cost + 0.1 and obj_record_best.iloc[0][1] >= initial_cost - 0.1:
    #     add_initial_best_time = Best_Running_Time
    # else:
    add_initial_best_time = Best_Running_Time + running_time_initial
    add_initial_total_time = Running_Time + running_time_initial
    Best_Running_Time_as_initial = add_initial_best_time
    if CP == 1:
        if heterogeneous_preferences == 1:
            served_r_number = check_served_R(1, all_routes[obj_record_best.index[0]])
            new_exp = [exp_number - 1, parallel_number, obj_number, request_number_in_R, served_r_number, k_number,
                       T_number,
                       iteration_number,
                       repeat, segment_number2, c,
                       percentage, bundle_or_not, obj_record_best.index[0], obj_record_best.iloc[0][1],
                       obj_record_best.iloc[0][6], obj_record_best.iloc[0][7], obj_record_best.iloc[0][8],
                       obj_record_best.iloc[0][9],
                       obj_record_best.iloc[0][10], obj_record_best.iloc[0][11], obj_record_best.iloc[0][12],
                       obj_record_best.iloc[0][13],
                       Best_Running_Time, Running_Time, running_time_initial, initial_cost, add_initial_best_time,
                       add_initial_total_time, node_number, processors_number, get_initial_bymyself,
                       by_wenjing, number_used_vehicles, barge_seved_r_portion, train_seved_r_portion,
                       truck_seved_r_portion,
                       note, barge_served_requests, train_served_requests, truck_served_requests, obj_record_best.iloc[0][15], obj_record_best.iloc[0][16], obj_record_best.iloc[0][17],
                       obj_record_best.iloc[0][18], obj_record_best.iloc[0][19], cost_per_container_per_km, time_ratio,
                       emissions_per_container_per_km, delay_time_ratio, transshipment_times,
                       heterogeneous_preferences_no_constraints, heterogeneous_preferences, fuzzy_constraints, CP, obj_record_best.iloc[0][28]]
        else:
            new_exp = [exp_number - 1, parallel_number, obj_number, request_number_in_R, k_number, T_number,
                       iteration_number,
                       repeat, segment_number2, c,
                       percentage, bundle_or_not, obj_record_best.index[0], obj_record_best.iloc[0][1],
                       obj_record_best.iloc[0][6], obj_record_best.iloc[0][7], obj_record_best.iloc[0][8],
                       obj_record_best.iloc[0][9],
                       obj_record_best.iloc[0][10], obj_record_best.iloc[0][11], obj_record_best.iloc[0][12],
                       obj_record_best.iloc[0][13],
                       Best_Running_Time, Running_Time, running_time_initial, initial_cost, add_initial_best_time,
                       add_initial_total_time, node_number, processors_number, get_initial_bymyself,
                       by_wenjing, number_used_vehicles, barge_seved_r_portion, train_seved_r_portion,
                       truck_seved_r_portion,
                       note, barge_served_requests, train_served_requests, truck_served_requests]
    else:
        if heterogeneous_preferences == 1:

            served_r_number = check_served_R(1, all_routes[obj_record_best.index[0]])
            new_exp = [exp_number - 1, parallel_number, obj_number, request_number_in_R, served_r_number, k_number, T_number,
                       iteration_number,
                       repeat, segment_number2, c,
                       percentage, bundle_or_not, obj_record_best.index[0], obj_record_best.iloc[0][1],
                       obj_record_best.iloc[0][6], obj_record_best.iloc[0][7], obj_record_best.iloc[0][8],
                       obj_record_best.iloc[0][9],
                       obj_record_best.iloc[0][10], obj_record_best.iloc[0][11], obj_record_best.iloc[0][12],
                       obj_record_best.iloc[0][13],
                       Best_Running_Time, Running_Time, running_time_initial, initial_cost, add_initial_best_time,
                       add_initial_total_time, node_number, processors_number, get_initial_bymyself,
                       by_wenjing, number_used_vehicles, barge_seved_r_portion, train_seved_r_portion,
                       truck_seved_r_portion,
                       note, obj_record_best.iloc[0][15], obj_record_best.iloc[0][16], obj_record_best.iloc[0][17], obj_record_best.iloc[0][18], obj_record_best.iloc[0][19], cost_per_container_per_km, time_ratio, emissions_per_container_per_km, delay_time_ratio, transshipment_times, heterogeneous_preferences_no_constraints, heterogeneous_preferences, fuzzy_constraints, CP, obj_record_best.iloc[0][25]]
        else:
            new_exp = [exp_number - 1, parallel_number, obj_number, request_number_in_R, k_number, T_number,
                       iteration_number,
                       repeat, segment_number2, c,
                       percentage, bundle_or_not, obj_record_best.index[0], obj_record_best.iloc[0][1],
                       obj_record_best.iloc[0][6], obj_record_best.iloc[0][7], obj_record_best.iloc[0][8],
                       obj_record_best.iloc[0][9],
                       obj_record_best.iloc[0][10], obj_record_best.iloc[0][11], obj_record_best.iloc[0][12],
                       obj_record_best.iloc[0][13],
                       Best_Running_Time, Running_Time, running_time_initial, initial_cost, add_initial_best_time,
                       add_initial_total_time, node_number, processors_number, get_initial_bymyself,
                       by_wenjing, number_used_vehicles, barge_seved_r_portion, train_seved_r_portion,
                       truck_seved_r_portion,
                       note]
    new_exp = pd.Series(new_exp, index=exps_record.columns)

    exps_record = exps_record.append(new_exp, ignore_index=True)
    with pd.ExcelWriter(exps_record_path) as writer:  # doctest: +SKIP
        exps_record.to_excel(writer, sheet_name='exps_record', index=False)
    # all_routes.to_excel("output.xlsx",sheet_name='Sheet_name_1')
    with pd.ExcelWriter(path + current_save + '/obj_record' + current_save + str(
            exp_number - 1) + '.xlsx') as writer:  # doctest: +SKIP
        obj_record_best.to_excel(writer, sheet_name='obj_record_best')
        obj_record.to_excel(writer, sheet_name='obj_record')

    with pd.ExcelWriter(path + current_save + '/best_routes' + current_save + '_' + str(
            exp_number - 1) + '.xlsx') as writer:  # doctest: +SKIP
        if CP == 1:
            global CP_best_routes
            CP_best_routes = {}
        for key, value in all_routes[obj_record_best.index[0]].items():
            if CP == 1:
                CP_best_routes[key] = value
            route_df = pd.DataFrame(value[0:4, :], columns=value[4])
            # revert_K = read_R_K(request_number_in_R, what='revert_K')
            k= list(revert_K.keys())[list(revert_K.values()).index(key)]
            route_df.to_excel(writer, k)
    if dynamic == 0:
        with pd.ExcelWriter(path + current_save + '/functions_time' + current_save + str(
                exp_number - 1) + '.xlsx') as writer:  # doctest: +SKIP
            functions_time.to_excel(writer, sheet_name='functions_time')

        if combination == 1:
            weight.to_excel(path + current_save + '/weight' + current_save + str(exp_number - 1) + '.xlsx',
                            sheet_name='weight')
        else:
            weight_insertion.to_excel(
                path + current_save + '/weight_insertion' + current_save + str(exp_number - 1) + '.xlsx',
                sheet_name='weight_insertion')
            weight_removal.to_excel(path + current_save + '/weight_removal' + current_save + str(exp_number - 1) + '.xlsx',
                                    sheet_name='weight_removal')

        all_Tem_df = pd.DataFrame(all_Tem, columns=['Temperature'])
        all_Tem_df.to_excel(path + current_save + '/all_Tem' + current_save + str(exp_number - 1) + '.xlsx',
                            sheet_name='tem')

        all_pro_df = pd.DataFrame(all_pro, columns=['Acceptance probability'])
        all_pro_df.to_excel(path + current_save + '/pro' + current_save + str(exp_number - 1) + '.xlsx', sheet_name='pro')

        draw_figures(obj_record_better, path, current_save)

    Graph(all_routes[obj_record_best.index[0]], 0)

    # obj_record.drop_duplicates(subset=['overall_cost'], inplace=True)
    # for index in obj_record.index:
    #     Graph(all_routes[index], 0)

    return obj_record_best, CPU_Time, Running_Time, Best_Running_Time

def calculate_CP_compare(before_or_after, CP_compare, obj_record, carrier, carriers_number):
    global carrier_number_no_serving
    # for column in ['overall_distance', 'overall_cost', 'overall_time', 'overall_profit',
    #                                            'overall_emission', 'served_requests', 'overall_request_cost',
    #                                            'overall_vehicle_cost', 'overall_wait_cost',
    #                                            'overall_transshipment_cost',
    #                                            'overall_un_load_cost', 'overall_emission_cost', 'overall_storage_cost',
    #                                            'overall_delay_penalty', 'barge_served_requests',
    #                                            'train_served_requests', 'truck_served_requests',
    #                                                'fuzzy_satisfy_or_not', 'hard_satisfy_or_not', 'overall_number_transshipment', 'transshipment_times']:
    #     CP_compare.loc[before_or_after][column] = CP_compare.loc[before_or_after][column] + obj_record.iloc[0][column]
    if carrier == 1:
        carrier_number_no_serving = 0
    if obj_record.iloc[0]['served_requests'] == 0:
        carrier_number_no_serving = carrier_number_no_serving + 1
    #CP_compare and obj_record have different number of columns, but dataframe will add the right comlumns together
    CP_compare.loc[before_or_after] = CP_compare.loc[before_or_after] + obj_record.iloc[0]
    for column in ['iteration_time']:
        if obj_record.iloc[0][column] > CP_compare.loc[before_or_after][column]:
            CP_compare.loc[before_or_after][column] = copy.copy(obj_record.iloc[0][column])

    if carrier == carriers_number:
        if heterogeneous_preferences == 1:
            for column in ['satisfactory_value', 'overall_average_time_ratio', 'cost_per_container_per_km', 'time_ratio', 'emissions_per_container_per_km', 'delay_time_ratio']:
                CP_compare.loc[before_or_after][column] = CP_compare.loc[before_or_after][column] / (carriers_number - carrier_number_no_serving)


    return CP_compare


def get_CP_compare(before_or_after, path):

    CP_compare_path = path + 'CP_compare' + str(exp_number) + 'R' + str(request_number_in_R) + '.xlsx'
    # get the results before/after collaboration
    # check whether all carriers finish initial optimization
    while True:
        # print(9160)
        optimization_done_number = 0
        for carrier in range(1, carriers_number + 1):
            current_save = 'percentage' + str(percentage) + 'parallel_number' + str(carrier)  + 'dynamic' + str(dynamic_t)
            CP_compare_check_path = path + current_save + '/obj_record' + current_save + str(exp_number) + '.xlsx'
            if os.path.exists(CP_compare_check_path):
                optimization_done_number = optimization_done_number + 1
            else:
                break
        if optimization_done_number == carriers_number:

            for carrier in range(1, carriers_number + 1):
                current_save = 'percentage' + str(percentage) + 'parallel_number' + str(carrier) + 'dynamic' + str(dynamic_t)
                CP_compare_check_path = path + current_save + '/obj_record' + current_save + str(
                    exp_number) + '.xlsx'
                while True:
                    try:
                        if before_or_after == 'before':
                            sheet = 'obj_record_best'
                        else:
                            sheet = 'final_obj'
                        obj_record = pd.read_excel(CP_compare_check_path, sheet_name=sheet, index_col=0)
                        break
                    except:
                        pass
                if before_or_after == 'before':
                    if carrier == 1:
                        CP_compare = pd.DataFrame(np.zeros((2, len(obj_record.columns))), index=['before', 'after'], columns=obj_record.columns)
                    # CP_compare.loc['before'] = CP_compare.loc['before'] + obj_record.iloc[0]
                    CP_compare = calculate_CP_compare('before', CP_compare, obj_record, carrier, carriers_number)
                else:
                    if carrier == 1:
                        CP_compare = pd.read_excel(CP_compare_path, sheet_name='CP_compare', index_col=0)
                    # CP_compare.loc['after'] = CP_compare.loc['after'] + obj_record.iloc[0]
                    CP_compare = calculate_CP_compare('after', CP_compare, obj_record, carrier, carriers_number)
            with pd.ExcelWriter(CP_compare_path) as writer:  # doctest: +SKIP
                CP_compare.to_excel(writer, sheet_name='CP_compare')
            break

def update_CP_R_pool(CP_R_pool_path, R_pool, add_or_delete = 1, success_r = -1):
    global path
    # print('update_CPR_pool_success1')
    # while True:
    #     if has_handle(CP_R_pool_path) == False:
    #         CP_R_pool = pd.read_excel(CP_R_pool_path, 'CP_R_pool', index_col=0)
    #         break

    print('update_CP_R_pool starts')
    if add_or_delete == 1 and not os.path.isfile(CP_R_pool_path):
        parallel_save_excel(CP_R_pool_path, pd.DataFrame(R_pool), 'CP_R_pool',index=True)
        return
    if add_or_delete == -1 and success_r == []:
        return
    my_CP_R_pool_path = path + 'CP_R_pool' + str(parallel_number) + '.xlsx'
    while True:
        try:
            if os.path.isfile(my_CP_R_pool_path):
                os.remove(my_CP_R_pool_path)
            shutil.copy(CP_R_pool_path, my_CP_R_pool_path)
            CP_R_pool = pd.read_excel(my_CP_R_pool_path, 'CP_R_pool', index_col=0)
            break
        except:
            print('oooooooooooooooo') #when bad zip file error try again

    CP_R_pool = CP_R_pool.values
    if add_or_delete == 1:
        #add
        # CP_R_pool.append(pd.DataFrame(R_pool))
        if R_pool.size != 0:
            # try:
            CP_R_pool = np.vstack([CP_R_pool, R_pool])
            # except:
            #     print('4')
            #     sys.exit(-1)
            if CP_R_pool.size != 0:
                CP_R_pool = CP_R_pool.astype(float)
                CP_R_pool = np.unique(CP_R_pool, axis=0)

    else:
        # #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
        #delete r that in CP_R_pool but not in CP_R_pool_this_carrier
        # print('update_CPR_pool_success2')
        if success_r:
            for original_r, original_carrier in success_r:
                # for key, value in map_change_index:
                #     if value == [original_r, original_carrier]:
                #         r = key
                #         break
                # r = map_change_index_reverse[tuple([original_r, original_carrier])]
                # if original_r not in R_pool[:, 7]:
                for index in range(len(CP_R_pool)):
                    if CP_R_pool[index, 7] == original_r and CP_R_pool[index, -1] == original_carrier:
                        CP_R_pool = np.delete(CP_R_pool, index, axis=0)
                        break
    # print('update_CPR_pool_success3')
    CP_R_pool = pd.DataFrame(CP_R_pool)
    while True:
        try:
            with pd.ExcelWriter(CP_R_pool_path) as writer:  # doctest: +SKIP
                CP_R_pool.to_excel(writer, sheet_name='CP_R_pool', index=True)
            pd.read_excel(CP_R_pool_path, 'CP_R_pool', index_col=0)
            break
        except:
            print(10000000003333333333333)
            pass
    print('update_CPR_pool_success')
    return


def parallel_save_excel(path, excel, sheet, index=False):
    while True:
        try:
            with pd.ExcelWriter(path) as writer: # doctest: +SKIP
                if index == False:
                    excel.to_excel(writer, sheet_name=sheet, index=index)
                else:
                    excel.to_excel(writer, sheet_name=sheet)
            pd.read_excel(path, sheet)
            break
        except:
            continue


def parallel_read_excel(path, sheet, index=1):
    while True:
        try:
            if index == 0:
                read_results = pd.read_excel(path, sheet, index_col=0)
            else:
                read_results = pd.read_excel(path, sheet)
            break
        except:
            continue
    return read_results

def CP_served_requests_mode():
    number_served_r = check_served_R()
    if parallel_number == 1:
        barge_served_requests = number_served_r
        train_served_requests = 0
        truck_served_requests = 0
    elif parallel_number == 2:
        barge_served_requests = 0
        train_served_requests = number_served_r
        truck_served_requests = 0
    else:
        barge_served_requests = 0
        train_served_requests = 0
        truck_served_requests = number_served_r
    return barge_served_requests, train_served_requests, truck_served_requests

def find_all_enumerate(s, c):
    return [i for i, x in enumerate(s) if x == c]

def add_carriers_to_segment(dict_segments_carriers, first_segment_r, carrier_1, carrier_2):
    if first_segment_r[7] not in dict_segments_carriers.keys():
        dict_segments_carriers[int(first_segment_r[7])] = [carrier_1, carrier_2]
    else:
        dict_segments_carriers[int(first_segment_r[7])].extend([carrier_1, carrier_2])
    return dict_segments_carriers

def VCP_coordinator():
    #the function of VCP_coordinator is to check whether the combinations of route segments are suitable and calculate objs
    #read routes first
    #carriers need to report which request and vehicles are influenced, then report the request's T and time at T that needs to be coordinated
    #chekc spatial

    #check time
    # for
    # carrier_
    return

def coordinator():
    global path
    path = '/data/yimeng/Figures/experiment' + str(exp_number) + '/'
    Path(path).mkdir(parents=True, exist_ok=True)
    CP_R_pool_path = path + 'CP_R_pool.xlsx'
    dict_segments_carriers_path = path + 'dict_segments_carriers.xlsx'
    CP_bids_path = path + 'CP_bids.xlsx'

    CP_R_pool_check_all_path = path + 'CP_R_pool_check' + '0.xlsx'

    CP_bids_check_all_path = path + 'CP_bids_check' + '0.xlsx'

    CP_round_end_check_all_path = path + 'CP_round_end_check_all' + '0.xlsx'

    get_CP_compare('before', path)

    # bids for a number of rounds,
    for round in range(auction_round_number):
        #wait for initial optimization results
        # check whether all carriers sumbitted their unserved requests
        while True:
            # print(9339)
            submitted_number = 0
            for carrier in range(1, carriers_number + 1):
                CP_R_pool_check_path = path + 'CP_R_pool_check' + str(carrier) + '.xlsx'
                if os.path.exists(CP_R_pool_check_path):
                    submitted_number = submitted_number + 1
                else:
                    break
            if submitted_number == carriers_number:

                for carrier in range(1, carriers_number + 1):
                    # add all submitted r to CP_R_pool
                    if os.path.exists(path + 'CP_R_pool_submit' + str(carrier) + '.xlsx'):
                        R_pool = parallel_read_excel(path + 'CP_R_pool_submit' + str(carrier) + '.xlsx', 'CP_R_pool', 0)
                        R_pool = R_pool.values
                        update_CP_R_pool(CP_R_pool_path, R_pool)
                
                if request_segment == 1:
                    #read CP_R_pool
                    CP_R_pool = pd.read_excel(CP_R_pool_path, 'CP_R_pool', index_col=0)
                    #segment all r in CP_R_pool by shared T
                    names = revert_names()
                    for carrier in range(1, carriers_number + 1):
                        #read each carrier's T
                        # data_path = 'C:\Intermodal\Case study\CP\instances\Intermodal_EGS_data_carrier' + str(carrier) + '.xlsx'
                        if different_companies == 1:
                            if carrier == 1:
                                data_path = "/data/yimeng/Case study/CP/instances/three_eco_labels/Intermodal_EGS_data_all" + note + ".xlsx"
                            elif carrier == 2:
                                data_path = "/data/yimeng/Case study/CP/instances/three_eco_labels/Intermodal_Contargo_data_all" + note + ".xlsx"
                            else:
                                data_path = "/data/yimeng/Case study/CP/instances/three_eco_labels/Intermodal_HSL_data_all" + note + ".xlsx"
                        else:
                            data_path = '/data/yimeng/Case study/CP/instances/three_eco_labels/Intermodal_EGS_data_carrier' + str(
                                carrier) + note + '.xlsx'
                        Data = pd.ExcelFile(data_path)
                        T_carrier = 'T'+str(carrier)
                        locals()[T_carrier] = pd.read_excel(Data, 'T')

                        locals()[T_carrier]['T'] = locals()[T_carrier]['T'].map(names).fillna(locals()[T_carrier]['T'])
                        locals()[T_carrier] = list(locals()[T_carrier]['T'])
                    # get the list of shared T
                    #for each two carriers, find sharing T
                    all_sharing_T = {}
                    for carrier_1 in range(1, carriers_number + 1):
                        for carrier_2 in range(1, carriers_number + 1):
                            if carrier_1 == carrier_2:
                                continue

                            T_carrier_1 = 'T' + str(carrier_1)
                            T_carrier_2 = 'T' + str(carrier_2)
                            T_carrier_2_as_set = set(locals()[T_carrier_2])
                            intersection = T_carrier_2_as_set.intersection(locals()[T_carrier_1])
                            all_sharing_T['carrier'+str(carrier_1)+'carrier'+str(carrier_2)] = list(intersection)
                    #segment request
                    #all requests in CP_R_pool are segmented by shared T of two carriers
                    #final aim is that add segmented r to relevant R_pool of carriers
                    #so the segmented rs should be under a key called carrier
                    #the upper key is the
                    CP_R_pool = CP_R_pool.values
                    CP_R_pool_2v = {}
                    # dict_CP_R_segment_pool = {}
                    # for carrier in range(1, carriers_number + 1):
                    #     dict_CP_R_segment_pool['carrier'] = np.array(np.empty(shape=(0,len(CP_R_pool[0]))))
                    #     dict_CP_R_segment_pool['carrier'][:]=np.NaN
                    dict_segments_carriers = {}
                    for index in range(len(CP_R_pool)):
                        #if it already be segmented, then do nothing
                        all_digits = get_all_digits_of_a_number(CP_R_pool[index, 7], [])
                        if all_digits[2] > 0:
                            continue
                        R_i = tuple(zip(CP_R_pool[index], ['p', 'd', 'ap', 'bp', 'ad', 'bd', 'qr', 'r']))
                        #for each R_i, find the suitable two/three carriers
                        if R_i in CP_R_pool_2v.keys():
                            pass
                        else:
                            CP_R_pool_2v[R_i] = {}
                        # R_change = R.copy()
                        for carriers in all_sharing_T.keys():
                            position_letter = find_all_enumerate(carriers, 'c')
                            carrier_1, carrier_2 = int(carriers[position_letter[0]+7: position_letter[1]]), int(carriers[position_letter[1]+7:])
                            for T_change in all_sharing_T[carriers]:

                                request = CP_R_pool[list(CP_R_pool[:, 7]).index(R_i[7][0])]
                                #reduce unnecessary computation
                                #T == p or d
                                if T_change == request[0] or T_change == request[1]:
                                    continue
                                first_segment_r, second_segment_r = segment_request(request, T_change)
                                #change name of r
                                first_segment_r[7] = first_segment_r[7] + big_r/100*1 + big_r/10000*T_change
                                second_segment_r[7] = second_segment_r[7] + big_r/100*2 + big_r/10000*T_change
                                # # add carrier who can transport this segment
                                # first_segment_r.extend([carrier_1,carrier_2])
                                # second_segment_r.extend([carrier_1, carrier_2])
                                if T_change not in CP_R_pool_2v[R_i].keys():
                                    CP_R_pool_2v[R_i][T_change] = pd.concat([first_segment_r, second_segment_r], axis=1).T
                                    # R_pool_2v[R_i][T_change].columns=['p','d','ap','bp','ad','bd','qr','r']
                                    # R_pool_2v[R_i][T_change].index = [0,1]
                                    CP_R_pool_2v[R_i][T_change] = CP_R_pool_2v[R_i][T_change].values
                                    #add these two segments to CP_R_pool
                                    CP_R_pool = np.vstack([CP_R_pool, CP_R_pool_2v[R_i][T_change]])
                                #add two carriers' number in a dict
                                dict_segments_carriers = add_carriers_to_segment(dict_segments_carriers, first_segment_r, carrier_1, carrier_2)
                                dict_segments_carriers = add_carriers_to_segment(dict_segments_carriers, second_segment_r, carrier_1, carrier_2)




                                #generate a dict for all carriers, each carrier has a table of segments
                                #improve: it would be better if coordinator considers the another terminal (except T) in carrier's network or not
                                # position_letter = find_all_enumerate(carriers, 'c')# if infeasible, then check the current operation at this terminal of this request
                                #                                             operation = new_getLetters(routes[k][5][col])
                                #                                             if operation == 'pickup':
                                #                                                 #this request can be removed fully
                                #                                                 routes, R_pool = remove_a_request(inserted_r, routes, R_pool)[
                                #                                                                  0:2]
                                #                                             elif operation == 'Td':
                                #                                                 #this request
                                #                                             if operation == 'delivery' or operation == 'Td' or operation == 'secondTd':
                                # dict_CP_R_segment_pool[carriers[0:(position_letter-1)]] = np.vstack([dict_CP_R_segment_pool[carriers[0:(position_letter-1)]], CP_R_pool_2v[R_i][T_change]])
                                # dict_CP_R_segment_pool[carriers[(position_letter-1):]] = np.vstack([dict_CP_R_segment_pool[carriers[(position_letter-1):]], CP_R_pool_2v[R_i][T_change]])

                    #save dict_CP_R_segment_pool
                    # with pd.ExcelWriter(dict_CP_R_segment_pool_path) as writer:  # doctest: +SKIP
                    #     for carrier, R_segment_pool in dict_CP_R_segment_pool.items():
                    #         R_segment_pool = pd.DataFrame(R_segment_pool)
                    #         R_segment_pool.to_excel(writer, carrier)
                    #save dict_segments_carriers
                    dict_segments_carriers_file = open(dict_segments_carriers_path, "w")
                    json.dump(dict_segments_carriers, dict_segments_carriers_file)
                    dict_segments_carriers_file.close()
                    #save CP_R_pool
                    parallel_save_excel(CP_R_pool_path, pd.DataFrame(CP_R_pool), 'CP_R_pool', True)
                # here must wait the adding new request finish, then delete, otherwise, may delete at one carrier, next carrier add it again
                # also wait for all CP_success_r_submit files are uploaded
                if round != 0:
                    while True:
                        submitted_number = 0
                        for carrier in range(1, carriers_number + 1):
                            # delete all success r in CP_R_pool
                            if os.path.exists(path + 'CP_success_r_submit' + str(carrier) + '.xlsx'):
                                submitted_number = submitted_number + 1
                            else:
                                break
                        if submitted_number == carriers_number:
                            break
                for carrier in range(1, carriers_number + 1):
                    # delete all success r in CP_R_pool
                    if os.path.exists(path + 'CP_success_r_submit' + str(carrier) + '.xlsx'):
                        success_r_df = parallel_read_excel(
                            path + 'CP_success_r_submit' + str(carrier) + '.xlsx', 'success_r')
                    else:
                        success_r_df = []
                    success_r = []
                    for i in range(len(success_r_df)):
                        success_r.append([success_r_df.iloc[i][0], success_r_df.iloc[i][1]])
                    #to do here need to consider segments of a r are all served
                    if request_segment == 1:
                        # set a dict that store all segments
                        seg_dict = {}
                        #detect whether there is a segment
                        for each_success_r in success_r:
                            all_digits = get_all_digits_of_a_number(each_success_r[0], [])
                            if all_digits[2] > 0:
                                #put it in the seg_dict
                                r_number = all_digits[-4] * 1000 + all_digits[-3] * 100 + all_digits[-2] * 10 + all_digits[-1]
                                seg_dict[all_digits[0]*big_r + r_number][all_digits[2]] = each_success_r
                        for each_r in seg_dict.keys():
                            #not two segments are served
                            if len(seg_dict[each_r]) < 2:
                                #delete it in success_r
                                for each_success_r in seg_dict[each_r]:
                                    success_r.remove(each_success_r)
                    update_CP_R_pool(CP_R_pool_path, -1, -1, success_r)

                CP_R_pool_check_all = pd.DataFrame([carriers_number])
                with pd.ExcelWriter(CP_R_pool_check_all_path) as writer:  # doctest: +SKIP
                    CP_R_pool_check_all.to_excel(writer, sheet_name='CP_R_pool_check', index=False)
                break
        #no need to clear all submit R_pool because submit R_pool will be refreshed automatically when save it, although submit R_pool will be used again when delete success r


        #all carriers start, so clear end marks
        try:
            os.remove(CP_round_end_check_all_path)
            for carrier in range(1, carriers_number + 1):
                CP_round_end_check_path = path + 'CP_round_end_check' + str(carrier) + '.xlsx'
                os.remove(CP_round_end_check_path)
        except:
            #first round has no these files
            pass
        #wait for bids from all carriers
        #detect whether all carriers submit bids or decide not submit any bids
        while True:
            submitted_number = 0
            for carrier in range(1, carriers_number + 1):
                CP_bids_check_path = path + 'CP_bids_check' + str(carrier) + '.xlsx'
                if os.path.exists(CP_bids_check_path):
                    submitted_number = submitted_number + 1
                else:
                    break
            if submitted_number == carriers_number:

                for carrier in range(1, carriers_number + 1):
                    CP_bids_submit_path = path + 'CP_bids' + 'submit' + str(carrier) + '.xlsx'
                    if os.path.isfile(CP_bids_submit_path):
                        bids = parallel_read_excel(CP_bids_submit_path, 'CP_bids')
                        if not os.path.isfile(CP_bids_path):
                            # Path(CP_bids_path).mkdir(parents=True, exist_ok=True)
                            parallel_save_excel(CP_bids_path, bids, 'CP_bids')
                            # with pd.ExcelWriter(CP_bids_path) as writer:  # doctest: +SKIP
                            #     bids.to_excel(writer, sheet_name='CP_bids', index=False)

                        else:

                            CP_bids = parallel_read_excel(CP_bids_path, 'CP_bids')
                            # CP_bids = pd.read_excel(CP_bids_path, 'CP_bids')
                            CP_bids = CP_bids.append(bids)
                            CP_bids = CP_bids.drop_duplicates()
                            parallel_save_excel(CP_bids_path, CP_bids,'CP_bids')
                CP_bids_check_all = pd.DataFrame([carriers_number])
                with pd.ExcelWriter(CP_bids_check_all_path) as writer:  # doctest: +SKIP
                    CP_bids_check_all.to_excel(writer, sheet_name='CP_bids_check', index=False)
                break


        # if there are bids, then group bids and rank the bids for the same request
        if os.path.exists(CP_bids_path):
            CP_bids = parallel_read_excel(CP_bids_path, 'CP_bids')
            group_bids = {}
            if request_segment == 1:
                #if there are request segments, then check whether all segments of a r are served.
                # if served, then sum costs, and compare with the cost of the case which not be segmented

                #set a dict to store all segments of each r
                seg_bids = {}
                for index in range(len(CP_bids)):
                    key = CP_bids['r'][index]
                    # check whether r (key) is a segment (if r is a segment, the third digit larger than 0)
                    all_digits = get_all_digits_of_a_number(key, [])
                    if all_digits[2] > 0:
                        r_number = all_digits[0] * big_r + all_digits[-4] * 1000 + all_digits[-3] * 100 + all_digits[-2] * 10 + all_digits[-1]
                        if all_digits[2] not in seg_bids.keys():
                            # the key of seg_bids = r

                            seg_bids[r_number] = {}
                        # if the first/second segment not in seg_bids[all_digits[-1]], then set a list and add its bid's information
                        if all_digits[2] not in seg_bids[r_number].keys():
                            seg_bids[r_number][all_digits[2]] = [CP_bids.iloc[index]]
                        else:
                            # add it
                            seg_bids[r_number][all_digits[2]].append(CP_bids.iloc[index])

                    if key not in group_bids.keys() and all_digits[2] == 0:
                        group_bids[key] = pd.DataFrame(columns=CP_bids.columns)
                    # the bid is added to group_bids directly only if it's not a segment
                    if key in group_bids.keys():
                        group_bids[key] = group_bids[key].append(CP_bids.iloc[index])
                #check whether all segments of a r are served, i.e., whether both two segment lists are not empty
                for key in seg_bids.keys():
                    if len(seg_bids[key])==2:
                        #sum the cost of the first and second segment of all seg bids
                        seg_bids[key]['both'] = pd.DataFrame(columns=CP_bids.columns)
                        for seg_bid_1 in seg_bids[key][1]:
                            for seg_bid_2 in seg_bids[key][2]:
                                #columns = ['r', 'cost', 'carrier', 'original_carrier']
                                #the carrier needs to add the original bid to it's sucess bids, so need a new format represent all segments in seg_bids['both'], and use this new format can find the orininal one
                                # which column should be added -> how to present the carrier? r by the name of this r, add costs, carrier1+00+carrier2+00, original carrier
                                # meaning of r: original_carrier00segment_number00transshipment_terminal00request_number
                                seg_bids[key]['both'].append([[seg_bid_1['r'], seg_bid_2['r']], seg_bid_1['r'] + seg_bid_2['r'], [seg_bid_1['carrier'], seg_bid_2['carrier']], seg_bid_1['original_carrier']])
                        seg_bids_min_cost = seg_bids.loc[seg_bids['cost'] == seg_bids['cost'].min()]
                        if key not in group_bids.keys():
                            group_bids[key] = pd.DataFrame(columns=CP_bids.columns)
                        group_bids[key] = group_bids[key].append(seg_bids_min_cost)
                # the bid with the lowest cost wins
                success_bids = pd.DataFrame(columns=CP_bids.columns)
                for key in group_bids.keys():
                    min_cost_bid = group_bids[key].loc[group_bids[key]['cost'] == group_bids[key]['cost'].min()]
                    if isinstance(min_cost_bid.iloc[0]['r'], list):
                        #if the segment is the lowest cost bid, then the success_bids should store segments
                        #find the segments
                        seg1 = seg_bids[key][1].loc[seg_bids[key][1]['r'] == min_cost_bid['r'][0][0]]
                        seg2 = seg_bids[key][2].loc[seg_bids[key][2]['r'] == min_cost_bid['r'][0][1]]
                        success_bids = success_bids.append(seg1)
                        success_bids = success_bids.append(seg2)
                    else:
                        success_bids = success_bids.append(min_cost_bid)

                with pd.ExcelWriter(path + 'success_bids.xlsx') as writer:  # doctest: +SKIP
                    success_bids.to_excel(writer, sheet_name='success_bids', index=False)
            else:
                for index in range(len(CP_bids)):
                    key = CP_bids['r'][index]
                    if key not in group_bids.keys():
                        group_bids[key] = pd.DataFrame(columns=CP_bids.columns)
                    group_bids[key] = group_bids[key].append(CP_bids.iloc[index])
                # the bid with the lowest cost wins
                success_bids = pd.DataFrame(columns=CP_bids.columns)
                for key in group_bids.keys():
                    success_bids = success_bids.append(group_bids[key].loc[group_bids[key]['cost'] == group_bids[key]['cost'].min()])
                with pd.ExcelWriter(path + 'success_bids.xlsx') as writer:  # doctest: +SKIP
                    success_bids.to_excel(writer, sheet_name='success_bids', index=False)



        #check whether all carriers end
        while True:
            # print(9617)
            end_number = 0
            for carrier in range(1, carriers_number + 1):
                CP_round_end_check_path = path + 'CP_round_end_check' + str(carrier) + '.xlsx'
                if os.path.exists(CP_round_end_check_path):
                    end_number = end_number + 1
                else:
                    break
            if end_number == carriers_number:

                CP_round_end_check_all = pd.DataFrame([carriers_number])
                # with pd.ExcelWriter(CP_round_end_check_all_path) as writer:  # doctest: +SKIP
                #     CP_round_end_check_all.to_excel(writer, sheet_name='CP_round_end_check', index=False)
                parallel_save_excel(CP_round_end_check_all_path, CP_round_end_check_all, 'CP_round_end_check')
                break

        # clear CP_R_pool_check_all, let this file empty to restart at next round
        os.remove(CP_R_pool_check_all_path)
        for carrier in range(1, carriers_number + 1):
            CP_R_pool_check_path = path + 'CP_R_pool_check' + str(carrier) + '.xlsx'
            while True:
                try:
                    os.remove(CP_R_pool_check_path)
                    break
                except:
                    continue
        for carrier in range(1, carriers_number + 1):
            CP_R_pool_submit_path = path + 'CP_R_pool_submit' + str(carrier) + '.xlsx'
            if os.path.exists(CP_R_pool_submit_path):
                while True:
                    # print(9647)
                    try:
                        os.remove(CP_R_pool_submit_path)
                        break
                    except:
                        continue
        #clear bids check
        os.remove(CP_bids_check_all_path)
        for carrier in range(1, carriers_number + 1):
            CP_bids_check_path = path + 'CP_bids_check' + str(carrier) + '.xlsx'
            os.remove(CP_bids_check_path)
        # clear bids
        if os.path.exists(CP_bids_path):
            os.remove(CP_bids_path)
        if os.path.exists(path + 'success_bids.xlsx'):
            os.remove(path + 'success_bids.xlsx')
    # get the results after collaboration
    get_CP_compare('after', path)


def save_results(round, routes_save, obj_record = -1):
    if CP == 1:
        # save each round's results
        Path(path + current_save + 'round' + str(round) + '/').mkdir(parents=True, exist_ok=True)
        best_routes_path_round = path + current_save + 'round' + str(
            round) + '/best_routes' + current_save + '_' + str(
            exp_number - 1) + '.xlsx'
    elif dynamic == 1:
        # save each round's results
        Path(path + current_save + '/').mkdir(parents=True, exist_ok=True)
        best_routes_path_round = path + current_save + '/best_routes' + current_save + '_' + str(
            exp_number - 1) + '.xlsx'
    # save routes, routes match, obj_record
    with pd.ExcelWriter(best_routes_path_round) as writer:  # doctest: +SKIP
        for key, value in routes_save.items():
            route_df = pd.DataFrame(value[0:4, :], columns=value[4])
            route_df.to_excel(writer, convert(key))
    if CP == 1:
        # save match
        path_round = path + current_save + 'round' + str(round) + '/routes_match' + current_save + str(
            exp_number - 1) + '.xlsx'
    elif dynamic == 1:
        path_round = path + current_save + '/routes_match' + current_save + str(
            exp_number - 1) + '.xlsx'
    Graph(routes_save, 0, path_round)
    if CP == 1:
        # save obj_record by copying
        copyfile(path + current_save + '/obj_record' + current_save + str(exp_number - 1) + '.xlsx',
                 path + current_save + 'round' + str(round) + '/obj_record' + current_save + str(
                     exp_number - 1) + '.xlsx')
    if not isinstance(obj_record, int):
        obj_record_path = path + current_save + '/obj_record' + current_save + '_' + str(
            exp_number - 1) + '.xlsx'
        with pd.ExcelWriter(obj_record_path) as writer:  # doctest: +SKIP
            obj_record.to_excel(writer, 'obj_record')

# @profile()
# @time_me()
def real_main(parallel_number2, dynamic_t2 = 0, distribution_name='default'):
    global waiting_times, only_check_this_influenced_r_in_dynamic_uncertainty, used_interrupt, ALNS_implement_start_RL_can_move, interrupt_by_implement_is_one_and_assign_action_once_only, time_dependent_truck_travel_time, dynamic_t_begin, ALNS_greedy_under_unknown_duration_assume_duration, number_of_training, number_of_implementation, re_plan_when_event_finishes_information, ALNS_end_flag, RL_is_trained_or_evaluated_or_ALNS_is_evaluated, delayed_time_table_uncertainty_index, RL_insertion_segment, congestion_nodes_at_begining, RL_removal_implementation_store, RL_insertion_implementation_store, ALNS_removal_implementation_store, ALNS_insertion_implementation_store, vessel_train, combine_insertion_and_removal_operators, state_reward_pairs_insertion, after_action_review, get_reward_by_cost_gap, ALNS_guides_RL, state_reward_pairs, congestion_nodes, congestion_links, path, stochastic, add_RL, request_segment_in_dynamic, delayed_time_table, unexpected_events, VCP_coordination, different_companies, during_iteration, emission_preference_constraints_after_iteration, wtw_emissions, dynamic, dynamic_t, big_r, request_flow_t, percentage, not_initial_in_CP, R, R_pool, parallel_number, carriers_number, auction_round_number, CP_try_r_of_other_carriers, use_speed, get_satisfactory_value_one_by_one, fuzzy_probability, only_eco_label, only_eco_label_add_cost, heterogeneous_preferences_no_constraints, request_segment, data_path, CP, parallel_ALNS, allow_infeasibility, swap_or_not, fuzzy_constraints,real_multi_obj,weight_interval,w1,w2,w3,Demir_barge_free,truck_fleet, forbid_much_delay, two_T, heterogeneous_preferences, Demir,old_current_save,parallel, parallel_thread, max_processors, start_from_best_at_begin_of_segement, belta, truck_time_free, functions_time, Fixed_Data, by_wenjing, T_number, k_number, node_number, processors_number, note, obj_number, exp_number, regret_k, service_time, transshipment_time, c_storage, fuel_cost, has_end_depot2, check_obj, exps_record_path, forbid_T_trucks, get_initial_bymyself, request_number_in_R, multi_obj, c_storage, b1, b2, b3, b4, b5, b6, b7, b8, b9, b10, alpha, bundle_or_not, c, devide_value, stop_time, regret_k, regular, insert_multiple_r, bi_obj_cost_emission, bi_obj_cost_time, K
    waiting_times = {}
    only_check_this_influenced_r_in_dynamic_uncertainty = -1
    ALNS_implement_start_RL_can_move = 0
    used_interrupt = 0
    time_dependent_truck_travel_time = 0
    vessel_train = 0
    stochastic = 1
    # zero 0, very big 1, random 2
    ALNS_greedy_under_unknown_duration_assume_duration =0 
    if stochastic == 1:
        import Dynamic_master34959
        global duration_type
        add_RL = Dynamic_master34959.add_RL
        while True:
            try:
                if add_RL == 0:
                    break
                if dynamic_RL34959.implement == 1 and dynamic_RL34959.stop_everything_in_learning_and_go_to_implementation_phase == 1:
                    break_it = 1
                    break
                if dynamic_RL34959.implement == 0:
                    interrupt_by_implement_is_one_and_assign_action_once_only = 0
                if dynamic_RL34959.implement == 0 or (dynamic_RL34959.implement == 1 and dynamic_RL34959.stop_everything_in_learning_and_go_to_implementation_phase == 0):
                    break
            except:
                continue
        if 'break_it' in locals():
            return
        #two_events_mu_5_1_80_5, three_events_mu_5_1_80_5_40_5, four_events_mu_5_1_80_5_40_5_5_1, five_events_mu_5_1_80_5_40_5_5_1_40_20, six_events_mu_5_1_80_5_40_5_5_1_40_20_80_40
        # 使用传入的分布名称，如果为default则使用原有配置
        if distribution_name == 'default':
            duration_type = 'mix_mu_5_40_terminal_dependent'  # 保持向后兼容
        else:
            duration_type = distribution_name
        RL_is_trained_or_evaluated_or_ALNS_is_evaluated = 0
        get_reward_by_cost_gap = 0
        ALNS_guides_RL = 0
        combine_insertion_and_removal_operators = 0
        congestion_links = []
        congestion_nodes = []
        congestion_nodes_at_begining = {}
        re_plan_when_event_finishes_information = pd.DataFrame(columns= ['k', 'request_number', 'vehicle_stop_time', 'T1', 'Td_time', 'r_number'])
        RL_insertion_segment = 1
        
        if add_RL == 1:
            ALNS_end_flag = 0
            # After-action review

            after_action_review = 1

            #dynamic_RL34959.state_action_reward_collect = np.array(np.empty(shape=(0, 9)))
            state_reward_pairs = pd.DataFrame(
                columns=['uncertainty_index', 'uncertainty_type', 'request', 'vehicle', 'delay_tolerance',
                         'passed_terminals',
                         'current_time', 'action', 'reward'])
            if combine_insertion_and_removal_operators == 1:
                state_reward_pairs_insertion = pd.DataFrame(
                    columns=['uncertainty_index', 'uncertainty_type', 'request', 'vehicle', 'delay_tolerance',
                             'passed_terminals',
                             'current_time', 'action', 'reward'])
            RL_removal_implementation_store = {}
            RL_insertion_implementation_store = {}
        else:
            after_action_review = 0
            ALNS_removal_implementation_store = {}
            ALNS_insertion_implementation_store = {}
            number_of_training, number_of_implementation = 500, 200
        # Dynamic_ALNS_RL34959.reward_list_in_implementation = []
    different_companies = 0
    during_iteration = 1


    wtw_emissions = 1
    dynamic = 1

    unexpected_events = 1
    dynamic_t = dynamic_t2; dynamic_t_begin = dynamic_t2

    auction_round_number = 3
    carriers_number = 3
    exp_number =34959 
    parallel_number = parallel_number2
    not_initial_in_CP = 0
    Demir = 0
    request_number_in_R =5 
    request_segment = 0
    big_r = 1000000000000
    heterogeneous_preferences = 0
    #when it is 1, the preference constraints work, no matter emission_preference_constraints_after_iteration is what
    heterogeneous_preferences_no_constraints = 0
    #when I use load_dependent_average_emission_rates, it must be 1, otherwise the fake average load (60%) is used in the end
    #only use fake loads during iteration when eco-label is A

    three_eco_labels = 0
    fuzzy_constraints = 1
    # fuzzy_probability is used to check which method is used in fuzzy preferences, if 1, then old method (before 20210320, which didn't use fuzzy rules & output and only use a membership function which created by myself), otherwise the new method.
    fuzzy_probability = 0
    if Demir == 1:
        #this percentage is fixed vehicles
        percentage = [0.716, 1]
        # percentage = 0
    else:
        # percentage of flexible vehicles, from the first one to the percentage one
        # percentage = 0.3, all truck are free; percentage = 0.72, all barge are free
        percentage = 0
    CP = 0
    VCP_coordination = 0
    if parallel_number == 0:
        #coordinator
        if VCP_coordination == 1:
            VCP_coordinator()
        else:
            coordinator()
        return

    functions_time = pd.DataFrame(index=range(2))

    # if 1, T
    # os.environ['PYTHONHASHSEED'] = '0'
    SA = 1
    combination = 0

    if vessel_train == 1:
        T_or_not = 0
    else:
        T_or_not = 1
    k_random_or = 1

    r2 = 0.5
    pro = 0.8
    segment_number2 = 1

    miu1_1 = 33
    miu2_1 = 9
    miu3_1 = 13

    only_T2 = 0
    has_end_depot2 = 1
    service_time = 1
    transshipment_time = 1
    transshipment_cost_per = 18
    fuel_cost = 0
    c_storage = 1

    regret_k = 5
    obj_number, T_number, k_number, node_number, processors_number, note = 1, 10, 116, 'laptop', 'laptop', '_sustainable_1'
    if note == '_sustainable_1':
        emission_preference_constraints_after_iteration = 1
    else:
        emission_preference_constraints_after_iteration = 0
    bundle_or_not = 1
    stop_time = 48

    weight_interval = 0


    use_speed = 0
    allow_infeasibility = 0
    get_satisfactory_value_one_by_one = 0
    two_T = 0
    swap_or_not = 1



    # auction_round_number = 3

    only_eco_label = 1
    # I will also consider cost when it is 1
    only_eco_label_add_cost = 0
    compare_with_gurobi = 0
    if VCP_coordination == 1:
        if parallel_number == 1:
            data_path = "/data/yimeng/Dynamic coordination using MARL/Instances/Intermodal_EGS_data_VCP_delay_carrier1.xlsx"
        elif parallel_number == 2:
            data_path = "/data/yimeng/Dynamic coordination using MARL/Instances/Intermodal_EGS_data_VCP_delay_carrier2.xlsx"
    else:
        if compare_with_gurobi == 1:
            data_path = "C:\Intermodal\Case study\Preferences\instances\Intermodal_EGS_data_Gurobi_low_cost.xlsx"
        else:
            if dynamic == 1:
                # data_path = '/data/yimeng/桌面/Intermodal_EGS_data_dynamic_new_requests.xlsx'
                # data_path = "/data/yimeng/Uncertainties Dynamic planning under unexpected events/Intermodal_EGS_data_dynamic_congestion.xlsx"
                # data_path = "/data/yimeng/Uncertainties Dynamic planning under unexpected events/newtargetInstances_high_impact/R" + str(request_number_in_R) + "/Intermodal_EGS_data_dynamic_congestion" + str(Dynamic_ALNS_RL34959.table_number) + ".xlsx"
                # data_path = "/data/yimeng/Uncertainties Dynamic planning under unexpected events/plot_distribution_targetInstances_high_medium_impact/R" + str(
                #     request_number_in_R) + "/Intermodal_EGS_data_dynamic_congestion" + str(
                #     Dynamic_ALNS_RL34959.table_number) + ".xlsx"
                # if ALNS_greedy_under_unknown_duration_assume_duration == 0 and add_RL == 0:
                #     table_number = 999 - Dynamic_ALNS_RL34959.table_number#if waiting and RL = 0, then implement start from the end
                # else:
                table_number = Dynamic_ALNS_RL34959.table_number; add_event_types =  0 
                if add_event_types == 1:
                    # 支持新旧两种文件格式
                    data_path = None
                    if distribution_name == 'default':
                        # 旧格式：向后兼容
                        old_path = "A:/MYpython/34959_RL/Uncertainties Dynamic planning under unexpected events/plot_distribution_targetInstances_disruption_" + duration_type + "_not_time_dependent/R" + str(
        request_number_in_R) + "/Intermodal_EGS_data_dynamic_congestion" + str(table_number) + ".xlsx"
                        data_path = old_path
                    else:
                        # 新格式：使用分布配置名称
                        new_path = "A:/MYpython/34959_RL/Uncertainties_Dynamic_Planning/plot_distribution_" + duration_type + "/R" + str(
        request_number_in_R) + "/Intermodal_EGS_data_dynamic_" + duration_type + "_table" + str(table_number) + ".xlsx"
                        data_path = new_path

                        # 如果新格式不存在，尝试回退到旧格式
                        if not os.path.exists(data_path):
                            old_path = "A:/MYpython/34959_RL/Uncertainties Dynamic planning under unexpected events/plot_distribution_targetInstances_disruption_" + duration_type + "_not_time_dependent/R" + str(
            request_number_in_R) + "/Intermodal_EGS_data_dynamic_congestion" + str(table_number) + ".xlsx"
                            if os.path.exists(old_path):
                                print(f"信息: 新格式文件不存在，回退到旧格式: {old_path}")
                                data_path = old_path
                            else:
                                print(f"警告: 新旧格式文件都不存在:")
                                print(f"  新格式: {new_path}")
                                print(f"  旧格式: {old_path}")
                else:
                    # 支持新旧两种文件格式
                    data_path = None
                    if distribution_name == 'default':
                        # 旧格式：向后兼容
                        old_path = "A:/MYpython/34959_RL/Uncertainties Dynamic planning under unexpected events/plot_distribution_targetInstances_disruption_" + duration_type + "_not_time_dependent/R" + str(
                            request_number_in_R) + "/Intermodal_EGS_data_dynamic_congestion" + str(table_number) + ".xlsx"
                        data_path = old_path
                    else:
                        # 新格式：使用分布配置名称
                        new_path = "A:/MYpython/34959_RL/Uncertainties_Dynamic_Planning/plot_distribution_" + duration_type + "/R" + str(
                            request_number_in_R) + "/Intermodal_EGS_data_dynamic_" + duration_type + "_table" + str(table_number) + ".xlsx"
                        data_path = new_path

                        # 如果新格式不存在，尝试回退到旧格式
                        if not os.path.exists(data_path):
                            old_path = "A:/MYpython/34959_RL/Uncertainties Dynamic planning under unexpected events/plot_distribution_targetInstances_disruption_" + duration_type + "_not_time_dependent/R" + str(
                                request_number_in_R) + "/Intermodal_EGS_data_dynamic_congestion" + str(table_number) + ".xlsx"
                            if os.path.exists(old_path):
                                print(f"信息: 新格式文件不存在，回退到旧格式: {old_path}")
                                data_path = old_path
                            else:
                                print(f"警告: 新旧格式文件都不存在:")
                                print(f"  新格式: {new_path}")
                                print(f"  旧格式: {old_path}")

                # 验证最终路径是否存在
                if data_path and not os.path.exists(data_path):
                    print(f"错误: 数据文件不存在: {data_path}")
                    print(f"请检查分布配置 '{distribution_name}' 的数据是否已生成")
                    raise FileNotFoundError(f"数据文件不存在: {data_path}")
                elif data_path:
                    print(f"调试: 使用数据文件: {data_path}")
                # data_path = "/data/yimeng/Uncertainties Dynamic planning under unexpected events/Instances/R" + str(request_number_in_R) + "/Intermodal_EGS_data_dynamic_congestion" + str(Dynamic_ALNS_RL34959.table_number) + ".xlsx"
            else:
                if CP == 1:
                    if three_eco_labels == 1:
                        if different_companies == 1:
                            if parallel_number == 1:
                                data_path = "C:\Intermodal\Case study\CP\EGS Contargo CTT\Intermodal_EGS_data_all" + note + ".xlsx"
                            elif parallel_number == 2:
                                data_path = "C:\Intermodal\Case study\CP\EGS Contargo CTT\Intermodal_Contargo_data_all" + note + ".xlsx"
                            else:
                                data_path = "C:\Intermodal\Case study\CP\EGS Contargo CTT\Intermodal_HSL_data_all" + note + ".xlsx"
                        else:
                            data_path = '/data/yimeng/Case study/CP/instances/three_eco_labels/Intermodal_EGS_data_carrier' + str(parallel_number) + note + '.xlsx'
                    else:
                        data_path = 'C:\Intermodal\Case study\CP\instances\Intermodal_EGS_data_carrier' + str(parallel_number) + note + '.xlsx'
                    # if parallel_number == 1:
                    #     data_path = 'C:\Intermodal\Case study\CP\Intermodal_EGS_data_barge.xlsx'
                    # elif parallel_number == 2:
                    #     data_path = 'C:\Intermodal\Case study\CP\Intermodal_EGS_data_train.xlsx'
                    # else:
                    #     data_path = 'C:\Intermodal\Case study\CP\Intermodal_EGS_data_truck.xlsx'
                else:
                    if Demir == 1:
                        data_path = 'C:\Intermodal\Case study\Demir\Intermodal_Demir_data - ' + str(request_number_in_R) + 'r.xlsx'
                    else:
                        if heterogeneous_preferences == 1:
                            # data_path = '/data/yimeng/Case study/Preferences/instances/Intermodal_EGS_data_all_' + note + '.xlsx'
                            if three_eco_labels == 1:
                                each_company = 0
                                if each_company == 1:
                                    if different_companies == 1:
                                        if parallel_number == 1:
                                            data_path = '/data/yimeng/Case study/CP/instances/noCP/three_eco_labels/Intermodal_EGS_data_all' + note + '.xlsx'
                                        elif parallel_number == 2:
                                            data_path = '/data/yimeng/Case study/CP/EGS Contargo CTT/Intermodal_Contargo_data_all' + note + '.xlsx'
                                        # data_path = "/data/yimeng/Case study/CP/instances/noCP/three_eco_labels/Intermodal_EGS_data_all_sustainable_2 - Copy.xlsx"
                                        else:
                                            data_path = '/data/yimeng/Case study/CP/EGS Contargo CTT/Intermodal_HSL_data_all' + note + '.xlsx'
                                    else:
                                        data_path = '/data/yimeng/Case study/CP/instances/noCP/three_eco_labels/Intermodal_EGS_data_all'  + note + '.xlsx'
                                else:
                                    if different_companies == 1:
                                        data_path = '/data/yimeng/Case study/CP/EGS Contargo CTT/Intermodal_Merge_data_all'  + note + '.xlsx'
                                    else:
                                        data_path = '/data/yimeng/Case study/CP/instances/noCP/three_eco_labels/Intermodal_EGS_data_all' + note + '.xlsx'
                            else:
                                data_path = '/data/yimeng/Case study/CP/instances/noCP/Intermodal_EGS_data_all' + note + '.xlsx'
                        else:
                            if three_eco_labels == 1:
                                if parallel_number == 1:
                                    data_path = '/data/yimeng/Case study/CP/instances/noCP/three_eco_labels/Intermodal_EGS_data_all' + note + '.xlsx'
                                elif parallel_number == 2:
                                    data_path = '/data/yimeng/Case study/CP/EGS Contargo CTT/Intermodal_Contargo_data_all' + note + '.xlsx'
                                # data_path = "/data/yimeng/Case study/CP/instances/noCP/three_eco_labels/Intermodal_EGS_data_all_sustainable_2 - Copy.xlsx"
                                else:
                                    data_path = '/data/yimeng/Case study/CP/EGS Contargo CTT/Intermodal_HSL_data_all' + note + '.xlsx'
                            else:
                                data_path = "/data/yimeng/Case study/vs. Wenjing/instances/Intermodal_EGS_data_all.xlsx"
                                # data_path = '/data/yimeng/Case study/CP/EGS Contargo CTT/Intermodal_Contargo_data_all' + note + '.xlsx'
                    # data_path = '/data/yimeng/Case study/Preferences/Intermodal_EGS_data_simple - ' + str(request_number_in_R) + 'r' + ' - test1r.xlsx'
                    # data_path = '/data/yimeng/Case study/Preferences/Intermodal_EGS_data_simple - ' + str(
                    #     request_number_in_R) + 'r.xlsx'
                    # data_path = '/data/yimeng/Case study/Small instance/Intermodal_EGS_data_simple -test.xlsx'
    Data = pd.ExcelFile(data_path)


    Demir_barge_free = 1
    w1,w2,w3=1,1,1
    r_number = 1

    Fixed = read_Fixed(request_number_in_R, percentage)
    exps_record_path = 'A:/MYpython/34959_RL/Uncertainties Dynamic planning under unexpected events/Figures/exps_record/exps_record_all_parallel' + 'exp' + str(
        exp_number) + 'parallel' + str(parallel_number) + '.xlsx'

    get_initial_bymyself = 1
    by_wenjing = 0
    step_by_step = 0
    # three situations for truck: 1. truck is free 2. truck time is free but route fixed, then set it as 1 3. truck is fixed, then set it as 0
    truck_time_free = 1
    b1, b2, b3, b4, b5, b6, b7, b8, b9, b10 = 0, 5, 7, 9, 13, 13, 17, 19, 21, 24
    alpha, belta = 2, 1.5
    if Demir == 0:
        forbid_T_trucks = 1
        forbid_much_delay = 0
        truck_fleet = 1
    else:
        forbid_T_trucks = 0
        forbid_much_delay = 0
        truck_fleet = 0
    insert_multiple_r = 0
    check_obj = 0
    start_from_best_at_begin_of_segement = 1
    # parallel inside ALNS
    parallel = 0
    parallel_thread = 0
    max_processors = 6
    #parallel between ALNS
    parallel_ALNS = 0

    #this is a mark for whether I use multi obj, only multi_obj is not enough because when regular = 1, the multi-obj is 0, which cause I have no regular non-dominated data when plotting, so add this real_
    real_multi_obj = 0
    bi_obj_cost_emission = 0
    bi_obj_cost_time = 1
    for j in [1600]:
        for iteration_number in [2]:
            for c in [0.99]:
                for repeat_number in [1]:

                    path = 'A:/MYpython/34959_RL/Uncertainties Dynamic planning under unexpected events/Figures/experiment' + str(exp_number) + '/'
                    exp_number = exp_number + 1

                    Path(path).mkdir(parents=True, exist_ok=True)
                    shutil.copy(data_path, path)
                    names = revert_names()
                    N = pd.read_excel(Data, 'N')
                    N['N'] = N['N'].map(names).fillna(N['N'])
                    N = N.values

                    o = pd.read_excel(Data, 'o')
                    o = o.set_index('K')
                    o['o'] = o['o'].map(names).fillna(o['o'])
                    o['o2'] = o['o2'].map(names).fillna(o['o2'])
                    o = o.values
                    if compare_with_gurobi == 1:
                        T = pd.read_excel(Data, 'T_5')
                    else:
                        T = pd.read_excel(Data, 'T')

                    T['T'] = T['T'].map(names).fillna(T['T'])
                    T = list(T['T'])
                    # N = N.set_index('N')
                    R, K, R_pool = read_R_K(request_number_in_R)
                    request_segment_in_dynamic = np.array(np.empty(shape=(0, len(R[0]) + 1)))
                    if dynamic == 1 and dynamic_t == 0:
                        delayed_time_table = pd.DataFrame(columns=range(len(K[:, 0])), index=range(len(N[:, 0])))
                        for k_ in delayed_time_table.columns:
                            if K[k_, 5] == 3:
                                for terminal_ in delayed_time_table.index:
                                    delayed_time_table[k_][terminal_] = {}
                            else:
                                for terminal_ in delayed_time_table.index:
                                    delayed_time_table[k_][terminal_] = 0
                        delayed_time_table_uncertainty_index = {}
                    comparison = pd.DataFrame(index=range((len(range(len(K))) - 3) * (len(R[:,7]) - 1)),
                                              columns=['Request number', 'Vehicle number', 'Repeat times of best obj',
                                                       'Cost of ALNS', 'Profit of ALNS', 'Served requests of ALNS',
                                                       'CPU time of ALNS', 'Running time of ALNS', 'Cost of Gurobi',
                                                       'Served requests of Gurobi', 'CPU time of Gurobi',
                                                       'Running time of Gurobi', 'Objective gap'])
                    # regular (no weight)
                    multi_obj = 0
                    regular = 1

                    devide_value = 5
                    if (add_RL == 1 and dynamic_RL34959.implement == 1) or (
                                            ALNS_greedy_under_unknown_duration_assume_duration == 3 and len(
                                        Dynamic_ALNS_RL34959.ALNS_reward_list_in_implementation) >= number_of_training):
                        current_save = 'percentage' + str(percentage) + 'parallel_number' + str(
                            parallel_number) + 'dynamic' + str(dynamic_t) + 'implement'
                    else:
                        current_save = 'percentage' + str(percentage) + 'parallel_number' + str(parallel_number) + 'dynamic' + str(dynamic_t)
                    # get_time_of_wenjing()
                    def second_main():

                        return main(R_pool, parallel_number, SA,combination, only_T2,has_end_depot2, T_or_not, path, N,
                                                                                      T, K, o, R, iteration_number,
                                                                                      current_save, len(K),
                                                                                      len(R[:,7]),
                                                                                      transshipment_time, service_time,
                                                                                      transshipment_cost_per,
                                                                                      fuel_cost, segment_number2, r2,
                                                                                      miu1_1, miu2_1, miu3_1, pro,
                                                                                      Fixed, percentage, k_random_or)

                    CP_try_r_of_other_carriers = 0
                    obj_record_best, CPU_Time, Running_Time, Best_Running_Time = second_main()
                    # with ProcessPoolExecutor() as executor:
                    #     executor.map(main, NUMBERS)

    if dynamic == 1:
        if VCP_coordination == 1:
            #save current_save to main folder with carrier_number
            # open text file
            text_file = open(path + "old_current_save" + str(parallel_number) + ".txt", "w")
            # write string to file
            text_file.write(current_save)
            # close file
            text_file.close()
        else:
            old_current_save = current_save
    if step_by_step == 1:
        get_initial_bymyself = 0
        old_current_save = current_save
        percentage = 0.3
        folder_name = 'compare' + str(r_number) + 'r_10000iteration_0620' + 'percentage' + str(percentage)
        current_save = folder_name + '_regular'
        obj_record_best, CPU_Time, Running_Time, Best_Running_Time = second_main()


        get_initial_bymyself = 0
        old_current_save = current_save
        percentage = 0.72
        folder_name = 'compare' + str(r_number) + 'r_10000iteration_0620' + 'percentage' + str(percentage)
        current_save = folder_name + '_regular'
        obj_record_best, CPU_Time, Running_Time, Best_Running_Time = second_main()

    if CP == 1:

        def change_r_index(R, R_pool, CP_R_pool):
            # change index of r from other carriers
            my_own_R = len(R)
            # create a map between changed index of r and unchanged r
            map_change_index = {}
            map_change_index_reverse = {}
            for i in range(len(R_pool)):
                changed_index = my_own_R + i
                # [original index, carrier]
                map_change_index[changed_index] = [R_pool[i, 7], CP_R_pool[i, -1]]
                map_change_index_reverse[tuple([R_pool[i, 7], CP_R_pool[i, -1]])] = changed_index
                R_pool[i, 7] = changed_index
            return map_change_index, R_pool, map_change_index_reverse

        def get_best_routes_pool():
            global CP_best_routes, routes
            routes = my_deepcopy(CP_best_routes)
            # get best R_pool

            R_pool = create_R_pool()

            return routes, R_pool
        def get_remove_list(remove_list,dict_segments_carriers,r):
            for segment, carriers in dict_segments_carriers.items():
                if r == segment and parallel_number not in carriers:
                    # remove this segment
                    remove_list.append(r)
            return remove_list
        #one round begins
        def CP_iteration(round):
            global request_flow_t, R_pool, routes,CP_try_r_of_other_carriers, R, not_initial_in_CP
            print('one round starts')

            not_initial_in_CP = 1
            # a = 5/0
            CP_try_r_of_other_carriers = 1
            CP_R_pool_path = path + 'CP_R_pool.xlsx'
            dict_segments_carriers_path = path + 'dict_segments_carriers.xlsx'
            CP_R_pool_submit_path = path + 'CP_R_pool_submit' + str(parallel_number) + '.xlsx'
            CP_R_pool_check_path = path + 'CP_R_pool_check'
            CP_bids_check_path = path + 'CP_bids_check'
            CP_success_r_submit_path = path + 'CP_success_r_submit' + str(parallel_number) + '.xlsx'
            if round == 0: #only use this at the first round, obtain from initial optimization
                routes, R_pool = get_best_routes_pool()
            initial_routes = my_deepcopy(routes)
            initial_R_pool = copy.copy(R_pool)
            initial_R = copy.copy(R)
            initial_request_flow_t = copy.copy(request_flow_t)
            if R_pool.size != 0:

                # #add carrier label
                # l = np.array(np.empty(shape=(len(R_pool),1)))
                # l[:] = parallel_number
                # R_pool = np.append(R_pool, l, axis=1)


                # if not os.path.isfile(CP_R_pool_path):
                    # Path(CP_R_pool_path).mkdir(parents=True, exist_ok=True)

                R_pool = pd.DataFrame(R_pool)
                parallel_save_excel(CP_R_pool_submit_path, R_pool, 'CP_R_pool', index=True)
                # with pd.ExcelWriter(CP_R_pool_submit_path) as writer:  # doctest: +SKIP
                #     R_pool.to_excel(writer, sheet_name='CP_R_pool')
                # else:
                #
                #     update_CP_R_pool(CP_R_pool_path, R_pool)
                print('add R_pool to CP')


            #mark that I have submitted unserved requests

            CP_R_pool_check = pd.DataFrame([1])


            while True:
                # print(9954)
                with pd.ExcelWriter(CP_R_pool_check_path + str(parallel_number) + '.xlsx') as writer:  # doctest: +SKIP
                    CP_R_pool_check.to_excel(writer, sheet_name='CP_R_pool_check', index=False)
                #very strange that the CP_R_pool_check can't be added to folder, and if check immediately, it exists
                # so need to sleep for a while, then check
                time.sleep(5)

                if os.path.exists(CP_R_pool_check_path +str(parallel_number) + '.xlsx'):
                    break
            print('told others that carrier' + str(parallel_number) + ' submitted R_pool')
            #check whether all carriers submitted their unserved requests
            while True:
                # print(9966)
                if os.path.exists(CP_R_pool_check_path + '0.xlsx'):
                    break
            # CP_R_pool_check_all = pd.read_excel(CP_R_pool_check_path + '0.xlsx', 'CP_R_pool_check')
            # while CP_R_pool_check_all[0][0] != carriers_number:
            #     time.sleep(1)
            #     CP_R_pool_check_all = pd.read_excel(CP_R_pool_check_path + '0.xlsx', 'CP_R_pool_check')
            print('checking bids starts')
            #read all r in CP_R_pool
            if os.path.isfile(CP_R_pool_path):
                CP_R_pool = parallel_read_excel(CP_R_pool_path, 'CP_R_pool', 0)
            else:
                # no bid, mark in the CP_bids_check that the bid is submitted
                CP_bids_check = pd.DataFrame([1])
                parallel_save_excel(CP_bids_check_path + str(parallel_number) + '.xlsx', CP_bids_check, 'CP_bids_check')
                return
            # CP_R_pool = pd.read_excel(, 'CP_R_pool', index_col=0)
            CP_R_pool = CP_R_pool.values


            # map_change_index, R_pool, map_change_index_reverse = change_r_index(R, R_pool, CP_R_pool)
            # get new R_pool from unserved requests of other carriers
            R_pool = CP_R_pool[~(CP_R_pool[:, -1] == parallel_number)]
            not_in_N = []

            for r_in_pool in range(len(R_pool)):

                if [R_pool[r_in_pool][0]] not in N or [R_pool[r_in_pool][1]] not in N:
                    not_in_N.append(r_in_pool)

            R_pool = np.delete(R_pool, not_in_N, axis=0)
            if request_segment == 1:
                #the carrier can also serve a segment of it's own request
                add_list = []
                for index in range(len(CP_R_pool)):
                    if CP_R_pool[index, -1] == parallel_number and int(str(CP_R_pool[index, 7])[2]) > 0:
                        add_list.append(index)
                R_pool = np.vstack([R_pool, CP_R_pool[add_list]])
                # read white list of carriers who can serve segments
                dict_segments_carriers_file = open(dict_segments_carriers_path, "r")
                dict_segments_carriers = eval(dict_segments_carriers_file.read())
                dict_segments_carriers_file.close()
                ## remove segments that this carrier can't serve or this carrier already own this segment in R_pool
                remove_list = []
                if R_pool.size > 17:
                    for index in range(len(R_pool)):
                        r = R_pool[index,7]
                        remove_list = get_remove_list(remove_list,dict_segments_carriers,r)
                else:
                    r = R_pool[7]
                    remove_list = get_remove_list(remove_list, dict_segments_carriers, r)
                #remove r
                R_pool  = np.delete(R_pool, remove_list, axis=0)

            R_pool_copy = copy.copy(R_pool)
            # R_pool = np.delete(R_pool, -1, axis=1)
            # R_pool_restore = copy.copy(R_pool)
            if len(R_pool) > 0:
                R_pool = np.unique(R_pool, axis=0)

            if R_pool.size == 0:
                # no bid, mark in the CP_bids_check that the bid is submitted
                CP_bids_check = pd.DataFrame([1])
                parallel_save_excel(CP_bids_check_path + str(parallel_number) + '.xlsx',CP_bids_check,'CP_bids_check')
                # success_r_submit = pd.DataFrame([1])
                # parallel_save_excel(CP_success_r_submit_path, success_r_submit, 'success_r')
                # this round of auction ends
                return
            else:
                # R also need to add r in R_pool, and after this def, R should be the original R
                # R_restore = copy.copy(R)
                #routes don't need to be restored because the last insertions of successful bids are based on the best routes in the try optimization step
                #my own r which hasn't be inserted should be deleted from R
                served_r = get_all_served_r()
                for r in R[:,7]:
                    if r not in served_r:
                        R = R[~(R[:, 7] == r)]

                #save this R as R_save, after this round, add success bids to R_save
                R_save = copy.copy(R)
                try:
                    R = np.vstack([R, R_pool])
                    T_k_record = np.insert(T_k_record, len(T_k_record), np.nan, axis=0)
                    request_flow_t = np.insert(request_flow_t, len(request_flow_t), np.nan, axis=0)
                except:
                    sys.exit(-1)

                #R_pool may repeat, causing duplication in R
                if R.size != 0:
                    R = np.unique(R, axis=0)
                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                # requests used for trying
                requests_try = R_pool[:,7]
                before_try_requests_number = len(R_pool)
                before_try_routes = my_deepcopy(routes)
                before_try_R_pool = copy.copy(R_pool)
                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                second_main()

                print('Tried')
                routes, R_pool = get_best_routes_pool()
                served_r = get_all_served_r()

                # if len(R_pool) < before_try_requests_number:
                #check_repeat_r_in_R_pool(), check_T_k_record_and_R()
                want_serve_r = []
                for r in requests_try:
                    if r not in R_pool[:,7]:
                        want_serve_r.append(r)
                if not want_serve_r:
                    #don't want to bid, then restore R_pool and routes
                    routes = my_deepcopy(initial_routes)
                    R_pool = copy.copy(initial_R_pool)
                    R = copy.copy(R_save)
                    request_flow_t = copy.copy(initial_request_flow_t)
                    # mark in the CP_bids_check that the bid is submitted

                    CP_bids_check = pd.DataFrame([1])
                    parallel_save_excel(CP_bids_check_path + str(parallel_number) + '.xlsx',CP_bids_check,'CP_bids_check')
                    #this round of auction ends
                    return

                # bids = np.array(np.empty(shape=(0,3)))
                # bids[:]=np.NaN
                bids = pd.DataFrame(columns = ['r', 'cost', 'carrier', 'original_carrier'])
                for r in want_serve_r:
                    # original_r, original_carrier = map_change_index[r]
                    #danger here restrict the number of carriers is less than 10
                    original_carrier = int(r/big_r)
                    bid_cost_r = get_r_cost_in_all_routes(r)[0]
                    bids.loc[len(bids)] = [r, bid_cost_r, parallel_number, original_carrier]
                    # bids = np.vstack([bids,[r, bid_cost_r, parallel_number]])
                #submit the bids

                CP_bids_path = path + 'CP_bids.xlsx'
                CP_bids_submit_path = path + 'CP_bids' + 'submit' + str(parallel_number) + '.xlsx'
                parallel_save_excel(CP_bids_submit_path, bids, 'CP_bids')


                #mark in the CP_bids_check that the bid is submitted

                CP_bids_check = pd.DataFrame([1])
                parallel_save_excel(CP_bids_check_path + str(parallel_number) + '.xlsx',CP_bids_check,'CP_bids_check')
                print('Submitted bids')
                #wait for the ranking by shipper/coordinator
                while True:
                    # print(10101)
                    if os.path.exists(path + 'success_bids.xlsx'):
                        break
                print(parallel_number)

                success_bids = parallel_read_excel(path + 'success_bids.xlsx', 'success_bids')
                # while True:
                #     try:
                #         success_bids = pd.read_excel(path + 'success_bids.xlsx', sheet_name='success_bids')
                #         break
                #     except:
                #         continue
                #check whether the bid is accepted
                success_r = []
                for index in success_bids.index:
                    if parallel_number == success_bids['carrier'][index]:
                        success_r.append([success_bids['r'][index],success_bids['original_carrier'][index]])
                #remove failed bids
                re_optimization = 0
                index_for_delete = []
                for r in want_serve_r:
                    # danger here restrict the number of carriers is less than 10
                    original_carrier = int(r / big_r)
                    if [r,original_carrier] not in success_r:
                        before_remove_length = len(R_pool)
                        routes, R_pool = remove_a_request(r, routes, R_pool)[0:2]
                        if len(R_pool) == before_remove_length:
                            #it didn't be removed because the constraints are violated, then do the optimization based on initial-routes, and the R_pool are the successful bids
                            re_optimization = 1
                            break
                        #here the failed bid should be deleted in R_pool and other carriers will serve it, otherwise it will be submitted to CP_R_pool in next round
                        R_pool = R_pool[~(R_pool[:,7] == r)]
                        # record the indexes of failed bids in R, then delete rows in request_flow_t accroding to this record
                        index_for_delete.append(list(R[:, 7]).index(r))


                if re_optimization == 1:
                    routes = my_deepcopy(initial_routes)
                    R_pool = create_R_pool()

                    R_pool[:] = np.NaN
                    #add successful bids to R_pool
                    for original_r, original_carrier in success_r:
                        # r = map_change_index_reverse[tuple([original_r, original_carrier])]
                        R_pool = np.vstack([R_pool, list(R[:, 7]).index(original_r)])
                    second_main()
                    routes, R_pool = get_best_routes_pool()
                    #here I don't need to delete uninseted r in R_pool, because it's not failed bid, no carrier serves it in this round, it will not be deleted in CP_R_pool
                    served_r = get_all_served_r()

                else:
                    # update request_flow_t
                    request_flow_t = np.delete(request_flow_t, index_for_delete, axis=0)
                    #here maybe not all r in success_r are inserted, therefore will add failed r in success_r in next round
                # success_index = []
                # for r, original_carrier in success_r:
                #     for index in range(len(R_pool_copy)):
                #         if R_pool_copy[index,7] != r or R_pool_copy[index,-1] != original_carrier:
                #             success_index.append(index)
                # #delete unsuccessful bids
                # R_pool_copy = np.delete(R_pool_copy, success_index, axis=0)
                #
                # R_pool = copy.copy(R_pool_copy)
                # #change index to avoid same index of r
                # map_change_index, R_pool, map_change_index_reverse = change_r_index(R, R_pool, CP_R_pool)
                # R_pool = np.delete(R_pool, -1, axis=1)
                # R_pool_copy2 = copy.copy(R_pool)
                # R = np.vstack([R, R_pool])


                #remove the unseccessful bids from routes, R, request_flow_t, then do the optimization

                index_for_delete = []
                #delete r in CP_R_pool if r is accepted
                for original_r, original_carrier in success_r:
                    # for key, value in map_change_index:
                    #     if value == [original_r, original_carrier]:
                    #         r = key
                    #         break
                    # r = map_change_index_reverse[tuple([original_r, original_carrier])]
                    #in parallel debugging, always check which carrier it is
                    if original_r not in R_pool[:,7]:
                        # for index in range(len(CP_R_pool)):
                        #     if CP_R_pool[index,7] == original_r and CP_R_pool[index,-1] == original_carrier:
                        #         CP_R_pool = np.delete(CP_R_pool, index, axis=0)
                        # CP_R_pool = np.delete(CP_R_pool, list(CP_R_pool[:,7]).index(r), axis=0)

                        #add this r to R
                        R_save = np.vstack([R_save, R[list(R[:,7]).index(original_r)]])
                    else:
                        # record the indexes of failed bids in R, then delete rows in request_flow_t accroding to this record
                        index_for_delete.append(list(R[:, 7]).index(original_r))


                R = copy.copy(R_save)
                # update request_flow_t
                request_flow_t = np.delete(request_flow_t, index_for_delete, axis=0)

                # # add carrier label
                # l = np.array(np.empty(shape=(len(R_pool), 1)))
                # l[:] = parallel_number
                # R_pool = np.append(R_pool, l, axis=1)
                #delete success r in CP_R_pool
                R_pool_submit = pd.DataFrame(R_pool)

                parallel_save_excel(CP_R_pool_submit_path, R_pool_submit, 'CP_R_pool', index=True)
                # with pd.ExcelWriter(CP_R_pool_submit_path) as writer:  # doctest: +SKIP
                #     R_pool_submit.to_excel(writer, sheet_name='CP_R_pool')
                success_r_submit = pd.DataFrame(success_r)
                parallel_save_excel(CP_success_r_submit_path, success_r_submit, 'success_r')
                # with pd.ExcelWriter(CP_success_r_submit_path) as writer:  # doctest: +SKIP
                #     success_r_submit.to_excel(writer, sheet_name='success_r')
                # update_CP_R_pool(CP_R_pool_path, R_pool, -1, success_r)
                print('delete served r')


        CP_round_end_check_path = path + 'CP_round_end_check'
        CP_round_end_check_all_path = path + 'CP_round_end_check_all' + '0.xlsx'

        save_results(-1, routes)

        begin_round_time = timeit.default_timer()
        for round in range(auction_round_number):
            
            # CP_try_r_of_other_carriers = 1
            # second_main()
            CP_iteration(round)

            #current is the best routes and R_pool after checking
            #add sheets because routes may changed
            #add routes, current obj_record, and update routes match
            best_routes_path = path + current_save + '/best_routes' + current_save + '_' + str(
                exp_number - 1) + '.xlsx'
            # write the current route to file
            routes_save = my_deepcopy(routes)
            with pd.ExcelWriter(best_routes_path) as writer:  # doctest: +SKIP
                for key, value in routes_save.items():
                    route_df = pd.DataFrame(value[0:4, :], columns=value[4])
                    route_df.to_excel(writer, convert(key))
            Graph(routes_save, 0)

            save_results(round, routes_save)
            #save the results after CP at the last round
            if round == auction_round_number - 1:
                # calculate total running time for this carrier
                end_round_time = timeit.default_timer()
                total_round_time = end_round_time - begin_round_time
                overall_distance, overall_cost, overall_time, overall_profit, overall_emission, served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost, overall_transshipment_cost, overall_un_load_cost, overall_emission_cost, overall_storage_cost, overall_delay_penalty, overall_number_transshipment, overall_average_speed, overall_average_time_ratio, overall_emission_transshipment = overall_obj(
                    routes)
                barge_served_requests, train_served_requests, truck_served_requests = CP_served_requests_mode()
                if heterogeneous_preferences == 1:

                    satisfactory_value, fuzzy_satisfy_or_not, hard_satisfy_or_not = overall_satisfactory_values(routes)

                    cost_per_container_per_km, time_ratio, emissions_per_container_per_km, delay_time_ratio, transshipment_times = overall_satisfactory_values(
                        routes, 1)
                    final_obj = pd.DataFrame(index=range(0, 1),
                                          columns=['overall_distance', 'overall_cost', 'overall_time', 'overall_profit',
                                                   'overall_emission', 'served_requests', 'overall_request_cost',
                                                   'overall_vehicle_cost', 'overall_wait_cost',
                                                   'overall_transshipment_cost',
                                                   'overall_un_load_cost', 'overall_emission_cost', 'overall_storage_cost',
                                                   'overall_delay_penalty', 'iteration_time', 'barge_served_requests', 'train_served_requests', 'truck_served_requests','satisfactory_value', 'fuzzy_satisfy_or_not', 'hard_satisfy_or_not',
                    'overall_number_transshipment', 'overall_average_time_ratio', 'cost_per_container_per_km', 'time_ratio', 'emissions_per_container_per_km', 'delay_time_ratio', 'transshipment_times', 'heterogeneous_preferences_no_constraints', 'heterogeneous_preferences', 'fuzzy_constraints', 'overall_emission_transshipment'])
                    final_obj.iloc[0] = [overall_distance, overall_cost, overall_time, overall_profit, overall_emission,
                                         served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost,
                                         overall_transshipment_cost, overall_un_load_cost, overall_emission_cost,
                                         overall_storage_cost, overall_delay_penalty, total_round_time, barge_served_requests,
                                         train_served_requests, truck_served_requests, satisfactory_value, fuzzy_satisfy_or_not, hard_satisfy_or_not,
                           -1, -1, cost_per_container_per_km, time_ratio,
                           emissions_per_container_per_km, delay_time_ratio, transshipment_times,
                           heterogeneous_preferences_no_constraints, heterogeneous_preferences, fuzzy_constraints, overall_emission_transshipment]
                else:
                    final_obj = pd.DataFrame(index=range(0, 1),
                                          columns=['overall_distance', 'overall_cost', 'overall_time', 'overall_profit',
                                                   'overall_emission', 'served_requests', 'overall_request_cost',
                                                   'overall_vehicle_cost', 'overall_wait_cost',
                                                   'overall_transshipment_cost',
                                                   'overall_un_load_cost', 'overall_emission_cost', 'overall_storage_cost',
                                                   'overall_delay_penalty', 'iteration_time', 'barge_served_requests', 'train_served_requests', 'truck_served_requests'])

                    final_obj.iloc[0] = [overall_distance, overall_cost, overall_time, overall_profit, overall_emission,
                                               served_requests, overall_request_cost, overall_vehicle_cost, overall_wait_cost,
                                               overall_transshipment_cost, overall_un_load_cost, overall_emission_cost,
                                               overall_storage_cost, overall_delay_penalty, 0, barge_served_requests, train_served_requests, truck_served_requests]



                # Defining the path which excel needs to be created
                # There must be a pre-existing excel sheet which can be updated
                FilePath = path + current_save + '/obj_record' + current_save + str(
                exp_number - 1) + '.xlsx'
                # Generating workbook
                ExcelWorkbook = load_workbook(FilePath)

                # Generating the writer engine
                writer = pd.ExcelWriter(FilePath, engine='openpyxl')

                # Assigning the workbook to the writer engine
                writer.book = ExcelWorkbook

                # Creating first dataframe

                # Adding the DataFrames to the excel as a new sheet
                final_obj.to_excel(writer, sheet_name='final_obj')

                writer.save()
                writer.close()
                # parallel_save_excel(path + current_save + '/obj_record' + current_save + str(
                # exp_number - 1) + '.xlsx', final_obj, 'final_obj')



            # for this carrier, one round end, mark it
            CP_round_end_check = pd.DataFrame([1])
            parallel_save_excel(CP_round_end_check_path + str(parallel_number) + '.xlsx', CP_round_end_check, 'CP_round_end_check')
            # with pd.ExcelWriter(CP_round_end_check_path + str(parallel_number) + '.xlsx') as writer:  # doctest: +SKIP
            #     CP_round_end_check.to_excel(writer, sheet_name='CP_round_end_check', index=False)
            # this round of auction ends when all carriers end
            while True:
                # print('CP_round_end_check_all')
                if os.path.exists(CP_round_end_check_all_path):
                    break
            if parallel_number == 1:
                print('carrier' + str(parallel_number) + ' one round finishes')
                print(get_all_served_r())
            if parallel_number == 2:
                print('carrier' + str(parallel_number) + ' one round finishes')
                print(get_all_served_r())
            if parallel_number == 3:
                print('carrier' + str(parallel_number) + ' one round finishes')
                print(get_all_served_r())
        return 'CP success!'
    folder_name = 'compare' + 'request_number' + str(request_number_in_R) + 'percentage' + str(percentage)
    # if not multi_obj, comment all follows
    multi_obj = 0
    regular = 0

    same_parameters = 0
    if multi_obj == 1:
        global weight_min_cost, weight_max_cost, weight_min_emission, weight_max_emission, weight_min_time, weight_max_time
        if bi_obj_cost_emission == 1 or bi_obj_cost_time == 1:
            if bi_obj_cost_emission == 1:
                if weight_interval == 1:
                    # global weight_min_cost,weight_max_cost,weight_min_emission,weight_max_emission,weight_min_time,weight_max_time
                    # s1c1
                    current_save = folder_name + '_s1c1'

                    weight_min_cost = 0.1
                    weight_max_cost = 0.9
                    weight_min_emission = 0.1
                    weight_max_emission = 0.9
                    weight_min_time = 0.1
                    weight_max_time = 0.9


                    second_main()

                    # s1c2
                    current_save = folder_name + '_s1c2'

                    weight_min_cost = 0.25
                    weight_max_cost = 0.75
                    weight_min_emission = 0.25
                    weight_max_emission = 0.75
                    weight_min_time = 0.25
                    weight_max_time = 0.75


                    second_main()

                    # s1c3
                    current_save = folder_name + '_s1c3'

                    weight_min_cost = 0.33
                    weight_max_cost = 0.66
                    weight_min_emission = 0.33
                    weight_max_emission = 0.66
                    weight_min_time = 0.33
                    weight_max_time = 0.66

                    second_main()

                    # s2c1
                    current_save = folder_name + '_s2c1'

                    weight_min_cost = 0.5
                    weight_max_cost = 1.0
                    # weight_min_time = 0.5
                    # weight_max_time = 1.0
                    weight_min_emission = 0.1
                    weight_max_emission = 0.5

                    second_main()

                    # s2c2
                    current_save = folder_name + '_s2c2'

                    weight_min_cost = 0.1
                    weight_max_cost = 0.5
                    # weight_min_time = 0.5
                    # weight_max_time = 1.0
                    weight_min_emission = 0.5
                    weight_max_emission = 1.0

                    second_main()
                else:
                    # s2c1
                    current_save = folder_name + '_s2c1'

                    weight_cost = 0.9

                    weight_emission = 0.1

                    second_main()

                    # s2c2
                    current_save = folder_name + '_s2c2'

                    weight_cost = 0.1

                    weight_emission = 0.9

                    second_main()

                    # s1c1
                    current_save = folder_name + '_s1c1'

                    weight_cost = 0.9
                    weight_emission = 0.9
                    weight_time = 0.9


                    second_main()

                    # s1c2
                    current_save = folder_name + '_s1c2'

                    weight_cost = 0.5
                    weight_emission = 0.5
                    weight_time = 0.5


                    second_main()

                    # s1c3
                    current_save = folder_name + '_s1c3'

                    weight_cost = 0.1
                    weight_emission = 0.1
                    weight_time = 0.1


                    second_main()
            else:
                if weight_interval == 1:

                    # s1c1
                    current_save = folder_name + '_s1c1'

                    weight_min_cost = 0.1
                    weight_max_cost = 0.9
                    weight_min_emission = 0.1
                    weight_max_emission = 0.9
                    weight_min_time = 0.1
                    weight_max_time = 0.9

                    second_main()

                    # s1c2
                    current_save = folder_name + '_s1c2'

                    weight_min_cost = 0.25
                    weight_max_cost = 0.75
                    weight_min_emission = 0.25
                    weight_max_emission = 0.75
                    weight_min_time = 0.25
                    weight_max_time = 0.75

                    second_main()

                    # s1c3
                    current_save = folder_name + '_s1c3'

                    weight_min_cost = 0.33
                    weight_max_cost = 0.66
                    weight_min_emission = 0.33
                    weight_max_emission = 0.66
                    weight_min_time = 0.33
                    weight_max_time = 0.66

                    second_main()

                    # s2c1
                    current_save = folder_name + '_s2c1'

                    weight_min_cost = 0.5
                    weight_max_cost = 1.0
                    weight_min_time = 0.1
                    weight_max_time = 0.5
                    # weight_min_emission = 0.1
                    # weight_max_emission = 0.5

                    second_main()

                    # s2c2
                    current_save = folder_name + '_s2c2'

                    weight_min_cost = 0.1
                    weight_max_cost = 0.5
                    weight_min_time = 0.5
                    weight_max_time = 1.0
                    # weight_min_emission = 0.5
                    # weight_max_emission = 1.0

                    second_main()
                else:
                    # s2c1
                    current_save = folder_name + '_s2c1'

                    weight_cost = 0.9

                    weight_time = 0.1

                    second_main()

                    # s2c2
                    current_save = folder_name + '_s2c2'

                    weight_cost = 0.1

                    weight_time = 0.9

                    second_main()

                    # s1c1
                    current_save = folder_name + '_s1c1'

                    weight_cost = 0.9
                    weight_emission = 0.9
                    weight_time = 0.9

                    second_main()

                    # s1c2
                    current_save = folder_name + '_s1c2'

                    weight_cost = 0.5
                    weight_emission = 0.5
                    weight_time = 0.5

                    second_main()

                    # s1c3
                    current_save = folder_name + '_s1c3'

                    weight_cost = 0.1
                    weight_emission = 0.1
                    weight_time = 0.1

                    second_main()
        else:
            if weight_interval == 1:
                #s1c1
                current_save = folder_name + '_s1c1'

                weight_min_cost = 0.1
                weight_max_cost = 0.9
                weight_min_emission = 0.1
                weight_max_emission = 0.9
                weight_min_time = 0.1
                weight_max_time = 0.9


                second_main()

                #s1c2
                current_save = folder_name + '_s1c2'

                weight_min_cost = 0.25
                weight_max_cost = 0.75
                weight_min_emission = 0.25
                weight_max_emission = 0.75
                weight_min_time = 0.25
                weight_max_time = 0.75


                second_main()

                #s1c3
                current_save = folder_name + '_s1c3'

                weight_min_cost = 0.33
                weight_max_cost = 0.66
                weight_min_emission = 0.33
                weight_max_emission = 0.66
                weight_min_time = 0.33
                weight_max_time = 0.66


                second_main()

                #s2c1
                current_save = folder_name + '_s2c1'

                weight_min_cost = 0.5
                weight_max_cost = 1.0
                weight_min_emission = 0.1
                weight_max_emission = 0.5
                weight_min_time = 0.1
                weight_max_time = 0.5


                second_main()


                #s2c2
                current_save = folder_name + '_s2c2'

                weight_min_cost = 0.1
                weight_max_cost = 0.5
                weight_min_emission = 0.5
                weight_max_emission = 1.0
                weight_min_time = 0.1
                weight_max_time = 0.5


                second_main()


                #s2c3
                current_save = folder_name + '_s2c3'

                weight_min_cost = 0.1
                weight_max_cost = 0.5
                weight_min_emission = 0.1
                weight_max_emission = 0.5
                weight_min_time = 0.5
                weight_max_time = 1.0


                second_main()
            else:
                # s2c1
                current_save = folder_name + '_s2c1'

                weight_cost = 0.8
                weight_emission = 0.1
                weight_time = 0.1


                second_main()

                # s2c2
                current_save = folder_name + '_s2c2'

                weight_cost = 0.1
                weight_emission = 0.8
                weight_time = 0.1


                second_main()

                # s2c3
                current_save = folder_name + '_s2c3'

                weight_cost = 0.1
                weight_emission = 0.1
                weight_time = 0.8

                second_main()

                # s1c1
                current_save = folder_name + '_s1c1'

                weight_cost = 0.9
                weight_time = 0.9
                weight_emission = 0.9

                second_main()

                # s1c2
                current_save = folder_name + '_s1c2'


                weight_cost = 0.5
                weight_time = 0.5
                weight_emission = 0.5

                second_main()

                # s1c3
                current_save = folder_name + '_s1c3'

                weight_cost = 0.1
                weight_time = 0.1
                weight_emission = 0.1

                second_main()
        r_number = r_number + 1

        save_file = "/data/yimeng/Figures" + '/500-547_c/'
        Path(save_file).mkdir(parents=True, exist_ok=True)
        if bi_obj_cost_emission == 1 or bi_obj_cost_time == 1:
            if bi_obj_cost_emission == 1:
                all_regular_Pareto_frontiers = pd.DataFrame(columns=['overall_cost', 'overall_emission'])
            else:
                all_regular_Pareto_frontiers = pd.DataFrame(columns=['overall_cost', 'overall_time'])
        else:
           all_regular_Pareto_frontiers = pd.DataFrame(columns=['overall_cost', 'overall_time', 'overall_emission'])

        if same_parameters == 1:
            same_parameters_Pareto_frontiers = {}
            for which in ['regular', 's1c1','s1c2','s1c3','s2c1','s2c3']:
                for iteration_number in [50, 100, 500, 1000]:
                   for c in [0.99, 0.999, 0.9999, 0.99999]:

                       if which == 'regular':
                           aa = "regular"
                       else:
                           aa = "preference"
            #                all_regular_Pareto_frontiers = {}
                       for exp_number in range(500,548):
                           path='/data/yimeng/Figures/experiment' + str(exp_number-1) + '_chu5_' + str(r_number) + 'r_2obj_interval_' + str(iteration_number) + 'iteration_number' + str(c) + 'cooling_rate/'
                           if os.path.exists(path):
                               non_dominated_solutions = pd.read_excel(path + 'compare' + str(r_number) + 'r_10000iteration_0620_'+ which +'/' + aa + 'compare' + str(r_number) + 'r_10000iteration_0620_' + which + '.xlsx',aa + '_non_dominated')
               #                    all_regular_Pareto_frontiers[which+str(exp_number-1)] = non_dominated_solutions
                               if which  == 'regular':
                                   all_regular_Pareto_frontiers = all_regular_Pareto_frontiers.append(non_dominated_solutions)
                                   if 'chu5_' + str(r_number) + 'r_2obj_interval_' + str(iteration_number) + 'iteration_number' + str(c) + 'cooling_rate' in same_parameters_Pareto_frontiers.keys():
                                       same_parameters_Pareto_frontiers['chu5_' + str(r_number) + 'r_2obj_interval_' + str(iteration_number) + 'iteration_number' + str(c) + 'cooling_rate'] = same_parameters_Pareto_frontiers['chu5_'+ str(r_number) + 'r_2obj_interval_' + str(iteration_number) + 'iteration_number' + str(c) + 'cooling_rate'].append(non_dominated_solutions)
                                   else:
                                       if bi_obj_cost_emission == 1 or bi_obj_cost_time == 1:
                                           if bi_obj_cost_emission == 1:
                                                same_parameters_Pareto_frontiers['chu5_' + str(r_number) + 'r_2obj_interval_' + str(iteration_number) + 'iteration_number' + str(c) + 'cooling_rate'] = pd.DataFrame(columns=['overall_cost', 'overall_emission'])
                                           else:
                                               same_parameters_Pareto_frontiers[
                                                   'chu5_' + str(r_number) + 'r_2obj_interval_' + str(
                                                       iteration_number) + 'iteration_number' + str(
                                                       c) + 'cooling_rate'] = pd.DataFrame(
                                                   columns=['overall_cost', 'overall_time'])
                                       else:
                                           same_parameters_Pareto_frontiers['chu5_' + str(r_number) + 'r_2obj_interval_' + str(iteration_number) + 'iteration_number' + str(c) + 'cooling_rate'] = pd.DataFrame(columns=['overall_cost', 'overall_time', 'overall_emission'])

                               label_str = 'c=' + str(c) +' ' + str(iteration_number) + 'iterations'
                               plt.scatter(non_dominated_solutions['overall_cost'], non_dominated_solutions['overall_emission'], marker='o', label= label_str)

                       # with pd.ExcelWriter(save_file + which + 'Compare on Pareto frontiers' + '.xlsx') as writer:  # doctest: +SKIP
                       #     all_regular_Pareto_frontiers.to_excel(writer, sheet_name='all_regular_Pareto_frontiers')

                   plt.xlabel('Cost (euro)')
                   plt.ylabel('Emissions (kg)')
                   plt.title('Compare on Pareto frontiers (' + which + ')')
                   plt.legend(loc='center left', bbox_to_anchor=(1, 0.5))
                   plt.ticklabel_format(useOffset=False, style='plain')
                   plt.savefig(
                       save_file + 'Compare on Pareto frontiers (' + which + ')' + '.pdf',
                       format='pdf', bbox_inches='tight')
               #plt.show()
                   plt.close()

    #            for frontier in same_parameters_Pareto_frontiers.keys():
    #
    #                for non_d_solution in same_parameters_Pareto_frontiers[keys].rows():
    #


    if multi_obj == 1 and same_parameters == 1:
        #for
        #all_regular_Pareto_frontiers.append(same_parameters_Pareto_frontiers[keys])

        #obj_record_save = copy.copy(obj_record)
        obj_record = copy.copy(all_regular_Pareto_frontiers)

        same_parameters_Pareto_frontiers_norm = my_deepcopy(same_parameters_Pareto_frontiers)

        for frontier_all_repeat in same_parameters_Pareto_frontiers_norm.keys():
           same_parameters_Pareto_frontiers_norm[frontier_all_repeat] = same_parameters_Pareto_frontiers_norm[frontier_all_repeat].reset_index(drop=True)
           for row_index in range(0, len(same_parameters_Pareto_frontiers_norm[frontier_all_repeat])):
               for column_name in same_parameters_Pareto_frontiers_norm[frontier_all_repeat].columns:
                   same_parameters_Pareto_frontiers_norm[frontier_all_repeat][column_name][row_index] = normalization(same_parameters_Pareto_frontiers_norm[frontier_all_repeat][column_name][row_index], column_name)

        compare_frontiers = pd.DataFrame(index = [0], columns = same_parameters_Pareto_frontiers_norm.keys())
        for frontier_all_repeat in same_parameters_Pareto_frontiers_norm.keys():
           sum_obj_all_repeat = same_parameters_Pareto_frontiers_norm[frontier_all_repeat].values.sum()
           average_obj = sum_obj_all_repeat / len(same_parameters_Pareto_frontiers_norm[frontier_all_repeat])
           compare_frontiers[frontier_all_repeat][0] = average_obj

        with pd.ExcelWriter(save_file + 'compare_frontiers.xlsx') as writer:  # doctest: +SKIP
           compare_frontiers.to_excel(writer, sheet_name='compare_frontiers')

        #min_frontier = compare_frontiers.min(axis=1)
        pd.options.display.max_colwidth = 100
        min_frontier_para = compare_frontiers.astype('float64').idxmin(axis=1)
        with open(save_file + "best_parameters.txt", "w") as text_file:
           print(min_frontier_para, file=text_file)
        print(min_frontier_para)


cprofile = 0
if cprofile == 1:
    import cProfile
    import pstats
    import io

    pr = cProfile.Profile()
    pr.enable()

if __name__ == '__main__':
    # #########trace########
    # import sys
    # import trace
    #
    # # create a Trace object, telling it what to ignore, and whether to
    # # do tracing or line-counting or both.
    # tracer = trace.Trace(
    #     ignoredirs=[sys.prefix, sys.exec_prefix],
    #     trace=1,
    #     count=1)
    #
    # # run the new command using the given tracer
    # tracer.run('real_main(0)')
    #
    # # make a report, placing output in the current directory
    # r = tracer.results()
    # r.write_results(show_missing=True, coverdir=".")
    # #########trace########
    # os.environ['PYTHONHASHSEED'] = '0'
    real_main(9)

if cprofile == 1:
    pr.disable()
    s = io.StringIO()
    ps = pstats.Stats(pr, stream=s).sort_stats('cumtime')
    ps.print_stats()

    with open('time/requestflow_array30r0.72_cumtime1205reduce.txt', 'w+', encoding="utf-8") as f:
        f.write(s.getvalue())
# real_main(0)
